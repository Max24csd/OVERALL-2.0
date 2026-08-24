from __future__ import annotations

from io import BytesIO
from pathlib import Path

from django.conf import settings
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage


MASTER_DIR = Path(settings.BASE_DIR) / "filtros_mm" / "masters"
SHEET_NAME = "Hoja1"

FA0201_MASTER = "20260208-0420 FLA0002_CARRILERIA_V1.xlsx"

MEASUREMENT_COLUMNS = {
    "espesor_a": "AE",
    "espesor_b": "AG",
    "espesor_c": "AI",
    "espesor_d": "AK",
    "ancho_izquierdo": "AQ",
    "ancho_derecho": "AU",
}

WHEEL_ROWS = {
    1: 492,
    2: 500,
    3: 508,
}

WHEEL_COLUMNS = {
    "izquierdo": {
        "a1": "AA",
        "a2": "AC",
        "a3": "AE",
        "b1": "AG",
        "b2": "AI",
        "b3": "AK",
    },
    "derecho": {
        "a1": "AN",
        "a2": "AP",
        "a3": "AR",
        "b1": "AT",
        "b2": "AV",
        "b3": "AX",
    },
}

UT_CELLS = {
    "ut_lh_marca": "AE75",
    "ut_lh_modelo": "AS75",
    "ut_lh_tipo_haz": "AE76",
    "ut_lh_ganancia": "AS76",
    "ut_lh_frecuencia": "AE77",
    "ut_lh_velocidad": "AS77",
    "ut_lh_ancho_banda": "AE78",
    "ut_lh_retardo": "AS78",
    "ut_lh_amortiguamiento": "AE79",
    "ut_lh_diametro": "AS79",
    "ut_rh_marca": "AE81",
    "ut_rh_modelo": "AS81",
    "ut_rh_tipo_haz": "AE82",
    "ut_rh_ganancia": "AS82",
    "ut_rh_frecuencia": "AE83",
    "ut_rh_velocidad": "AS83",
    "ut_rh_ancho_banda": "AE84",
    "ut_rh_retardo": "AS84",
    "ut_rh_amortiguamiento": "AE85",
    "ut_rh_diametro": "AS85",
}

PHOTO_SLOTS = {
    ("zonas_soldadura", 1): {
        "anchor": "C248",
        "comment": "C269",
        "max_width": 520,
        "max_height": 235,
        "input": "foto_zonas_1",
    },
    ("zonas_soldadura", 2): {
        "anchor": "AB248",
        "comment": "AB269",
        "max_width": 520,
        "max_height": 235,
        "input": "foto_zonas_2",
    },
    ("zonas_soldadura", 3): {
        "anchor": "C319",
        "comment": "C340",
        "max_width": 520,
        "max_height": 235,
        "input": "foto_zonas_3",
    },
    ("zonas_soldadura", 4): {
        "anchor": "AB319",
        "comment": "AB340",
        "max_width": 520,
        "max_height": 235,
        "input": "foto_zonas_4",
    },
    ("zonas_soldadura", 5): {
        "anchor": "C342",
        "comment": "C363",
        "max_width": 520,
        "max_height": 235,
        "input": "foto_zonas_5",
    },
    ("zonas_soldadura", 6): {
        "anchor": "AB342",
        "comment": "AB363",
        "max_width": 520,
        "max_height": 235,
        "input": "foto_zonas_6",
    },
    ("zonas_soldadura", 7): {
        "anchor": "C366",
        "comment": "C387",
        "max_width": 520,
        "max_height": 235,
        "input": "foto_zonas_7",
    },
    ("zonas_soldadura", 8): {
        "anchor": "AB366",
        "comment": "AB387",
        "max_width": 520,
        "max_height": 235,
        "input": "foto_zonas_8",
    },
    ("zonas_soldadura", 9): {
        "anchor": "C389",
        "comment": "C410",
        "max_width": 520,
        "max_height": 235,
        "input": "foto_zonas_9",
    },
    ("zonas_soldadura", 10): {
        "anchor": "AB389",
        "comment": "AB410",
        "max_width": 520,
        "max_height": 235,
        "input": "foto_zonas_10",
    },
    ("zonas_soldadura", 11): {
        "anchor": "C412",
        "comment": "C433",
        "max_width": 520,
        "max_height": 235,
        "input": "foto_zonas_11",
    },
    ("zonas_soldadura", 12): {
        "anchor": "AB412",
        "comment": "AB433",
        "max_width": 520,
        "max_height": 235,
        "input": "foto_zonas_12",
    },
    ("ruedas", 1): {
        "anchor": "C518",
        "comment": "C537",
        "max_width": 520,
        "max_height": 210,
        "input": "foto_ruedas_1",
    },
    ("ruedas", 2): {
        "anchor": "AB518",
        "comment": "AB537",
        "max_width": 520,
        "max_height": 210,
        "input": "foto_ruedas_2",
    },
    ("ruedas", 3): {
        "anchor": "C539",
        "comment": "C558",
        "max_width": 520,
        "max_height": 210,
        "input": "foto_ruedas_3",
    },
    ("ruedas", 4): {
        "anchor": "AB539",
        "comment": "AB558",
        "max_width": 520,
        "max_height": 210,
        "input": "foto_ruedas_4",
    },
    ("ruedas", 5): {
        "anchor": "C561",
        "comment": "C580",
        "max_width": 520,
        "max_height": 210,
        "input": "foto_ruedas_5",
    },
    ("ruedas", 6): {
        "anchor": "AB561",
        "comment": "AB580",
        "max_width": 520,
        "max_height": 210,
        "input": "foto_ruedas_6",
    },
    ("ruedas", 7): {
        "anchor": "C582",
        "comment": "C601",
        "max_width": 520,
        "max_height": 210,
        "input": "foto_ruedas_7",
    },
    ("ruedas", 8): {
        "anchor": "AB582",
        "comment": "AB601",
        "max_width": 520,
        "max_height": 210,
        "input": "foto_ruedas_8",
    },
    ("ruedas", 9): {
        "anchor": "C604",
        "comment": "C623",
        "max_width": 520,
        "max_height": 210,
        "input": "foto_ruedas_9",
    },
    ("ruedas", 10): {
        "anchor": "AB604",
        "comment": "AB623",
        "max_width": 520,
        "max_height": 210,
        "input": "foto_ruedas_10",
    },
    ("ruedas", 11): {
        "anchor": "C626",
        "comment": "C645",
        "max_width": 520,
        "max_height": 210,
        "input": "foto_ruedas_11",
    },
    ("ruedas", 12): {
        "anchor": "AB626",
        "comment": "AB645",
        "max_width": 520,
        "max_height": 210,
        "input": "foto_ruedas_12",
    },
}

PHOTO_ROW_RANGES = (
    range(248, 457),
    range(518, 646),
)


def _master_path(config):
    filename = config.get("excel_master") or FA0201_MASTER
    path = MASTER_DIR / filename
    if path.is_file():
        return path

    fallback = MASTER_DIR / FA0201_MASTER
    if fallback.is_file():
        return fallback

    raise FileNotFoundError(f"No existe el master Excel: {path}")


def _set(ws, coordinate, value):
    cell = ws[coordinate]
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if coordinate in merged_range:
                cell = ws.cell(merged_range.min_row, merged_range.min_col)
                break
    cell.value = value if value not in (None, "") else None


def _text(datos, key):
    return (datos or {}).get(key, "")


def _responsables(reporte):
    datos = reporte.datos or {}
    responsables = datos.get("responsables") or {}
    return {
        "inspectores_campo": responsables.get("inspectores_campo", ""),
        "supervisores_campo": responsables.get("supervisores_campo", ""),
        "analistas": responsables.get("analistas", ""),
        "inspector_elabora": responsables.get("inspector_elabora", ""),
    }


def _number_or_blank(value):
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return value


def _condition_text(reporte):
    if reporte.condicion_general == reporte.Condicion.NO_MEDIDO:
        return ""
    return reporte.get_condicion_general_display().upper()


def _tag_con_guiones(tag):
    tag = (tag or "").strip().upper()
    if tag.startswith("FLA") and len(tag) == 7:
        return f"FLA-{tag[-4:]}"
    return tag


def _fill_header(ws, reporte):
    datos = reporte.datos or {}
    responsables = _responsables(reporte)

    _set(
        ws,
        "L1",
        (
            "REPORTE DE INSPECCION NDT-"
            f"{reporte.codigo_reporte}_{_tag_con_guiones(reporte.tag)}_CARRILERIA"
        ),
    )
    _set(ws, "J7", _condition_text(reporte))
    _set(ws, "J10", "FILTROS")
    _set(ws, "X10", "TRANSPORTE DE COBRE")
    _set(ws, "J12", "FILTRO PRENSA")
    _set(ws, "X12", reporte.tag)
    _set(ws, "J14", "Operaciones")
    _set(ws, "X14", "EN OPERACION")
    _set(ws, "J16", reporte.fecha_inspeccion)
    _set(ws, "X16", reporte.fecha_reporte)
    _set(ws, "J18", responsables["inspectores_campo"])
    _set(ws, "X18", responsables["supervisores_campo"])
    _set(ws, "J20", responsables["inspector_elabora"])
    _set(ws, "X20", responsables["analistas"])
    _set(ws, "J23", _text(datos, "circunstancias"))
    _set(ws, "J24", _text(datos, "antecedentes"))
    _set(ws, "J25", _text(datos, "observaciones"))
    _set(ws, "J46", _text(datos, "recomendaciones"))


def _fill_ut_parameters(ws, reporte):
    parametros = (reporte.datos or {}).get("parametros_ut") or {}
    for key, coordinate in UT_CELLS.items():
        _set(ws, coordinate, parametros.get(key, ""))


def _fill_measurements(ws, reporte):
    for row in range(112, 184):
        for col in MEASUREMENT_COLUMNS.values():
            _set(ws, f"{col}{row}", None)

    for item in reporte.mediciones or []:
        punto = item.get("punto")
        if punto is None:
            continue

        try:
            row = 112 + (36 - int(punto))
        except (TypeError, ValueError):
            continue

        if not 112 <= row <= 183:
            continue

        _set(ws, f"AC{row}", int(punto))
        _set(ws, f"AO{row}", int(punto))

        for key, col in MEASUREMENT_COLUMNS.items():
            _set(ws, f"{col}{row}", _number_or_blank(item.get(key)))


def _fill_wheels(ws, reporte):
    ruedas = (reporte.datos or {}).get("ruedas") or {}

    for columns in WHEEL_COLUMNS.values():
        for row in WHEEL_ROWS.values():
            for col in columns.values():
                _set(ws, f"{col}{row}", None)

    for lado, columns in WHEEL_COLUMNS.items():
        for item in ruedas.get(lado, []):
            numero = item.get("numero")
            row = WHEEL_ROWS.get(numero)
            if not row:
                continue

            for key, col in columns.items():
                _set(ws, f"{col}{row}", _number_or_blank(item.get(key)))

    _set(ws, "C513", _text(reporte.datos, "observaciones_ruedas"))


def _image_start_row(image):
    try:
        return image.anchor._from.row + 1
    except AttributeError:
        return None


def _clear_example_photos(ws):
    preserved = []
    for image in getattr(ws, "_images", []):
        row = _image_start_row(image)
        if row is not None and any(row in rows for rows in PHOTO_ROW_RANGES):
            continue
        preserved.append(image)
    ws._images = preserved


def _fit_size(path, max_width, max_height):
    with PILImage.open(path) as image:
        width, height = image.size
    if not width or not height:
        return max_width, max_height
    scale = min(max_width / width, max_height / height)
    return int(width * scale), int(height * scale)


def _fill_photos(ws, reporte):
    _clear_example_photos(ws)
    ws.print_area = "B1:BA646"

    comentarios = (reporte.datos or {}).get("comentarios_fotos") or {}
    fotos = {}
    if reporte.pk:
        fotos = {
            (foto.seccion, foto.orden): foto
            for foto in reporte.fotografias.filter(
                seccion__in=["zonas_soldadura", "ruedas"],
            )
        }

    for key, slot in PHOTO_SLOTS.items():
        _set(ws, slot["comment"], comentarios.get(slot["input"], ""))
        foto = fotos.get(key)
        if not foto or not foto.imagen:
            continue

        path = Path(foto.imagen.path)
        if not path.is_file():
            continue

        image = ExcelImage(str(path))
        image.width, image.height = _fit_size(
            path,
            slot["max_width"],
            slot["max_height"],
        )
        ws.add_image(image, slot["anchor"])


def generar_excel_carrileria(reporte, config):
    workbook = load_workbook(
        _master_path(config),
        data_only=False,
        read_only=False,
    )
    worksheet = workbook[SHEET_NAME]

    _fill_header(worksheet, reporte)
    _fill_ut_parameters(worksheet, reporte)
    _fill_measurements(worksheet, reporte)
    _fill_wheels(worksheet, reporte)
    _fill_photos(worksheet, reporte)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
