from __future__ import annotations

from io import BytesIO
from pathlib import Path

from django.conf import settings
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage


MASTER_DIR = Path(settings.BASE_DIR) / "filtros_mm" / "masters"
SHEET_NAME = "Hoja1"

PHOTO_SLOTS = [
    ("C116", "C136"),
    ("AE116", "AE136"),
    ("C152", "C172"),
    ("AE152", "AE172"),
    ("C185", "C205"),
    ("AE185", "AE205"),
    ("C225", "C245"),
    ("AE225", "AE245"),
    ("C268", "C288"),
    ("AE268", "AE288"),
]
PHOTO_SECTIONS = ["faja_empalme", "polea_cola", "polea_cabeza", "faja_poleas"]

PHOTO_START_ROW = 100
BELTING_CELLS = {
    "item": "D53",
    "qty": "F53",
    "manufacturer": "I53",
    "type": "N53",
    "rated": "R53",
    "belt_width": "U53",
    "plys": "X53",
    "top_cover": "AA53",
    "bottom_cover": "AF53",
    "cover_type": "AJ53",
    "splice_type": "AN53",
    "remarks": "AR53",
}
SECTION_LAYOUTS = {
    "top_cover": {
        "procedimiento": "K84",
        "material": "K85",
        "componente": "K86",
        "ut": {
            "marca": "AE84",
            "modelo": "AR84",
            "tipo_haz": "AE85",
            "ganancia": "AR85",
            "frecuencia": "AE86",
            "velocidad": "AR86",
            "ancho_banda": "AE87",
            "retardo": "AR87",
            "amortiguamiento": "AE88",
            "diametro": "AR88",
        },
        "rows": {"1": 93, "2": 94, "3": 95},
        "comments": "C110",
    },
    "empalme": {
        "rows": {"antes": 103, "despues": 105},
        "comments": "C110",
    },
    "polea_cola": {
        "procedimiento": "K211",
        "material": "K212",
        "componente": "K213",
        "ut": {
            "marca": "AE211",
            "modelo": "AR211",
            "tipo_haz": "AE212",
            "ganancia": "AR212",
            "frecuencia": "AE213",
            "velocidad": "AR213",
            "ancho_banda": "AE214",
            "retardo": "AR214",
            "amortiguamiento": "AE215",
            "diametro": "AR215",
        },
        "rows": {"1": 220, "2": 221, "3": 222},
        "comments": "C244",
    },
    "polea_cabeza": {
        "procedimiento": "K250",
        "material": "K251",
        "componente": "K252",
        "ut": {
            "marca": "AE250",
            "modelo": "AR250",
            "tipo_haz": "AE251",
            "ganancia": "AR251",
            "frecuencia": "AE252",
            "velocidad": "AR252",
            "ancho_banda": "AE253",
            "retardo": "AR253",
            "amortiguamiento": "AE254",
            "diametro": "AR254",
        },
        "rows": {"1": 259, "2": 260, "3": 261},
        "comments": "C266",
    },
}
MEASUREMENT_COLUMNS = {
    "a": "T",
    "b": "V",
    "c": "X",
    "d": "Z",
    "e": "AB",
    "f": "AD",
    "g": "AF",
}

HEADER_LAYOUTS = {
    "CV2401": {
        "title": "L1",
        "condition": "J7",
        "plant": "J10",
        "process": "W10",
        "equipment": "J12",
        "tag": "W12",
        "stage": "J14",
        "equipment_condition": "W14",
        "inspection_date": "J16",
        "report_date": "W16",
        "circumstances": "J23",
        "background": "J26",
        "observations": "J27",
    },
    "CV2501": {
        "title": "L1",
        "condition": "I7",
        "plant": "I9",
        "process": "V9",
        "equipment": "I11",
        "tag": "V11",
        "stage": "I13",
        "equipment_condition": "V13",
        "inspection_date": "I15",
        "report_date": "V15",
        "circumstances": "I22",
        "background": "I27",
        "observations": "I29",
    },
    "CV2601": {
        "title": "K1",
        "condition": "H7",
        "plant": "H9",
        "process": "U9",
        "equipment": "H11",
        "tag": "U11",
        "stage": "H13",
        "equipment_condition": "U13",
        "inspection_date": "H15",
        "report_date": "U15",
        "circumstances": "H22",
        "background": "H26",
        "observations": "H28",
    },
}


def _master_path(config):
    filename = config.get("excel_master")
    path = MASTER_DIR / filename
    if not path.is_file():
        raise FileNotFoundError(f"No existe el master Excel: {path}")
    return path


def _set(ws, coordinate, value):
    cell = ws[coordinate]
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if coordinate in merged_range:
                cell = ws.cell(merged_range.min_row, merged_range.min_col)
                break
    cell.value = value if value not in (None, "") else None


def _text(datos, key):
    return (datos or {}).get(key, "")


def _maybe_number(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip().replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return value


def _condition_text(reporte):
    if reporte.condicion_general == reporte.Condicion.NO_MEDIDO:
        return ""
    return reporte.get_condicion_general_display().upper()


def _fill_header(ws, reporte):
    datos = reporte.datos or {}
    layout = HEADER_LAYOUTS.get(reporte.codigo_catalogo, HEADER_LAYOUTS["CV2401"])
    _set(
        ws,
        layout["title"],
        f"REPORTE INSPECCION {reporte.codigo_reporte}-{reporte.tag}-FAJA Y POLEAS",
    )
    _set(ws, layout["condition"], _condition_text(reporte))
    _set(ws, layout["plant"], "FILTROS DE COBRE")
    _set(ws, layout["process"], "TRANSPORTE DE CONCENTRADO")
    _set(ws, layout["equipment"], f"FAJA N° {reporte.tag[-2:]}")
    _set(ws, layout["tag"], f"0420-CVB-{reporte.tag[-4:]}")
    _set(ws, layout["stage"], "Operaciones")
    _set(ws, layout["equipment_condition"], "EN USO")
    _set(ws, layout["inspection_date"], reporte.fecha_inspeccion)
    _set(ws, layout["report_date"], reporte.fecha_reporte)
    _set(ws, layout["circumstances"], _text(datos, "circunstancias"))
    _set(ws, layout["background"], _text(datos, "antecedentes"))
    _set(ws, layout["observations"], _text(datos, "observaciones"))


def _fill_technical_sections(ws, reporte):
    datos = (reporte.datos or {}).get("faja_poleas") or {}

    for key, cell in BELTING_CELLS.items():
        value = (datos.get("belting") or {}).get(key)
        if value not in (None, ""):
            _set(ws, cell, _maybe_number(value))

    for section_key, section_data in (datos.get("secciones") or {}).items():
        layout = SECTION_LAYOUTS.get(section_key)
        if not layout:
            continue

        for field in ("procedimiento", "material", "componente"):
            cell = layout.get(field)
            value = section_data.get(field)
            if cell and value not in (None, ""):
                _set(ws, cell, value)

        for field, cell in (layout.get("ut") or {}).items():
            value = (section_data.get("ut") or {}).get(field)
            if value not in (None, ""):
                _set(ws, cell, _maybe_number(value))

        for item in section_data.get("mediciones") or []:
            row = layout.get("rows", {}).get(str(item.get("punto")))
            if not row:
                continue
            for column_key, column_letter in MEASUREMENT_COLUMNS.items():
                value = (item.get("valores") or {}).get(column_key)
                if value not in (None, ""):
                    _set(ws, f"{column_letter}{row}", _maybe_number(value))

        if section_data.get("comentarios"):
            _set(ws, layout["comments"], section_data.get("comentarios"))


def _image_start_row(image):
    try:
        return image.anchor._from.row + 1
    except AttributeError:
        return None


def _clear_example_photos(ws):
    ws._images = [
        image
        for image in getattr(ws, "_images", [])
        if (_image_start_row(image) or 0) < PHOTO_START_ROW
    ]


def _fit_size(path, max_width=500, max_height=220):
    with PILImage.open(path) as image:
        width, height = image.size
    if not width or not height:
        return max_width, max_height
    scale = min(max_width / width, max_height / height)
    return int(width * scale), int(height * scale)


def _fill_photos(ws, reporte):
    _clear_example_photos(ws)
    comentarios = (reporte.datos or {}).get("comentarios_faja_poleas") or {}
    fotos_ordenadas = []
    if reporte.pk:
        fotos_ordenadas = sorted(
            reporte.fotografias.filter(seccion__in=PHOTO_SECTIONS),
            key=lambda foto: (PHOTO_SECTIONS.index(foto.seccion), foto.orden),
        )

    for index, (anchor, comment_cell) in enumerate(PHOTO_SLOTS, start=1):
        foto = fotos_ordenadas[index - 1] if index <= len(fotos_ordenadas) else None
        comentario = ""
        if foto:
            comentario = comentarios.get(f"foto_{foto.seccion}_{foto.orden}", "")
        _set(ws, comment_cell, comentario)
        if not foto or not foto.imagen:
            continue

        path = Path(foto.imagen.path)
        if not path.is_file():
            continue

        image = ExcelImage(str(path))
        image.width, image.height = _fit_size(path)
        ws.add_image(image, anchor)


def generar_excel_faja_poleas(reporte, config):
    workbook = load_workbook(
        _master_path(config),
        data_only=False,
        read_only=False,
    )
    worksheet = workbook[SHEET_NAME]

    _fill_header(worksheet, reporte)
    _fill_technical_sections(worksheet, reporte)
    _fill_photos(worksheet, reporte)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
