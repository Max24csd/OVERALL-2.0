from __future__ import annotations

import hashlib
import json
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from django.conf import settings
from django.core.files.storage import default_storage
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.db.models import Q
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from inspecciones.models import (
    FaseCampana,
    FotoFajaCVB0003,
    Inspeccion,
    LifeShaftInspeccion,
    MedicionEmpalmeCVB0003,
    MedicionLifeShaftCampana,
    MedicionPolea,
    MedicionPoleaCampana,
    MedicionTramoCVB0003,
    PoleaInspeccion,
    TipoMedicionComponente,
)


FECHA_INSPECCION = date(2026, 8, 3)
FECHA_REPORTE = date(2026, 8, 8)
CAMPOS = tuple("abcdefg")

DEFAULT_POLEAS = Path.home() / "Downloads" / "20260807-VTUT-CVB0003-POLEAS (1).xlsx"
DEFAULT_LIFE_SHAFT = (
    Path.home() / "Downloads" / "20260806-VTUT-0220CVB003-LIVESHAFT.xlsx"
)
DEFAULT_FAJA = (
    Path.home() / "Downloads" / "20260808-VTUT-0220CVB0003_INSPECCION_FAJA_CVB003.xlsx"
)


def decimal_excel(value):
    if value in (None, "", "-"):
        return None
    return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def serializar(value):
    if isinstance(value, (date, datetime, Decimal, Path)):
        return str(value)
    return value


def sha256(path):
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def valores_modelo(medicion):
    return {campo.upper(): getattr(medicion, campo) for campo in CAMPOS}


def datos_campana(queryset):
    return [
        {
            "id": medicion.id,
            "fase": medicion.fase,
            "punto": medicion.punto,
            "valores": valores_modelo(medicion),
            "minimo": medicion.minimo,
            "promedio": medicion.promedio,
        }
        for medicion in queryset.order_by("fase", "orden", "punto")
    ]


def snapshot(inspeccion_faja, inspeccion_poleas, inspeccion_life):
    polea_3 = PoleaInspeccion.objects.get(inspeccion=inspeccion_poleas, numero=3)
    polea_7 = PoleaInspeccion.objects.get(inspeccion=inspeccion_poleas, numero=7)
    shaft_1 = LifeShaftInspeccion.objects.get(inspeccion=inspeccion_life, numero=1)
    return {
        "faja": {
            "id": inspeccion_faja.id,
            "codigo_reporte": inspeccion_faja.codigo_reporte,
            "fecha_inspeccion": inspeccion_faja.fecha_inspeccion,
            "fecha_reporte": inspeccion_faja.fecha_reporte,
            "empalmes": MedicionEmpalmeCVB0003.objects.filter(
                inspeccion=inspeccion_faja
            ).count(),
            "empalmes_con_datos": MedicionEmpalmeCVB0003.objects.filter(
                inspeccion=inspeccion_faja
            ).exclude(a=None, b=None, c=None, d=None, e=None, f=None, g=None).count(),
            "tramos": MedicionTramoCVB0003.objects.filter(
                inspeccion=inspeccion_faja
            ).count(),
            "tramos_con_datos": MedicionTramoCVB0003.objects.filter(
                inspeccion=inspeccion_faja
            ).exclude(a=None, b=None, c=None, d=None, e=None, f=None, g=None).count(),
            "fotografias": FotoFajaCVB0003.objects.filter(
                inspeccion=inspeccion_faja
            ).count(),
        },
        "polea_03": {
            "id": polea_3.id,
            "tipo_medicion": polea_3.tipo_medicion,
            "normal_conservado": polea_3.mediciones.count(),
            "campana": datos_campana(polea_3.mediciones_campana.all()),
        },
        "life_shaft_01": {
            "id": shaft_1.id,
            "tipo_medicion": shaft_1.tipo_medicion,
            "normal_conservado": shaft_1.mediciones.count(),
            "campana": datos_campana(shaft_1.mediciones_campana.all()),
        },
        "polea_07": [
            {
                "id": medicion.id,
                "punto": medicion.punto,
                "valores": valores_modelo(medicion),
                "minimo": medicion.minimo,
                "promedio": medicion.promedio,
            }
            for medicion in polea_7.mediciones.order_by("orden", "punto")
        ],
        "poleas_04_05": list(
            PoleaInspeccion.objects.filter(
                inspeccion=inspeccion_poleas, numero__in=(4, 5)
            )
            .order_by("numero")
            .values("numero", "tipo_medicion")
        ),
    }


def buscar_inspeccion(tipo):
    inspecciones = Inspeccion.objects.filter(
        faja__tag__icontains="CVB0003",
        tipo=tipo,
    ).filter(
        fecha_inspeccion=FECHA_INSPECCION
    )
    if tipo == Inspeccion.Tipo.FAJA:
        inspecciones = Inspeccion.objects.filter(
            faja__tag__icontains="CVB0003", tipo=tipo
        ).filter(Q(fecha_inspeccion=FECHA_INSPECCION) | Q(fecha_inspeccion__isnull=True))
    coincidencias = list(inspecciones.order_by("id"))
    if len(coincidencias) != 1:
        raise CommandError(
            f"Se esperaba una inspección CVB003 {tipo}; se encontraron "
            f"{len(coincidencias)}: {[obj.id for obj in coincidencias]}"
        )
    return coincidencias[0]


def leer_filas_componente(ws, filas, columnas, puntos_esperados):
    resultado = []
    for fila, punto in zip(filas, puntos_esperados):
        valores = {
            campo: decimal_excel(ws[f"{columna}{fila}"].value)
            for campo, columna in zip(CAMPOS, columnas)
        }
        resultado.append({"punto": punto, "valores": valores})
    return resultado


def leer_fuentes(poleas_path, life_path, faja_path):
    wb_poleas = load_workbook(poleas_path, data_only=True, read_only=False)
    ws_poleas = wb_poleas[wb_poleas.sheetnames[0]]
    titulo_polea = str(ws_poleas["W139"].value or "").upper()
    if "INICIO DE CAMPAÑA" not in titulo_polea:
        raise CommandError("El Excel de Poleas no confirma INICIO DE CAMPAÑA en W139.")
    polea_3 = leer_filas_componente(
        ws_poleas,
        range(142, 147),
        ("AC", "AE", "AG", "AI", "AK", "AM", "AO"),
        range(1, 6),
    )
    polea_7 = leer_filas_componente(
        ws_poleas,
        range(283, 288),
        ("AB", "AD", "AF", "AH", "AJ", "AL", "AN"),
        range(1, 6),
    )

    wb_life = load_workbook(life_path, data_only=True, read_only=False)
    ws_life = wb_life[wb_life.sheetnames[0]]
    texto_campana = str(ws_life["C103"].value or "").upper()
    if "INICIO DE CAMPAÑA" not in texto_campana:
        raise CommandError(
            "El Excel de Life Shaft no confirma INICIO DE CAMPAÑA en C103."
        )
    shaft_1 = leer_filas_componente(
        ws_life,
        range(73, 77),
        ("AB", "AD", "AF", "AH", "AJ", "AL", "AN"),
        range(1, 5),
    )

    wb_faja_values = load_workbook(faja_path, data_only=True, read_only=False)
    ws_faja = wb_faja_values["REPORTE DE INSPECCION CV0003"]
    empalmes = leer_empalmes(ws_faja)
    carga = leer_tramos(ws_faja, "CARGA", 893, 1012)
    retorno = leer_tramos(ws_faja, "RETORNO", 1022, 1123)
    wb_faja_images = load_workbook(faja_path, data_only=False, read_only=False)
    fotos = leer_fotos(wb_faja_images["REPORTE DE INSPECCION CV0003"])

    return {
        "polea_3": polea_3,
        "polea_7": polea_7,
        "shaft_1": shaft_1,
        "empalmes": empalmes,
        "carga": carga,
        "retorno": retorno,
        "fotos": fotos,
    }


def leer_empalmes(ws):
    resultado = []
    zona = empalme = bastidor = None
    for fila in range(222, 312):
        zona = ws[f"C{fila}"].value or zona
        empalme = ws[f"D{fila}"].value or empalme
        bastidor = ws[f"E{fila}"].value or bastidor
        posicion = ws[f"F{fila}"].value
        if not all((zona, empalme, bastidor, posicion)):
            raise CommandError(f"Fila de empalme incompleta en Excel: {fila}")
        resultado.append(
            {
                "fila_excel": fila,
                "zona": str(zona).strip(),
                "empalme": str(empalme).strip(),
                "bastidor_lado": " ".join(str(bastidor).split()),
                "posicion": str(posicion).strip(),
                "espesor_nominal": decimal_excel(ws[f"G{fila}"].value),
                "valores": {
                    campo: decimal_excel(ws.cell(fila, columna).value)
                    for campo, columna in zip(CAMPOS, range(8, 15))
                },
            }
        )
    return resultado


def leer_tramos(ws, tipo, inicio, fin):
    resultado = []
    tramo = None
    for fila in range(inicio, fin + 1):
        tramo = ws[f"D{fila}"].value or tramo
        medicion = ws[f"E{fila}"].value
        bastidor = ws[f"F{fila}"].value
        if tramo is None or medicion is None or bastidor is None:
            raise CommandError(f"Fila de tramo incompleta en Excel: {fila}")
        resultado.append(
            {
                "fila_excel": fila,
                "tipo": tipo,
                "tramo": " ".join(str(tramo).split()),
                "medicion": int(medicion),
                "bastidor": str(int(bastidor)) if isinstance(bastidor, (int, float)) else str(bastidor),
                "espesor_nominal": decimal_excel(ws[f"G{fila}"].value),
                "valores": {
                    campo: decimal_excel(ws.cell(fila, columna).value)
                    for campo, columna in zip(CAMPOS, range(8, 15))
                },
            }
        )
    return resultado


def descripcion_foto(ws, row, anchor):
    textos = []
    for fila in range(max(315, row - 3), min(ws.max_row, row + 4) + 1):
        for columna in range(3, 22):
            value = ws.cell(fila, columna).value
            if isinstance(value, str):
                value = " ".join(value.split())
                if value and value not in textos:
                    textos.append(value)
    detalle = " | ".join(textos)[:900]
    base = f"Fotografía histórica del reporte final CVB003 (anclaje {anchor})."
    return f"{base} {detalle}".strip()


def leer_fotos(ws):
    fotos = []
    for image in ws._images:
        row = image.anchor._from.row + 1
        column = image.anchor._from.col + 1
        if row < 315:
            continue
        seccion = (
            FotoFajaCVB0003.Seccion.EMPALMES
            if row < 888
            else FotoFajaCVB0003.Seccion.CARGA
        )
        anchor = f"{get_column_letter(column)}{row}"
        data = image._data()
        digest = hashlib.sha256(data).hexdigest()
        extension = (image.format or "png").lower().replace("jpeg", "jpg")
        filename = f"20260803_{seccion.lower()}_{anchor.lower()}_{digest[:12]}.{extension}"
        relative_path = f"inspecciones/faja/cvb0003/historico_20260803/{filename}"
        fotos.append(
            {
                "seccion": seccion,
                "anchor": anchor,
                "sha256": digest,
                "data": data,
                "relative_path": relative_path,
                "descripcion": descripcion_foto(ws, row, anchor),
            }
        )
    if len(fotos) != 79 or len({foto["sha256"] for foto in fotos}) != 79:
        raise CommandError(
            "La clasificación de fotografías variables no coincide con la auditoría "
            f"aprobada: {len(fotos)} slots / {len({f['sha256'] for f in fotos})} hashes."
        )
    return fotos


def aplicar_campana(componente, filas_excel, modelo, relacion, audit):
    if componente.tipo_medicion != TipoMedicionComponente.CAMPANA:
        type(componente).objects.filter(pk=componente.pk).update(
            tipo_medicion=TipoMedicionComponente.CAMPANA
        )
        audit["componentes_tipo_actualizado"].append(
            {"componente": str(componente), "antes": componente.tipo_medicion, "despues": "CAMPANA"}
        )
    normales = {fila.punto: fila for fila in componente.mediciones.order_by("orden", "punto")}
    for fuente in filas_excel:
        normal = normales.get(fuente["punto"])
        if normal is None:
            raise CommandError(
                f"No existe fila NORMAL de respaldo para {componente}, punto {fuente['punto']}."
            )
        lookup = {relacion: componente, "fase": FaseCampana.INICIO, "punto": normal.punto}
        defaults = {
            "orden": normal.orden,
            "observacion": normal.observacion,
            **fuente["valores"],
        }
        if modelo is MedicionPoleaCampana:
            defaults["posicion"] = normal.posicion
        else:
            defaults["ubicacion"] = normal.ubicacion
        existente = modelo.objects.filter(**lookup).first()
        antes = valores_modelo(existente) if existente else None
        objeto, creado = modelo.objects.update_or_create(defaults=defaults, **lookup)
        despues = valores_modelo(objeto)
        if creado:
            audit["mediciones_campana_creadas"] += 1
        elif antes != despues:
            audit["mediciones_campana_actualizadas"] += 1
        audit["campana_detalle"].append(
            {
                "componente": str(componente),
                "fase": "INICIO DE CAMPAÑA",
                "punto": normal.punto,
                "creado": creado,
                "antes": antes,
                "despues": despues,
            }
        )


def aplicar_polea_7(polea, filas_excel, audit):
    existentes = {fila.punto: fila for fila in polea.mediciones.order_by("orden", "punto")}
    for fuente in filas_excel:
        medicion = existentes.get(fuente["punto"])
        if medicion is None:
            raise CommandError(f"No existe Polea 07 / punto {fuente['punto']}.")
        cambios = {}
        for campo, aprobado in fuente["valores"].items():
            actual = getattr(medicion, campo)
            if actual != aprobado:
                cambios[campo] = {"antes": actual, "despues": aprobado}
                setattr(medicion, campo, aprobado)
        if cambios:
            medicion.save(update_fields=[*cambios.keys(), "minimo", "promedio"])
            audit["polea_07_coordenadas_actualizadas"] += len(cambios)
            audit["polea_07_detalle"].append(
                {"punto": medicion.punto, "cambios": cambios}
            )


def actualizar_filas(queryset, fuentes, campos_estructura, audit, categoria):
    filas = list(queryset.order_by("orden", "id"))
    if len(filas) != len(fuentes):
        raise CommandError(
            f"Cantidad incompatible para {categoria}: Django={len(filas)}, Excel={len(fuentes)}"
        )
    for objeto, fuente in zip(filas, fuentes):
        cambios = {}
        for campo in campos_estructura:
            aprobado = fuente[campo]
            actual = getattr(objeto, campo)
            if actual != aprobado:
                cambios[campo] = {"antes": actual, "despues": aprobado}
                setattr(objeto, campo, aprobado)
        for campo, aprobado in fuente["valores"].items():
            actual = getattr(objeto, campo)
            if actual != aprobado:
                cambios[campo] = {"antes": actual, "despues": aprobado}
                setattr(objeto, campo, aprobado)
        if cambios:
            objeto.save(update_fields=list(cambios))
            audit["faja_registros_modificados"] += 1
            audit["faja_coordenadas_actualizadas"] += sum(
                1 for campo in cambios if campo in CAMPOS
            )
            audit["faja_detalle"].append(
                {
                    "categoria": categoria,
                    "id": objeto.id,
                    "fila_excel": fuente["fila_excel"],
                    "cambios": cambios,
                }
            )


def importar_fotos(inspeccion, fotos, audit, dry_run):
    creados_archivo = []
    for foto in fotos:
        if FotoFajaCVB0003.objects.filter(
            inspeccion=inspeccion, imagen=foto["relative_path"]
        ).exists():
            continue
        if dry_run:
            audit["fotografias_importadas"] += 1
            continue
        if not default_storage.exists(foto["relative_path"]):
            from django.core.files.base import ContentFile

            saved_name = default_storage.save(
                foto["relative_path"], ContentFile(foto["data"])
            )
            if saved_name != foto["relative_path"]:
                raise CommandError(
                    f"El almacenamiento alteró el nombre determinista: {saved_name}"
                )
            creados_archivo.append(saved_name)
        FotoFajaCVB0003.objects.create(
            inspeccion=inspeccion,
            seccion=foto["seccion"],
            imagen=foto["relative_path"],
            codigo_dano="",
            descripcion=foto["descripcion"],
            subida_por=inspeccion.inspector,
        )
        audit["fotografias_importadas"] += 1
        audit["fotografias_detalle"].append(
            {
                "seccion": foto["seccion"],
                "anchor": foto["anchor"],
                "sha256": foto["sha256"],
                "imagen": foto["relative_path"],
            }
        )
    return creados_archivo


class Command(BaseCommand):
    help = "Corrige de forma idempotente el historial CVB003 del 03/08/2026."

    def add_arguments(self, parser):
        parser.add_argument("--apply", action="store_true", help="Confirma los cambios.")
        parser.add_argument("--poleas", type=Path, default=DEFAULT_POLEAS)
        parser.add_argument("--life-shaft", type=Path, default=DEFAULT_LIFE_SHAFT)
        parser.add_argument("--faja", type=Path, default=DEFAULT_FAJA)
        parser.add_argument(
            "--output-dir",
            type=Path,
            default=Path(settings.BASE_DIR) / "auditorias" / "cvb0003_20260803",
        )

    def handle(self, *args, **options):
        dry_run = not options["apply"]
        rutas = {
            "poleas": options["poleas"].resolve(),
            "life_shaft": options["life_shaft"].resolve(),
            "faja": options["faja"].resolve(),
        }
        for label, path in rutas.items():
            if not path.is_file():
                raise CommandError(f"No existe la fuente {label}: {path}")
        fuentes_hash = {label: sha256(path) for label, path in rutas.items()}
        fuentes = leer_fuentes(rutas["poleas"], rutas["life_shaft"], rutas["faja"])

        inspeccion_faja = buscar_inspeccion(Inspeccion.Tipo.FAJA)
        inspeccion_poleas = buscar_inspeccion(Inspeccion.Tipo.POLEAS)
        inspeccion_life = buscar_inspeccion(Inspeccion.Tipo.LIFE_SHAFT)
        antes = snapshot(inspeccion_faja, inspeccion_poleas, inspeccion_life)
        codigo_faja_original = inspeccion_faja.codigo_reporte
        tipos_04_05 = antes["poleas_04_05"]
        archivos_nuevos = []
        audit = {
            "modo": "DRY_RUN" if dry_run else "APPLY",
            "fecha_objetivo": str(FECHA_INSPECCION),
            "fuentes": {
                label: {"ruta": str(rutas[label]), "sha256": fuentes_hash[label]}
                for label in rutas
            },
            "componentes_tipo_actualizado": [],
            "mediciones_campana_creadas": 0,
            "mediciones_campana_actualizadas": 0,
            "campana_detalle": [],
            "polea_07_coordenadas_actualizadas": 0,
            "polea_07_detalle": [],
            "faja_registros_modificados": 0,
            "faja_coordenadas_actualizadas": 0,
            "faja_detalle": [],
            "fotografias_importadas": 0,
            "fotografias_detalle": [],
        }

        try:
            with transaction.atomic():
                PoleaInspeccion.objects.select_for_update().filter(
                    inspeccion=inspeccion_poleas, numero__in=(3, 4, 5, 7)
                ).count()
                LifeShaftInspeccion.objects.select_for_update().filter(
                    inspeccion=inspeccion_life, numero=1
                ).count()
                Inspeccion.objects.select_for_update().filter(
                    pk__in=(inspeccion_faja.pk, inspeccion_poleas.pk, inspeccion_life.pk)
                ).count()

                if (
                    inspeccion_faja.fecha_inspeccion != FECHA_INSPECCION
                    or inspeccion_faja.fecha_reporte != FECHA_REPORTE
                ):
                    Inspeccion.objects.filter(pk=inspeccion_faja.pk).update(
                        fecha_inspeccion=FECHA_INSPECCION,
                        fecha_reporte=FECHA_REPORTE,
                    )
                    audit["faja_fechas_actualizadas"] = True
                else:
                    audit["faja_fechas_actualizadas"] = False

                polea_3 = PoleaInspeccion.objects.get(
                    inspeccion=inspeccion_poleas, numero=3
                )
                aplicar_campana(
                    polea_3,
                    fuentes["polea_3"],
                    MedicionPoleaCampana,
                    "polea",
                    audit,
                )
                shaft_1 = LifeShaftInspeccion.objects.get(
                    inspeccion=inspeccion_life, numero=1
                )
                aplicar_campana(
                    shaft_1,
                    fuentes["shaft_1"],
                    MedicionLifeShaftCampana,
                    "life_shaft",
                    audit,
                )
                polea_7 = PoleaInspeccion.objects.get(
                    inspeccion=inspeccion_poleas, numero=7
                )
                aplicar_polea_7(polea_7, fuentes["polea_7"], audit)

                actualizar_filas(
                    MedicionEmpalmeCVB0003.objects.filter(inspeccion=inspeccion_faja),
                    fuentes["empalmes"],
                    ("zona", "empalme", "bastidor_lado", "posicion", "espesor_nominal"),
                    audit,
                    "EMPALMES",
                )
                actualizar_filas(
                    MedicionTramoCVB0003.objects.filter(
                        inspeccion=inspeccion_faja, tipo=MedicionTramoCVB0003.Tipo.CARGA
                    ),
                    fuentes["carga"],
                    ("tipo", "tramo", "medicion", "bastidor", "espesor_nominal"),
                    audit,
                    "CARGA",
                )
                actualizar_filas(
                    MedicionTramoCVB0003.objects.filter(
                        inspeccion=inspeccion_faja, tipo=MedicionTramoCVB0003.Tipo.RETORNO
                    ),
                    fuentes["retorno"],
                    ("tipo", "tramo", "medicion", "bastidor", "espesor_nominal"),
                    audit,
                    "RETORNO",
                )
                archivos_nuevos = importar_fotos(
                    inspeccion_faja, fuentes["fotos"], audit, dry_run
                )

                inspeccion_faja.refresh_from_db()
                if inspeccion_faja.codigo_reporte != codigo_faja_original:
                    raise CommandError("Se intentó alterar el código histórico de Faja.")
                actuales_04_05 = list(
                    PoleaInspeccion.objects.filter(
                        inspeccion=inspeccion_poleas, numero__in=(4, 5)
                    )
                    .order_by("numero")
                    .values("numero", "tipo_medicion")
                )
                if actuales_04_05 != tipos_04_05:
                    raise CommandError("Se alteró indebidamente Polea 04 o Polea 05.")
                if MedicionPoleaCampana.objects.filter(
                    polea=polea_3, fase=FaseCampana.FIN
                ).exists():
                    raise CommandError("Polea 03 contiene un FIN no aprobado.")
                if MedicionLifeShaftCampana.objects.filter(
                    life_shaft=shaft_1, fase=FaseCampana.FIN
                ).exists():
                    raise CommandError("Life Shaft 01 contiene un FIN no aprobado.")

                despues = snapshot(inspeccion_faja, inspeccion_poleas, inspeccion_life)
                if dry_run:
                    transaction.set_rollback(True)
        except Exception:
            for name in archivos_nuevos:
                default_storage.delete(name)
            raise

        if dry_run:
            despues = antes
            despues_proyectado = {
                "nota": "Los cambios se revirtieron deliberadamente por DRY_RUN.",
                "conteos_proyectados": {
                    key: audit[key]
                    for key in (
                        "mediciones_campana_creadas",
                        "mediciones_campana_actualizadas",
                        "polea_07_coordenadas_actualizadas",
                        "faja_registros_modificados",
                        "faja_coordenadas_actualizadas",
                        "fotografias_importadas",
                    )
                },
            }
        else:
            despues_proyectado = None

        output_dir = options["output_dir"].resolve()
        output_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        reporte = {
            "ejecutado_en": datetime.now().isoformat(),
            "antes": antes,
            "despues": despues,
            "proyeccion": despues_proyectado,
            "auditoria": audit,
        }
        if not dry_run:
            original = output_dir / "auditoria_antes_original.json"
            if not original.exists():
                original.write_text(
                    json.dumps(antes, ensure_ascii=False, indent=2, default=serializar),
                    encoding="utf-8",
                )
            (output_dir / "auditoria_despues.json").write_text(
                json.dumps(despues, ensure_ascii=False, indent=2, default=serializar),
                encoding="utf-8",
            )
        report_path = output_dir / f"ejecucion_{timestamp}_{audit['modo'].lower()}.json"
        report_path.write_text(
            json.dumps(reporte, ensure_ascii=False, indent=2, default=serializar),
            encoding="utf-8",
        )
        self.stdout.write(self.style.SUCCESS(f"Auditoría: {report_path}"))
        self.stdout.write(json.dumps(audit, ensure_ascii=False, indent=2, default=serializar))
        if dry_run:
            self.stdout.write(self.style.WARNING("DRY_RUN: no se guardó ningún cambio."))
