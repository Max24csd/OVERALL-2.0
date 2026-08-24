from copy import copy
from re import match

from openpyxl.cell.cell import MergedCell
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter


def limpiar_textos_historicos_campana(worksheet):
    frases = ("INICIO DE CAMPAÑA", "FIN DE CAMPAÑA", "INICIO DE CAMPANA", "FIN DE CAMPANA")
    for row in worksheet.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            texto = cell.value
            for frase in frases:
                texto = texto.replace(f" / {frase}", "").replace(f" - {frase}", "").replace(frase, "")
            cell.value = " ".join(texto.split()) if texto != cell.value else cell.value


def _numero(valor):
    return float(valor) if valor is not None else None


def _promedio(valores):
    valores = [float(valor) for valor in valores if valor is not None]
    return sum(valores) / len(valores) if valores else None


def _minimo(valores):
    valores = [float(valor) for valor in valores if valor is not None]
    return min(valores) if valores else None


def _primero_con_titulo(layouts):
    if isinstance(layouts, dict):
        grupos = layouts.items()
    else:
        grupos = enumerate(((layout,) for layout in layouts), 1)
    for numero, grupo in grupos:
        if isinstance(grupo, dict):
            grupo = (grupo,)
        for layout in grupo:
            if "title" in layout:
                return layout, numero
    raise ValueError("La plantilla no contiene un bloque técnico normal para clonar.")


def _layout_componente(layouts, numero, indice, respaldo):
    if isinstance(layouts, dict):
        grupo = layouts.get(numero, ())
        if isinstance(grupo, dict):
            grupo = (grupo,)
        layout = next((layout for layout in grupo if "title" in layout), None)
        return (layout, numero) if layout else respaldo
    if indice < len(layouts) and "title" in layouts[indice]:
        return layouts[indice], numero
    return respaldo


def _copiar_bloque(origen, destino, fila_inicio, fila_fin, fila_destino, columna_fin):
    desplazamiento = fila_destino - fila_inicio
    for indice_columna in range(1, columna_fin + 1):
        letra = get_column_letter(indice_columna)
        destino.column_dimensions[letra].width = origen.column_dimensions[letra].width
        destino.column_dimensions[letra].hidden = origen.column_dimensions[letra].hidden
    for fila in range(fila_inicio, fila_fin + 1):
        fila_nueva = fila + desplazamiento
        destino.row_dimensions[fila_nueva].height = origen.row_dimensions[fila].height
        destino.row_dimensions[fila_nueva].hidden = False
        for columna in range(1, columna_fin + 1):
            celda_origen = origen.cell(fila, columna)
            if isinstance(celda_origen, MergedCell):
                continue
            celda_destino = destino.cell(fila_nueva, columna)
            valor = celda_origen.value
            if isinstance(valor, str) and valor.startswith("="):
                try:
                    valor = Translator(valor, origin=celda_origen.coordinate).translate_formula(celda_destino.coordinate)
                except Exception:
                    pass
            celda_destino.value = valor
            if celda_origen.has_style:
                celda_destino._style = copy(celda_origen._style)
            celda_destino.number_format = celda_origen.number_format
            celda_destino.alignment = copy(celda_origen.alignment)
            celda_destino.protection = copy(celda_origen.protection)
    for rango in origen.merged_cells.ranges:
        if rango.min_row >= fila_inicio and rango.max_row <= fila_fin:
            destino.merge_cells(
                start_row=rango.min_row + desplazamiento,
                start_column=rango.min_col,
                end_row=rango.max_row + desplazamiento,
                end_column=rango.max_col,
            )
    return desplazamiento


def _columna_fila(valor, columna_predeterminada):
    if isinstance(valor, str):
        encontrado = match(r"([A-Z]+)(\d+)", valor)
        if encontrado:
            return encontrado.group(1), int(encontrado.group(2))
    return columna_predeterminada, int(valor)


def _escribir_mediciones(ws, layout, desplazamiento, mediciones, config):
    mediciones = list(mediciones)
    cantidad = layout.get("data_end", layout["average"] - 1) - layout["data"] + 1
    mediciones = mediciones[:cantidad]
    columnas = layout.get("columns", config["measurement_columns"])
    columna_punto = layout.get("point", config["point_column"])
    columna_resultado = layout.get("result", config["result_column"])
    columna_residual = layout.get("residual", config["residual_column"])
    for offset in range(cantidad):
        fila = layout["data"] + desplazamiento + offset
        medicion = mediciones[offset] if offset < len(mediciones) else None
        ws[f"{columna_punto}{fila}"] = medicion.punto if medicion else None
        for campo, columna in zip("abcdefg", columnas):
            ws[f"{columna}{fila}"] = _numero(getattr(medicion, campo)) if medicion else None
        ws[f"{columna_resultado}{fila}"] = _numero(medicion.promedio) if medicion else None
        ws[f"{columna_residual}{fila}"] = _numero(medicion.minimo) if medicion else None
    fila_promedio = layout["average"] + desplazamiento
    fila_minimo = layout["minimum"] + desplazamiento
    for campo, columna in zip("abcdefg", columnas):
        valores = [getattr(medicion, campo) for medicion in mediciones]
        ws[f"{columna}{fila_promedio}"] = _promedio(valores)
        ws[f"{columna}{fila_minimo}"] = _minimo(valores)
    promedios = [medicion.promedio for medicion in mediciones]
    minimos = [medicion.minimo for medicion in mediciones]
    ws[f"{columna_resultado}{fila_promedio}"] = _promedio(promedios)
    ws[f"{columna_residual}{fila_promedio}"] = _promedio(minimos)
    ws[f"{columna_resultado}{fila_minimo}"] = _minimo(promedios)
    ws[f"{columna_residual}{fila_minimo}"] = _minimo(minimos)
    candidatos = []
    for medicion in mediciones:
        for campo in "abcdefg":
            valor = getattr(medicion, campo)
            if valor is not None:
                candidatos.append((float(valor), campo.upper(), medicion.punto))
    if candidatos:
        valor, punto, radial = min(candidatos, key=lambda item: item[0])
        nota = f"El espesor mínimo encontrado es de {valor:.2f} mm en el punto {punto}, medición {radial}."
    else:
        nota = "No existen mediciones registradas para esta fase."
    columna_nota, fila_nota = _columna_fila(layout["note"], config["note_column"])
    ws[f"{columna_nota}{fila_nota + desplazamiento}"] = nota


def _escribir_calibracion(ws, layout, desplazamiento, componente, config):
    columna, fila = _columna_fila(layout["calibration"], config["calibration_column"])
    valores = (
        componente.marca_equipo,
        getattr(componente, "tipo_haz", None) or getattr(componente, "modelo_equipo", None),
        getattr(componente, "frecuencia_mhz", None),
        getattr(componente, "ancho_banda", None) or getattr(componente, "rango_mm", None),
        getattr(componente, "amortiguamiento", None) or getattr(componente, "metodo_empleado", None),
        getattr(componente, "velocidad_ms", None),
        getattr(componente, "retardo_us", None),
    )
    for offset, valor in enumerate(valores):
        ws[f"{columna}{fila + desplazamiento + offset}"] = valor or "-"


def _reemplazar_etiquetas(ws, fila_inicio, fila_fin, numero_origen, numero_destino, es_ceramica):
    for fila in ws.iter_rows(min_row=fila_inicio, max_row=fila_fin):
        for celda in fila:
            if isinstance(celda.value, str):
                texto = celda.value
                for prefijo in ("POLEA #", "POLEA ", "LIFE SHAFT #", "LIFE SHAFT ", "LIVESHAFT #", "LIVESHAFT "):
                    texto = texto.replace(f"{prefijo}{numero_origen:02d}", f"{prefijo}{numero_destino:02d}")
                    texto = texto.replace(f"{prefijo}{numero_origen}", f"{prefijo}{numero_destino}")
                if es_ceramica:
                    texto = texto.replace("DE CAUCHO", "CERÁMICO").replace("CAUCHO", "CERÁMICA")
                    texto = texto.replace("LAGGING DE LA POLEA", "LAGGING CERÁMICO DE LA POLEA")
                celda.value = texto


def agregar_hoja_mediciones_campana(workbook, bloques, clave_componente, titulo, *, config):
    bloques_campana = [bloque for bloque in bloques if bloque.get("es_campana")]
    nombre = "CAMPAÑA"
    if nombre in workbook.sheetnames:
        del workbook[nombre]
    if "MEDICIONES CAMPANA" in workbook.sheetnames:
        del workbook["MEDICIONES CAMPANA"]
    if not bloques_campana:
        return
    origen = workbook[config.get("sheet_name", "Hoja1")]
    destino = workbook.create_sheet(nombre)
    destino.sheet_view.showGridLines = False
    destino.page_setup = copy(origen.page_setup)
    destino.page_margins = copy(origen.page_margins)
    destino.sheet_properties = copy(origen.sheet_properties)
    destino.sheet_properties.pageSetUpPr.fitToPage = True
    respaldo = _primero_con_titulo(config["layouts"])
    fila_destino = 1
    for indice, bloque in enumerate(bloques_campana):
        componente = bloque[clave_componente]
        layout, numero_origen = _layout_componente(config["layouts"], componente.numero, indice, respaldo)
        fila_fin = layout.get("note")
        _, fila_fin = _columna_fila(fila_fin, config["note_column"])
        for fase, clave in (("INICIO", "mediciones_inicio"), ("FIN", "mediciones_fin")):
            desplazamiento = _copiar_bloque(
                origen, destino, layout["title"], fila_fin, fila_destino,
                config.get("max_column", 50),
            )
            titulo_celda = f"{config['title_column']}{layout['title'] + desplazamiento}"
            es_ceramica = componente.numero in config.get("ceramic_numbers", ())
            material = "CERÁMICA" if es_ceramica else "CAUCHO"
            destino[titulo_celda] = (
                f"{titulo} - #{componente.numero:02d} - MATERIAL {material} - {fase} DE CAMPAÑA"
            )
            _escribir_calibracion(destino, layout, desplazamiento, componente, config)
            _escribir_mediciones(destino, layout, desplazamiento, bloque.get(clave, []), config)
            inicio_bloque = layout["title"] + desplazamiento
            fin_bloque = fila_fin + desplazamiento
            _reemplazar_etiquetas(
                destino,
                inicio_bloque,
                fin_bloque,
                numero_origen,
                componente.numero,
                es_ceramica,
            )
            fila_destino = fin_bloque + 3
    destino.print_area = f"A1:{get_column_letter(config.get('max_column', 50))}{fila_destino - 2}"
    destino.page_setup.fitToWidth = 1
    destino.page_setup.fitToHeight = 0
