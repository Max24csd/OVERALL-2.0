from copy import copy
from io import BytesIO
from pathlib import Path

from django.conf import settings
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.worksheet.pagebreak import Break
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.drawing.spreadsheet_drawing import (
    AnchorMarker,
    OneCellAnchor,
)
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from PIL import Image as PILImage, ImageOps
from pathlib import Path

TEMPLATE_PATH = (
    Path(settings.BASE_DIR)
    / "inspecciones"
    / "reportes"
    / "cvb0001"
    / "assets"
    / "reporte_faja_cvb0001.xlsx"
)
SOURCE_SHEET = "REPORTE DE INSPECCION CV0001"
OUTPUT_SHEET = "REPORTE DE INSPECCION"


def _copy_style(source_cell, target_cell):
    target_cell._style = copy(source_cell._style)
    target_cell.number_format = source_cell.number_format
    target_cell.protection = copy(source_cell.protection)

def _resolve_media_path(file_field):
    """
    Devuelve un Path real de un FileField/ImageField de Django.
    """
    if not file_field:
        return None

    try:
        path = Path(file_field.path)
    except Exception:
        return None

    return path if path.exists() else None

def _style_range(source_cell, worksheet, cell_range):
    for row in worksheet[cell_range]:
        for cell in row:
            _copy_style(source_cell, cell)


def _merge_and_write(worksheet, cell_range, value, source_cell):
    worksheet.merge_cells(cell_range)
    cell = worksheet[cell_range.split(":", 1)[0]]
    _copy_style(source_cell, cell)
    cell.value = value
    return cell


def _number(value):
    return float(value) if value is not None else None


def _percentage(value):
    return float(value) / 100 if value is not None else None


def _safe_image_path(photo):
    try:
        path = Path(photo.imagen.path)
    except (AttributeError, NotImplementedError, OSError, ValueError):
        return None
    return path if path.is_file() else None


def _column_width_pixels(worksheet, column):
    """Convierte el ancho de una columna de Excel a píxeles aprox."""
    letter = get_column_letter(column)
    width = worksheet.column_dimensions[letter].width
    if width is None:
        width = 8.43
    if width < 1:
        return int(width * 12)
    return int(width * 7 + 5)


def _row_height_pixels(worksheet, row):
    """Convierte la altura de una fila de Excel a píxeles aprox."""
    height = worksheet.row_dimensions[row].height
    if height is None:
        height = 15
    return int(height * 96 / 72)


def _box_pixels(worksheet, min_col, max_col, min_row, max_row):
    width = sum(
        _column_width_pixels(worksheet, col)
        for col in range(min_col, max_col + 1)
    )
    height = sum(
        _row_height_pixels(worksheet, row)
        for row in range(min_row, max_row + 1)
    )
    return width, height


def _add_image_in_box(
    worksheet,
    path,
    min_col,
    max_col,
    min_row,
    max_row,
    margin=6,
    allow_upscale=False,
):
    """
    Inserta una imagen centrada dentro de un cuadro fijo.
    Conserva proporción y nunca cambia filas ni columnas.
    """
    if not path:
        return False

    try:
        with PILImage.open(path) as original:
            normalized = ImageOps.exif_transpose(original)
            img_width, img_height = normalized.size
    except (FileNotFoundError, OSError, ValueError):
        return False

    box_width, box_height = _box_pixels(
        worksheet, min_col, max_col, min_row, max_row
    )
    available_width = max(box_width - margin * 2, 1)
    available_height = max(box_height - margin * 2, 1)

    scale = min(
        available_width / img_width,
        available_height / img_height,
    )
    if not allow_upscale:
        scale = min(scale, 1)

    new_width = max(int(img_width * scale), 1)
    new_height = max(int(img_height * scale), 1)
    offset_x = max((box_width - new_width) // 2, 0)
    offset_y = max((box_height - new_height) // 2, 0)

    try:
        image = ExcelImage(str(path))
    except (FileNotFoundError, OSError, ValueError):
        return False

    marker = AnchorMarker(
        col=min_col - 1,
        colOff=pixels_to_EMU(offset_x),
        row=min_row - 1,
        rowOff=pixels_to_EMU(offset_y),
    )
    size = XDRPositiveSize2D(
        cx=pixels_to_EMU(new_width),
        cy=pixels_to_EMU(new_height),
    )
    image.anchor = OneCellAnchor(_from=marker, ext=size)
    worksheet.add_image(image)
    return True


def _frame_range(worksheet, cell_range, color="7F7F7F", style="thin"):
    """Dibuja solo el marco exterior de un rango sin alterar su contenido."""
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    side = Side(style=style, color=color)

    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = worksheet.cell(row=row, column=col)
            current = cell.border
            cell.border = Border(
                left=side if col == min_col else current.left,
                right=side if col == max_col else current.right,
                top=side if row == min_row else current.top,
                bottom=side if row == max_row else current.bottom,
                diagonal=current.diagonal,
                diagonal_direction=current.diagonal_direction,
                diagonalUp=current.diagonalUp,
                diagonalDown=current.diagonalDown,
                outline=current.outline,
                vertical=current.vertical,
                horizontal=current.horizontal,
            )


def _prepare_photo_box(worksheet, cell_range):
    """Crea un cuadro limpio y visible para una foto sin cambiar dimensiones."""
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    worksheet.merge_cells(cell_range)
    cell = worksheet.cell(min_row, min_col)
    cell.fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    _frame_range(worksheet, cell_range, color="A6A6A6", style="thin")


def _copy_header_template(source, worksheet):
    for row in range(1, 22):
        worksheet.row_dimensions[row].height = source.row_dimensions[row].height
        for column in range(1, 24):
            _copy_style(source.cell(row, column), worksheet.cell(row, column))
    for merged in source.merged_cells.ranges:
        if merged.max_row <= 21 and merged.max_col <= 23:
            worksheet.merge_cells(str(merged))


def _write_header(source, worksheet, inspeccion):
    """
    Escribe la cabecera del reporte CVB0001.

    Incluye:
    - datos generales;
    - condición;
    - logos;
    - imagen fija principal del equipo;
    - marco exterior;
    - sin congelar paneles.
    """

    # ========================================================
    # 1. COPIAR ESTRUCTURA ORIGINAL DE CABECERA
    # ========================================================

    _copy_header_template(
        source,
        worksheet,
    )

    # ========================================================
    # 2. DATOS DEL REPORTE
    # ========================================================

    values = {
        "F2": (
            f"REPORTE DE INSPECCIÓN "
            f"{inspeccion.codigo_reporte}"
        ),
        "F5": (
            "Ingeniería de Confiabilidad "
            "- Operaciones Procesos"
        ),
        "C7": "CONDICIÓN",
        "F7": (
            inspeccion
            .get_condicion_general_display()
            .upper()
        ),

        "C9": "DATOS DEL EQUIPO",

        "C10": "PLANTA",
        "F10": inspeccion.planta,

        "I10": "PROCESO",
        "L10": inspeccion.proceso,

        "C12": "EQUIPO",
        "F12": inspeccion.faja.nombre,

        "I12": "TAG-NUMBER",
        "L12": inspeccion.faja.tag,

        "C14": "ETAPA",
        "F14": inspeccion.etapa,

        "I14": "CONDICIÓN DEL EQUIPO",
        "L14": inspeccion.condicion_equipo,

        "C16": "FECHA DE INSPECCIÓN",
        "F16": inspeccion.fecha_inspeccion,

        "I16": "FECHA DE REPORTE",
        "L16": inspeccion.fecha_reporte,

        "C18": "INSPECTOR DE CAMPO",
        "F18": inspeccion.inspector_campo_nombre,

        "I18": "SUPERVISOR DE CAMPO",
        "L18": inspeccion.supervisor_campo_nombre,

        "C20": "ANALISTA QUE ELABORA",
        "F20": inspeccion.analista_elabora_nombre,

        "I20": "ANALISTA QUE VALIDA",
        "L20": inspeccion.analista_valida_nombre,
    }

    for coordinate, value in values.items():
        worksheet[coordinate] = (
            value
            if value not in (None, "")
            else "—"
        )

    # ========================================================
    # 3. FORMATO DE FECHAS
    # ========================================================

    worksheet["F16"].number_format = (
        "dd-mm-yyyy"
    )

    worksheet["L16"].number_format = (
        "dd-mm-yyyy"
    )

    # ========================================================
    # 4. COLOR DE CONDICIÓN
    # ========================================================

    for row_cells in worksheet[
        "F7:N8"
    ]:
        for cell in row_cells:
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor="00B050",
            )

    # ========================================================
    # 5. DIRECTORIO DE RECURSOS ESTÁTICOS
    # ========================================================

    static_dir = (
        Path(settings.BASE_DIR)
        / "static"
        / "inspecciones"
        / "faja"
        / "cvb0001"
    )

    # ========================================================
    # 6. LOGO OVERALL SOLUTIONS
    # ========================================================

    logo_overall = (
        static_dir
        / "logo_overall.jpeg"
    )

    if logo_overall.exists():
        _add_image_in_box(
            worksheet,
            logo_overall,
            min_col=2,    # B
            max_col=5,    # E
            min_row=2,
            max_row=5,
            margin=6,
            allow_upscale=True,
        )

    # ========================================================
    # 7. LOGO MMG LAS BAMBAS
    # ========================================================

    logo_mmg = (
        static_dir
        / "logo_mmg_las_bambas.jpeg"
    )

    if logo_mmg.exists():
        _add_image_in_box(
            worksheet,
            logo_mmg,
            min_col=19,   # S
            max_col=22,   # V
            min_row=2,
            max_row=5,
            margin=6,
            allow_upscale=True,
        )

    # ========================================================
    # 8. IMAGEN FIJA PRINCIPAL DEL EQUIPO
    # ========================================================

    imagen_equipo = (
        static_dir
        / "imagen_equipo.jpeg"
    )

    if imagen_equipo.exists():
        _add_image_in_box(
            worksheet,
            imagen_equipo,
            min_col=15,   # O
            max_col=22,   # V
            min_row=7,
            max_row=20,
            margin=5,
            allow_upscale=True,
        )

    else:
        print(
            "CVB0001: No se encontró la imagen principal:",
            imagen_equipo,
            flush=True,
        )

    # ========================================================
    # 9. MARCO DEL ÁREA DE LA IMAGEN PRINCIPAL
    # ========================================================

    # Solo se copian bordes.
    # NO copiar fill aquí porque podría pintar de blanco
    # encima del diseño original.

    donor_border = copy(
        source["V10"].border
    )

    for row_cells in worksheet.iter_rows(
        min_row=7,
        max_row=20,
        min_col=15,
        max_col=22,
    ):
        for cell in row_cells:
            cell.border = copy(
                donor_border
            )

    # ========================================================
    # 10. MARCO EXTERIOR DE CABECERA
    # ========================================================

    _frame_range(
        worksheet,
        "B2:V21",
        color="000000",
        style="thin",
    )


def _write_section_title(source, worksheet, row, title):
    cell = _merge_and_write(
        worksheet,
        f"C{row}:U{row}",
        title,
        source["C114"],
    )
    cell.alignment = copy(source["C114"].alignment)
    worksheet.row_dimensions[row].height = 21
    _frame_range(worksheet, f"C{row}:U{row}", color="000000", style="thin")
    return row + 2


def _write_empalme(source, worksheet, row, title, measurements, summary):
    row = _write_section_title(source, worksheet, row, title)
    header_top = row
    header_bottom = row + 1
    fixed_headers = [
        ("D", "Empalme"),
        ("E", "Bastidor"),
        ("F", "Lado"),
        ("G", "Posición"),
        ("H", "Espesor nominal (mm)"),
    ]
    for column, label in fixed_headers:
        _merge_and_write(
            worksheet,
            f"{column}{header_top}:{column}{header_bottom}",
            label,
            source["D116"],
        )
    _merge_and_write(
        worksheet,
        f"I{header_top}:O{header_top}",
        "Espesor residual (mm)",
        source["I116"],
    )
    _merge_and_write(
        worksheet,
        f"P{header_top}:Q{header_top}",
        "Espesor (mm)",
        source["P116"],
    )
    _merge_and_write(
        worksheet,
        f"R{header_top}:T{header_top}",
        "Desgaste (mínimo)",
        source["R116"],
    )
    subheaders = list("ABCDEFG") + [
        "Mínimo",
        "Promedio",
        "Desgaste (mm)",
        "% desgaste",
        "% residual",
    ]
    for index, label in enumerate(subheaders, start=9):
        cell = worksheet.cell(header_bottom, index, label)
        _copy_style(source["I117"], cell)

    data_start = header_bottom + 1
    if measurements:
        data_end = data_start + len(measurements) - 1
        for column in "DEF":
            worksheet.merge_cells(
                start_row=data_start,
                start_column=ord(column) - 64,
                end_row=data_end,
                end_column=ord(column) - 64,
            )
        first = measurements[0]
        worksheet[f"D{data_start}"] = first.seccion.replace("EMPALME ", "")
        worksheet[f"E{data_start}"] = first.bastidor
        worksheet[f"F{data_start}"] = first.lado
        for coordinate in (f"D{data_start}", f"E{data_start}", f"F{data_start}"):
            _copy_style(source["D119"], worksheet[coordinate])
        for offset, measurement in enumerate(measurements):
            current = data_start + offset
            values = [
                measurement.posicion,
                _number(measurement.espesor_nominal),
                *[_number(getattr(measurement, letter)) for letter in "abcdefg"],
                _number(measurement.minimo),
                _number(measurement.promedio),
                _number(measurement.desgaste),
                _percentage(measurement.porcentaje_desgaste),
                _percentage(measurement.porcentaje_residual),
            ]
            for column, value in enumerate(values, start=7):
                cell = worksheet.cell(current, column, value)
                donor = source["P119"] if column >= 16 else source["I119"]
                _copy_style(donor, cell)
                cell.number_format = "0.00%" if column >= 19 else "0.00"
            worksheet.row_dimensions[current].height = 22
    else:
        data_end = data_start
        _merge_and_write(
            worksheet,
            f"D{data_start}:T{data_start}",
            "No existen mediciones registradas para este empalme.",
            source["D121"],
        )

    summary_start = data_end + 2
    summary_end = summary_start + 2
    cell = _merge_and_write(
        worksheet,
        f"D{summary_start}:T{summary_end}",
        summary["texto"],
        source["D121"],
    )
    cell.alignment = copy(source["D121"].alignment)
    cell.alignment = copy(cell.alignment)
    cell.alignment = cell.alignment.copy(wrapText=True, vertical="center")
    for item in range(summary_start, summary_end + 1):
        worksheet.row_dimensions[item].height = 20
    return summary_end + 2


def _write_photos(
    source,
    worksheet,
    row,
    title,
    photos,
):
    """
    Sección de inspección visual ordenada.

    - 1 foto: centrada.
    - 2 fotos: una a cada lado.
    - 3 o más: dos por fila.
    - Todas conservan proporción.
    - Los cuadros nunca cambian de tamaño por la imagen.
    """
    row = _write_section_title(source, worksheet, row, title)

    valid_photos = []
    for photo in photos:
        path = _safe_image_path(photo)
        if path:
            valid_photos.append((photo, path))

    if not valid_photos:
        cell = _merge_and_write(
            worksheet,
            f"C{row}:U{row + 1}",
            "No se registraron fotografías en esta sección.",
            source["C150"],
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        worksheet.row_dimensions[row].height = 22
        worksheet.row_dimensions[row + 1].height = 22
        _frame_range(worksheet, f"C{row}:U{row + 1}", color="A6A6A6")
        return row + 3

    for index in range(0, len(valid_photos), 2):
        pair = valid_photos[index:index + 2]

        image_start_row = row
        image_end_row = row + 12
        text_start_row = image_end_row + 1
        text_end_row = text_start_row + 2

        # Alturas estables para que todas las inspecciones luzcan iguales.
        for current_row in range(image_start_row, image_end_row + 1):
            worksheet.row_dimensions[current_row].height = 18
        for current_row in range(text_start_row, text_end_row + 1):
            worksheet.row_dimensions[current_row].height = 18

        if len(pair) == 1:
            # Una sola foto: centrada en el reporte.
            boxes = [(7, 14, f"G{text_start_row}:N{text_end_row}")]
        else:
            boxes = [
                (4, 11, f"D{text_start_row}:K{text_end_row}"),
                (13, 20, f"M{text_start_row}:T{text_end_row}"),
            ]

        for (photo, path), (min_col, max_col, text_range) in zip(pair, boxes):
            image_range = (
                f"{get_column_letter(min_col)}{image_start_row}:"
                f"{get_column_letter(max_col)}{image_end_row}"
            )
            _prepare_photo_box(worksheet, image_range)

            _add_image_in_box(
                worksheet,
                path,
                min_col,
                max_col,
                image_start_row,
                image_end_row,
                margin=8,
                allow_upscale=True,
            )

            details = []
            titulo_foto = (getattr(photo, "titulo", "") or "").strip()
            descripcion = (getattr(photo, "descripcion", "") or "").strip()
            if titulo_foto:
                details.append(titulo_foto)
            if descripcion:
                details.append(descripcion)

            if details:
                cell = _merge_and_write(
                    worksheet,
                    text_range,
                    "\n".join(details),
                    source["C150"],
                )
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
                _frame_range(worksheet, text_range, color="A6A6A6")

        # Mantener marco visual del ancho general de la sección.
        _frame_range(
            worksheet,
            f"C{image_start_row}:U{text_end_row}",
            color="D9D9D9",
            style="thin",
        )
        row = text_end_row + 2

    return row


def _write_tramos(source, worksheet, row, measurements, summary):
    row = _write_section_title(
        source,
        worksheet,
        row,
        "DATOS DE MEDICIÓN DE ESPESORES POR UT EN PUNTOS DE MEDICIÓN",
    )
    headers = [
        "Tramo",
        "Medición",
        "Bastidor",
        "Nominal",
        "A",
        "B",
        "C",
        "D",
        "E",
        "F",
        "G",
        "Mínimo",
        "Promedio",
        "Desgaste",
        "% desgaste",
        "% residual",
    ]
    for index, label in enumerate(headers, start=4):
        cell = worksheet.cell(row, index, label)
        _copy_style(source["D155"], cell)
    data_start = row + 1
    for offset, measurement in enumerate(measurements):
        current = data_start + offset
        values = [
            measurement.seccion,
            measurement.punto,
            measurement.bastidor,
            _number(measurement.espesor_nominal),
            *[_number(getattr(measurement, letter)) for letter in "abcdefg"],
            _number(measurement.minimo),
            _number(measurement.promedio),
            _number(measurement.desgaste),
            _percentage(measurement.porcentaje_desgaste),
            _percentage(measurement.porcentaje_residual),
        ]
        for column, value in enumerate(values, start=4):
            cell = worksheet.cell(current, column, value)
            donor = source["O158"] if column >= 15 else source["H158"]
            _copy_style(donor, cell)
            if column <= 6:
                cell.number_format = "@"
            elif column >= 18:
                cell.number_format = "0.00%"
            else:
                cell.number_format = "0.00"
        worksheet.row_dimensions[current].height = 20

    summary_row = data_start + len(measurements)
    for label, key, current in (
        ("ESPESOR MÍNIMO (mm)", "minimos", summary_row),
        ("ESPESOR PROMEDIO (mm)", "promedios", summary_row + 1),
    ):
        _merge_and_write(
            worksheet,
            f"D{current}:G{current}",
            label,
            source["D173"],
        )
        for offset, letter in enumerate("abcdefg", start=8):
            cell = worksheet.cell(current, offset, summary[key].get(letter))
            _copy_style(source["H173"], cell)
            cell.number_format = "0.00"
        for column in range(15, 20):
            _copy_style(source["O173"], worksheet.cell(current, column))
    return summary_row + 3


def _write_notes(source, worksheet, row, inspeccion):
    for title, value in (
        ("OBSERVACIONES", inspeccion.observaciones),
        ("CONCLUSIONES Y RECOMENDACIONES", inspeccion.recomendaciones),
    ):
        row = _write_section_title(source, worksheet, row, title)
        cell = _merge_and_write(
            worksheet,
            f"C{row}:U{row + 4}",
            value or "—",
            source["F40"],
        )
        cell.alignment = cell.alignment.copy(wrapText=True, vertical="top", horizontal="left")
        for current in range(row, row + 5):
            worksheet.row_dimensions[current].height = 20
        row += 6
    return row


def generar_reporte_faja_cvb0001_excel(
    inspeccion,
    empalme_e01,
    empalme_e02,
    tramos,
    resumen_e01,
    resumen_e02,
    resumen_tramos,
    fotos_e01,
    fotos_e02,
    fotos_puntos,
):
    workbook = load_workbook(TEMPLATE_PATH)
    source = workbook[SOURCE_SHEET]
    worksheet = workbook.create_sheet(OUTPUT_SHEET, 0)

    for column in range(1, 24):
        letter = get_column_letter(column)
        worksheet.column_dimensions[letter].width = source.column_dimensions[letter].width
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_view.zoomScale = 80

    _write_header(source, worksheet, inspeccion)
    row = 24
    row = _write_empalme(
        source,
        worksheet,
        row,
        "DATOS DE MEDICIÓN – EMPALME E-01 TOP COVER",
        empalme_e01,
        resumen_e01,
    )
    if any(_safe_image_path(photo) for photo in fotos_e01):
        worksheet.row_breaks.append(Break(id=row - 1))
    row = _write_photos(
        source,
        worksheet,
        row,
        "INSPECCIÓN VISUAL – EMPALME E-01 TOP COVER",
        fotos_e01,
    )
    worksheet.row_breaks.append(Break(id=row - 1))
    row = _write_empalme(
        source,
        worksheet,
        row,
        "DATOS DE MEDICIÓN – EMPALME E-02 BOTTOM COVER",
        empalme_e02,
        resumen_e02,
    )
    row = _write_photos(
        source,
        worksheet,
        row,
        "INSPECCIÓN VISUAL – EMPALME E-02 BOTTOM COVER",
        fotos_e02,
    )
    worksheet.row_breaks.append(Break(id=row - 1))
    row = _write_tramos(source, worksheet, row, tramos, resumen_tramos)
    row = _write_photos(
        source,
        worksheet,
        row,
        "INSPECCIÓN VISUAL – PUNTOS DE MEDICIÓN",
        fotos_puntos,
    )
    row = _write_notes(source, worksheet, row, inspeccion)

    for old_sheet in list(workbook.worksheets):
        if old_sheet is not worksheet:
            workbook.remove(old_sheet)
    workbook.active = 0
    worksheet.freeze_panes = None
    worksheet.print_area = f"B1:V{row}"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.print_options.horizontalCentered = True
    worksheet.page_margins.left = 0.25
    worksheet.page_margins.right = 0.25
    worksheet.page_margins.top = 0.3
    worksheet.page_margins.bottom = 0.3
    worksheet.page_margins.header = 0.15
    worksheet.page_margins.footer = 0.15

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
