from copy import copy
from io import BytesIO
from pathlib import Path

from django.conf import settings
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from .life_shaft_image_utils import insertar_imagen_ajustada, tamano_comun_cuadros_px
from inspecciones.reportes.campaign_excel import agregar_hoja_mediciones_campana, limpiar_textos_historicos_campana
from .text_utils import (
    conclusiones_life_shaft,
    minimo_componente,
    observaciones_life_shaft,
)


TEMPLATE_PATH = Path(settings.BASE_DIR) / "inspecciones/reportes/cvb0001/templates/20260513-VTUT-CVB0001-LIFE SHAFT.xlsx"
COLUMNAS = ("AC", "AE", "AG", "AI", "AK", "AM", "AO")
LAYOUTS = (
    dict(title=57, calibration=60, data=62, average=65, minimum=66, note="Z67", visual=70, photos=(71, 86), caption="E88"),
    dict(title=92, calibration=95, data=97, average=100, minimum=101, note="Z102", visual=105, photos=(106, 121), caption="D123"),
)


def _nombre(usuario):
    return (usuario.get_full_name() or usuario.username or "").strip() if usuario else ""


def _campo(inspeccion, nombre, usuario):
    return (getattr(inspeccion, nombre, "") or "").strip() or _nombre(usuario)


def _numero(valor):
    return float(valor) if valor is not None else None


def _promedio(valores):
    valores = [float(v) for v in valores if v is not None]
    return sum(valores) / len(valores) if valores else None


def _minimo(valores):
    valores = [float(v) for v in valores if v is not None]
    return min(valores) if valores else None


def _ruta_foto(foto):
    try:
        ruta = Path(foto.imagen.path)
    except (AttributeError, NotImplementedError, OSError, ValueError):
        return None
    return ruta if ruta.is_file() else None


def _cabecera(ws, inspeccion, shafts):
    ws["M3"] = f"REPORTE INSPECCIÓN {inspeccion.codigo_reporte}"
    ws["K9"] = inspeccion.get_condicion_general_display().upper()
    ws["K9"].fill = PatternFill("solid", fgColor="00B050")
    ws["K9"].font = Font(name=ws["K9"].font.name, size=ws["K9"].font.sz, bold=True, color="FFFFFF")
    valores = {
        "K12": inspeccion.planta, "Y12": inspeccion.proceso,
        "K14": "LIFE SHAFT", "Y14": inspeccion.faja.tag,
        "K16": inspeccion.etapa, "Y16": inspeccion.condicion_equipo,
        "K18": inspeccion.fecha_inspeccion, "Y18": inspeccion.fecha_reporte,
        "K20": _campo(inspeccion, "inspector_campo_nombre", inspeccion.inspector),
        "Y20": _campo(inspeccion, "supervisor_campo_nombre", inspeccion.supervisor),
        "K22": _campo(inspeccion, "analista_elabora_nombre", inspeccion.analista),
        "Y22": _campo(inspeccion, "analista_valida_nombre", inspeccion.analista),
        "K25": inspeccion.circunstancias, "K28": inspeccion.antecedentes,
        "K30": observaciones_life_shaft(shafts, inspeccion.observaciones),
        "D43": f"ESQUEMA DE UBICACIÓN DE LIFE SHAFT DE LA FAJA {inspeccion.faja.tag}",
    }
    for celda, valor in valores.items():
        ws[celda] = valor or "-"
    ws["K18"].number_format = ws["Y18"].number_format = "dd mmmm yyyy"
    lineas = conclusiones_life_shaft(shafts, inspeccion.recomendaciones).splitlines()
    for indice, fila in enumerate(range(37, 41)):
        ws[f"K{fila}"] = lineas[indice] if indice < len(lineas) else None


def _rangos_fotos(inicio, fin, cantidad):
    if cantidad <= 1:
        return [f"D{inicio}:AT{fin}"]
    if cantidad == 2:
        return [f"E{inicio}:Z{fin}", f"AD{inicio}:AS{fin}"]
    mitad = inicio + (fin - inicio + 1) // 2
    rangos = [f"D{inicio}:R{mitad-1}", f"S{inicio}:AF{mitad-1}", f"AG{inicio}:AT{mitad-1}", f"E{mitad}:Z{fin}", f"AD{mitad}:AS{fin}"]
    return rangos[:cantidad]


def _mediciones(ws, layout, mediciones):
    mediciones = list(mediciones)[:3]
    for offset in range(3):
        fila = layout["data"] + offset
        medicion = mediciones[offset] if offset < len(mediciones) else None
        ws[f"Z{fila}"] = medicion.punto if medicion else None
        for campo, columna in zip("abcdefg", COLUMNAS):
            ws[f"{columna}{fila}"] = _numero(getattr(medicion, campo)) if medicion else None
        ws[f"AQ{fila}"] = _numero(medicion.promedio) if medicion else None
        ws[f"AT{fila}"] = _numero(medicion.minimo) if medicion else None
    for campo, columna in zip("abcdefg", COLUMNAS):
        valores = [getattr(m, campo) for m in mediciones]
        ws[f"{columna}{layout['average']}"] = _promedio(valores)
        ws[f"{columna}{layout['minimum']}"] = _minimo(valores)
    ws[f"AQ{layout['average']}"] = _promedio([m.promedio for m in mediciones])
    ws[f"AT{layout['average']}"] = _promedio([m.minimo for m in mediciones])
    ws[f"AQ{layout['minimum']}"] = _minimo([m.promedio for m in mediciones])
    ws[f"AT{layout['minimum']}"] = _minimo([m.minimo for m in mediciones])


def _shaft(ws, layout, bloque, inspeccion):
    shaft = bloque["life_shaft"]
    ws[f"D{layout['visual']}"] = f"INSPECCIÓN VISUAL DEL LIFE SHAFT #{shaft.numero:02d}"
    if bloque.get("es_campana"):
        for fila in range(layout["title"], layout["visual"]):
            ws.row_dimensions[fila].hidden = True
    else:
        ws[f"D{layout['title']}"] = f"MEDICIÓN DE ESPESORES DEL LIFE SHAFT #{shaft.numero:02d} - {inspeccion.faja.tag}"
        parametros = (shaft.marca_equipo, shaft.tipo_haz, shaft.frecuencia_mhz, shaft.ancho_banda, shaft.amortiguamiento, shaft.velocidad_ms, shaft.retardo_us)
        for offset, valor in enumerate(parametros):
            ws[f"N{layout['calibration'] + offset}"] = valor or "-"
        _mediciones(ws, layout, bloque["mediciones"])
        minimo = minimo_componente(shaft)
        ws[layout["note"]] = f"El espesor mínimo encontrado es de {minimo[0]:.2f} mm en el punto {minimo[1]}." if minimo else ""
    fotos = [(foto, ruta) for foto in bloque["fotografias"] if (ruta := _ruta_foto(foto))][:5]
    rangos = _rangos_fotos(*layout["photos"], len(fotos))
    ancho, alto = tamano_comun_cuadros_px(ws, rangos)
    for (_foto, ruta), rango in zip(fotos, rangos):
        insertar_imagen_ajustada(ws, ruta, rango, ancho, alto, 4)
    partes = []
    for etiqueta, valor in (("Nombre", shaft.nombre), ("TAG", shaft.tag), ("Ubicación", shaft.ubicacion), ("Condición", shaft.get_condicion_display()), ("Observación visual", shaft.observacion_visual), ("Observación de medición", shaft.observacion_medicion), ("Recomendaciones", shaft.recomendaciones)):
        if (valor or "").strip():
            partes.append(f"{etiqueta}: {valor.strip()}")
    titulos = [f.codigo_dano.strip() for f, _r in fotos if (f.codigo_dano or "").strip()]
    descripciones = [f.descripcion.strip() for f, _r in fotos if (f.descripcion or "").strip()]
    if titulos:
        partes.append("Fotografías: " + " | ".join(titulos))
    if descripciones:
        partes.append("Observaciones fotográficas: " + " | ".join(descripciones))
    celda = ws[layout["caption"]]
    celda.value = "\n".join(partes) if any((shaft.observacion_visual, shaft.observacion_medicion, shaft.recomendaciones, titulos, descripciones)) else None
    if celda.value:
        alineacion = copy(celda.alignment); alineacion.wrap_text = True; celda.alignment = alineacion


def generar_excel_life_shaft_cvb0001(inspeccion, bloques):
    workbook = load_workbook(BytesIO(TEMPLATE_PATH.read_bytes()))
    ws = workbook["Hoja1"]
    limpiar_textos_historicos_campana(ws)
    shafts = [b["life_shaft"] for b in bloques]
    _cabecera(ws, inspeccion, shafts)
    ws._images = [imagen for imagen in ws._images if imagen.anchor._from.row + 1 < 57]
    for layout, bloque in zip(LAYOUTS, bloques):
        _shaft(ws, layout, bloque, inspeccion)
    agregar_hoja_mediciones_campana(
        workbook, bloques, "life_shaft", f"MEDICIÓN DE ESPESORES DEL LIFE SHAFT {inspeccion.faja.tag}",
        config={"layouts": LAYOUTS, "title_column": "D", "calibration_column": "N", "measurement_columns": COLUMNAS, "point_column": "Z", "result_column": "AQ", "residual_column": "AT", "note_column": "Z"},
    )
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    salida = BytesIO(); workbook.save(salida); salida.seek(0)
    return salida
