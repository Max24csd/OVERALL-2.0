from copy import copy
from io import BytesIO
from pathlib import Path

from django.conf import settings
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from .life_shaft_image_utils import insertar_imagen_ajustada, tamano_comun_cuadros_px
from inspecciones.reportes.campaign_excel import agregar_hoja_mediciones_campana, limpiar_textos_historicos_campana
from .poleas_text_utils import conclusiones_poleas, observaciones_poleas
from .text_utils import minimo_componente


TEMPLATE_PATH = Path(settings.BASE_DIR) / "inspecciones/reportes/cvb0001/templates/20260513-VTUT-CVB0001-POLEAS.xlsx"
COLUMNAS = ("AC", "AE", "AG", "AI", "AK", "AM", "AO")
LAYOUTS = {
    1: dict(visual=78, photos=(80, 101), caption="D103"),
    2: dict(title=105, calibration=108, data=110, average=115, minimum=116, note="Z117", visual=120, photos=(121, 135), caption="D137"),
    3: dict(title=139, calibration=142, data=144, average=149, minimum=150, note="Z151", visual=154, photos=(155, 169), caption="E171"),
    4: dict(title=174, calibration=177, data=179, average=184, minimum=185, note="W186", visual=188, photos=(189, 205), caption="E207"),
    5: dict(title=210, calibration=213, data=215, average=220, minimum=221, note="Z222", visual=224, photos=(225, 244), caption="E246"),
}


def _nombre(usuario):
    return (usuario.get_full_name() or usuario.username or "").strip() if usuario else ""


def _campo(inspeccion, nombre, usuario):
    return (getattr(inspeccion, nombre, "") or "").strip() or _nombre(usuario)


def _numero(valor): return float(valor) if valor is not None else None
def _promedio(valores):
    valores=[float(v) for v in valores if v is not None]; return sum(valores)/len(valores) if valores else None
def _minimo(valores):
    valores=[float(v) for v in valores if v is not None]; return min(valores) if valores else None
def _ruta_foto(foto):
    try: ruta=Path(foto.imagen.path)
    except (AttributeError,NotImplementedError,OSError,ValueError): return None
    return ruta if ruta.is_file() else None


def _cabecera(ws, inspeccion, poleas):
    ws["M3"] = f"REPORTE INSPECCIÓN {inspeccion.codigo_reporte}"
    ws["K9"] = inspeccion.get_condicion_general_display().upper()
    ws["K9"].fill = PatternFill("solid", fgColor="00B050")
    ws["K9"].font = Font(name=ws["K9"].font.name,size=ws["K9"].font.sz,bold=True,color="FFFFFF")
    valores={"K12":inspeccion.planta,"Y12":inspeccion.proceso,"K14":"POLEAS","Y14":inspeccion.faja.tag,"K16":inspeccion.etapa,"Y16":inspeccion.condicion_equipo,"K18":inspeccion.fecha_inspeccion,"Y18":inspeccion.fecha_reporte,"K20":_campo(inspeccion,"inspector_campo_nombre",inspeccion.inspector),"Y20":_campo(inspeccion,"supervisor_campo_nombre",inspeccion.supervisor),"K22":_campo(inspeccion,"analista_elabora_nombre",inspeccion.analista),"Y22":_campo(inspeccion,"analista_valida_nombre",inspeccion.analista),"K25":inspeccion.circunstancias,"K28":inspeccion.antecedentes,"K30":observaciones_poleas(poleas,inspeccion.observaciones),"D64":f"ESQUEMA DE UBICACIÓN DE POLEAS DE LA FAJA {inspeccion.faja.tag}"}
    for celda,valor in valores.items(): ws[celda]=valor or "-"
    ws["K18"].number_format=ws["Y18"].number_format="dd mmmm yyyy"
    lineas=conclusiones_poleas(poleas,inspeccion.recomendaciones).splitlines()
    for indice,fila in enumerate(range(56,62)): ws[f"K{fila}"]=lineas[indice] if indice<len(lineas) else None


def _mediciones(ws,layout,mediciones):
    mediciones=list(mediciones)[:5]
    for offset in range(5):
        fila=layout["data"]+offset;med=mediciones[offset] if offset<len(mediciones) else None
        ws[f"Z{fila}"]=med.punto if med else None
        for campo,columna in zip("abcdefg",COLUMNAS): ws[f"{columna}{fila}"]=_numero(getattr(med,campo)) if med else None
        ws[f"AQ{fila}"]=_numero(med.promedio) if med else None;ws[f"AT{fila}"]=_numero(med.minimo) if med else None
    for campo,columna in zip("abcdefg",COLUMNAS):
        valores=[getattr(m,campo) for m in mediciones];ws[f"{columna}{layout['average']}"]=_promedio(valores);ws[f"{columna}{layout['minimum']}"]=_minimo(valores)
    ws[f"AQ{layout['average']}"]=_promedio([m.promedio for m in mediciones]);ws[f"AT{layout['average']}"]=_promedio([m.minimo for m in mediciones]);ws[f"AQ{layout['minimum']}"]=_minimo([m.promedio for m in mediciones]);ws[f"AT{layout['minimum']}"]=_minimo([m.minimo for m in mediciones])


def _rangos(inicio,fin,cantidad):
    if cantidad<=1:return [f"D{inicio}:AT{fin}"]
    if cantidad==2:return [f"E{inicio}:Z{fin}",f"AB{inicio}:AT{fin}"]
    mitad=inicio+(fin-inicio+1)//2
    return [f"D{inicio}:R{mitad-1}",f"S{inicio}:AF{mitad-1}",f"AG{inicio}:AT{mitad-1}",f"E{mitad}:Z{fin}",f"AB{mitad}:AT{fin}"][:cantidad]


def _polea(ws,layout,bloque,inspeccion,vistos):
    polea=bloque["polea"]
    if bloque.get("es_campana") and "title" in layout:
        for fila in range(layout["title"],layout["visual"]):ws.row_dimensions[fila].hidden=True
    elif "title" in layout:
        ws[f"D{layout['title']}"]=f"MEDICIÓN DE ESPESORES DEL LAGGING DE LA POLEA #{polea.numero:02d} - {inspeccion.faja.tag}"
        parametros=(polea.marca_equipo,polea.modelo_equipo,polea.frecuencia_mhz,polea.rango_mm,polea.metodo_empleado,polea.velocidad_ms,polea.retardo_us)
        for offset,valor in enumerate(parametros):ws[f"N{layout['calibration']+offset}"]=valor or "-"
        _mediciones(ws,layout,bloque["mediciones"]);minimo=minimo_componente(polea);ws[layout["note"]]=f"El espesor mínimo encontrado es de {minimo[0]:.2f} mm en el punto {minimo[1]}." if minimo else ""
    ws[f"D{layout['visual']}"]=f"INSPECCIÓN VISUAL DE LA POLEA #{polea.numero:02d} - {inspeccion.faja.tag}"
    fotos=[]
    for foto in bloque["fotografias"]:
        ruta=_ruta_foto(foto)
        if ruta is None:continue
        clave=str(ruta.resolve()).casefold()
        if clave in vistos:continue
        vistos.add(clave);fotos.append((foto,ruta))
    fotos=fotos[:5];rangos=_rangos(*layout["photos"],len(fotos));ancho,alto=tamano_comun_cuadros_px(ws,rangos)
    for (_foto,ruta),rango in zip(fotos,rangos):insertar_imagen_ajustada(ws,ruta,rango,ancho,alto,4)
    partes=[]
    for etiqueta,valor in (("Nombre",polea.nombre),("TAG",polea.tag),("Ubicación",polea.ubicacion),("Condición",polea.get_condicion_display()),("Observación visual",polea.observacion_visual),("Observación de medición",polea.observacion_medicion),("Recomendaciones",polea.recomendaciones)):
        if (valor or "").strip():partes.append(f"{etiqueta}: {valor.strip()}")
    descripciones=[f.descripcion.strip() for f,_r in fotos if (f.descripcion or "").strip()]
    if descripciones:partes.append("Fotografías: "+" | ".join(descripciones))
    celda=ws[layout["caption"]];celda.value=" | ".join(partes) if any((polea.observacion_visual,polea.observacion_medicion,polea.recomendaciones,descripciones)) else None
    if celda.value:alineacion=copy(celda.alignment);alineacion.wrap_text=True;celda.alignment=alineacion


def generar_excel_poleas_cvb0001(inspeccion,bloques):
    workbook=load_workbook(BytesIO(TEMPLATE_PATH.read_bytes()));ws=workbook["Hoja1"];limpiar_textos_historicos_campana(ws);poleas=[b["polea"] for b in bloques];_cabecera(ws,inspeccion,poleas)
    ws._images=[imagen for imagen in ws._images if imagen.anchor._from.row+1<78];vistos=set()
    for bloque in bloques:_polea(ws,LAYOUTS[bloque["polea"].numero],bloque,inspeccion,vistos)
    agregar_hoja_mediciones_campana(
        workbook, bloques, "polea", f"MEDICIÓN DE ESPESORES DEL LAGGING DE LA POLEA {inspeccion.faja.tag}",
        config={"layouts": LAYOUTS, "title_column": "D", "calibration_column": "N", "measurement_columns": COLUMNAS, "point_column": "Z", "result_column": "AQ", "residual_column": "AT", "note_column": "Z", "ceramic_numbers": {1}},
    )
    workbook.calculation.fullCalcOnLoad=True;workbook.calculation.forceFullCalc=True;workbook.calculation.calcMode="auto"
    salida=BytesIO();workbook.save(salida);salida.seek(0);return salida
