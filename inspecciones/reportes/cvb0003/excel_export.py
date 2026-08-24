from copy import copy
from io import BytesIO
from pathlib import Path
from textwrap import wrap

from django.conf import settings
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, PatternFill

from .image_utils import insertar_imagen_ajustada, tamano_comun_cuadros_px
from inspecciones.reportes.campaign_excel import agregar_hoja_mediciones_campana, limpiar_textos_historicos_campana


TEMPLATE_PATH = (
    Path(settings.BASE_DIR)
    / "inspecciones"
    / "reportes"
    / "cvb0003"
    / "assets"
    / "reporte_life_shaft_cvb0003.xlsx"
)

SHEET_NAME = "Hoja1"
MEASUREMENT_COLUMNS = ("AB", "AD", "AF", "AH", "AJ", "AL", "AN")
SHAFT_LAYOUTS = (
    {
        "title": 68, "calibration": 71, "data": 73, "average": 77,
        "minimum": 78, "note": 80, "visual": 83, "photos": 84,
        "caption": 103,
    },
    {
        "title": 105, "calibration": 108, "data": 110, "average": 114,
        "minimum": 115, "note": 117, "visual": 120, "photos": 121,
        "caption": 140,
    },
    {
        "title": 142, "calibration": 145, "data": 147, "average": 151,
        "minimum": 152, "note": 154, "visual": 157, "photos": 158,
        "caption": 177,
    },
    {
        "title": 179, "calibration": 182, "data": 184, "average": 188,
        "minimum": 189, "note": 191, "visual": 194, "photos": 195,
        "caption": 215,
    },
    {
        "title": 217, "calibration": 220, "data": 222, "average": 226,
        "minimum": 227, "note": 229, "visual": 232, "photos": 233,
        "caption": 254,
    },
)


def _user_name(user):
    if user is None:
        return ""
    return (user.get_full_name() or user.username or "").strip()


def _field_name(inspeccion, field_name, fallback_user):
    return (getattr(inspeccion, field_name, "") or "").strip() or _user_name(
        fallback_user
    )


def _number(value):
    return float(value) if value is not None else None


def _averages(values):
    clean = [float(value) for value in values if value is not None]
    return sum(clean) / len(clean) if clean else None


def _minimum(values):
    clean = [float(value) for value in values if value is not None]
    return min(clean) if clean else None


def _safe_image_path(photo):
    try:
        path = Path(photo.imagen.path)
    except (AttributeError, NotImplementedError, OSError, ValueError):
        return None
    return path if path.is_file() else None


def _write_wrapped_rows(worksheet, rows, value):
    text = (value or "").strip()
    parts = []
    for line in text.splitlines() or [""]:
        parts.extend(wrap(line, width=150) or [""])
    if len(parts) > len(rows):
        parts = parts[: len(rows) - 1] + [" ".join(parts[len(rows) - 1 :])]
    for index, row in enumerate(rows):
        worksheet[f"I{row}"] = parts[index] if index < len(parts) else None


def _write_header(worksheet, inspeccion):
    worksheet["L1"] = f"REPORTE INSPECCIÓN {inspeccion.codigo_reporte}"
    worksheet["I7"] = inspeccion.get_condicion_general_display().upper()
    worksheet["I7"].fill = PatternFill("solid", fgColor="00B050")
    worksheet["I7"].font = Font(
        name=worksheet["I7"].font.name,
        size=worksheet["I7"].font.sz,
        bold=True,
        color="FFFFFF",
    )
    values = {
        "I10": inspeccion.planta,
        "V10": inspeccion.proceso,
        "I12": inspeccion.get_tipo_display(),
        "V12": inspeccion.faja.tag,
        "I14": inspeccion.etapa,
        "V14": inspeccion.condicion_equipo,
        "I16": inspeccion.fecha_inspeccion,
        "V16": inspeccion.fecha_reporte,
        "I18": _field_name(inspeccion, "inspector_campo_nombre", inspeccion.inspector),
        "V18": _field_name(inspeccion, "supervisor_campo_nombre", inspeccion.supervisor),
        "I20": _field_name(inspeccion, "analista_elabora_nombre", inspeccion.analista),
        "V20": _field_name(inspeccion, "analista_valida_nombre", inspeccion.analista),
        "I23": inspeccion.circunstancias,
        "I26": inspeccion.antecedentes,
        "I28": inspeccion.observaciones,
        "C54": f"ESQUEMA DE UBICACIÓN DE LIFE SHAFT DE LA FAJA {inspeccion.faja.tag}",
    }
    for coordinate, value in values.items():
        worksheet[coordinate] = value if value not in (None, "") else "-"
    worksheet["I16"].number_format = "dd mmmm yyyy"
    worksheet["V16"].number_format = "dd mmmm yyyy"
    _write_wrapped_rows(worksheet, range(45, 52), inspeccion.recomendaciones)


def _clear_original_photos(worksheet):
    worksheet._images = [
        image
        for image in worksheet._images
        if image.anchor._from.row + 1 < 68
    ]


def _photo_ranges(layout, count):
    start_row = layout["photos"]
    end_row = layout["caption"] - 1
    if count <= 1:
        return [f"C{start_row}:AX{end_row}"]
    if count == 2:
        return [
            f"E{start_row}:AA{end_row}",
            f"AD{start_row}:AW{end_row}",
        ]
    if count == 3:
        return [
            f"C{start_row}:T{end_row}",
            f"U{start_row}:AI{end_row}",
            f"AJ{start_row}:AX{end_row}",
        ]

    second_row = start_row + 9
    top_end = second_row - 1
    ranges = [
        f"C{start_row}:T{top_end}",
        f"U{start_row}:AI{top_end}",
        f"AJ{start_row}:AX{top_end}",
    ]
    if count == 4:
        ranges.append(f"U{second_row}:AI{end_row}")
    else:
        ranges.extend(
            [
                f"E{second_row}:AA{end_row}",
                f"AD{second_row}:AW{end_row}",
            ]
        )
    return ranges


def _write_measurement_summary(worksheet, layout, measurements):
    for offset in range(4):
        row = layout["data"] + offset
        measurement = measurements[offset] if offset < len(measurements) else None
        worksheet[f"Y{row}"] = measurement.punto if measurement else None
        for letter, column in zip("abcdefg", MEASUREMENT_COLUMNS):
            worksheet[f"{column}{row}"] = (
                _number(getattr(measurement, letter)) if measurement else None
            )
        worksheet[f"AP{row}"] = _number(measurement.promedio) if measurement else None
        worksheet[f"AS{row}"] = _number(measurement.minimo) if measurement else None

    locations = []
    for measurement in measurements:
        location = (measurement.ubicacion or "").strip()
        if location and location not in locations:
            locations.append(location)
    worksheet[f"V{layout['data']}"] = " / ".join(locations) or "-"

    for letter, column in zip("abcdefg", MEASUREMENT_COLUMNS):
        values = [getattr(measurement, letter) for measurement in measurements]
        worksheet[f"{column}{layout['average']}"] = _averages(values)
        worksheet[f"{column}{layout['minimum']}"] = _minimum(values)

    row_averages = [measurement.promedio for measurement in measurements]
    row_minimums = [measurement.minimo for measurement in measurements]
    worksheet[f"AP{layout['average']}"] = _averages(row_averages)
    worksheet[f"AS{layout['average']}"] = _averages(row_minimums)
    worksheet[f"AP{layout['minimum']}"] = _minimum(row_averages)
    worksheet[f"AS{layout['minimum']}"] = _minimum(row_minimums)

    candidates = []
    for measurement in measurements:
        for letter in "abcdefg":
            value = getattr(measurement, letter)
            if value is not None:
                candidates.append((value, letter.upper(), measurement.punto))
    if candidates:
        value, letter, point = min(candidates, key=lambda item: item[0])
        note = (
            f"El espesor mínimo encontrado es de {value:.2f} mm "
            f"en el punto {letter}, medición {point}."
        )
    else:
        note = "No existen mediciones de espesor registradas."
    worksheet[f"V{layout['note']}"] = note


def _write_shaft(worksheet, layout, block, inspeccion):
    shaft = block["life_shaft"] if block else None
    measurements = block["mediciones"] if block else []
    photos = block["fotografias"] if block else []
    number = shaft.numero if shaft else None
    number_label = f"{number:02d}" if number is not None else "-"

    worksheet[f"C{layout['visual']}"] = (
        f"INSPECCIÓN VISUAL DE LIFE SHAFT #{number_label}"
    )

    calibration = (
        shaft.marca_equipo if shaft else None,
        shaft.tipo_haz if shaft else None,
        shaft.frecuencia_mhz if shaft else None,
        shaft.ancho_banda if shaft else None,
        shaft.amortiguamiento if shaft else None,
        shaft.velocidad_ms if shaft else None,
        shaft.retardo_us if shaft else None,
    )
    if block and block.get("es_campana"):
        for row in range(layout["title"], layout["visual"]):
            worksheet.row_dimensions[row].hidden = True
    else:
        worksheet[f"C{layout['title']}"] = f"MEDICIÓN DE ESPESORES A LIFE SHAFT #{number_label} / {inspeccion.faja.tag}"
        for offset, value in enumerate(calibration):
            worksheet[f"M{layout['calibration'] + offset}"] = value or "-"
        _write_measurement_summary(worksheet, layout, measurements)

    for row in range(layout["photos"], layout["caption"]):
        for cell in worksheet[row][:51]:
            if not isinstance(cell, MergedCell) and cell.value not in (None, ""):
                cell.value = None

    details = []
    if shaft:
        details.extend(
            [
                (
                    f"Nombre: {shaft.nombre or '-'} | TAG: {shaft.tag or '-'} | "
                    f"Ubicación: {shaft.ubicacion or '-'} | "
                    f"Condición: {shaft.get_condicion_display()}"
                ),
                (
                    f"Observación visual: {shaft.observacion_visual or '-'} | "
                    f"Observación de medición: {shaft.observacion_medicion or '-'}"
                ),
                f"Recomendaciones: {shaft.recomendaciones or '-'}",
            ]
        )
    descriptions = [
        (photo.descripcion or "").strip()
        for photo in photos
        if (photo.descripcion or "").strip()
    ]
    if descriptions:
        details[-1] += " | Fotografías: " + " | ".join(descriptions)
    caption_cell = worksheet[f"C{layout['caption']}"]
    caption_cell.value = "\n".join(details) or "Sin datos registrados."
    alignment = copy(caption_cell.alignment)
    alignment.wrap_text = True
    alignment.shrink_to_fit = True
    alignment.vertical = "center"
    caption_cell.alignment = alignment
    worksheet.row_dimensions[layout["caption"]].height = 32
    worksheet.row_dimensions[layout["caption"] + 1].height = 22

    valid_photos = [
        (photo, path)
        for photo in photos
        if (path := _safe_image_path(photo)) is not None
    ][:5]
    ranges = _photo_ranges(layout, len(valid_photos))
    common_width, common_height = tamano_comun_cuadros_px(
        worksheet, ranges
    )
    for (_photo, path), cell_range in zip(valid_photos, ranges):
        insertar_imagen_ajustada(
            worksheet,
            path,
            cell_range,
            ancho_max_px=common_width,
            alto_max_px=common_height,
            margen_px=4,
        )


def generar_excel_life_shaft_cvb0003(inspeccion, blocks):
    workbook = load_workbook(TEMPLATE_PATH)
    worksheet = workbook[SHEET_NAME]
    limpiar_textos_historicos_campana(worksheet)
    _write_header(worksheet, inspeccion)
    _clear_original_photos(worksheet)

    for index, layout in enumerate(SHAFT_LAYOUTS):
        block = blocks[index] if index < len(blocks) else None
        _write_shaft(worksheet, layout, block, inspeccion)
    agregar_hoja_mediciones_campana(
        workbook, blocks, "life_shaft", f"MEDICIÓN DE ESPESORES DEL LIFE SHAFT {inspeccion.faja.tag}",
        config={"layouts": SHAFT_LAYOUTS, "title_column": "C", "calibration_column": "M", "measurement_columns": MEASUREMENT_COLUMNS, "point_column": "Y", "result_column": "AP", "residual_column": "AS", "note_column": "V"},
    )

    workbook.active = workbook.sheetnames.index(SHEET_NAME)
    if hasattr(workbook, "calculation"):
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
