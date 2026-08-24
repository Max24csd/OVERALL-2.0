from __future__ import annotations

import hashlib
import re
from collections import defaultdict
from copy import copy
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill
from openpyxl.utils.cell import range_boundaries

from inspecciones.models import FotoFajaCVB0003, Inspeccion
from inspecciones.reportes.cvb0003.image_utils import (
    insertar_imagen_ajustada,
    insertar_fotografia_ajustada,
    rangos_fotos_adaptativos,
    rangos_fotos_fijos,
)
from inspecciones.reportes.cvb0003.mappings.faja import (
    CARGA_DATA_ROWS,
    CONTINUATION_PAGE_COUNT,
    CONTINUATION_ROWS_PER_PAGE,
    CONTINUATION_SHEET_NAME,
    DIAGRAM_IMAGE_PATH,
    DIAGRAM_IMAGE_RANGE,
    EMPALME_DATA_ROWS,
    EMPALME_PHOTO_SLOTS,
    GENERAL_PHOTO_ROWS,
    GENERAL_PHOTO_SLOTS,
    HEADER_CELLS,
    MASTER_PATH,
    MASTER_SHA256,
    MEASUREMENT_FIELDS,
    RETORNO_DATA_ROWS,
    SHEET_NAME,
    STATIC_IMAGE_MAX_ROW,
    UT_CALIBRATION_BLOCKS,
    VALUE_COLUMNS,
)
from inspecciones.reportes.cvb0003.photo_summary import paginas_fotograficas_faja


def _sha256(path):
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest().upper()


def _assert_master():
    if not MASTER_PATH.is_file():
        raise FileNotFoundError(f"No existe la plantilla maestra: {MASTER_PATH}")
    current_hash = _sha256(MASTER_PATH)
    if current_hash != MASTER_SHA256:
        raise ValueError(
            "La plantilla maestra CVB003 Faja fue modificada. "
            f"Esperado {MASTER_SHA256}; actual {current_hash}."
        )


def _structure_signature(workbook):
    return tuple(
        (
            worksheet.title,
            worksheet.sheet_state,
            worksheet.calculate_dimension(),
            tuple(sorted(str(item) for item in worksheet.merged_cells.ranges)),
            tuple(
                sorted(
                    (key, item.min, item.max, item.width, item.hidden, item.outlineLevel)
                    for key, item in worksheet.column_dimensions.items()
                )
            ),
            tuple(
                sorted(
                    (row, item.height)
                    for row, item in worksheet.row_dimensions.items()
                    if item.height is not None
                )
            ),
            str(worksheet.print_area),
            tuple(item.id for item in worksheet.row_breaks.brk),
            (
                worksheet.page_setup.orientation,
                worksheet.page_setup.paperSize,
                worksheet.page_setup.fitToWidth,
                worksheet.page_setup.fitToHeight,
                worksheet.page_setup.scale,
            ),
            (
                worksheet.page_margins.left,
                worksheet.page_margins.right,
                worksheet.page_margins.top,
                worksheet.page_margins.bottom,
                worksheet.page_margins.header,
                worksheet.page_margins.footer,
            ),
        )
        for worksheet in workbook.worksheets
    )


def _user_name(user):
    if user is None:
        return ""
    return (user.get_full_name() or user.username or "").strip()


def _field_name(inspection, field, fallback):
    return (getattr(inspection, field, "") or "").strip() or _user_name(fallback)


def _condition_text(instance):
    value = getattr(instance, "condicion_general", Inspeccion.Condicion.NORMAL)
    if value == Inspeccion.Condicion.NO_MEDIDO:
        return "TOLERABLE"
    return instance.get_condicion_general_display().upper()


def _style_condition(cell, inspection):
    if inspection.condicion_general != Inspeccion.Condicion.CRITICO:
        return
    cell.fill = PatternFill(fill_type="solid", fgColor="FFFF0000")
    font = copy(cell.font)
    font.color = "FF000000"
    font.bold = True
    cell.font = font


def _number(value):
    return float(value) if value is not None else None


def _values(measurement):
    return [getattr(measurement, field) for field in MEASUREMENT_FIELDS]


def _minimum_detail(measurements):
    candidates = []
    for measurement in measurements:
        for field in MEASUREMENT_FIELDS:
            value = getattr(measurement, field)
            if value is not None:
                candidates.append((value, measurement, field.upper()))
    return min(candidates, key=lambda item: item[0]) if candidates else None


def _automatic_text(empalmes, carga, retorno):
    lines = []
    for label, measurements in (
        ("EMPALMES", empalmes),
        ("TRAMOS DE CARGA", carga),
        ("TRAMOS DE RETORNO", retorno),
    ):
        detail = _minimum_detail(measurements)
        if detail is None:
            lines.append(f"{label}: sin mediciones registradas.")
            continue
        value, measurement, column = detail
        identity = (
            getattr(measurement, "empalme", "")
            or getattr(measurement, "tramo", "")
            or "componente"
        )
        point = (
            getattr(measurement, "posicion", "")
            or getattr(measurement, "medicion", "")
        )
        lines.append(
            f"{label}: espesor mínimo {value:.2f} mm en {identity}, "
            f"{point}, punto {column}."
        )
    return "\n".join(lines)


def _photo_summary(inspection, measurement_text):
    visual = (inspection.observaciones or "").strip() or "-"
    measurement = (measurement_text or "").strip().replace("\n", " ") or "-"
    recommendations = (inspection.recomendaciones or "").strip() or "-"
    return (
        f"Condición: {_condition_text(inspection)} | "
        f"Observación visual: {visual} | "
        f"Observación de medición: {measurement} | "
        f"Recomendaciones: {recommendations}"
    )


def _write_header(worksheet, inspection, empalmes, carga, retorno):
    automatic = _automatic_text(empalmes, carga, retorno)
    observations = "\n".join(
        value for value in ((inspection.observaciones or "").strip(), automatic)
        if value
    )
    recommendations = (inspection.recomendaciones or "").strip()
    if not recommendations:
        recommendations = (
            f"Condición general {_condition_text(inspection)}. "
            "Continuar con el plan de inspecciones programadas."
        )
    values = {
        HEADER_CELLS["title"]: f"REPORTE DE INSPECCIÓN {inspection.codigo_reporte}",
        HEADER_CELLS["condition"]: _condition_text(inspection),
        HEADER_CELLS["plant"]: inspection.planta or "-",
        HEADER_CELLS["process"]: inspection.proceso or "-",
        HEADER_CELLS["equipment"]: inspection.faja.nombre or inspection.get_tipo_display(),
        HEADER_CELLS["tag"]: inspection.faja.tag,
        HEADER_CELLS["stage"]: inspection.etapa or "-",
        HEADER_CELLS["equipment_condition"]: inspection.condicion_equipo or "-",
        HEADER_CELLS["inspection_date"]: inspection.fecha_inspeccion,
        HEADER_CELLS["report_date"]: inspection.fecha_reporte,
        HEADER_CELLS["inspector"]: _field_name(
            inspection, "inspector_campo_nombre", inspection.inspector
        ),
        HEADER_CELLS["supervisor"]: _field_name(
            inspection, "supervisor_campo_nombre", inspection.supervisor
        ),
        HEADER_CELLS["author"]: _field_name(
            inspection, "analista_elabora_nombre", inspection.analista
        ),
        HEADER_CELLS["validator"]: _field_name(
            inspection, "analista_valida_nombre", inspection.analista
        ),
        HEADER_CELLS["circumstances"]: inspection.circunstancias or "-",
        HEADER_CELLS["background"]: inspection.antecedentes or "-",
        HEADER_CELLS["observations"]: observations or "-",
        HEADER_CELLS["recommendations"]: recommendations,
    }
    for coordinate, value in values.items():
        worksheet[coordinate] = value if value not in (None, "") else "-"
    worksheet[HEADER_CELLS["inspection_date"]].number_format = "dd/mm/yyyy"
    worksheet[HEADER_CELLS["report_date"]].number_format = "dd/mm/yyyy"
    _style_condition(worksheet[HEADER_CELLS["condition"]], inspection)


def _write_ut_calibrations(worksheet, inspection):
    calibrations = {
        item.numero: item
        for item in inspection.calibraciones_ut_faja_cvb0003.order_by("numero")
    }
    for block in UT_CALIBRATION_BLOCKS:
        calibration = calibrations.get(block["number"])
        if calibration is None:
            continue
        for field, coordinate in block.items():
            if field == "number":
                continue
            worksheet[coordinate] = getattr(calibration, field) or "-"


def _write_if_cell(worksheet, coordinate, value):
    cell = worksheet[coordinate]
    if not isinstance(cell, MergedCell):
        cell.value = value


def _write_empalmes(worksheet, measurements):
    rows = EMPALME_DATA_ROWS
    for index, row in enumerate(rows):
        measurement = measurements[index] if index < len(measurements) else None
        if measurement is None:
            for column in ("G",) + VALUE_COLUMNS:
                _write_if_cell(worksheet, f"{column}{row}", None)
            continue
        for column, value in (
            ("C", measurement.zona),
            ("D", measurement.empalme),
            ("E", measurement.bastidor_lado),
            ("F", measurement.posicion),
            ("G", _number(measurement.espesor_nominal)),
        ):
            _write_if_cell(worksheet, f"{column}{row}", value)
        for column, value in zip(VALUE_COLUMNS, _values(measurement)):
            worksheet[f"{column}{row}"] = _number(value)


def _write_tramos(worksheet, measurements, rows):
    for index, row in enumerate(rows):
        measurement = measurements[index] if index < len(measurements) else None
        if measurement is None:
            for column in ("G",) + VALUE_COLUMNS:
                _write_if_cell(worksheet, f"{column}{row}", None)
            continue
        for column, value in (
            ("D", measurement.tramo),
            ("E", measurement.medicion),
            ("F", measurement.bastidor),
            ("G", _number(measurement.espesor_nominal)),
        ):
            _write_if_cell(worksheet, f"{column}{row}", value)
        for column, value in zip(VALUE_COLUMNS, _values(measurement)):
            worksheet[f"{column}{row}"] = _number(value)


def _format_percentages(worksheet):
    for row in EMPALME_DATA_ROWS + CARGA_DATA_ROWS + RETORNO_DATA_ROWS:
        for column in ("R", "S"):
            cell = worksheet[f"{column}{row}"]
            cell.number_format = "0.0%"
            alignment = copy(cell.alignment)
            alignment.shrink_to_fit = True
            cell.alignment = alignment


def _safe_image_path(photo):
    try:
        path = Path(photo.imagen.path)
    except (AttributeError, NotImplementedError, OSError, ValueError):
        return None
    return path if path.is_file() else None


def _anchor_from_photo(photo):
    description = photo.descripcion or ""
    match = re.search(r"anclaje\s+([A-Z]+\d+)", description, flags=re.IGNORECASE)
    if match:
        return match.group(1).upper()
    filename = Path(photo.imagen.name or "").name
    match = re.search(r"_(?:empalmes|carga|retorno)_([a-z]+\d+)_", filename)
    return match.group(1).upper() if match else None


def _description(photo, prefix=""):
    text = re.sub(
        r"Fotografía histórica del reporte final CVB003\s*", "",
        (photo.descripcion or "").strip(), flags=re.IGNORECASE,
    )
    text = re.sub(r"\(?anclaje\s+[A-Z]+\d+\)?\.?", "", text, flags=re.IGNORECASE)
    parts = [
        value for value in (
            prefix,
            (photo.codigo_dano or "").strip(),
            text.strip(" .-"),
        ) if value
    ]
    return " - ".join(parts) or prefix or "Fotografía de inspección visual"


def _assign_slots(photos, slots):
    by_anchor = {cell_range.split(":", 1)[0].upper(): index for index, (cell_range, _caption) in enumerate(slots)}
    assigned = []
    used = set()
    pending = []
    for photo in photos:
        path = _safe_image_path(photo)
        if path is None:
            continue
        anchor = _anchor_from_photo(photo)
        index = by_anchor.get(anchor) if anchor else None
        if index is None or index in used:
            pending.append((photo, path))
            continue
        used.add(index)
        assigned.append((index, photo, path))
    available = (index for index in range(len(slots)) if index not in used)
    for item in pending:
        try:
            index = next(available)
        except StopIteration as exc:
            raise ValueError(
                f"La plantilla admite {len(slots)} fotografías en esta sección."
            ) from exc
        used.add(index)
        assigned.append((index, *item))
    return sorted(assigned)


def _clear_variable_images(worksheet):
    worksheet._images = [
        image for image in worksheet._images
        if image.anchor._from.row + 1 <= STATIC_IMAGE_MAX_ROW
    ]


def _replace_diagram_image(worksheet):
    """Reemplaza solo la imagen del diagrama por la usada en la vista web."""
    if not DIAGRAM_IMAGE_PATH.is_file():
        raise FileNotFoundError(
            f"No existe el diagrama final CVB003 Faja: {DIAGRAM_IMAGE_PATH}"
        )

    min_col, min_row, max_col, max_row = range_boundaries(DIAGRAM_IMAGE_RANGE)
    preserved_images = []
    for image in worksheet._images:
        try:
            anchor_col = image.anchor._from.col + 1
            anchor_row = image.anchor._from.row + 1
        except AttributeError:
            preserved_images.append(image)
            continue
        if min_col <= anchor_col <= max_col and min_row <= anchor_row <= max_row:
            continue
        preserved_images.append(image)
    worksheet._images = preserved_images

    if not insertar_imagen_ajustada(
        worksheet,
        DIAGRAM_IMAGE_PATH,
        DIAGRAM_IMAGE_RANGE,
        margen_px=4,
    ):
        raise ValueError("No fue posible insertar el diagrama final CVB003 Faja.")


def _write_empalme_photos(worksheet, photos, summary):
    for _range, caption in EMPALME_PHOTO_SLOTS:
        worksheet[caption] = None
    assigned = _assign_slots(photos, EMPALME_PHOTO_SLOTS)
    groups = defaultdict(list)
    for index, photo, path in assigned:
        cell_range, caption = EMPALME_PHOTO_SLOTS[index]
        groups[caption].append((cell_range, photo, path))
    for caption, items in groups.items():
        source_ranges = [
            cell_range
            for cell_range, source_caption in EMPALME_PHOTO_SLOTS
            if source_caption == caption
        ]
        bounds = [range_boundaries(cell_range) for cell_range in source_ranges]
        start_row = min(item[1] for item in bounds)
        caption_row = int(re.search(r"\d+", caption).group())
        end_row = min(max(item[3] for item in bounds), caption_row - 1)
        slots = rangos_fotos_adaptativos(
            start_row,
            end_row,
            len(items),
            columna_inicio="C",
            columna_fin="U",
        )
        for (_cell_range, _photo, path), slot in zip(items, slots):
            insertar_fotografia_ajustada(worksheet, path, slot)
        worksheet[caption] = summary


def _write_general_photos(worksheet, photos, summary):
    for cell_range, description_cell in GENERAL_PHOTO_SLOTS:
        _write_if_cell(worksheet, cell_range.split(":", 1)[0], None)
        worksheet[description_cell] = None
    assigned = _assign_slots(photos, GENERAL_PHOTO_SLOTS)
    bands = [
        (start, description - 1, f"C{description}")
        for start, description in GENERAL_PHOTO_ROWS
    ] + [
        (1524, 1545, "C1546"),
        (1549, 1571, "C1572"),
    ]
    slots = []
    remaining = len(assigned)
    used_caption = None
    for start, end, caption in bands:
        if remaining <= 0:
            break
        count = min(3, remaining)
        slots.extend(
            rangos_fotos_adaptativos(
                start,
                end,
                count,
                columna_inicio="C",
                columna_fin="U",
            )
        )
        used_caption = caption
        remaining -= count
    for (_index, _photo, path), slot in zip(assigned, slots):
        insertar_fotografia_ajustada(worksheet, path, slot)
    if used_caption:
        worksheet[used_caption] = summary


def _show_main_photo_areas(worksheet):
    """
    Mantiene visibles las zonas fotográficas que ya existen en el master
    de Faja CVB003.

    - 315:886   -> fotografías debajo de la tabla de empalmes.
    - 1129:1574 -> registro fotográfico debajo de las tablas de carga/retorno.

    No crea filas, no cambia anchos de columnas y no mueve las tablas.
    """
    for start, end in ((315, 886), (1129, 1574)):
        for row in range(start, end + 1):
            worksheet.row_dimensions[row].hidden = False


def _hide_legacy_photo_areas(worksheet):
    # Se conserva por compatibilidad, pero ya NO se usa en la exportación final.
    for start, end in ((315, 886), (1129, 1574)):
        for row in range(start, end + 1):
            worksheet.row_dimensions[row].hidden = True
            for cell in worksheet[row]:
                if cell.data_type == "e":
                    cell.value = None


def _valid_unique_photos(photos, seen):
    valid = []
    for photo in photos:
        path = _safe_image_path(photo)
        if path is None:
            continue
        key = str(path.resolve()).casefold()
        if key in seen:
            continue
        seen.add(key)
        valid.append(photo)
    return valid


def _reset_continuation_sheet(worksheet):
    worksheet._images = []
    for page_index in range(CONTINUATION_PAGE_COUNT):
        start = 1 + (page_index * CONTINUATION_ROWS_PER_PAGE)
        end = start + CONTINUATION_ROWS_PER_PAGE - 1
        worksheet[f"C{start}"] = None
        worksheet[f"C{start + 37}"] = None
        for row in range(start, end + 1):
            worksheet.row_dimensions[row].hidden = True


def _write_paginated_photos(workbook, inspection, photos, summary):
    worksheet = workbook[CONTINUATION_SHEET_NAME]
    _reset_continuation_sheet(worksheet)
    seen = set()
    pages = []
    for section, label in (
        (FotoFajaCVB0003.Seccion.EMPALMES, "EMPALMES"),
        (FotoFajaCVB0003.Seccion.CARGA, "CARGA"),
        (FotoFajaCVB0003.Seccion.RETORNO, "RETORNO"),
    ):
        section_photos = _valid_unique_photos(
            [photo for photo in photos if photo.seccion == section],
            seen,
        )
        pages.extend(paginas_fotograficas_faja(section_photos, label))
    if len(pages) > CONTINUATION_PAGE_COUNT:
        raise ValueError(
            f"La plantilla admite {CONTINUATION_PAGE_COUNT * 9} fotografías "
            "paginadas en total para Faja CVB003."
        )

    for page_index, page in enumerate(pages):
        start = 1 + (page_index * CONTINUATION_ROWS_PER_PAGE)
        end = start + CONTINUATION_ROWS_PER_PAGE - 1
        for row in range(start, end + 1):
            worksheet.row_dimensions[row].hidden = False
        worksheet[f"C{start}"] = page["titulo"]
        slots = rangos_fotos_fijos(
            start + 1,
            start + 36,
            len(page["fotos"]),
            columna_inicio="C",
            columna_fin="U",
            filas_maximas=3,
            worksheet=worksheet,
        )
        for photo, slot in zip(page["fotos"], slots):
            insertar_fotografia_ajustada(
                worksheet,
                _safe_image_path(photo),
                slot,
            )
        if page["mostrar_resumen"]:
            worksheet[f"C{start + 37}"] = summary


def generar_excel_faja_cvb0003_master(inspection):
    """Completa una copia del master aprobado sin modificar su estructura."""
    _assert_master()
    workbook = load_workbook(BytesIO(MASTER_PATH.read_bytes()))
    structure_before = _structure_signature(workbook)
    worksheet = workbook[SHEET_NAME]
    empalmes = list(inspection.empalmes_cvb0003.order_by("orden", "id"))
    carga = list(
        inspection.tramos_cvb0003.filter(tipo="CARGA").order_by("orden", "id")
    )
    retorno = list(
        inspection.tramos_cvb0003.filter(tipo="RETORNO").order_by("orden", "id")
    )
    photos = list(inspection.fotografias_cvb0003.order_by("seccion", "creada_en", "id"))

    _write_header(worksheet, inspection, empalmes, carga, retorno)
    _write_ut_calibrations(worksheet, inspection)
    _write_empalmes(worksheet, empalmes)
    _write_tramos(worksheet, carga, CARGA_DATA_ROWS)
    _write_tramos(worksheet, retorno, RETORNO_DATA_ROWS)
    _format_percentages(worksheet)
    _clear_variable_images(worksheet)
    _replace_diagram_image(worksheet)
    summary = _photo_summary(
        inspection,
        _automatic_text(empalmes, carga, retorno),
    )
    # -----------------------------------------------------
    # FOTOGRAFÍAS EN LA MISMA HOJA DEL REPORTE
    # -----------------------------------------------------
    # El master ya contiene zonas fotográficas debajo de las tablas.
    # No deben enviarse a la hoja "FOTOS CONTINUACION".
    _show_main_photo_areas(worksheet)

    seen = set()

    fotos_empalmes = _valid_unique_photos(
        [
            photo
            for photo in photos
            if photo.seccion == FotoFajaCVB0003.Seccion.EMPALMES
        ],
        seen,
    )

    fotos_carga_retorno = _valid_unique_photos(
        [
            photo
            for photo in photos
            if photo.seccion in (
                FotoFajaCVB0003.Seccion.CARGA,
                FotoFajaCVB0003.Seccion.RETORNO,
            )
        ],
        seen,
    )

    # Empalmes: inmediatamente debajo de la tabla de empalmes.
    _write_empalme_photos(
        worksheet,
        fotos_empalmes,
        summary,
    )

    # Carga + Retorno: en el bloque fotográfico de la hoja principal,
    # debajo de las tablas de medición, respetando el orden guardado.
    _write_general_photos(
        worksheet,
        fotos_carga_retorno,
        summary,
    )

    # La hoja de continuación queda vacía; ya no recibe fotografías.
    if CONTINUATION_SHEET_NAME in workbook.sheetnames:
        _reset_continuation_sheet(
            workbook[CONTINUATION_SHEET_NAME]
        )

    if _structure_signature(workbook) != structure_before:
        raise ValueError(
            "La exportación intentó modificar hojas, filas, columnas, merges o impresión."
        )
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
