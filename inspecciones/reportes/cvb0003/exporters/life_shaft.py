from __future__ import annotations

import hashlib
from copy import copy
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill

from inspecciones.models import Inspeccion, TipoMedicionComponente
from inspecciones.reportes.cvb0003.image_utils import (
    bandas_filas_fotos,
    insertar_imagen_ajustada,
    insertar_fotografia_ajustada,
    rangos_fotos_fijos,
)
from inspecciones.reportes.cvb0003.mappings.life_shaft import (
    HEADER_CELLS,
    LIFE_SHAFT_BLOCKS,
    MASTER_PATH,
    MASTER_SHA256,
    MEASUREMENT_FIELDS,
    SHEET_NAME,
    STATIC_IMAGE_MAX_ROW,
    TECHNICAL_LAYOUT,
)


PHASE_NORMAL = "NORMAL"
PHASE_START = "INICIO DE CAMPAÑA"
PHASE_END = "FIN DE CAMPAÑA"
MAX_PHOTOS = 12


def _sha256(path):
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest().upper()


def _assert_master():
    if not MASTER_PATH.is_file():
        raise FileNotFoundError(f"No existe la plantilla maestra: {MASTER_PATH}")
    current_hash = _sha256(MASTER_PATH)
    if current_hash != MASTER_SHA256:
        raise ValueError(
            "La plantilla maestra CVB003 Life Shaft fue modificada. "
            f"Esperado {MASTER_SHA256}; actual {current_hash}."
        )


def _structure_signature(workbook):
    return tuple(
        (
            worksheet.title,
            worksheet.sheet_state,
            worksheet.calculate_dimension(),
            tuple(sorted(str(item) for item in worksheet.merged_cells.ranges)),
            tuple(
                sorted(
                    (key, item.min, item.max, item.width, item.hidden, item.outlineLevel)
                    for key, item in worksheet.column_dimensions.items()
                )
            ),
            tuple(
                sorted(
                    (row, item.height)
                    for row, item in worksheet.row_dimensions.items()
                    if item.height is not None
                )
            ),
            str(worksheet.print_area),
            tuple(item.id for item in worksheet.row_breaks.brk),
            (
                worksheet.page_setup.orientation,
                worksheet.page_setup.paperSize,
                worksheet.page_setup.fitToWidth,
                worksheet.page_setup.fitToHeight,
                worksheet.page_setup.scale,
            ),
            (
                worksheet.page_margins.left,
                worksheet.page_margins.right,
                worksheet.page_margins.top,
                worksheet.page_margins.bottom,
                worksheet.page_margins.header,
                worksheet.page_margins.footer,
            ),
        )
        for worksheet in workbook.worksheets
    )


def _user_name(user):
    if user is None:
        return ""
    return (user.get_full_name() or user.username or "").strip()


def _field_name(inspection, field, fallback):
    return (getattr(inspection, field, "") or "").strip() or _user_name(fallback)


def _condition_text(instance):
    value = getattr(instance, "condicion_general", None)
    if value is None:
        value = getattr(instance, "condicion", None)
    if value == Inspeccion.Condicion.NO_MEDIDO:
        return "TOLERABLE"
    getter = (
        getattr(instance, "get_condicion_general_display", None)
        or getattr(instance, "get_condicion_display", None)
    )
    return (getter() if getter else value or "NORMAL").upper()


def _style_condition(cell, inspection):
    if inspection.condicion_general != Inspeccion.Condicion.CRITICO:
        return
    cell.fill = PatternFill(fill_type="solid", fgColor="FFFF0000")
    font = copy(cell.font)
    font.color = "FF000000"
    font.bold = True
    cell.font = font


def _number(value):
    return float(value) if value is not None else None


def _values(measurement):
    return [getattr(measurement, field) for field in MEASUREMENT_FIELDS]


def _average(values):
    available = [float(value) for value in values if value is not None]
    return round(sum(available) / len(available), 2) if available else None


def _minimum(values):
    available = [float(value) for value in values if value is not None]
    return min(available) if available else None


def _minimum_detail(measurements):
    candidates = []
    for measurement in measurements:
        for field in MEASUREMENT_FIELDS:
            value = getattr(measurement, field)
            if value is not None:
                candidates.append((value, field.upper(), measurement.punto))
    return min(candidates, key=lambda item: item[0]) if candidates else None


def _phase_note(measurements, phase, shaft):
    detail = _minimum_detail(measurements)
    if detail:
        value, column, point = detail
        parts = [
            f"{phase}: espesor mínimo {value:.2f} mm en el punto {column}, "
            f"medición radial {point}."
        ]
    else:
        parts = [f"{phase}: sin mediciones registradas."]
    manual = []
    for measurement in measurements:
        text = (measurement.observacion or "").strip()
        if text and text not in manual:
            manual.append(text)
    shaft_text = (shaft.observacion_medicion or "").strip()
    if shaft_text and shaft_text not in manual:
        manual.append(shaft_text)
    if manual:
        parts.append(" ".join(manual))
    return " ".join(parts)


def _write_header(worksheet, inspection, blocks):
    values = {
        HEADER_CELLS["title"]: f"REPORTE INSPECCIÓN {inspection.codigo_reporte}",
        HEADER_CELLS["condition"]: _condition_text(inspection),
        HEADER_CELLS["plant"]: inspection.planta or "-",
        HEADER_CELLS["process"]: inspection.proceso or "-",
        HEADER_CELLS["equipment"]: inspection.get_tipo_display(),
        HEADER_CELLS["tag"]: inspection.faja.tag,
        HEADER_CELLS["stage"]: inspection.etapa or "-",
        HEADER_CELLS["equipment_condition"]: inspection.condicion_equipo or "-",
        HEADER_CELLS["inspection_date"]: inspection.fecha_inspeccion,
        HEADER_CELLS["report_date"]: inspection.fecha_reporte,
        HEADER_CELLS["inspector"]: _field_name(
            inspection, "inspector_campo_nombre", inspection.inspector
        ),
        HEADER_CELLS["supervisor"]: _field_name(
            inspection, "supervisor_campo_nombre", inspection.supervisor
        ),
        HEADER_CELLS["author"]: _field_name(
            inspection, "analista_elabora_nombre", inspection.analista
        ),
        HEADER_CELLS["validator"]: _field_name(
            inspection, "analista_valida_nombre", inspection.analista
        ),
        HEADER_CELLS["circumstances"]: inspection.circunstancias or "-",
        HEADER_CELLS["background"]: inspection.antecedentes or "-",
        HEADER_CELLS["diagram_title"]: (
            f"ESQUEMA DE UBICACIÓN DE LIFE SHAFT DE LA FAJA {inspection.faja.tag}"
        ),
    }
    for coordinate, value in values.items():
        worksheet[coordinate] = value if value not in (None, "") else "-"
    worksheet[HEADER_CELLS["inspection_date"]].number_format = "dd/mm/yyyy"
    worksheet[HEADER_CELLS["report_date"]].number_format = "dd/mm/yyyy"
    _style_condition(worksheet[HEADER_CELLS["condition"]], inspection)

    observations = []
    if (inspection.observaciones or "").strip():
        observations.append(inspection.observaciones.strip())
    recommendations = []
    for block in blocks:
        shaft = block["life_shaft"]
        phases = (
            (
                (PHASE_START, list(block.get("mediciones_inicio", []))),
                (PHASE_END, list(block.get("mediciones_fin", []))),
            )
            if shaft.tipo_medicion == TipoMedicionComponente.CAMPANA
            else ((PHASE_NORMAL, list(block.get("mediciones", []))),)
        )
        for phase, measurements in phases:
            observations.append(
                f"LIFE SHAFT #{shaft.numero:02d}: {_phase_note(measurements, phase, shaft)}"
            )
        line = f"LIFE SHAFT #{shaft.numero:02d} condición {_condition_text(shaft)}"
        if (shaft.recomendaciones or "").strip():
            line += f". {shaft.recomendaciones.strip()}"
        recommendations.append(line)
    worksheet[HEADER_CELLS["observations"]] = "\n".join(observations) or "-"
    if (inspection.recomendaciones or "").strip():
        recommendations.append(inspection.recomendaciones.strip())
    rows = HEADER_CELLS["recommendation_rows"]
    if len(recommendations) > len(rows):
        recommendations = recommendations[: len(rows) - 1] + [
            " ".join(recommendations[len(rows) - 1 :])
        ]
    for index, coordinate in enumerate(rows):
        worksheet[coordinate] = recommendations[index] if index < len(recommendations) else None


def _set_hidden(worksheet, rows, hidden):
    for row in range(rows[0], rows[1] + 1):
        worksheet.row_dimensions[row].hidden = hidden


def _write_measurements(worksheet, start, measurements):
    mapping = TECHNICAL_LAYOUT
    count = mapping["measurement_count"]
    measurements = list(measurements)[:count]
    data_start = start + mapping["data_offset"]
    for offset in range(count):
        row = data_start + offset
        measurement = measurements[offset] if offset < len(measurements) else None
        worksheet[f"{mapping['point_column']}{row}"] = (
            measurement.punto if measurement else offset + 1
        )
        values = _values(measurement) if measurement else [None] * 7
        for column, value in zip(mapping["measurement_columns"], values):
            worksheet[f"{column}{row}"] = _number(value)
        worksheet[f"{mapping['average_column']}{row}"] = _average(values)
        worksheet[f"{mapping['minimum_column']}{row}"] = _minimum(values)

    average_row = data_start + count
    minimum_row = average_row + 1
    for field, column in zip(MEASUREMENT_FIELDS, mapping["measurement_columns"]):
        values = [getattr(item, field) for item in measurements]
        worksheet[f"{column}{average_row}"] = _average(values)
        worksheet[f"{column}{minimum_row}"] = _minimum(values)
    all_values = [value for item in measurements for value in _values(item)]
    worksheet[f"{mapping['average_column']}{average_row}"] = _average(all_values)
    worksheet[f"{mapping['minimum_column']}{average_row}"] = _minimum(all_values)
    row_averages = [_average(_values(item)) for item in measurements]
    row_minimums = [_minimum(_values(item)) for item in measurements]
    worksheet[f"{mapping['average_column']}{minimum_row}"] = _average(row_averages)
    worksheet[f"{mapping['minimum_column']}{minimum_row}"] = _minimum(row_minimums)


def _write_phase(worksheet, mapping, slot, shaft, phase, measurements, tag):
    start, _end = mapping[slot]
    technical = TECHNICAL_LAYOUT
    worksheet[f"{technical['title_column']}{start}"] = (
        f"MEDICIÓN DE ESPESORES A LIFE SHAFT #{shaft.numero:02d} / {tag} {phase}"
    )
    worksheet[
        f"{technical['inner_title_column']}"
        f"{start + technical['inner_title_offset']}"
    ] = f"MEDICIÓN DE ESPESORES DE LIFE SHAFT {shaft.numero} {phase}"
    calibration = (
        shaft.marca_equipo,
        shaft.tipo_haz,
        shaft.frecuencia_mhz,
        shaft.ancho_banda,
        shaft.amortiguamiento,
        shaft.velocidad_ms,
        shaft.retardo_us,
    )
    row = start + technical["calibration_offset"]
    for offset, value in enumerate(calibration):
        worksheet[f"{technical['calibration_column']}{row + offset}"] = value or "-"
    _write_measurements(worksheet, start, measurements)
    worksheet[
        f"{technical['note_column']}{start + technical['note_offset']}"
    ] = _phase_note(list(measurements), phase, shaft)


def _safe_image_path(photo):
    try:
        path = Path(photo.imagen.path)
    except (AttributeError, NotImplementedError, OSError, ValueError):
        return None
    return path if path.is_file() else None


def _photo_slots(worksheet, mapping, count):
    if count <= 0:
        return []
    if count > MAX_PHOTOS:
        raise ValueError(f"La plantilla admite hasta {MAX_PHOTOS} fotos por Life Shaft.")
    start, end = mapping["photo_rows"]
    return rangos_fotos_fijos(start, end, count, worksheet=worksheet)


def _set_photo_rows_visibility(worksheet, mapping, count):
    used_bands = (count + 2) // 3
    for index, (start, end) in enumerate(
        bandas_filas_fotos(*mapping["photo_rows"], 4)
    ):
        hidden = index >= used_bands
        for row in range(start, end + 1):
            worksheet.row_dimensions[row].hidden = hidden


def _clear_variable_images(worksheet):
    worksheet._images = [
        image for image in worksheet._images
        if image.anchor._from.row + 1 <= STATIC_IMAGE_MAX_ROW
    ]


def _clear_photo_cells(worksheet, mapping):
    start, end = mapping["photo_rows"]
    for row in range(start, end + 1):
        for column in range(3, 51):
            cell = worksheet.cell(row, column)
            if not isinstance(cell, MergedCell):
                cell.value = None


def _write_photos(worksheet, mapping, shaft, photos, seen):
    valid = []
    for photo in photos:
        path = _safe_image_path(photo)
        if path is None:
            continue
        key = str(path.resolve()).casefold()
        if key in seen:
            continue
        seen.add(key)
        valid.append((photo, path))
    slots = _photo_slots(worksheet, mapping, len(valid))
    _set_photo_rows_visibility(worksheet, mapping, len(valid))
    for (_photo, path), slot in zip(valid, slots):
        insertar_fotografia_ajustada(worksheet, path, slot)

    caption = [f"Condición: {_condition_text(shaft)}"]
    for label, value in (
        ("Observación visual", shaft.observacion_visual),
        ("Observación de medición", shaft.observacion_medicion),
        ("Recomendaciones", shaft.recomendaciones),
    ):
        caption.append(f"{label}: {(value or '').strip() or '-'}")
    worksheet[mapping["caption"]] = " | ".join(caption)


def _write_shaft(worksheet, block, tag, seen):
    shaft = block["life_shaft"]
    mapping = LIFE_SHAFT_BLOCKS[shaft.numero]
    if shaft.tipo_medicion == TipoMedicionComponente.CAMPANA:
        _set_hidden(worksheet, mapping["slot_a"], False)
        _set_hidden(worksheet, mapping["slot_b"], False)
        _write_phase(
            worksheet, mapping, "slot_a", shaft, PHASE_START,
            list(block.get("mediciones_inicio", [])), tag,
        )
        _write_phase(
            worksheet, mapping, "slot_b", shaft, PHASE_END,
            list(block.get("mediciones_fin", [])), tag,
        )
    else:
        _set_hidden(worksheet, mapping["slot_a"], False)
        _set_hidden(worksheet, mapping["slot_b"], True)
        _write_phase(
            worksheet, mapping, "slot_a", shaft, PHASE_NORMAL,
            list(block.get("mediciones", [])), tag,
        )
        _write_phase(worksheet, mapping, "slot_b", shaft, PHASE_END, [], tag)
    worksheet[f"C{mapping['visual']}"] = (
        f"INSPECCIÓN VISUAL DE LIFE SHAFT #{shaft.numero:02d}"
    )
    _clear_photo_cells(worksheet, mapping)
    _write_photos(
        worksheet, mapping, shaft,
        list(block.get("fotografias", [])), seen,
    )


def generar_excel_life_shaft_cvb0003_master(inspection, blocks):
    """Completa una copia del master sin cambiar su estructura fija."""
    _assert_master()
    workbook = load_workbook(BytesIO(MASTER_PATH.read_bytes()))
    structure_before = _structure_signature(workbook)
    worksheet = workbook[SHEET_NAME]
    blocks = sorted(blocks, key=lambda item: item["life_shaft"].numero)
    _write_header(worksheet, inspection, blocks)
    _clear_variable_images(worksheet)
    seen = set()
    by_number = {block["life_shaft"].numero: block for block in blocks}
    for number in range(1, 6):
        block = by_number.get(number)
        if block is not None:
            _write_shaft(worksheet, block, inspection.faja.tag, seen)
    if _structure_signature(workbook) != structure_before:
        raise ValueError(
            "La exportación intentó modificar filas, columnas, merges o impresión."
        )
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
