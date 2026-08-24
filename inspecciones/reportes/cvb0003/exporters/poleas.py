from __future__ import annotations

import hashlib
from copy import copy
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from inspecciones.models import FaseCampana, Inspeccion, TipoMedicionComponente
from inspecciones.reportes.cvb0003.image_utils import (
    bandas_filas_fotos,
    insertar_imagen_ajustada,
    insertar_fotografia_ajustada,
    rangos_fotos_fijos,
)
from inspecciones.reportes.cvb0003.mappings.poleas import (
    HEADER_CELLS,
    HIDDEN_SHEET_POLEA_8,
    MASTER_PATH,
    MASTER_SHA256,
    MEASUREMENT_FIELDS,
    POLEA_BLOCKS,
    SHEET_NAME,
    STATIC_IMAGE_MAX_ROW,
)


PHASE_NORMAL = "NORMAL"
PHASE_START = "INICIO DE CAMPAÑA"
PHASE_END = "FIN DE CAMPAÑA"
MAX_PHOTOS_PER_POLEA = 12


def _sha256(path):
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest().upper()


def _assert_master():
    if not MASTER_PATH.is_file():
        raise FileNotFoundError(f"No existe la plantilla maestra: {MASTER_PATH}")
    current_hash = _sha256(MASTER_PATH)
    if current_hash != MASTER_SHA256:
        raise ValueError(
            "La plantilla maestra CVB003 Poleas fue modificada. "
            f"Esperado {MASTER_SHA256}; actual {current_hash}."
        )


def _structure_signature(workbook):
    sheets = []
    for worksheet in workbook.worksheets:
        sheets.append(
            {
                "name": worksheet.title,
                "state": worksheet.sheet_state,
                "dimension": worksheet.calculate_dimension(),
                "merges": tuple(sorted(str(item) for item in worksheet.merged_cells.ranges)),
                "columns": tuple(
                    sorted(
                        (
                            key,
                            dimension.min,
                            dimension.max,
                            dimension.width,
                            dimension.hidden,
                            dimension.outlineLevel,
                        )
                        for key, dimension in worksheet.column_dimensions.items()
                    )
                ),
                "row_heights": tuple(
                    sorted(
                        (row, dimension.height)
                        for row, dimension in worksheet.row_dimensions.items()
                        if dimension.height is not None
                    )
                ),
                "print_area": str(worksheet.print_area),
                "print_title_rows": worksheet.print_title_rows,
                "row_breaks": tuple(item.id for item in worksheet.row_breaks.brk),
                "page_setup": (
                    worksheet.page_setup.orientation,
                    worksheet.page_setup.paperSize,
                    worksheet.page_setup.fitToWidth,
                    worksheet.page_setup.fitToHeight,
                    worksheet.page_setup.scale,
                ),
                "margins": (
                    worksheet.page_margins.left,
                    worksheet.page_margins.right,
                    worksheet.page_margins.top,
                    worksheet.page_margins.bottom,
                    worksheet.page_margins.header,
                    worksheet.page_margins.footer,
                ),
            }
        )
    return tuple(sheets)


def _user_name(user):
    if user is None:
        return ""
    return (user.get_full_name() or user.username or "").strip()


def _field_name(inspection, field, fallback):
    return (getattr(inspection, field, "") or "").strip() or _user_name(fallback)


def _condition_text(instance):
    value = getattr(instance, "condicion_general", None)
    if value is None:
        value = getattr(instance, "condicion", None)
    if value == Inspeccion.Condicion.NO_MEDIDO:
        return "TOLERABLE"
    getter = (
        getattr(instance, "get_condicion_general_display", None)
        or getattr(instance, "get_condicion_display", None)
    )
    return (getter() if getter else value or "NORMAL").upper()


def _style_condition(cell, inspection):
    if inspection.condicion_general != Inspeccion.Condicion.CRITICO:
        return
    cell.fill = PatternFill(fill_type="solid", fgColor="FFFF0000")
    font = copy(cell.font)
    font.color = "FF000000"
    font.bold = True
    cell.font = font


def _number(value):
    return float(value) if value is not None else None


def _values(measurement):
    return [getattr(measurement, field) for field in MEASUREMENT_FIELDS]


def _average(values):
    available = [float(value) for value in values if value is not None]
    return round(sum(available) / len(available), 2) if available else None


def _minimum(values):
    available = [float(value) for value in values if value is not None]
    return min(available) if available else None


def _minimum_detail(measurements):
    candidates = []
    for measurement in measurements:
        for field in MEASUREMENT_FIELDS:
            value = getattr(measurement, field)
            if value is not None:
                candidates.append((value, field.upper(), measurement.punto))
    return min(candidates, key=lambda item: item[0]) if candidates else None


def _phase_note(measurements, phase, polea):
    detail = _minimum_detail(measurements)
    if detail:
        value, column, point = detail
        parts = [
            f"{phase}: espesor mínimo {value:.2f} mm en punto {column}, "
            f"medición radial {point}."
        ]
    else:
        parts = [f"{phase}: sin mediciones registradas."]
    manual = []
    for measurement in measurements:
        text = (measurement.observacion or "").strip()
        if text and text not in manual:
            manual.append(text)
    polea_text = (polea.observacion_medicion or "").strip()
    if polea_text and polea_text not in manual:
        manual.append(polea_text)
    if manual:
        parts.append(" ".join(manual))
    return " ".join(parts)


def _write_header(worksheet, inspection, blocks):
    values = {
        HEADER_CELLS["title"]: f"REPORTE INSPECCIÓN {inspection.codigo_reporte}",
        HEADER_CELLS["condition"]: _condition_text(inspection),
        HEADER_CELLS["plant"]: inspection.planta or "-",
        HEADER_CELLS["process"]: inspection.proceso or "-",
        HEADER_CELLS["equipment"]: inspection.get_tipo_display(),
        HEADER_CELLS["tag"]: inspection.faja.tag,
        HEADER_CELLS["stage"]: inspection.etapa or "-",
        HEADER_CELLS["equipment_condition"]: inspection.condicion_equipo or "-",
        HEADER_CELLS["inspection_date"]: inspection.fecha_inspeccion,
        HEADER_CELLS["report_date"]: inspection.fecha_reporte,
        HEADER_CELLS["inspector"]: _field_name(
            inspection, "inspector_campo_nombre", inspection.inspector
        ),
        HEADER_CELLS["supervisor"]: _field_name(
            inspection, "supervisor_campo_nombre", inspection.supervisor
        ),
        HEADER_CELLS["author"]: _field_name(
            inspection, "analista_elabora_nombre", inspection.analista
        ),
        HEADER_CELLS["validator"]: _field_name(
            inspection, "analista_valida_nombre", inspection.analista
        ),
        HEADER_CELLS["circumstances"]: inspection.circunstancias or "-",
        HEADER_CELLS["background"]: inspection.antecedentes or "-",
        HEADER_CELLS["diagram_title"]: (
            f"ESQUEMA DE UBICACIÓN DE POLEAS DE LA FAJA {inspection.faja.tag}"
        ),
    }
    for coordinate, value in values.items():
        worksheet[coordinate] = value if value not in (None, "") else "-"
    worksheet[HEADER_CELLS["inspection_date"]].number_format = "dd/mm/yyyy"
    worksheet[HEADER_CELLS["report_date"]].number_format = "dd/mm/yyyy"
    _style_condition(worksheet[HEADER_CELLS["condition"]], inspection)

    observations = []
    if (inspection.observaciones or "").strip():
        observations.append(inspection.observaciones.strip())
    recommendations = []
    for block in blocks:
        polea = block["polea"]
        if polea.tipo_medicion == TipoMedicionComponente.CAMPANA:
            phases = (
                (PHASE_START, list(block.get("mediciones_inicio", []))),
                (PHASE_END, list(block.get("mediciones_fin", []))),
            )
        else:
            phases = ((PHASE_NORMAL, list(block.get("mediciones", []))),)
        for phase, measurements in phases:
            observations.append(f"POLEA #{polea.numero:02d}: {_phase_note(measurements, phase, polea)}")
        condition = f"Polea #{polea.numero:02d} condición {_condition_text(polea)}"
        if (polea.recomendaciones or "").strip():
            condition += f". {polea.recomendaciones.strip()}"
        recommendations.append(condition)
    worksheet[HEADER_CELLS["observations"]] = "\n".join(observations) or "-"
    general = (inspection.recomendaciones or "").strip()
    if general:
        recommendations.append(general)
    rows = HEADER_CELLS["recommendation_rows"]
    if len(recommendations) > len(rows):
        recommendations = recommendations[: len(rows) - 1] + [
            " ".join(recommendations[len(rows) - 1 :])
        ]
    for index, coordinate in enumerate(rows):
        worksheet[coordinate] = recommendations[index] if index < len(recommendations) else None


def _set_hidden(worksheet, block, hidden):
    for row in range(block[0], block[1] + 1):
        worksheet.row_dimensions[row].hidden = hidden


def _write_calibration(worksheet, start, mapping, polea):
    values = (
        polea.marca_equipo,
        polea.modelo_equipo,
        polea.frecuencia_mhz,
        polea.rango_mm,
        polea.metodo_empleado,
        polea.velocidad_ms,
        polea.retardo_us,
    )
    row = start + mapping["calibration_offset"]
    column = mapping["calibration_column"]
    for offset, value in enumerate(values):
        worksheet[f"{column}{row + offset}"] = value or "-"


def _write_measurements(worksheet, start, mapping, measurements):
    measurements = list(measurements)[:5]
    data_start = start + mapping["data_offset"]
    row_averages = []
    row_minimums = []
    for offset in range(5):
        row = data_start + offset
        measurement = measurements[offset] if offset < len(measurements) else None
        worksheet[f"{mapping['point_column']}{row}"] = (
            measurement.punto if measurement else offset + 1
        )
        values = _values(measurement) if measurement else [None] * 7
        for column, value in zip(mapping["measurement_columns"], values):
            worksheet[f"{column}{row}"] = _number(value)
        row_average = _average(values)
        row_minimum = _minimum(values)
        row_averages.append(row_average)
        row_minimums.append(row_minimum)
        worksheet[f"{mapping['average_column']}{row}"] = row_average
        worksheet[f"{mapping['minimum_column']}{row}"] = row_minimum

    average_row = data_start + 5
    minimum_row = data_start + 6
    for field, column in zip(MEASUREMENT_FIELDS, mapping["measurement_columns"]):
        values = [getattr(item, field) for item in measurements]
        worksheet[f"{column}{average_row}"] = _average(values)
        worksheet[f"{column}{minimum_row}"] = _minimum(values)
    all_values = [value for item in measurements for value in _values(item)]
    worksheet[f"{mapping['average_column']}{average_row}"] = _average(all_values)
    worksheet[f"{mapping['minimum_column']}{average_row}"] = _minimum(all_values)
    worksheet[f"{mapping['average_column']}{minimum_row}"] = _average(row_minimums)
    worksheet[f"{mapping['minimum_column']}{minimum_row}"] = _minimum(row_minimums)


def _write_phase(worksheet, mapping, slot, polea, phase, measurements, tag):
    start, _end = mapping[slot]
    title = (
        f"MEDICIÓN DE ESPESORES DEL LAGGING DE LA POLEA #{polea.numero:02d} "
        f"/ {tag} {phase}"
    )
    worksheet[f"{mapping['title_column']}{start}"] = title
    worksheet[
        f"{mapping['inner_title_column']}{start + mapping['inner_title_offset']}"
    ] = f"MEDICIÓN DE ESPESORES DEL LAGGING DE LA POLEA {polea.numero} {phase}"
    _write_calibration(worksheet, start, mapping, polea)
    _write_measurements(worksheet, start, mapping, measurements)
    note_row = start + mapping["data_offset"] + 7
    worksheet[f"{mapping['note_column']}{note_row}"] = _phase_note(
        list(measurements), phase, polea
    )
    if polea.numero == 7:
        for row in range(start, mapping[slot][1] + 1):
            for column in range(62, 71):
                worksheet.cell(row, column).value = None


def _safe_image_path(photo):
    try:
        path = Path(photo.imagen.path)
    except (AttributeError, NotImplementedError, OSError, ValueError):
        return None
    return path if path.is_file() else None


def _photo_slots(worksheet, start, end, count, preferred=None):
    if count <= 0:
        return []
    if count > MAX_PHOTOS_PER_POLEA:
        raise ValueError(
            f"La plantilla admite hasta {MAX_PHOTOS_PER_POLEA} fotos por Polea; recibió {count}."
        )
    return rangos_fotos_fijos(
        start,
        end,
        count,
        columna_inicio="D",
        columna_fin="AX",
        worksheet=worksheet,
    )


def _set_photo_rows_visibility(worksheet, mapping, count):
    used_bands = (count + 2) // 3
    for index, (start, end) in enumerate(
        bandas_filas_fotos(*mapping["photo_rows"], 4)
    ):
        hidden = index >= used_bands
        for row in range(start, end + 1):
            worksheet.row_dimensions[row].hidden = hidden


def _clear_variable_images(worksheet):
    worksheet._images = [
        image
        for image in worksheet._images
        if image.anchor._from.row + 1 <= STATIC_IMAGE_MAX_ROW
    ]


def _write_photos(worksheet, mapping, polea, photos, seen):
    valid = []
    for photo in photos:
        path = _safe_image_path(photo)
        if path is None:
            continue
        key = str(path.resolve()).casefold()
        if key in seen:
            continue
        seen.add(key)
        valid.append((photo, path))
    slots = _photo_slots(
        worksheet,
        *mapping["photo_rows"],
        len(valid),
        preferred=mapping["photo_layout"],
    )
    _set_photo_rows_visibility(worksheet, mapping, len(valid))
    for (_photo, path), slot in zip(valid, slots):
        insertar_fotografia_ajustada(worksheet, path, slot)
    caption_parts = [f"Condición: {_condition_text(polea)}"]
    for label, value in (
        ("Observación visual", polea.observacion_visual),
        ("Observación de medición", polea.observacion_medicion),
        ("Recomendaciones", polea.recomendaciones),
    ):
        caption_parts.append(f"{label}: {(value or '').strip() or '-'}")
    worksheet[mapping["caption"]] = " | ".join(caption_parts)


def _write_polea(worksheet, block, tag, seen_paths):
    polea = block["polea"]
    mapping = POLEA_BLOCKS[polea.numero]
    if polea.tipo_medicion == TipoMedicionComponente.CAMPANA:
        _set_hidden(worksheet, mapping["slot_a"], False)
        _set_hidden(worksheet, mapping["slot_b"], False)
        _write_phase(
            worksheet, mapping, "slot_a", polea, PHASE_START,
            list(block.get("mediciones_inicio", [])), tag,
        )
        _write_phase(
            worksheet, mapping, "slot_b", polea, PHASE_END,
            list(block.get("mediciones_fin", [])), tag,
        )
    else:
        _set_hidden(worksheet, mapping["slot_a"], False)
        _set_hidden(worksheet, mapping["slot_b"], True)
        _write_phase(
            worksheet, mapping, "slot_a", polea, PHASE_NORMAL,
            list(block.get("mediciones", [])), tag,
        )
        _write_phase(worksheet, mapping, "slot_b", polea, PHASE_END, [], tag)
    worksheet[f"D{mapping['visual']}"] = (
        f"INSPECCIÓN VISUAL DE LA POLEA #{polea.numero:02d} / {tag}"
    )
    _write_photos(
        worksheet, mapping, polea, list(block.get("fotografias", [])), seen_paths
    )


def _write_hidden_polea_8(workbook, block, tag):
    if HIDDEN_SHEET_POLEA_8["sheet"] not in workbook.sheetnames or block is None:
        return
    worksheet = workbook[HIDDEN_SHEET_POLEA_8["sheet"]]
    polea = block["polea"]
    if polea.tipo_medicion == TipoMedicionComponente.CAMPANA:
        phase = PHASE_START
        measurements = list(block.get("mediciones_inicio", []))
    else:
        phase = PHASE_NORMAL
        measurements = list(block.get("mediciones", []))
    worksheet[HIDDEN_SHEET_POLEA_8["title"]] = (
        f"MEDICIÓN DE ESPESORES DEL LAGGING DE LA POLEA #08 / {tag} {phase}"
    )
    worksheet[HIDDEN_SHEET_POLEA_8["inner_title"]] = (
        f"MEDICIÓN DE ESPESORES DEL LAGGING DE LA POLEA 8 {phase}"
    )
    mapping = {
        "data_offset": 0,
        "point_column": HIDDEN_SHEET_POLEA_8["point_column"],
        "measurement_columns": HIDDEN_SHEET_POLEA_8["measurement_columns"],
        "average_column": HIDDEN_SHEET_POLEA_8["average_column"],
        "minimum_column": HIDDEN_SHEET_POLEA_8["minimum_column"],
    }
    _write_measurements(
        worksheet, HIDDEN_SHEET_POLEA_8["data_start"], mapping, measurements
    )
    values = (
        polea.marca_equipo, polea.modelo_equipo, polea.frecuencia_mhz,
        polea.rango_mm, polea.metodo_empleado, polea.velocidad_ms, polea.retardo_us,
    )
    for offset, value in enumerate(values):
        worksheet[
            f"{HIDDEN_SHEET_POLEA_8['calibration_column']}"
            f"{HIDDEN_SHEET_POLEA_8['calibration_start'] + offset}"
        ] = value or "-"
    worksheet[HIDDEN_SHEET_POLEA_8["note"]] = _phase_note(
        measurements, phase, polea
    )


def generar_excel_poleas_cvb0003_master(inspection, blocks):
    """Completa una copia del master sin insertar/eliminar filas o columnas."""
    _assert_master()
    workbook = load_workbook(BytesIO(MASTER_PATH.read_bytes()))
    structure_before = _structure_signature(workbook)
    worksheet = workbook[SHEET_NAME]
    blocks = sorted(blocks, key=lambda item: item["polea"].numero)
    _write_header(worksheet, inspection, blocks)
    _clear_variable_images(worksheet)
    by_number = {block["polea"].numero: block for block in blocks}
    seen_paths = set()
    for number in range(1, 10):
        block = by_number.get(number)
        if block is not None:
            _write_polea(worksheet, block, inspection.faja.tag, seen_paths)
    _write_hidden_polea_8(workbook, by_number.get(8), inspection.faja.tag)
    if _structure_signature(workbook) != structure_before:
        raise ValueError(
            "La exportación intentó modificar filas, columnas, merges o impresión del master."
        )
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
