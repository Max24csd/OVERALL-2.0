from io import BytesIO
from math import ceil, floor

from PIL import Image as PILImage
from PIL import ImageDraw, ImageFont
from PIL import ImageOps
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.cell import (
    column_index_from_string,
    get_column_letter,
    range_boundaries,
)
from openpyxl.utils.units import pixels_to_EMU, points_to_pixels


DEFAULT_COLUMN_WIDTH = 8.43
DEFAULT_ROW_HEIGHT_POINTS = 15


def _column_width_to_pixels(width, include_default_padding=False):
    width = DEFAULT_COLUMN_WIDTH if width is None else width
    pixels = floor(((256 * width + floor(128 / 7)) / 256) * 7)
    # El ancho predeterminado no trae una ColumnDimension explícita; Excel
    # agrega a ese caso el padding visual de la celda. Los anchos explícitos
    # del XML ya están normalizados y no deben recibirlo otra vez.
    if include_default_padding:
        pixels += 5
    return max(pixels, 1)


def _row_height_to_pixels(height_points):
    height_points = (
        DEFAULT_ROW_HEIGHT_POINTS if height_points is None else height_points
    )
    return max(int(points_to_pixels(height_points)), 1)


def medir_rango_celdas_px(worksheet, rango_celdas):
    min_col, min_row, max_col, max_row = range_boundaries(rango_celdas)
    default_column_width = (
        worksheet.sheet_format.defaultColWidth or DEFAULT_COLUMN_WIDTH
    )
    default_row_height = (
        worksheet.sheet_format.defaultRowHeight or DEFAULT_ROW_HEIGHT_POINTS
    )

    width = 0
    for column in range(min_col, max_col + 1):
        dimension = next(
            (
                item
                for item in worksheet.column_dimensions.values()
                if item.min <= column <= item.max
            ),
            None,
        )
        explicit_width = dimension is not None and dimension.width is not None
        width += _column_width_to_pixels(
            dimension.width if explicit_width else default_column_width,
            include_default_padding=not explicit_width,
        )

    height = 0
    for row in range(min_row, max_row + 1):
        dimension = worksheet.row_dimensions.get(row)
        height += _row_height_to_pixels(
            dimension.height
            if dimension is not None and dimension.height is not None
            else default_row_height
        )
    return width, height


def _column_width_px(worksheet, column):
    default_width = worksheet.sheet_format.defaultColWidth or DEFAULT_COLUMN_WIDTH
    dimension = next(
        (
            item
            for item in worksheet.column_dimensions.values()
            if item.min <= column <= item.max
        ),
        None,
    )
    explicit_width = dimension is not None and dimension.width is not None
    return _column_width_to_pixels(
        dimension.width if explicit_width else default_width,
        include_default_padding=not explicit_width,
    )


def _column_partitions_by_width(worksheet, column_min, column_max, parts):
    widths = [
        _column_width_px(worksheet, column)
        for column in range(column_min, column_max + 1)
    ]
    total_width = sum(widths)
    boundaries = []
    search_start = 0
    for part in range(1, parts):
        target = total_width * part / parts
        last_candidate = len(widths) - (parts - part)
        best = min(
            range(search_start, last_candidate),
            key=lambda index: abs(sum(widths[: index + 1]) - target),
        )
        boundaries.append(best + 1)
        search_start = best + 1
    boundaries.append(len(widths))
    partitions = []
    start_offset = 0
    for end_offset in boundaries:
        partitions.append(
            (column_min + start_offset, column_min + end_offset - 1)
        )
        start_offset = end_offset
    return partitions


def tamano_comun_cuadros_px(worksheet, rangos_celdas):
    sizes = [
        medir_rango_celdas_px(worksheet, rango)
        for rango in rangos_celdas
    ]
    if not sizes:
        return 0, 0
    return min(width for width, _height in sizes), min(
        height for _width, height in sizes
    )


def rangos_fotos_adaptativos(
    fila_inicio,
    fila_fin,
    cantidad,
    columna_inicio="C",
    columna_fin="AX",
    maximo_por_fila=3,
):
    """Distribuye fotos en filas de hasta tres slots grandes y simétricos."""
    if cantidad <= 0:
        return []
    filas_necesarias = ceil(cantidad / maximo_por_fila)
    bandas = []
    total_filas = fila_fin - fila_inicio + 1
    for indice in range(filas_necesarias):
        inicio = fila_inicio + round((total_filas * indice) / filas_necesarias)
        fin = fila_inicio + round(
            (total_filas * (indice + 1)) / filas_necesarias
        ) - 1
        bandas.append((inicio, max(inicio, fin)))

    columna_min = column_index_from_string(columna_inicio)
    columna_max = column_index_from_string(columna_fin)
    total_columnas = columna_max - columna_min + 1
    rangos = []
    restantes = cantidad
    for fila_inicio_banda, fila_fin_banda in bandas:
        columnas_en_fila = min(maximo_por_fila, restantes)
        for indice in range(columnas_en_fila):
            inicio_columna = columna_min + round(
                (total_columnas * indice) / columnas_en_fila
            )
            fin_columna = columna_min + round(
                (total_columnas * (indice + 1)) / columnas_en_fila
            ) - 1
            rangos.append(
                f"{get_column_letter(inicio_columna)}{fila_inicio_banda}:"
                f"{get_column_letter(max(inicio_columna, fin_columna))}{fila_fin_banda}"
            )
        restantes -= columnas_en_fila
    return rangos


def bandas_filas_fotos(fila_inicio, fila_fin, cantidad_bandas):
    total_filas = fila_fin - fila_inicio + 1
    bandas = []
    for indice in range(cantidad_bandas):
        inicio = fila_inicio + round((total_filas * indice) / cantidad_bandas)
        fin = fila_inicio + round(
            (total_filas * (indice + 1)) / cantidad_bandas
        ) - 1
        bandas.append((inicio, max(inicio, fin)))
    return bandas


def rangos_fotos_fijos(
    fila_inicio,
    fila_fin,
    cantidad,
    columna_inicio="C",
    columna_fin="AX",
    filas_maximas=4,
    columnas=3,
    worksheet=None,
):
    """Devuelve una cuadrícula fija; el tamaño de cada foto no depende del total."""
    capacidad = filas_maximas * columnas
    if cantidad < 0 or cantidad > capacidad:
        raise ValueError(f"La cuadrícula admite entre 0 y {capacidad} fotografías.")
    bandas = bandas_filas_fotos(fila_inicio, fila_fin, filas_maximas)
    columna_min = column_index_from_string(columna_inicio)
    columna_max = column_index_from_string(columna_fin)
    total_columnas = columna_max - columna_min + 1
    rangos = []
    restantes = cantidad
    for inicio_fila, fin_fila in bandas:
        columnas_en_fila = min(columnas, restantes)
        if columnas_en_fila <= 0:
            break
        if worksheet is not None:
            particiones = _column_partitions_by_width(
                worksheet,
                columna_min,
                columna_max,
                columnas_en_fila,
            )
        else:
            particiones = [
                (
                    columna_min + round((total_columnas * indice) / columnas_en_fila),
                    columna_min + round((total_columnas * (indice + 1)) / columnas_en_fila) - 1,
                )
                for indice in range(columnas_en_fila)
            ]
        for inicio_columna, fin_columna in particiones:
            rangos.append(
                f"{get_column_letter(inicio_columna)}{inicio_fila}:"
                f"{get_column_letter(max(inicio_columna, fin_columna))}{fin_fila}"
            )
        restantes -= columnas_en_fila
    return rangos


def _excel_image_with_orientation(path):
    with PILImage.open(path) as source:
        original_orientation = source.getexif().get(274, 1)
        adjusted = ImageOps.exif_transpose(source)
        original_width, original_height = adjusted.size

        if original_orientation == 1:
            return ExcelImage(str(path)), original_width, original_height

        normalized = adjusted.copy()
        if normalized.mode not in {"RGB", "RGBA"}:
            normalized = normalized.convert("RGB")
        buffer = BytesIO()
        normalized.save(buffer, format="PNG")
        buffer.seek(0)
        excel_image = ExcelImage(buffer)
        excel_image._cvb0003_buffer = buffer
        return excel_image, original_width, original_height


def _caption_font(size=11):
    try:
        return ImageFont.truetype("arial.ttf", size=size)
    except OSError:
        return ImageFont.load_default()


def _wrapped_caption(draw, text, font, max_width, max_lines=3):
    words = (text or "").split()
    if not words:
        return []
    lines = []
    current = ""
    for word in words:
        candidate = f"{current} {word}".strip()
        width = draw.textbbox((0, 0), candidate, font=font)[2]
        if current and width > max_width:
            lines.append(current)
            current = word
            if len(lines) == max_lines:
                break
        else:
            current = candidate
    if len(lines) < max_lines and current:
        lines.append(current)
    if len(lines) == max_lines and words:
        last = lines[-1]
        while last and draw.textbbox((0, 0), f"{last}…", font=font)[2] > max_width:
            last = last[:-1]
        lines[-1] = f"{last.rstrip()}…" if last else "…"
    return lines


def _excel_photo_card(path, width, height, description=""):
    with PILImage.open(path) as source:
        photo = ImageOps.exif_transpose(source).convert("RGB")
        card = PILImage.new("RGB", (width, height), "white")
        draw = ImageDraw.Draw(card)
        caption_height = max(14, min(48, int(round(height * 0.20))))
        image_height = max(height - caption_height - 2, 1)
        fitted = ImageOps.contain(photo, (max(width - 8, 1), max(image_height - 8, 1)))
        x = max((width - fitted.width) // 2, 0)
        y = max((image_height - fitted.height) // 2, 0)
        card.paste(fitted, (x, y))
        draw.line((0, image_height, width, image_height), fill="#A7B4C2", width=1)
        draw.rectangle((0, 0, width - 1, height - 1), outline="#A7B4C2", width=1)
        font = _caption_font(11 if caption_height >= 30 else 9)
        line_y = image_height + 3
        for line in _wrapped_caption(draw, description, font, max(width - 8, 1)):
            draw.text((4, line_y), line, fill="#111827", font=font)
            line_y += max(draw.textbbox((0, 0), line, font=font)[3] + 1, 10)
            if line_y >= height - 2:
                break
    buffer = BytesIO()
    card.save(buffer, format="PNG")
    buffer.seek(0)
    excel_image = ExcelImage(buffer)
    excel_image._cvb0003_buffer = buffer
    return excel_image


def _anchor_centered(worksheet, excel_image, rango_celdas, width, height):
    range_width, range_height = medir_rango_celdas_px(worksheet, rango_celdas)
    horizontal_offset = max(int(round((range_width - width) / 2)), 0)
    vertical_offset = max(int(round((range_height - height) / 2)), 0)
    min_col, min_row, max_col, max_row = range_boundaries(rango_celdas)
    default_column_width = (
        worksheet.sheet_format.defaultColWidth or DEFAULT_COLUMN_WIDTH
    )
    default_row_height = (
        worksheet.sheet_format.defaultRowHeight or DEFAULT_ROW_HEIGHT_POINTS
    )
    anchor_col = min_col
    remaining_x = horizontal_offset
    while anchor_col < max_col:
        dimension = next(
            (
                item for item in worksheet.column_dimensions.values()
                if item.min <= anchor_col <= item.max
            ),
            None,
        )
        explicit_width = dimension is not None and dimension.width is not None
        column_pixels = _column_width_to_pixels(
            dimension.width if explicit_width else default_column_width,
            include_default_padding=not explicit_width,
        )
        if remaining_x < column_pixels:
            break
        remaining_x -= column_pixels
        anchor_col += 1

    anchor_row = min_row
    remaining_y = vertical_offset
    while anchor_row < max_row:
        dimension = worksheet.row_dimensions.get(anchor_row)
        row_pixels = _row_height_to_pixels(
            dimension.height
            if dimension is not None and dimension.height is not None
            else default_row_height
        )
        if remaining_y < row_pixels:
            break
        remaining_y -= row_pixels
        anchor_row += 1
    marker = AnchorMarker(
        col=anchor_col - 1,
        row=anchor_row - 1,
        colOff=pixels_to_EMU(remaining_x),
        rowOff=pixels_to_EMU(remaining_y),
    )
    extent = XDRPositiveSize2D(
        cx=pixels_to_EMU(width),
        cy=pixels_to_EMU(height),
    )
    excel_image.width = width
    excel_image.height = height
    excel_image.anchor = OneCellAnchor(_from=marker, ext=extent)
    worksheet.add_image(excel_image)


def insertar_imagen_ajustada(
    worksheet,
    ruta_imagen,
    rango_celdas,
    ancho_max_px=None,
    alto_max_px=None,
    margen_px=4,
    factor_ancho=1.0,
    factor_alto=1.0,
):
    if not ruta_imagen:
        return False

    try:
        excel_image, original_width, original_height = (
            _excel_image_with_orientation(ruta_imagen)
        )
    except (FileNotFoundError, OSError, ValueError):
        return False

    range_width, range_height = medir_rango_celdas_px(
        worksheet, rango_celdas
    )
    box_width = min(ancho_max_px or range_width, range_width)
    box_height = min(alto_max_px or range_height, range_height)
    available_width = max(
        int(floor(box_width * min(max(factor_ancho, 0.01), 1.0)))
        - (margen_px * 2),
        1,
    )
    available_height = max(
        int(floor(box_height * min(max(factor_alto, 0.01), 1.0)))
        - (margen_px * 2),
        1,
    )
    scale = min(
        available_width / original_width,
        available_height / original_height,
    )
    final_width = max(int(round(original_width * scale)), 1)
    final_height = max(int(round(original_height * scale)), 1)

    _anchor_centered(
        worksheet, excel_image, rango_celdas, final_width, final_height
    )
    return True


def insertar_fotografia_ajustada(worksheet, ruta_imagen, rango_celdas):
    """Ocupa casi todo el slot fotográfico sin alterar su geometría."""
    return insertar_imagen_ajustada(
        worksheet,
        ruta_imagen,
        rango_celdas,
        margen_px=0,
        factor_ancho=0.95,
        factor_alto=0.93,
    )


def insertar_tarjeta_foto_ajustada(
    worksheet,
    ruta_imagen,
    rango_celdas,
    descripcion="",
    ancho_max_px=None,
    alto_max_px=None,
    margen_px=4,
):
    """Inserta una tarjeta uniforme: foto contenida y descripción debajo."""
    if not ruta_imagen:
        return False
    range_width, range_height = medir_rango_celdas_px(worksheet, rango_celdas)
    box_width = min(ancho_max_px or range_width, range_width)
    box_height = min(alto_max_px or range_height, range_height)
    final_width = max(int(box_width - (margen_px * 2)), 1)
    final_height = max(int(box_height - (margen_px * 2)), 1)
    try:
        excel_image = _excel_photo_card(
            ruta_imagen, final_width, final_height, descripcion
        )
    except (FileNotFoundError, OSError, ValueError):
        return False
    _anchor_centered(
        worksheet, excel_image, rango_celdas, final_width, final_height
    )
    return True
