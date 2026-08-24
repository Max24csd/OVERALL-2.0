from copy import copy
from io import BytesIO
from pathlib import Path
from textwrap import wrap

from django.conf import settings
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from .image_utils import insertar_imagen_ajustada, tamano_comun_cuadros_px
from inspecciones.reportes.campaign_excel import agregar_hoja_mediciones_campana, limpiar_textos_historicos_campana


TEMPLATE_PATH = (
    Path(settings.BASE_DIR)
    / "inspecciones"
    / "reportes"
    / "cvb0003"
    / "assets"
    / "reporte_poleas_cvb0003.xlsx"
)

MEASUREMENT_COLUMNS = ("AC", "AE", "AG", "AI", "AK", "AM", "AO")

# La plantilla contiene dos bloques para las poleas 03 y 07. Ambos se
# actualizan con los datos actuales para evitar que sobrevivan valores del
# libro maestro; las fotografias se insertan una sola vez en el bloque visual
# de inicio de campana.
POLEA_LAYOUTS = {
    1: ({"title": 87, "calibration": 90, "data": 92, "average": 97,
         "minimum": 98, "note": 99, "visual": 101},),
    2: ({"title": 125, "calibration": 127, "data": 129, "average": 134,
         "minimum": 135, "note": 136, "visual": 138},),
    3: (
        {"title": 156, "calibration": 159, "data": 161, "average": 166,
         "minimum": 167, "note": 168, "visual": 171,
         "suffix": " FIN DE CAMPAÑA"},
        {"title": 190, "calibration": 193, "data": 195, "average": 200,
         "minimum": 201, "note": 202, "visual": 205,
         "suffix": " INICIO DE CAMPAÑA"},
    ),
    4: ({"visual": 244},),
    5: ({"visual": 264},),
    6: ({"title": 284, "calibration": 287, "data": 289, "average": 294,
         "minimum": 295, "note": 296, "visual": 298},),
    7: (
        {"title": 317, "calibration": 320, "data": 322, "average": 327,
         "minimum": 328, "note": 329, "visual": 332,
         "suffix": " FIN DE CAMPAÑA"},
        {"title": 355, "calibration": 358, "data": 360, "average": 365,
         "minimum": 366, "note": 367, "visual": 370,
         "suffix": " INICIO DE CAMPAÑA"},
    ),
    8: ({"title": 393, "calibration": 396, "data": 398, "average": 403,
         "minimum": 404, "note": 405, "visual": 408},),
    9: ({"title": 429, "calibration": 432, "data": 434, "average": 439,
         "minimum": 440, "note": 441, "visual": 443},),
}

PHOTO_ZONES = {
    1: (102, 122, "E123"),
    2: (139, 154, "D155"),
    3: (206, 222, "D243"),
    4: (245, 262, "O263"),
    5: (265, 281, "U282"),
    6: (299, 315, "D316"),
    7: (372, 391, "D392"),
    8: (409, 427, "D428"),
    9: (445, 462, "D463"),
}


def _user_name(user):
    if user is None:
        return ""
    return (user.get_full_name() or user.username or "").strip()


def _field_name(inspeccion, field_name, fallback_user):
    return (getattr(inspeccion, field_name, "") or "").strip() or _user_name(
        fallback_user
    )


def _number(value):
    return float(value) if value is not None else None


def _average(values):
    clean = [float(value) for value in values if value is not None]
    return sum(clean) / len(clean) if clean else None


def _minimum(values):
    clean = [float(value) for value in values if value is not None]
    return min(clean) if clean else None


def _safe_image_path(photo):
    try:
        path = Path(photo.imagen.path)
    except (AttributeError, NotImplementedError, OSError, ValueError):
        return None
    return path if path.is_file() else None


def _write_wrapped_rows(worksheet, column, rows, value, width=145):
    text = str(value or "").strip()
    parts = []
    for line in text.splitlines() or [""]:
        parts.extend(wrap(line, width=width) or [""])
    if len(parts) > len(rows):
        parts = parts[: len(rows) - 1] + [" ".join(parts[len(rows) - 1 :])]
    for index, row in enumerate(rows):
        worksheet[f"{column}{row}"] = parts[index] if index < len(parts) else None


def _write_header(worksheet, inspeccion):
    worksheet["M3"] = f"REPORTE INSPECCIÓN {inspeccion.codigo_reporte}"
    worksheet["K9"] = inspeccion.get_condicion_general_display().upper()
    worksheet["K9"].fill = PatternFill("solid", fgColor="00B050")
    worksheet["K9"].font = Font(
        name=worksheet["K9"].font.name,
        size=worksheet["K9"].font.sz,
        bold=True,
        color="FFFFFF",
    )

    values = {
        "K12": inspeccion.planta,
        "Y12": inspeccion.proceso,
        "K14": inspeccion.get_tipo_display(),
        "Y14": inspeccion.faja.tag,
        "K16": inspeccion.etapa,
        "Y16": inspeccion.condicion_equipo,
        "K18": inspeccion.fecha_inspeccion,
        "Y18": inspeccion.fecha_reporte,
        "K20": _field_name(
            inspeccion, "inspector_campo_nombre", inspeccion.inspector
        ),
        "Y20": _field_name(
            inspeccion, "supervisor_campo_nombre", inspeccion.supervisor
        ),
        "K22": _field_name(
            inspeccion, "analista_elabora_nombre", inspeccion.analista
        ),
        "Y22": _field_name(
            inspeccion, "analista_valida_nombre", inspeccion.analista
        ),
        "K25": inspeccion.circunstancias,
        "K27": inspeccion.antecedentes,
        "K29": inspeccion.observaciones,
        "D73": (
            "ESQUEMA DE UBICACIÓN DE POLEAS DE LA FAJA "
            f"{inspeccion.faja.tag}"
        ),
    }
    worksheet["R22"] = "ANALISTA QUE VALIDA:"
    for coordinate, value in values.items():
        worksheet[coordinate] = value if value not in (None, "") else "-"

    worksheet["K18"].number_format = "dd mmmm yyyy"
    worksheet["Y18"].number_format = "dd mmmm yyyy"
    for coordinate in ("K25", "K27", "K29"):
        alignment = copy(worksheet[coordinate].alignment)
        alignment.wrap_text = True
        alignment.vertical = "top"
        worksheet[coordinate].alignment = alignment

    _write_wrapped_rows(
        worksheet,
        "K",
        list(range(60, 71)),
        inspeccion.recomendaciones,
        width=135,
    )


def _clear_master_photos(worksheet):
    worksheet._images = [
        image
        for image in worksheet._images
        if image.anchor._from.row + 1 < 87
    ]


def _write_measurements(worksheet, layout, measurements):
    measurements = list(measurements[:5])
    for offset in range(5):
        row = layout["data"] + offset
        measurement = measurements[offset] if offset < len(measurements) else None
        worksheet[f"Z{row}"] = measurement.punto if measurement else None
        for field_name, column in zip("abcdefg", MEASUREMENT_COLUMNS):
            worksheet[f"{column}{row}"] = (
                _number(getattr(measurement, field_name)) if measurement else None
            )
        worksheet[f"AQ{row}"] = (
            _number(measurement.promedio) if measurement else None
        )
        worksheet[f"AT{row}"] = (
            _number(measurement.minimo) if measurement else None
        )

    for field_name, column in zip("abcdefg", MEASUREMENT_COLUMNS):
        values = [getattr(item, field_name) for item in measurements]
        worksheet[f"{column}{layout['average']}"] = _average(values)
        worksheet[f"{column}{layout['minimum']}"] = _minimum(values)

    row_averages = [item.promedio for item in measurements]
    row_minimums = [item.minimo for item in measurements]
    worksheet[f"AQ{layout['average']}"] = _average(row_averages)
    worksheet[f"AT{layout['average']}"] = _average(row_minimums)
    worksheet[f"AQ{layout['minimum']}"] = _minimum(row_averages)
    worksheet[f"AT{layout['minimum']}"] = _minimum(row_minimums)

    candidates = []
    for measurement in measurements:
        for field_name in "abcdefg":
            value = getattr(measurement, field_name)
            if value is not None:
                candidates.append((value, field_name.upper(), measurement.punto))
    if candidates:
        value, point_letter, radial_point = min(candidates, key=lambda item: item[0])
        note = (
            f"El espesor mínimo encontrado es de {value:.2f} mm en el punto "
            f"{point_letter}, medición radial {radial_point}."
        )
    else:
        note = "No existen mediciones de espesor registradas."
    worksheet[f"Z{layout['note']}"] = note


def _write_calibration(worksheet, layout, polea):
    values = (
        polea.marca_equipo,
        polea.modelo_equipo,
        polea.frecuencia_mhz,
        polea.rango_mm,
        polea.metodo_empleado,
        polea.velocidad_ms,
        polea.retardo_us,
    )
    for offset, value in enumerate(values):
        worksheet[f"N{layout['calibration'] + offset}"] = value or "-"


def _write_polea_blocks(worksheet, polea, block, tag):
    number_label = f"{polea.numero:02d}"
    layouts = POLEA_LAYOUTS[polea.numero]
    if block.get("es_campana"):
        for layout in layouts:
            if "title" in layout:
                for row in range(layout["title"], layout["visual"]):
                    worksheet.row_dimensions[row].hidden = True
        worksheet[f"D{layouts[0]['visual']}"] = f"INSPECCIÓN VISUAL DE LA POLEA #{number_label} / {tag}"
        return
    measurements = block.get("mediciones", [])
    for index, layout in enumerate(layouts):
        if index:
            if "title" in layout:
                for row in range(layout["title"], layout["visual"]):
                    worksheet.row_dimensions[row].hidden = True
            continue
        suffix = ""
        if "title" in layout:
            worksheet[f"D{layout['title']}"] = (
                "MEDICIÓN DE ESPESORES DEL LAGGING DE LA POLEA "
                f"#{number_label} / {tag}{suffix}"
            )
            _write_calibration(worksheet, layout, polea)
            _write_measurements(worksheet, layout, measurements)
        worksheet[f"D{layout['visual']}"] = (
            f"INSPECCIÓN VISUAL DE LA POLEA #{number_label} / {tag}{suffix}"
        )


def _photo_ranges(start_row, end_row, count):
    if count <= 1:
        return [f"D{start_row}:AX{end_row}"]
    if count == 2:
        return [
            f"E{start_row}:AB{end_row}",
            f"AD{start_row}:AW{end_row}",
        ]
    if count == 3:
        return [
            f"D{start_row}:R{end_row}",
            f"S{start_row}:AG{end_row}",
            f"AH{start_row}:AX{end_row}",
        ]
    if count == 4:
        return [
            f"D{start_row}:O{end_row}",
            f"P{start_row}:AA{end_row}",
            f"AB{start_row}:AM{end_row}",
            f"AN{start_row}:AX{end_row}",
        ]
    middle_row = start_row + ((end_row - start_row + 1) // 2)
    first_end = middle_row - 1
    ranges = [
        f"D{start_row}:R{first_end}",
        f"S{start_row}:AG{first_end}",
        f"AH{start_row}:AX{first_end}",
    ]
    if count == 5:
        ranges.extend(
            [
                f"E{middle_row}:AB{end_row}",
                f"AD{middle_row}:AW{end_row}",
            ]
        )
    else:
        ranges.extend(
            [
                f"D{middle_row}:R{end_row}",
                f"S{middle_row}:AG{end_row}",
                f"AH{middle_row}:AX{end_row}",
            ]
        )
    return ranges


def _write_photos_and_caption(worksheet, polea, photos, seen_paths):
    start_row, end_row, caption_coordinate = PHOTO_ZONES[polea.numero]
    unique_photos = []
    for photo in photos:
        path = _safe_image_path(photo)
        if path is None:
            continue
        key = str(path.resolve()).casefold()
        if key in seen_paths:
            continue
        seen_paths.add(key)
        unique_photos.append((photo, path))

    unique_photos = unique_photos[:6]
    ranges = _photo_ranges(start_row, end_row, len(unique_photos))
    common_width, common_height = tamano_comun_cuadros_px(
        worksheet, ranges
    )
    for (_photo, path), cell_range in zip(unique_photos, ranges):
        insertar_imagen_ajustada(
            worksheet,
            path,
            cell_range,
            ancho_max_px=common_width,
            alto_max_px=common_height,
            margen_px=4,
        )

    photo_details = []
    for photo, _path in unique_photos:
        detail = " - ".join(
            value
            for value in (
                (photo.codigo_dano or "").strip(),
                (photo.descripcion or "").strip(),
            )
            if value
        )
        if detail:
            photo_details.append(detail)
    parts = [
        f"Condición: {polea.get_condicion_display()}",
        f"Observación visual: {polea.observacion_visual or '-'}",
        f"Observación de medición: {polea.observacion_medicion or '-'}",
        f"Recomendaciones: {polea.recomendaciones or '-'}",
    ]
    if photo_details:
        parts.append("Fotografías: " + " | ".join(photo_details))
    caption = worksheet[caption_coordinate]
    caption.value = " | ".join(parts)
    alignment = copy(caption.alignment)
    is_merged = any(
        merged.min_row <= caption.row <= merged.max_row
        and merged.min_col <= caption.column <= merged.max_col
        for merged in worksheet.merged_cells.ranges
    )
    alignment.wrap_text = is_merged
    alignment.shrink_to_fit = not is_merged
    alignment.horizontal = "center"
    alignment.vertical = "center"
    caption.alignment = alignment
    font = copy(caption.font)
    font.sz = min(font.sz or 7, 7)
    caption.font = font


def _write_hoja2(worksheet, polea, measurements, tag):
    worksheet["D8"] = f"MEDICIÓN DE ESPESORES DEL LAGGING DE LA POLEA #08 - {tag}"
    layout = {"calibration": 11, "data": 13, "average": 18, "minimum": 19,
              "note": 20}
    _write_calibration(worksheet, layout, polea)
    _write_measurements(worksheet, layout, measurements)


def generar_excel_poleas_cvb0003(inspeccion, bloques):
    template_data = BytesIO(TEMPLATE_PATH.read_bytes())
    workbook = load_workbook(template_data)
    worksheet = workbook["Hoja1"]
    limpiar_textos_historicos_campana(worksheet)
    _write_header(worksheet, inspeccion)
    _clear_master_photos(worksheet)

    # Elimina textos descriptivos variables pertenecientes al reporte maestro.
    for coordinate in ("E123", "D155", "D189", "D243", "O263", "U282",
                       "D316", "D354", "D392", "D428", "D463"):
        worksheet[coordinate] = None

    blocks_by_number = {
        block["polea"].numero: block
        for block in bloques
        if block.get("polea") is not None
    }
    seen_paths = set()
    for number in range(1, 10):
        block = blocks_by_number.get(number)
        if block is None:
            continue
        polea = block["polea"]
        measurements = list(block.get("mediciones", []))
        photos = list(block.get("fotografias", []))
        _write_polea_blocks(worksheet, polea, block, inspeccion.faja.tag)
        _write_photos_and_caption(worksheet, polea, photos, seen_paths)

    polea_8 = blocks_by_number.get(8)
    if polea_8 is not None and not polea_8.get("es_campana") and "Hoja2" in workbook.sheetnames:
        _write_hoja2(
            workbook["Hoja2"],
            polea_8["polea"],
            list(polea_8.get("mediciones", [])),
            inspeccion.faja.tag,
        )
    agregar_hoja_mediciones_campana(
        workbook, bloques, "polea", f"MEDICIÓN DE ESPESORES DEL LAGGING DE LA POLEA {inspeccion.faja.tag}",
        config={"layouts": POLEA_LAYOUTS, "title_column": "D", "calibration_column": "N", "measurement_columns": MEASUREMENT_COLUMNS, "point_column": "Z", "result_column": "AQ", "residual_column": "AT", "note_column": "Z", "ceramic_numbers": {4, 5}},
    )

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
