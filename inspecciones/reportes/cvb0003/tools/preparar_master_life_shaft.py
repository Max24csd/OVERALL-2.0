"""Prepara una sola vez el master fijo de Life Shaft CVB003.

Este script no forma parte de la descarga. Duplica verticalmente cada bloque
técnico aprobado para disponer de Slot A y Slot B sin insertar filas durante
la exportación.
"""

from __future__ import annotations

import hashlib
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.dimensions import RowDimension
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.cell_range import CellRange


SOURCE_SHA256 = "3FBDDDEFE1A59D1592CE12D938BDEE557D4039F1487D3977EFAF8CC420BBFF61"
SOURCE_NAME = "approved_life_shaft_cvb0003.xlsx"
OUTPUT_NAME = "master_life_shaft_cvb0003.xlsx"
SHEET_NAME = "Hoja1"

# Se procesa de abajo hacia arriba para que cada bloque fuente conserve sus
# coordenadas aprobadas hasta el momento de clonarlo.
INSERTIONS = (
    (232, 217, 231),
    (194, 179, 193),
    (157, 142, 156),
    (120, 105, 119),
    (83, 68, 82),
)


def _sha256(path):
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest().upper()


def _copy_cell(source, target):
    if not isinstance(source, MergedCell):
        target.value = source.value
    target._style = copy(source._style)
    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def _shift_row_dimensions(worksheet, start, amount, previous_max):
    saved = {
        row: copy(worksheet.row_dimensions[row])
        for row in range(start, previous_max + 1)
        if row in worksheet.row_dimensions
    }
    for row in range(start, previous_max + amount + 1):
        if row in worksheet.row_dimensions:
            del worksheet.row_dimensions[row]
    for row, dimension in saved.items():
        moved = copy(dimension)
        moved.index = row + amount
        worksheet.row_dimensions[row + amount] = moved


def _snapshot_merge_styles(worksheet, ranges):
    snapshots = {}
    for merged_range in ranges:
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for column in range(merged_range.min_col, merged_range.max_col + 1):
                cell = worksheet._cells.get((row, column))
                if cell is not None:
                    snapshots[(row, column)] = copy(cell._style)
    return snapshots


def _shift_images(worksheet, start, amount):
    for image in worksheet._images:
        anchor = image.anchor
        if anchor._from.row + 1 >= start:
            anchor._from.row += amount
        if hasattr(anchor, "_to") and anchor._to.row + 1 >= start:
            anchor._to.row += amount


def _insert_and_clone(worksheet, insert_at, source_start, source_end):
    amount = source_end - source_start + 1
    previous_max = worksheet.max_row
    all_merges = [CellRange(str(item)) for item in worksheet.merged_cells.ranges]
    crossing = [
        item for item in all_merges
        if item.min_row < insert_at <= item.max_row
    ]
    if crossing:
        raise ValueError(f"Hay merges que cruzan la inserción {insert_at}: {crossing}")
    affected = [item for item in all_merges if item.min_row >= insert_at]
    merge_styles = _snapshot_merge_styles(worksheet, affected)
    for item in affected:
        worksheet.unmerge_cells(str(item))

    previous_breaks = [item.id for item in worksheet.row_breaks.brk]
    _shift_images(worksheet, insert_at, amount)
    worksheet.insert_rows(insert_at, amount)
    _shift_row_dimensions(worksheet, insert_at, amount, previous_max)

    for item in affected:
        item.shift(row_shift=amount)
        worksheet.merge_cells(str(item))
    for (row, column), style in merge_styles.items():
        worksheet._cells[(row + amount, column)]._style = copy(style)

    for offset in range(amount):
        source_row = source_start + offset
        target_row = insert_at + offset
        source_dimension = worksheet.row_dimensions.get(source_row)
        if source_dimension is not None:
            target_dimension = copy(source_dimension)
            target_dimension.index = target_row
            worksheet.row_dimensions[target_row] = target_dimension
        for column in range(1, worksheet.max_column + 1):
            source_cell = worksheet._cells.get((source_row, column))
            if source_cell is None:
                continue
            target_cell = worksheet.cell(target_row, column)
            _copy_cell(source_cell, target_cell)

    source_merges = [
        CellRange(str(item)) for item in worksheet.merged_cells.ranges
        if source_start <= item.min_row and item.max_row <= source_end
    ]
    for item in source_merges:
        cloned = CellRange(str(item))
        cloned.shift(row_shift=insert_at - source_start)
        worksheet.merge_cells(str(cloned))

    worksheet.row_breaks = type(worksheet.row_breaks)()
    for break_id in previous_breaks:
        worksheet.row_breaks.append(
            Break(id=break_id + amount if break_id >= insert_at else break_id)
        )


def prepare(source_path=None, output_path=None):
    assets = Path(__file__).resolve().parents[1] / "assets"
    source = Path(source_path) if source_path else assets / SOURCE_NAME
    output = Path(output_path) if output_path else assets / OUTPUT_NAME
    if _sha256(source) != SOURCE_SHA256:
        raise ValueError("La copia aprobada de Life Shaft no coincide con su SHA256.")

    workbook = load_workbook(source)
    worksheet = workbook[SHEET_NAME]
    for insertion in INSERTIONS:
        _insert_and_clone(worksheet, *insertion)

    worksheet.print_area = "A1:AY327"
    workbook.active = workbook.sheetnames.index(SHEET_NAME)
    workbook.save(output)
    return output


if __name__ == "__main__":
    print(prepare())
