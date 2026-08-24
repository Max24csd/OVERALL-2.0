"""Exportación aislada de Faja CVB004 basada en su master aprobado."""

from __future__ import annotations

import hashlib
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from inspecciones.models import FotoFajaCVB0003
from inspecciones.reportes.cvb0003.image_utils import insertar_imagen_ajustada


MASTER_PATH = Path(__file__).resolve().parent / "assets" / "master_faja_cvb0004.xlsx"
SHEET_NAME = "REPORTE DE INSPECCION CV0004"
PHOTO_SLOTS = {
    "EMPALMES": ("C811:H828", "J811:O828", "Q811:V828", "C832:H849", "J832:O849", "Q832:V849"),
    "CARGA": ("C951:H968", "J951:O968", "Q951:V968", "C972:H989", "J972:O989", "Q972:V989"),
}


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def _bytes_sha256(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest().upper()


MASTER_SHA256 = _sha256(MASTER_PATH)


def _set(ws, coordinate: str, value) -> None:
    cell = ws[coordinate]
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if coordinate in merged_range:
                cell = ws.cell(merged_range.min_row, merged_range.min_col)
                break
    if isinstance(cell, MergedCell):
        return
    cell.value = "" if value is None else value


def _signature(ws):
    return (
        tuple(sorted(str(item) for item in ws.merged_cells.ranges)),
        tuple((key, item.width, item.hidden) for key, item in ws.column_dimensions.items()),
        tuple((row, item.height, item.hidden) for row, item in ws.row_dimensions.items() if item.height is not None),
        str(ws.print_area),
        tuple(item.id for item in ws.row_breaks.brk),
        (ws.page_setup.orientation, ws.page_setup.paperSize, ws.page_setup.fitToWidth, ws.page_setup.fitToHeight),
    )


def _range_anchor(cell_range: str) -> str:
    return cell_range.split(":", 1)[0]


def _write_header(ws, inspection) -> None:
    faja = inspection.faja
    _set(ws, "F2", f"REPORTE DE INSPECCIÓN {inspection.codigo_reporte or ''}".strip())
    _set(ws, "F7", inspection.get_condicion_general_display())
    _set(ws, "F10", inspection.planta)
    _set(ws, "L10", inspection.proceso)
    _set(ws, "F12", faja.nombre)
    _set(ws, "L12", faja.tag)
    _set(ws, "F14", inspection.etapa)
    _set(ws, "L14", inspection.condicion_equipo)
    _set(ws, "F16", inspection.fecha_inspeccion)
    _set(ws, "L16", inspection.fecha_reporte)
    _set(ws, "F18", inspection.inspector_campo_nombre or getattr(inspection.inspector, "get_full_name", lambda: "")())
    _set(ws, "L18", inspection.supervisor_campo_nombre or getattr(inspection.supervisor, "get_full_name", lambda: "")())
    _set(ws, "F21", inspection.analista_elabora_nombre or getattr(inspection.analista, "get_full_name", lambda: "")())
    _set(ws, "F24", inspection.circunstancias)
    _set(ws, "F27", inspection.antecedentes)
    _set(ws, "F30", inspection.observaciones)
    _set(ws, "F54", inspection.recomendaciones)


def _normalizar(value) -> str:
    return "".join(character for character in str(value or "").upper() if character.isalnum())


def _write_measurements(ws, inspection) -> None:
    for excel_row in range(171, 724):
        for column in "HIJKLMNT":
            _set(ws, f"{column}{excel_row}", None)
    for excel_row in range(728, 789):
        for column in "HIJKLMNT":
            _set(ws, f"{column}{excel_row}", None)

    empalmes_por_clave = {
        (_normalizar(row.empalme), _normalizar(row.posicion)): row
        for row in inspection.empalmes_cvb0003.all()
    }
    empalme_actual = ""
    for excel_row in range(171, 724):
        if ws[f"D{excel_row}"].value:
            empalme_actual = _normalizar(ws[f"D{excel_row}"].value)
        posicion = _normalizar(ws[f"F{excel_row}"].value)
        if not empalme_actual or not posicion:
            continue
        row = empalmes_por_clave.get((empalme_actual, posicion))
        for column, value in zip("HIJKLMN", (row.a, row.b, row.c, row.d, row.e, row.f, row.g) if row else (None,) * 7):
            _set(ws, f"{column}{excel_row}", value)
        _set(ws, f"T{excel_row}", row.observacion if row else None)

    tramos_por_clave = {
        (_normalizar(row.tramo), row.medicion): row
        for row in inspection.tramos_cvb0003.filter(tipo="CARGA")
    }
    tramo_actual = ""
    for excel_row in range(728, 789):
        if ws[f"D{excel_row}"].value:
            tramo_actual = _normalizar(ws[f"D{excel_row}"].value)
        medicion = ws[f"E{excel_row}"].value
        if not tramo_actual or not isinstance(medicion, int):
            continue
        row = tramos_por_clave.get((tramo_actual, medicion))
        for column, value in zip("HIJKLMN", (row.a, row.b, row.c, row.d, row.e, row.f, row.g) if row else (None,) * 7):
            _set(ws, f"{column}{excel_row}", value)
        _set(ws, f"T{excel_row}", row.observacion if row else None)


def _clear_variable_images(ws) -> None:
    """Conserva los seis recursos estáticos ubicados antes de las tablas."""
    ws._images = [
        image
        for image in ws._images
        if getattr(getattr(image, "anchor", None), "_from", None) is not None
        and image.anchor._from.row < 169
    ]


def _write_photos(ws, inspection) -> None:
    for section, slots in PHOTO_SLOTS.items():
        photos = list(
            inspection.fotografias_cvb0003.filter(seccion=section).order_by("creada_en", "id")
        )
        for photo, slot in zip(photos, slots):
            if photo.imagen and Path(photo.imagen.path).is_file():
                insertar_imagen_ajustada(
                    ws,
                    photo.imagen.path,
                    slot,
                    margen_px=5,
                    factor_ancho=0.95,
                    factor_alto=0.93,
                )


def generar_excel_faja_cvb0004(inspection) -> BytesIO:
    if _sha256(MASTER_PATH) != MASTER_SHA256:
        raise ValueError("La plantilla maestra de Faja CVB004 fue modificada.")

    workbook = load_workbook(MASTER_PATH)
    worksheet = workbook[SHEET_NAME]
    before = _signature(worksheet)

    _write_header(worksheet, inspection)
    _write_measurements(worksheet, inspection)
    _clear_variable_images(worksheet)
    _write_photos(worksheet, inspection)

    if _signature(worksheet) != before:
        raise ValueError("La exportación CVB004 Faja intentó alterar la estructura del master.")

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
