from copy import copy
from io import BytesIO
from pathlib import Path

from django.conf import settings
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from .image_utils import insertar_imagen_ajustada, tamano_comun_cuadros_px
from inspecciones.reportes.campaign_excel import agregar_hoja_mediciones_campana, limpiar_textos_historicos_campana
from .utils import (
    etiqueta_condicion,
    generar_conclusiones_life_shaft,
    generar_observaciones_life_shaft,
    minimo_componente,
    nombre_campo,
)


TEMPLATE_PATH = (
    Path(settings.BASE_DIR)
    / "inspecciones/reportes/cvb0004/templates/20260515-VTUT-0220CVB004-LIVESHAFT.xlsx"
)
SHEET_NAME = "Hoja1"
LAYOUTS = (
    dict(title=63, calibration="L66", data=68, data_end=71, point="Y", columns=("AB", "AD", "AF", "AH", "AJ", "AL", "AN"), average=72, minimum=73, result="AP", residual="AS", note="Y74", visual=77, photos=(78, 94), caption="C95"),
    dict(title=97, calibration="M100", data=102, data_end=105, point="Y", columns=("AB", "AD", "AF", "AH", "AJ", "AL", "AN"), average=106, minimum=107, result="AP", residual="AS", note="Y108", visual=111, photos=(112, 128), caption="C129"),
    dict(title=130, calibration="M134", data=136, data_end=138, point="Z", columns=("AC", "AE", "AG", "AI", "AK", "AM", "AO"), average=139, minimum=140, result="AQ", residual="AT", note="AC141", visual=144, photos=(145, 163), caption="C164"),
    dict(title=165, calibration="M168", data=170, data_end=172, point="Y", columns=("AB", "AD", "AF", "AH", "AJ", "AL", "AN"), average=173, minimum=174, result="AP", residual="AS", note="AB175", visual=178, photos=(179, 197), caption="C198"),
)


def _numero(valor):
    return float(valor) if valor is not None else None


def _promedio(valores):
    valores = [float(v) for v in valores if v is not None]
    return sum(valores) / len(valores) if valores else None


def _minimo(valores):
    valores = [float(v) for v in valores if v is not None]
    return min(valores) if valores else None


def _ruta_foto(foto):
    try:
        ruta = Path(foto.imagen.path)
    except (AttributeError, NotImplementedError, OSError, ValueError):
        return None
    return ruta if ruta.is_file() else None


def _escribir_cabecera(ws, inspeccion, shafts):
    ws["L1"] = f"REPORTE INSPECCION {inspeccion.codigo_reporte}"
    ws["I7"] = inspeccion.get_condicion_general_display().upper()
    ws["I7"].fill = PatternFill("solid", fgColor="00B050")
    ws["I7"].font = Font(name=ws["I7"].font.name, size=ws["I7"].font.sz, bold=True, color="FFFFFF")
    valores = {
        "I10": inspeccion.planta,
        "V10": inspeccion.proceso,
        "I12": "LIFE SHAFT",
        "V12": inspeccion.faja.tag,
        "I14": inspeccion.etapa,
        "V14": inspeccion.condicion_equipo,
        "I16": inspeccion.fecha_inspeccion,
        "V16": inspeccion.fecha_reporte,
        "I18": nombre_campo(inspeccion, "inspector_campo_nombre", inspeccion.inspector),
        "V18": nombre_campo(inspeccion, "supervisor_campo_nombre", inspeccion.supervisor),
        "I20": nombre_campo(inspeccion, "analista_elabora_nombre", inspeccion.analista),
        "V20": nombre_campo(inspeccion, "analista_valida_nombre", inspeccion.analista),
        "I23": inspeccion.circunstancias,
        "I26": inspeccion.antecedentes,
        "I28": inspeccion.observaciones or generar_observaciones_life_shaft(shafts),
        "C49": f"ESQUEMA DE UBICACION DE LIVESHAFT DE LA FAJA {inspeccion.faja.tag}",
    }
    for celda, valor in valores.items():
        ws[celda] = valor or "-"
    ws["I16"].number_format = ws["V16"].number_format = "dd mmmm yyyy"
    conclusiones = (inspeccion.recomendaciones or generar_conclusiones_life_shaft(shafts)).splitlines()
    for indice, fila in enumerate(range(41, 47)):
        ws[f"I{fila}"] = conclusiones[indice] if indice < len(conclusiones) else None


def _rangos_fotos(inicio, fin, cantidad):
    if cantidad <= 1:
        return [f"C{inicio}:AX{fin}"]
    if cantidad == 2:
        return [f"E{inicio}:AA{fin}", f"AD{inicio}:AW{fin}"]
    if cantidad == 3:
        return [f"C{inicio}:T{fin}", f"U{inicio}:AI{fin}", f"AJ{inicio}:AX{fin}"]
    mitad = inicio + max((fin - inicio + 1) // 2, 1)
    rangos = [f"C{inicio}:T{mitad - 1}", f"U{inicio}:AI{mitad - 1}", f"AJ{inicio}:AX{mitad - 1}"]
    rangos.extend([f"E{mitad}:AA{fin}", f"AD{mitad}:AW{fin}"])
    return rangos[:cantidad]


def _escribir_mediciones(ws, layout, mediciones):
    mediciones = list(mediciones)
    cantidad = layout["data_end"] - layout["data"] + 1
    for offset in range(cantidad):
        fila = layout["data"] + offset
        medicion = mediciones[offset] if offset < len(mediciones) else None
        ws[f"{layout['point']}{fila}"] = medicion.punto if medicion else None
        for campo, columna in zip("abcdefg", layout["columns"]):
            ws[f"{columna}{fila}"] = _numero(getattr(medicion, campo)) if medicion else None
        ws[f"{layout['result']}{fila}"] = _numero(medicion.promedio) if medicion else None
        ws[f"{layout['residual']}{fila}"] = _numero(medicion.minimo) if medicion else None
    usadas = mediciones[:cantidad]
    for campo, columna in zip("abcdefg", layout["columns"]):
        valores = [getattr(m, campo) for m in usadas]
        ws[f"{columna}{layout['average']}"] = _promedio(valores)
        ws[f"{columna}{layout['minimum']}"] = _minimo(valores)
    ws[f"{layout['result']}{layout['average']}"] = _promedio([m.promedio for m in usadas])
    ws[f"{layout['residual']}{layout['average']}"] = _promedio([m.minimo for m in usadas])
    ws[f"{layout['result']}{layout['minimum']}"] = _minimo([m.promedio for m in usadas])
    ws[f"{layout['residual']}{layout['minimum']}"] = _minimo([m.minimo for m in usadas])


def _escribir_shaft(ws, layout, bloque, inspeccion):
    shaft = bloque["life_shaft"]
    ws[f"C{layout['visual']}"] = f"INSPECCION VISUAL DE LIVESHAFT #{shaft.numero:02d}"
    columna = ''.join(filter(str.isalpha, layout["calibration"]))
    fila = int(''.join(filter(str.isdigit, layout["calibration"])))
    parametros = (shaft.marca_equipo, shaft.tipo_haz, shaft.frecuencia_mhz, shaft.ancho_banda, shaft.amortiguamiento, shaft.velocidad_ms, shaft.retardo_us)
    if bloque.get("es_campana"):
        for numero_fila in range(layout["title"], layout["visual"]):
            ws.row_dimensions[numero_fila].hidden = True
    else:
        ws[f"C{layout['title']}"] = f"MEDICION DE ESPESORES A LIVESHAFT #{shaft.numero:02d} / {inspeccion.faja.tag}"
        for offset, valor in enumerate(parametros):
            ws[f"{columna}{fila + offset}"] = valor or "-"
        _escribir_mediciones(ws, layout, bloque["mediciones"])
        minimo = minimo_componente(shaft)
        ws[layout["note"]] = f"El espesor minimo hallado fue de {minimo[0]:.2f} mm en el punto {minimo[1]}. {etiqueta_condicion(shaft)}." if minimo else ""
    fotos = [(foto, ruta) for foto in bloque["fotografias"] if (ruta := _ruta_foto(foto))][:5]
    rangos = _rangos_fotos(*layout["photos"], len(fotos))
    ancho, alto = tamano_comun_cuadros_px(ws, rangos)
    for (_foto, ruta), rango in zip(fotos, rangos):
        insertar_imagen_ajustada(ws, ruta, rango, ancho, alto, 4)
    detalles = [
        f"Nombre: {shaft.nombre} | TAG: {shaft.tag} | Ubicacion: {shaft.ubicacion} | Condicion: {etiqueta_condicion(shaft)}"
    ]
    for etiqueta, valor in (("Observacion visual", shaft.observacion_visual), ("Observacion de medicion", shaft.observacion_medicion), ("Recomendaciones", shaft.recomendaciones)):
        if (valor or "").strip():
            detalles.append(f"{etiqueta}: {valor.strip()}")
    ws[layout["caption"]] = "\n".join(detalles) if len(detalles) > 1 else None
    if ws[layout["caption"]].value:
        alineacion = copy(ws[layout["caption"]].alignment)
        alineacion.wrap_text = True
        ws[layout["caption"]].alignment = alineacion


def generar_excel_life_shaft_cvb0004(inspeccion, bloques):
    workbook = load_workbook(BytesIO(TEMPLATE_PATH.read_bytes()))
    ws = workbook[SHEET_NAME]
    limpiar_textos_historicos_campana(ws)
    shafts = [bloque["life_shaft"] for bloque in bloques]
    _escribir_cabecera(ws, inspeccion, shafts)
    ws._images = [imagen for imagen in ws._images if imagen.anchor._from.row + 1 < 63]
    for layout, bloque in zip(LAYOUTS, bloques):
        _escribir_shaft(ws, layout, bloque, inspeccion)
    agregar_hoja_mediciones_campana(
        workbook, bloques, "life_shaft", f"MEDICIÓN DE ESPESORES DEL LIFE SHAFT {inspeccion.faja.tag}",
        config={"layouts": LAYOUTS, "title_column": "C", "calibration_column": "M", "measurement_columns": ("AB", "AD", "AF", "AH", "AJ", "AL", "AN"), "point_column": "Y", "result_column": "AP", "residual_column": "AS", "note_column": "Y"},
    )
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    salida = BytesIO()
    workbook.save(salida)
    salida.seek(0)
    return salida
