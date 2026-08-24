from copy import copy
from io import BytesIO
from math import floor
from pathlib import Path

from django.conf import settings

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.drawing.spreadsheet_drawing import (
    AnchorMarker,
    TwoCellAnchor,
)
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.utils.units import (
    EMU_to_pixels,
    pixels_to_EMU,
)
from openpyxl.worksheet.cell_range import MultiCellRange

from PIL import Image as PILImage
from PIL import ImageOps

from inspecciones.reportes.campaign_excel import (
    limpiar_textos_historicos_campana,
)

from .utils import (
    generar_conclusiones_poleas,
    generar_observaciones_poleas,
    nombre_campo,
)


# ============================================================
# CONFIGURACIÓN
# ============================================================

TEMPLATE_PATH = (
    Path(settings.BASE_DIR)
    / "inspecciones/reportes/cvb0004/templates/"
    / "20260517-VTUT-0220CVB004-POLEAS.xlsx"
)

# Columnas donde están A-G en la tabla original.
COLUMNAS = (
    "AC",
    "AE",
    "AG",
    "AI",
    "AK",
    "AM",
    "AO",
)

# El reporte real trabaja hasta AX.
# Nunca copiar más allá de esta columna.
MAX_COLUMNA_REPORTE = column_index_from_string("AX")


LAYOUTS = {
    1: (
        dict(
            title=82,
            calibration=85,
            data=87,
            average=92,
            minimum=93,
            note="AC94",
            visual=96,
        ),
    ),

    2: (
        dict(
            title=116,
            calibration=119,
            data=121,
            average=126,
            minimum=127,
            note="AC128",
            visual=130,
        ),
    ),

    3: (
        dict(
            title=146,
            calibration=149,
            data=151,
            average=156,
            minimum=157,
            note="Z158",
            visual=160,
        ),
        dict(
            title=183,
            calibration=188,
            data=190,
            average=195,
            minimum=196,
            note="Z197",
            visual=200,
        ),
    ),

    # 04 y 05 utilizan su estructura cerámica.
    4: (
        dict(
            visual=222,
        ),
    ),

    5: (
        dict(
            visual=245,
        ),
    ),

    6: (
        dict(
            title=266,
            calibration=269,
            data=271,
            average=276,
            minimum=277,
            note="AC278",
            visual=280,
        ),
    ),

    7: (
        dict(
            title=307,
            calibration=310,
            data=312,
            average=317,
            minimum=318,
            note="Z319",
            visual=321,
        ),
    ),

    8: (
        dict(
            title=344,
            calibration=347,
            data=349,
            average=354,
            minimum=355,
            note="Z356",
            visual=358,
        ),
        dict(
            title=380,
            calibration=383,
            data=385,
            average=390,
            minimum=391,
            note="Z392",
            visual=394,
        ),
    ),

    9: (
        dict(
            title=416,
            calibration=419,
            data=421,
            average=426,
            minimum=427,
            note="AC428",
            visual=430,
        ),
    ),
}


CAPTIONS = {
    1: "D115",
    2: "D145",
    3: "D221",
    4: "D243",
    5: "D264",
    6: "D305",
    7: "D343",
    8: "D415",
    9: "D451",
}


# ============================================================
# COLORES
# ============================================================

AMARILLO_MINIMO = PatternFill(
    fill_type="solid",
    fgColor="FFF200",
)

GRIS_DESCRIPCION = PatternFill(
    fill_type="solid",
    fgColor="D9D9D9",
)


# ============================================================
# UTILIDADES
# ============================================================

def _numero(valor):
    if valor is None:
        return None

    return float(valor)


def _promedio(valores):
    datos = [
        float(valor)
        for valor in valores
        if valor is not None
    ]

    if not datos:
        return None

    return sum(datos) / len(datos)


def _minimo(valores):
    datos = [
        float(valor)
        for valor in valores
        if valor is not None
    ]

    if not datos:
        return None

    return min(datos)


def _minimo_global_mediciones(mediciones):

    candidatos = []

    for medicion in mediciones:

        for campo in "abcdefg":

            valor = getattr(
                medicion,
                campo,
                None,
            )

            if valor is None:
                continue

            candidatos.append(
                (
                    float(valor),
                    campo.upper(),
                )
            )

    if not candidatos:
        return None, None

    return min(
        candidatos,
        key=lambda item: item[0],
    )


# ============================================================
# FOTOGRAFÍAS
# ============================================================

def _ruta_foto(foto):

    try:
        ruta = Path(
            foto.imagen.path
        )

    except (
        AttributeError,
        NotImplementedError,
        OSError,
        ValueError,
    ):
        return None

    if not ruta.is_file():
        return None

    return ruta


def _anclas_fotos_originales(ws):
    """
    Guarda las posiciones originales de las fotografías del Excel.

    Esas posiciones serán reutilizadas para las fotografías
    cargadas desde la aplicación web.
    """

    anclas = {
        numero: []
        for numero in LAYOUTS
    }

    limites = {}

    for numero, layouts in LAYOUTS.items():

        fila_visual = min(
            layout["visual"]
            for layout in layouts
        )

        _, fila_caption = (
            coordinate_from_string(
                CAPTIONS[numero]
            )
        )

        limites[numero] = (
            fila_visual,
            fila_caption,
        )

    for imagen in ws._images:

        anchor = getattr(
            imagen,
            "anchor",
            None,
        )

        if not hasattr(
            anchor,
            "_from",
        ):
            continue

        if not hasattr(
            anchor,
            "to",
        ):
            continue

        fila = (
            anchor._from.row
            + 1
        )

        for numero, (
            inicio,
            final,
        ) in limites.items():

            if inicio < fila <= final:

                anclas[numero].append(
                    copy(anchor)
                )

                break

    for lista in anclas.values():

        lista.sort(
            key=lambda anchor: (
                anchor._from.row,
                anchor._from.col,
            )
        )

    return anclas


def _es_foto_historica_polea(imagen):
    """
    Identifica únicamente fotografías históricas.

    NO elimina:
    - logos
    - diagramas
    - esquema de ubicación
    - imágenes técnicas fijas
    """

    anchor = getattr(
        imagen,
        "anchor",
        None,
    )

    if not hasattr(
        anchor,
        "_from",
    ):
        return False

    fila = (
        anchor._from.row
        + 1
    )

    for numero, layouts in LAYOUTS.items():

        visual = min(
            layout["visual"]
            for layout in layouts
        )

        _, caption = (
            coordinate_from_string(
                CAPTIONS[numero]
            )
        )

        if visual < fila <= caption:
            return True

    return False


# ============================================================
# DIMENSIONES DEL EXCEL
# ============================================================

def _firma_dimensiones_columnas(ws):

    return tuple(
        (
            clave,
            dimension.width,
            dimension.hidden,
            dimension.min,
            dimension.max,
            dimension.bestFit,
            dimension.outlineLevel,
            dimension.collapsed,
        )
        for clave, dimension
        in ws.column_dimensions.items()
    )


def _dimension_columna(ws, numero):

    for dimension in ws.column_dimensions.values():

        minimo = (
            dimension.min
            if dimension.min is not None
            else numero
        )

        maximo = (
            dimension.max
            if dimension.max is not None
            else minimo
        )

        if minimo <= numero <= maximo:
            return dimension

    return None


def _ancho_columna_px(
    ws,
    numero,
):

    dimension = _dimension_columna(
        ws,
        numero,
    )

    if (
        dimension
        and dimension.width
        is not None
    ):
        ancho = dimension.width

    else:
        ancho = (
            ws.sheet_format.defaultColWidth
            or 8.43
        )

    return max(
        floor(
            (
                (
                    256 * ancho
                    + floor(128 / 7)
                )
                / 256
            )
            * 7
        ),
        1,
    )


def _alto_fila_px(
    ws,
    numero,
):

    dimension = (
        ws.row_dimensions.get(
            numero
        )
    )

    if (
        dimension
        and dimension.height
        is not None
    ):
        alto = dimension.height

    else:
        alto = (
            ws.sheet_format.defaultRowHeight
            or 15
        )

    return max(
        int(
            round(
                alto * 96 / 72
            )
        ),
        1,
    )


# ============================================================
# POSICIÓN EXACTA DE IMÁGENES
# ============================================================

def _posicion_marcador_px(
    ws,
    marcador,
):

    x = sum(
        _ancho_columna_px(
            ws,
            columna,
        )
        for columna in range(
            1,
            marcador.col + 1,
        )
    )

    y = sum(
        _alto_fila_px(
            ws,
            fila,
        )
        for fila in range(
            1,
            marcador.row + 1,
        )
    )

    return (
        x + EMU_to_pixels(
            marcador.colOff
        ),
        y + EMU_to_pixels(
            marcador.rowOff
        ),
    )


def _marcador_desde_px(
    ws,
    x,
    y,
):

    columna = 1

    restante_x = max(
        float(x),
        0,
    )

    while (
        restante_x
        >= _ancho_columna_px(
            ws,
            columna,
        )
    ):

        restante_x -= (
            _ancho_columna_px(
                ws,
                columna,
            )
        )

        columna += 1

    fila = 1

    restante_y = max(
        float(y),
        0,
    )

    while (
        restante_y
        >= _alto_fila_px(
            ws,
            fila,
        )
    ):

        restante_y -= (
            _alto_fila_px(
                ws,
                fila,
            )
        )

        fila += 1

    return AnchorMarker(
        col=columna - 1,
        row=fila - 1,
        colOff=pixels_to_EMU(
            restante_x
        ),
        rowOff=pixels_to_EMU(
            restante_y
        ),
    )


def _imagen_normalizada(ruta):

    with PILImage.open(
        ruta
    ) as origen:

        orientacion = (
            origen
            .getexif()
            .get(274, 1)
        )

        ajustada = (
            ImageOps.exif_transpose(
                origen
            )
        )

        ancho, alto = (
            ajustada.size
        )

        if orientacion == 1:

            return (
                ExcelImage(
                    str(ruta)
                ),
                ancho,
                alto,
            )

        normalizada = (
            ajustada.copy()
        )

        if normalizada.mode not in {
            "RGB",
            "RGBA",
        }:

            normalizada = (
                normalizada.convert(
                    "RGB"
                )
            )

    buffer = BytesIO()

    if normalizada.mode == "RGBA":

        normalizada.save(
            buffer,
            format="PNG",
        )

    else:

        normalizada.save(
            buffer,
            format="JPEG",
            quality=90,
            optimize=True,
        )

    buffer.seek(0)

    imagen = ExcelImage(
        buffer
    )

    # Evita que Python cierre el buffer antes de guardar.
    imagen._cvb0004_buffer = (
        buffer
    )

    return (
        imagen,
        ancho,
        alto,
    )


def _insertar_en_ancla_original(
    ws,
    ruta,
    ancla_original,
    margen_px=4,
):
    """
    La imagen se adapta al cuadro.
    El cuadro NO cambia por la imagen.
    """

    try:

        (
            imagen,
            ancho_original,
            alto_original,
        ) = _imagen_normalizada(
            ruta
        )

    except (
        FileNotFoundError,
        OSError,
        ValueError,
    ):
        return False

    x1, y1 = (
        _posicion_marcador_px(
            ws,
            ancla_original._from,
        )
    )

    x2, y2 = (
        _posicion_marcador_px(
            ws,
            ancla_original.to,
        )
    )

    ancho_disponible = max(
        x2
        - x1
        - margen_px * 2,
        1,
    )

    alto_disponible = max(
        y2
        - y1
        - margen_px * 2,
        1,
    )

    escala = min(
        ancho_disponible
        / ancho_original,
        alto_disponible
        / alto_original,
    )

    ancho_final = (
        ancho_original
        * escala
    )

    alto_final = (
        alto_original
        * escala
    )

    inicio_x = (
        x1
        + (
            x2
            - x1
            - ancho_final
        )
        / 2
    )

    inicio_y = (
        y1
        + (
            y2
            - y1
            - alto_final
        )
        / 2
    )

    fin_x = (
        inicio_x
        + ancho_final
    )

    fin_y = (
        inicio_y
        + alto_final
    )

    imagen.anchor = TwoCellAnchor(
        editAs="twoCell",

        _from=_marcador_desde_px(
            ws,
            inicio_x,
            inicio_y,
        ),

        to=_marcador_desde_px(
            ws,
            fin_x,
            fin_y,
        ),
    )

    ws.add_image(
        imagen
    )

    return True


# ============================================================
# DESPLAZAMIENTO VERTICAL
# ============================================================

def _desplazar_ancla_vertical(
    anchor,
    fila_insercion,
    cantidad,
):

    if (
        hasattr(
            anchor,
            "_from",
        )
        and (
            anchor._from.row
            + 1
            >= fila_insercion
        )
    ):

        anchor._from.row += (
            cantidad
        )

    if (
        hasattr(
            anchor,
            "to",
        )
        and (
            anchor.to.row
            + 1
            >= fila_insercion
        )
    ):

        anchor.to.row += (
            cantidad
        )


def _desplazar_celda_vertical(
    coordenada,
    fila_insercion,
    cantidad,
):

    columna, fila = (
        coordinate_from_string(
            coordenada
        )
    )

    if fila >= fila_insercion:
        fila += cantidad

    return (
        f"{columna}{fila}"
    )


def _desplazar_layouts(
    layouts,
    captions,
    anclas_fotos,
    fila_insercion,
    cantidad,
):

    for lista in layouts.values():

        for layout in lista:

            for clave, valor in tuple(
                layout.items()
            ):

                if clave == "note":

                    layout[clave] = (
                        _desplazar_celda_vertical(
                            valor,
                            fila_insercion,
                            cantidad,
                        )
                    )

                elif (
                    isinstance(
                        valor,
                        int,
                    )
                    and valor
                    >= fila_insercion
                ):

                    layout[clave] = (
                        valor
                        + cantidad
                    )

    for numero, coordenada in tuple(
        captions.items()
    ):

        captions[numero] = (
            _desplazar_celda_vertical(
                coordenada,
                fila_insercion,
                cantidad,
            )
        )

    for lista in (
        anclas_fotos.values()
    ):

        for anchor in lista:

            _desplazar_ancla_vertical(
                anchor,
                fila_insercion,
                cantidad,
            )


# ============================================================
# DUPLICAR TABLA NORMAL VERTICALMENTE
# ============================================================

def _insertar_copia_vertical(
    ws,
    fila_inicio,
    fila_fin,
    fila_insercion,
):
    """
    Duplica un bloque técnico verticalmente.

    OBJETIVO:
    - Mantener la estructura original.
    - No modificar anchos de columnas.
    - No crear columnas nuevas.
    - No romper merges.
    - No cortar encabezados.
    - No mover horizontalmente las tablas.
    - Crecimiento únicamente VERTICAL.
    """

    cantidad = fila_fin - fila_inicio + 1

    # El ancho útil del reporte original llega hasta AX.
    max_columna = column_index_from_string("AX")

    # =========================================================
    # 1. GUARDAR CELDAS DEL BLOQUE ORIGINAL
    # =========================================================

    celdas_origen = {}

    for fila in range(
        fila_inicio,
        fila_fin + 1,
    ):
        for columna in range(
            1,
            max_columna + 1,
        ):
            celda = ws.cell(
                row=fila,
                column=columna,
            )

            celdas_origen[
                (
                    fila - fila_inicio,
                    columna,
                )
            ] = {
                "value": celda.value,
                "style": copy(celda._style),
                "alignment": copy(celda.alignment),
                "protection": copy(celda.protection),
            }

    # =========================================================
    # 2. GUARDAR ALTURAS DE FILAS
    # =========================================================

    dimensiones_originales = {
        numero: copy(dimension)
        for numero, dimension
        in ws.row_dimensions.items()
    }

    # =========================================================
    # 3. GUARDAR MERGES
    # =========================================================

    merges_originales = list(
        ws.merged_cells.ranges
    )

    merges_del_bloque = []

    for rango in merges_originales:

        if (
            fila_inicio <= rango.min_row
            and rango.max_row <= fila_fin
            and rango.max_col <= max_columna
        ):
            merges_del_bloque.append(
                (
                    rango.min_row - fila_inicio,
                    rango.max_row - fila_inicio,
                    rango.min_col,
                    rango.max_col,
                )
            )

    # =========================================================
    # 4. INSERTAR FILAS
    # =========================================================

    ws.insert_rows(
        fila_insercion,
        amount=cantidad,
    )

    # =========================================================
    # 5. RECONSTRUIR ALTURAS
    # =========================================================

    ws.row_dimensions.clear()

    for numero, dimension in (
        dimensiones_originales.items()
    ):

        if numero >= fila_insercion:
            destino = numero + cantidad
        else:
            destino = numero

        nueva_dimension = copy(
            dimension
        )

        nueva_dimension.index = destino

        ws.row_dimensions[
            destino
        ] = nueva_dimension

    # Copiar las alturas originales al bloque nuevo.
    for offset in range(cantidad):

        origen = fila_inicio + offset
        destino = fila_insercion + offset

        if origen in dimensiones_originales:

            nueva_dimension = copy(
                dimensiones_originales[
                    origen
                ]
            )

            nueva_dimension.index = destino

            ws.row_dimensions[
                destino
            ] = nueva_dimension

    # =========================================================
    # 6. COPIAR CELDAS
    # =========================================================

    for (
        offset,
        columna,
    ), datos in celdas_origen.items():

        destino = ws.cell(
            row=fila_insercion + offset,
            column=columna,
        )

        destino.value = datos[
            "value"
        ]

        destino._style = copy(
            datos["style"]
        )

        destino.alignment = copy(
            datos["alignment"]
        )

        destino.protection = copy(
            datos["protection"]
        )

    # =========================================================
    # 7. RECONSTRUIR MERGES SIN ESTIRARLOS
    # =========================================================

    merges_nuevos = []

    for rango in merges_originales:

        # Merge completamente encima.
        if rango.max_row < fila_insercion:

            merges_nuevos.append(
                (
                    rango.min_row,
                    rango.max_row,
                    rango.min_col,
                    rango.max_col,
                )
            )

        # Merge completamente debajo.
        elif rango.min_row >= fila_insercion:

            merges_nuevos.append(
                (
                    rango.min_row + cantidad,
                    rango.max_row + cantidad,
                    rango.min_col,
                    rango.max_col,
                )
            )

        # Merge que cruza el punto de inserción.
        # NO se estira.
        else:

            merges_nuevos.append(
                (
                    rango.min_row,
                    rango.max_row,
                    rango.min_col,
                    rango.max_col,
                )
            )

    # =========================================================
    # 8. COPIAR MERGES DEL BLOQUE NUEVO
    # =========================================================

    for (
        min_offset,
        max_offset,
        min_col,
        max_col,
    ) in merges_del_bloque:

        merges_nuevos.append(
            (
                fila_insercion + min_offset,
                fila_insercion + max_offset,
                min_col,
                max_col,
            )
        )

    # Quitar duplicados.
    merges_unicos = []
    vistos = set()

    for merge in merges_nuevos:

        if merge in vistos:
            continue

        vistos.add(merge)
        merges_unicos.append(
            merge
        )

    # Eliminar los merges actuales y reconstruir.
    ws.merged_cells = MultiCellRange()

    for (
        min_row,
        max_row,
        min_col,
        max_col,
    ) in merges_unicos:

        try:
            ws.merge_cells(
                start_row=min_row,
                end_row=max_row,
                start_column=min_col,
                end_column=max_col,
            )

        except ValueError:
            continue

    # =========================================================
    # 9. DESPLAZAR IMÁGENES
    # =========================================================

    for imagen in ws._images:

        anchor = getattr(
            imagen,
            "anchor",
            None,
        )

        if anchor is None:
            continue

        _desplazar_ancla_vertical(
            anchor,
            fila_insercion,
            cantidad,
        )

    # =========================================================
    # 10. DESPLAZAR SALTOS DE PÁGINA
    # =========================================================

    for salto in ws.row_breaks.brk:

        if salto.id >= fila_insercion:
            salto.id += cantidad

    return cantidad
def _preparar_bloques_verticales_campana(
    ws,
    bloques,
    anclas_fotos,
):

    layouts = {
        numero: [
            dict(layout)
            for layout in lista
        ]
        for numero, lista
        in LAYOUTS.items()
    }

    captions = dict(
        CAPTIONS
    )

    fila_final_original = (
        ws.max_row
    )

    filas_agregadas = 0

    # Trabajar de abajo hacia arriba.
    bloques_campana = [
        bloque
        for bloque in bloques
        if bloque.get(
            "es_campana"
        )
    ]

    bloques_campana.sort(
        key=lambda bloque: max(
            (
                layout.get(
                    "title",
                    0,
                )
                for layout
                in layouts[
                    bloque["polea"].numero
                ]
            ),
            default=0,
        ),
        reverse=True,
    )

    for bloque in (
        bloques_campana
    ):

        polea = bloque[
            "polea"
        ]

        tecnicos = [
            layout
            for layout
            in layouts[
                polea.numero
            ]
            if "title"
            in layout
        ]

        if not tecnicos:
            continue

        # Si ya existen dos tablas técnicas en la plantilla,
        # NO crear una tercera.
        if len(tecnicos) >= 2:
            continue

        inicio = tecnicos[0]

        _, fila_nota = (
            coordinate_from_string(
                inicio["note"]
            )
        )

        fila_insercion = (
            fila_nota + 1
        )

        cantidad = (
            _insertar_copia_vertical(
                ws,
                inicio["title"],
                fila_nota,
                fila_insercion,
            )
        )

        _desplazar_layouts(
            layouts,
            captions,
            anclas_fotos,
            fila_insercion,
            cantidad,
        )

        fin = {}

        for clave, valor in (
            inicio.items()
        ):

            if clave == "visual":
                continue

            if clave == "note":

                columna, fila = (
                    coordinate_from_string(
                        valor
                    )
                )

                fin[clave] = (
                    f"{columna}"
                    f"{fila + cantidad}"
                )

            else:

                fin[clave] = (
                    valor + cantidad
                )

        posicion = (
            layouts[
                polea.numero
            ].index(inicio)
            + 1
        )

        layouts[
            polea.numero
        ].insert(
            posicion,
            fin,
        )

        filas_agregadas += (
            cantidad
        )

    if filas_agregadas:

        ws.print_area = (
            f"A1:AX"
            f"{fila_final_original + filas_agregadas}"
        )

    return (
        layouts,
        captions,
    )


# ============================================================
# CABECERA
# ============================================================

def _escribir_cabecera(
    ws,
    inspeccion,
    poleas,
):

    ws["M3"] = (
        "REPORTE INSPECCION "
        f"{inspeccion.codigo_reporte}"
    )

    ws["K8"] = (
        inspeccion
        .get_condicion_general_display()
        .upper()
    )

    ws["K8"].fill = (
        PatternFill(
            fill_type="solid",
            fgColor="00B050",
        )
    )

    fuente = ws["K8"].font

    ws["K8"].font = Font(
        name=fuente.name,
        size=fuente.sz,
        bold=True,
        color="FFFFFF",
    )

    valores = {

        "K11":
            inspeccion.planta,

        "Y11":
            inspeccion.proceso,

        "K13":
            "POLEAS",

        "Y13":
            inspeccion.faja.tag,

        "K15":
            inspeccion.etapa,

        "Y15":
            inspeccion.condicion_equipo,

        "K17":
            inspeccion.fecha_inspeccion,

        "Y17":
            inspeccion.fecha_reporte,

        "K19":
            nombre_campo(
                inspeccion,
                "inspector_campo_nombre",
                inspeccion.inspector,
            ),

        "Y19":
            nombre_campo(
                inspeccion,
                "supervisor_campo_nombre",
                inspeccion.supervisor,
            ),

        "K21":
            nombre_campo(
                inspeccion,
                "analista_elabora_nombre",
                inspeccion.analista,
            ),

        "Y21":
            nombre_campo(
                inspeccion,
                "analista_valida_nombre",
                inspeccion.analista,
            ),

        "K24":
            inspeccion.circunstancias,

        "K27":
            inspeccion.antecedentes,

        "K29":
            (
                inspeccion.observaciones
                or generar_observaciones_poleas(
                    poleas
                )
            ),

        "D68":
            (
                "ESQUEMA DE UBICACION "
                "DE POLEAS DE LA FAJA "
                f"{inspeccion.faja.tag}"
            ),
    }

    for celda, valor in (
        valores.items()
    ):

        ws[celda] = (
            valor
            if valor not in (
                None,
                "",
            )
            else "-"
        )

    ws["K17"].number_format = (
        "dd mmmm yyyy"
    )

    ws["Y17"].number_format = (
        "dd mmmm yyyy"
    )

    conclusiones = (
        inspeccion.recomendaciones
        or generar_conclusiones_poleas(
            poleas
        )
    )

    lineas = (
        conclusiones.splitlines()
    )

    for indice, fila in enumerate(
        range(55, 66)
    ):

        ws[f"K{fila}"] = (
            lineas[indice]
            if indice
            < len(lineas)
            else None
        )


# ============================================================
# MEDICIONES + MÍNIMO DESTACADO
# ============================================================

def _escribir_mediciones(
    ws,
    layout,
    mediciones,
):

    mediciones = list(
        mediciones
    )[:5]

    candidatos = []

    for offset in range(5):

        fila = (
            layout["data"]
            + offset
        )

        medicion = (
            mediciones[offset]
            if offset
            < len(mediciones)
            else None
        )

        ws[f"Z{fila}"] = (
            medicion.punto
            if medicion
            else None
        )

        for campo, columna in zip(
            "abcdefg",
            COLUMNAS,
        ):

            valor = (
                getattr(
                    medicion,
                    campo,
                    None,
                )
                if medicion
                else None
            )

            ws[
                f"{columna}{fila}"
            ] = _numero(
                valor
            )

            if valor is not None:

                candidatos.append(
                    (
                        float(valor),
                        campo.upper(),
                        f"{columna}{fila}",
                    )
                )

        ws[f"AQ{fila}"] = (
            _numero(
                medicion.promedio
            )
            if medicion
            else None
        )

        ws[f"AT{fila}"] = (
            _numero(
                medicion.minimo
            )
            if medicion
            else None
        )

    # Promedio y mínimo por letra.
    for campo, columna in zip(
        "abcdefg",
        COLUMNAS,
    ):

        valores = [
            getattr(
                medicion,
                campo,
            )
            for medicion
            in mediciones
        ]

        ws[
            f"{columna}"
            f"{layout['average']}"
        ] = _promedio(
            valores
        )

        ws[
            f"{columna}"
            f"{layout['minimum']}"
        ] = _minimo(
            valores
        )

    ws[
        f"AQ{layout['average']}"
    ] = _promedio(
        [
            medicion.promedio
            for medicion
            in mediciones
        ]
    )

    ws[
        f"AT{layout['average']}"
    ] = _promedio(
        [
            medicion.minimo
            for medicion
            in mediciones
        ]
    )

    ws[
        f"AQ{layout['minimum']}"
    ] = _minimo(
        [
            medicion.promedio
            for medicion
            in mediciones
        ]
    )

    # Si no hay datos.
    if not candidatos:

        ws[
            f"AT{layout['minimum']}"
        ] = None

        return None, None

    # Mínimo REAL de toda la tabla.
    valor_minimo = min(
        valor
        for (
            valor,
            _punto,
            _celda,
        ) in candidatos
    )

    primer_minimo = next(
        item
        for item in candidatos
        if abs(
            item[0]
            - valor_minimo
        ) < 1e-9
    )

    punto_minimo = (
        primer_minimo[1]
    )

    # Resaltar en la propia tabla.
    for (
        valor,
        _punto,
        coordenada,
    ) in candidatos:

        if abs(
            valor
            - valor_minimo
        ) < 1e-9:

            celda = ws[
                coordenada
            ]

            celda.fill = copy(
                AMARILLO_MINIMO
            )

            fuente = (
                celda.font
            )

            celda.font = Font(
                name=(
                    fuente.name
                    or "Arial"
                ),
                size=fuente.sz,
                bold=True,
                color="FF0000",
            )

    # Resumen mínimo inferior derecho.
    celda_resumen = ws[
        f"AT{layout['minimum']}"
    ]

    celda_resumen.value = (
        valor_minimo
    )

    celda_resumen.fill = copy(
        AMARILLO_MINIMO
    )

    fuente = (
        celda_resumen.font
    )

    celda_resumen.font = Font(
        name=(
            fuente.name
            or "Arial"
        ),
        size=fuente.sz,
        bold=True,
        color="FF0000",
    )

    return (
        valor_minimo,
        punto_minimo,
    )


# ============================================================
# TABLA TÉCNICA
# ============================================================

def _escribir_layout(
    ws,
    layout,
    polea,
    mediciones,
    tag,
    fase="",
):
    """
    Escribe el bloque técnico de una polea.

    Ajustes visuales:
    - título centrado;
    - encabezado técnico centrado;
    - mantiene estilo original;
    - mensaje inferior con altura uniforme;
    - no modifica columnas ni merges.
    """

    # ========================================================
    # TÍTULO GENERAL
    # ========================================================

    if fase:
        titulo = (
            "MEDICION DE ESPESORES "
            "DEL LAGGING DE LA POLEA "
            f"#{polea.numero:02d} / "
            f"{tag} - "
            f"{fase} DE CAMPAÑA"
        )
    else:
        titulo = (
            "MEDICION DE ESPESORES "
            "DEL LAGGING DE LA POLEA "
            f"#{polea.numero:02d} / "
            f"{tag}"
        )

    celda_titulo = ws[
        f"D{layout['title']}"
    ]

    celda_titulo.value = titulo

    alineacion_titulo = copy(
        celda_titulo.alignment
    )

    alineacion_titulo.horizontal = "center"
    alineacion_titulo.vertical = "center"

    celda_titulo.alignment = (
        alineacion_titulo
    )

    # ========================================================
    # ENCABEZADO DE TABLA
    # ========================================================

    encabezado = (
        "MEDICIÓN DE ESPESORES "
        "DEL LAGGING DE LA POLEA "
        f"{polea.numero}"
    )

    if fase:
        encabezado += (
            f" - {fase} DE CAMPAÑA"
        )

    celda_encabezado = ws[
        f"W{layout['calibration'] - 1}"
    ]

    celda_encabezado.value = (
        encabezado
    )

    alineacion_encabezado = copy(
        celda_encabezado.alignment
    )

    alineacion_encabezado.horizontal = (
        "center"
    )

    alineacion_encabezado.vertical = (
        "center"
    )

    celda_encabezado.alignment = (
        alineacion_encabezado
    )

    # ========================================================
    # PARÁMETROS DE CALIBRACIÓN
    # ========================================================

    parametros = (
        polea.marca_equipo,
        polea.modelo_equipo,
        polea.frecuencia_mhz,
        polea.rango_mm,
        polea.metodo_empleado,
        polea.velocidad_ms,
        polea.retardo_us,
    )

    for offset, valor in enumerate(
        parametros
    ):

        ws[
            f"N"
            f"{layout['calibration'] + offset}"
        ] = (
            valor
            if valor not in (
                None,
                "",
            )
            else "-"
        )

    # ========================================================
    # MEDICIONES
    # ========================================================

    (
        minimo,
        punto,
    ) = _escribir_mediciones(
        ws,
        layout,
        mediciones,
    )

    # ========================================================
    # MENSAJE AUTOMÁTICO
    # ========================================================

    celda_nota = ws[
        layout["note"]
    ]

    if minimo is None:
        celda_nota.value = ""

    else:

        if fase:
            mensaje = (
                f"{fase.capitalize()} de campaña: "
                f"el espesor mínimo encontrado "
                f"es de {minimo:.2f} mm "
                f"en el punto {punto}."
            )

        else:
            mensaje = (
                "El espesor mínimo encontrado "
                f"es de {minimo:.2f} mm "
                f"en el punto {punto}."
            )

        celda_nota.value = mensaje

    # ========================================================
    # FORMATO DEL MENSAJE
    # ========================================================

    alineacion_nota = copy(
        celda_nota.alignment
    )

    alineacion_nota.horizontal = (
        "center"
    )

    alineacion_nota.vertical = (
        "center"
    )

    alineacion_nota.wrap_text = True

    celda_nota.alignment = (
        alineacion_nota
    )

    _, fila_nota = (
        coordinate_from_string(
            layout["note"]
        )
    )

    # Separación visual uniforme.
    ws.row_dimensions[
        fila_nota
    ].height = 18


def _limpiar_layout_variable(
    ws,
    layout,
):

    ws[
        f"D{layout['title']}"
    ] = None

    for offset in range(7):

        ws[
            f"N"
            f"{layout['calibration'] + offset}"
        ] = None

    _escribir_mediciones(
        ws,
        layout,
        [],
    )

    ws[
        layout["note"]
    ] = None


# ============================================================
# NORMAL / CAMPAÑA
# ============================================================

def _escribir_bloques(
    ws,
    polea,
    bloque,
    tag,
    layouts_por_polea,
):
    """
    Escribe NORMAL o CAMPAÑA.

    NORMAL:
        una sola tabla.

    CAMPAÑA:
        Inicio + Fin.

    También uniformiza la barra de INSPECCIÓN VISUAL.
    """

    layouts = (
        layouts_por_polea[
            polea.numero
        ]
    )

    tecnicos = [
        layout
        for layout in layouts
        if "title" in layout
    ]

    # ========================================================
    # CAMPAÑA
    # ========================================================

    if bloque.get(
        "es_campana"
    ):

        if len(tecnicos) >= 2:

            _escribir_layout(
                ws,
                tecnicos[0],
                polea,
                bloque.get(
                    "mediciones_inicio",
                    [],
                ),
                tag,
                "INICIO",
            )

            _escribir_layout(
                ws,
                tecnicos[1],
                polea,
                bloque.get(
                    "mediciones_fin",
                    [],
                ),
                tag,
                "FIN",
            )

            # Si hubiera algún bloque técnico sobrante,
            # limpiarlo.
            for layout in tecnicos[2:]:

                _limpiar_layout_variable(
                    ws,
                    layout,
                )

    # ========================================================
    # NORMAL
    # ========================================================

    elif tecnicos:

        _escribir_layout(
            ws,
            tecnicos[0],
            polea,
            bloque.get(
                "mediciones",
                [],
            ),
            tag,
        )

        # Limpiar cualquier bloque técnico adicional.
        for layout in tecnicos[1:]:

            _limpiar_layout_variable(
                ws,
                layout,
            )

            if "note" in layout:
                ws[
                    layout["note"]
                ] = None

    # ========================================================
    # BARRA INSPECCIÓN VISUAL
    # ========================================================

    fila_visual = (
        layouts[0]["visual"]
    )

    celda_visual = ws[
        f"D{fila_visual}"
    ]

    celda_visual.value = (
        "INSPECCION VISUAL "
        "DE LA POLEA "
        f"#{polea.numero:02d} / "
        f"{tag}"
    )

    alineacion_visual = copy(
        celda_visual.alignment
    )

    alineacion_visual.horizontal = (
        "center"
    )

    alineacion_visual.vertical = (
        "center"
    )

    celda_visual.alignment = (
        alineacion_visual
    )

    # Altura uniforme de la barra.
    ws.row_dimensions[
        fila_visual
    ].height = 22

def _observacion_medicion_automatica(
    bloque,
):

    if bloque.get(
        "es_campana"
    ):

        inicio = list(
            bloque.get(
                "mediciones_inicio",
                [],
            )
        )

        fin = list(
            bloque.get(
                "mediciones_fin",
                [],
            )
        )

        minimo_inicio, punto_inicio = (
            _minimo_global_mediciones(
                inicio
            )
        )

        minimo_fin, punto_fin = (
            _minimo_global_mediciones(
                fin
            )
        )

        textos = []

        if minimo_inicio is not None:

            textos.append(
                "Inicio de campaña: "
                "espesor mínimo "
                f"{minimo_inicio:.2f} mm "
                f"en el punto "
                f"{punto_inicio}."
            )

        if minimo_fin is not None:

            textos.append(
                "Fin de campaña: "
                "espesor mínimo "
                f"{minimo_fin:.2f} mm "
                f"en el punto "
                f"{punto_fin}."
            )

        return " ".join(
            textos
        )

    minimo, punto = (
        _minimo_global_mediciones(
            list(
                bloque.get(
                    "mediciones",
                    [],
                )
            )
        )
    )

    if minimo is None:
        return ""

    return (
        "Espesor mínimo hallado fue "
        f"de {minimo:.2f} mm "
        f"en el punto {punto}."
    )


# ============================================================
# RECUADRO GRIS DE INSPECCIÓN VISUAL
# ============================================================

def _aplicar_estilo_caption(
    ws,
    coordenada,
):
    """
    Formato final del recuadro gris de Inspección Visual.

    No modifica columnas ni merges.
    """

    caption = ws[
        coordenada
    ]

    # Fondo gris.
    caption.fill = PatternFill(
        fill_type="solid",
        fgColor="D9D9D9",
    )

    # Mantener tipografía compatible con la plantilla.
    fuente_actual = (
        caption.font
    )

    caption.font = Font(
        name=(
            fuente_actual.name
            or "Arial"
        ),
        size=9,
        bold=False,
        color="000000",
    )

    # Texto centrado y ajustado.
    alineacion = copy(
        caption.alignment
    )

    alineacion.horizontal = (
        "center"
    )

    alineacion.vertical = (
        "center"
    )

    alineacion.wrap_text = True

    caption.alignment = (
        alineacion
    )

    # Altura del recuadro.
    _, fila = (
        coordinate_from_string(
            coordenada
        )
    )

    ws.row_dimensions[
        fila
    ].height = 30

def _escribir_fotos(
    ws,
    polea,
    bloque,
    fotos,
    vistos,
    anclas_originales,
    captions,
):
    """
    Inserta las fotografías cargadas desde la aplicación web
    dentro de los cuadros originales de la plantilla.

    REGLA:
    La fotografía se adapta al cuadro.
    El cuadro NO cambia de tamaño.

    Además genera el recuadro gris inferior con:
    - condición
    - observación visual
    - observación de medición
    - recomendaciones
    """

    fotos_validas = []

    # ========================================================
    # 1. OBTENER FOTOS VÁLIDAS Y EVITAR DUPLICADOS
    # ========================================================

    for foto in fotos:

        ruta = _ruta_foto(
            foto
        )

        if ruta is None:
            continue

        clave = str(
            ruta.resolve()
        ).casefold()

        if clave in vistos:
            continue

        vistos.add(clave)

        fotos_validas.append(
            ruta
        )

    # ========================================================
    # 2. OBTENER LOS CUADROS ORIGINALES
    # ========================================================

    espacios = (
        anclas_originales.get(
            polea.numero,
            [],
        )
    )

    # ========================================================
    # 3. INSERTAR LAS FOTOS
    # ========================================================

    for ruta, ancla in zip(
        fotos_validas,
        espacios,
    ):

        _insertar_en_ancla_original(
            ws,
            ruta,
            ancla,
            margen_px=5,
        )

    # ========================================================
    # 4. OBSERVACIÓN VISUAL
    # ========================================================

    observacion_visual = (
        polea.observacion_visual
        or ""
    ).strip()

    if not observacion_visual:
        observacion_visual = (
            "Sin observaciones visuales relevantes."
        )

    # ========================================================
    # 5. OBSERVACIÓN DE MEDICIÓN
    # ========================================================

    observacion_medicion = (
        polea.observacion_medicion
        or ""
    ).strip()

    if not observacion_medicion:

        observacion_medicion = (
            _observacion_medicion_automatica(
                bloque
            )
        )

    # ========================================================
    # 6. RECOMENDACIONES
    # ========================================================

    recomendaciones = (
        polea.recomendaciones
        or ""
    ).strip()

    if not recomendaciones:
        recomendaciones = (
            "Continuar con las inspecciones "
            "según el programa de mantenimiento."
        )

    # ========================================================
    # 7. CONDICIÓN
    # ========================================================

    condicion = (
        polea.get_condicion_display()
    )

    # ========================================================
    # 8. TEXTO DEL RECUADRO
    # ========================================================

    linea_1 = (
        f"Condición: {condicion}"
        f" | Observación visual: "
        f"{observacion_visual}"
    )

    if observacion_medicion:

        linea_1 += (
            " | Observación de medición: "
            f"{observacion_medicion}"
        )

    linea_2 = (
        "Recomendaciones: "
        f"{recomendaciones}"
    )

    texto = (
        f"{linea_1}\n"
        f"{linea_2}"
    )

    # ========================================================
    # 9. ESCRIBIR EN EL CAPTION CORRECTO
    # ========================================================

    coordenada = captions[
        polea.numero
    ]

    ws[coordenada] = texto

    _aplicar_estilo_caption(
        ws,
        coordenada,
    )

def generar_excel_poleas_cvb0004(
    inspeccion,
    bloques,
):
    """
    Exporta POLEAS CVB0004 en UNA SOLA HOJA: Hoja1.

    Reglas:
    - Siempre parte de la plantilla Excel original.
    - Nunca crea una hoja CAMPAÑA.
    - NORMAL queda con un solo bloque técnico.
    - CAMPAÑA inserta INICIO + FIN verticalmente en Hoja1.
    - No modifica anchos de columnas.
    - El área de impresión queda limitada a A:AX.
    - Conserva logos/diagramas y sustituye solo fotos históricas.
    - Las fotos nuevas se ajustan dentro de los cuadros existentes.
    - Escribe descripción gris y resalta mínimo en amarillo/rojo.
    """

    # --------------------------------------------------------
    # 1. SIEMPRE abrir la plantilla ORIGINAL.
    # --------------------------------------------------------
    workbook = load_workbook(
        BytesIO(
            TEMPLATE_PATH.read_bytes()
        )
    )

    # --------------------------------------------------------
    # 2. CVB004 POLEAS DEBE TENER UNA SOLA HOJA.
    #    Si por cualquier motivo la plantilla/copias anteriores
    #    contienen CAMPAÑA u otra hoja, se eliminan aquí.
    # --------------------------------------------------------
    for nombre_hoja in list(workbook.sheetnames):
        if nombre_hoja != "Hoja1":
            del workbook[nombre_hoja]

    ws = workbook["Hoja1"]
    workbook.active = 0

    # Guardar propiedades estructurales originales para verificar
    # que la exportación NO ensanche el documento.
    dimensiones_originales = _firma_dimensiones_columnas(ws)
    max_columna_original = ws.max_column

    # Vista cómoda al abrir el Excel. Esto NO cambia la impresión.
    # Evita abrir directamente en Page Break Preview con "Página 1".
    ws.sheet_view.view = "normal"
    ws.sheet_view.zoomScale = 80
    ws.sheet_view.zoomScaleNormal = 80

    # --------------------------------------------------------
    # 3. Limpiar textos históricos variables.
    # --------------------------------------------------------
    limpiar_textos_historicos_campana(ws)

    poleas = [
        bloque["polea"]
        for bloque in bloques
    ]

    # --------------------------------------------------------
    # 4. Capturar las posiciones de fotos ANTES de borrar
    #    las fotografías históricas de la plantilla.
    # --------------------------------------------------------
    anclas_originales = _anclas_fotos_originales(ws)

    # --------------------------------------------------------
    # 5. Eliminar SOLO fotografías históricas de inspección.
    #    Se conservan logos, esquemas y diagramas fijos.
    # --------------------------------------------------------
    ws._images = [
        imagen
        for imagen in ws._images
        if not _es_foto_historica_polea(imagen)
    ]

    # --------------------------------------------------------
    # 6. Preparar los bloques de CAMPAÑA EN Hoja1.
    #    Aquí se inserta verticalmente Inicio + Fin.
    # --------------------------------------------------------
    layouts, captions = _preparar_bloques_verticales_campana(
        ws,
        bloques,
        anclas_originales,
    )

    # El crecimiento permitido es SOLO vertical.
    # El ancho de impresión siempre queda A:AX.
    ws.print_area = f"A1:AX{ws.max_row}"

    # --------------------------------------------------------
    # 7. Cabecera general.
    # --------------------------------------------------------
    _escribir_cabecera(
        ws,
        inspeccion,
        poleas,
    )

    # Limpiar captions históricos antes de escribir los actuales.
    for coordenada in captions.values():
        ws[coordenada] = None

    # --------------------------------------------------------
    # 8. Escribir tablas, fotos y descripción visual.
    # --------------------------------------------------------
    vistos = set()

    for bloque in bloques:
        polea = bloque["polea"]

        _escribir_bloques(
            ws,
            polea,
            bloque,
            inspeccion.faja.tag,
            layouts,
        )

        _escribir_fotos(
            ws,
            polea,
            bloque,
            bloque.get("fotografias", []),
            vistos,
            anclas_originales,
            captions,
        )

    # --------------------------------------------------------
    # 9. Mantener el área final dentro del ancho original.
    # --------------------------------------------------------
    ws.print_area = f"A1:AX{ws.max_row}"

    # --------------------------------------------------------
    # 10. Recalcular fórmulas al abrir.
    # --------------------------------------------------------
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"

    # --------------------------------------------------------
    # 11. PROTECCIONES ESTRUCTURALES.
    # --------------------------------------------------------
    if _firma_dimensiones_columnas(ws) != dimensiones_originales:
        raise RuntimeError(
            "CVB004 Poleas modificó los anchos de columna "
            "de la plantilla original."
        )

    # Protección final absoluta: el archivo debe salir solo con Hoja1.
    for nombre_hoja in list(workbook.sheetnames):
        if nombre_hoja != "Hoja1":
            del workbook[nombre_hoja]

    if workbook.sheetnames != ["Hoja1"]:
        raise RuntimeError(
            "CVB004 Poleas debe exportarse únicamente en Hoja1."
        )

    workbook.active = 0

    # --------------------------------------------------------
    # 12. PROTECCIÓN FINAL Y GUARDAR
    # --------------------------------------------------------

    # Eliminar cualquier hoja auxiliar.
    for nombre_hoja in list(workbook.sheetnames):
        if nombre_hoja != "Hoja1":
            del workbook[nombre_hoja]

    # Debe existir una sola hoja.
    if workbook.sheetnames != ["Hoja1"]:
        raise RuntimeError(
            "ERROR CVB004: el archivo intenta "
            "salir con más de una hoja: "
            f"{workbook.sheetnames}"
        )

    workbook.active = 0

    print(
        ">>> HOJAS FINALES CVB004:",
        workbook.sheetnames,
    )

    print(
        ">>> AREA IMPRESION:",
        ws.print_area,
    )

    salida = BytesIO()

    workbook.save(salida)

    salida.seek(0)

    return salida