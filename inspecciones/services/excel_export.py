from copy import copy
from io import BytesIO
from pathlib import Path

from django.conf import settings
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import PatternFill
from openpyxl.worksheet.pagebreak import Break
from openpyxl.utils import get_column_letter


TEMPLATE_PATH = (
    Path(settings.BASE_DIR)
    / "inspecciones"
    / "assets"
    / "cvb0001_report_template.xlsx"
)
SOURCE_SHEET = "REPORTE DE INSPECCION CV0001"
OUTPUT_SHEET = "REPORTE DE INSPECCION"


def _copy_style(source_cell, target_cell):
    target_cell._style = copy(source_cell._style)
    target_cell.number_format = source_cell.number_format
    target_cell.protection = copy(source_cell.protection)


def _style_range(source_cell, worksheet, cell_range):
    for row in worksheet[cell_range]:
        for cell in row:
            _copy_style(source_cell, cell)


def _merge_and_write(worksheet, cell_range, value, source_cell):
    worksheet.merge_cells(cell_range)
    cell = worksheet[cell_range.split(":", 1)[0]]
    _copy_style(source_cell, cell)
    cell.value = value
    return cell


def _number(value):
    return float(value) if value is not None else None


def _percentage(value):
    return float(value) / 100 if value is not None else None


def _safe_image_path(photo):
    try:
        path = Path(photo.imagen.path)
    except (AttributeError, NotImplementedError, OSError, ValueError):
        return None
    return path if path.is_file() else None


def _add_image(worksheet, path, anchor, max_width, max_height):
    try:
        image = ExcelImage(str(path))
    except (FileNotFoundError, OSError, ValueError):
        return False
    scale = min(max_width / image.width, max_height / image.height, 1)
    image.width = int(image.width * scale)
    image.height = int(image.height * scale)
    worksheet.add_image(image, anchor)
    return True


def _copy_header_template(source, worksheet):
    for row in range(1, 22):
        worksheet.row_dimensions[row].height = source.row_dimensions[row].height
        for column in range(1, 24):
            _copy_style(source.cell(row, column), worksheet.cell(row, column))
    for merged in source.merged_cells.ranges:
        if merged.max_row <= 21 and merged.max_col <= 23:
            worksheet.merge_cells(str(merged))


def _write_header(source, worksheet, inspeccion):
    _copy_header_template(source, worksheet)
    values = {
        "F2": f"REPORTE DE INSPECCIÓN {inspeccion.codigo_reporte}",
        "F5": "Ingeniería de Confiabilidad - Operaciones Procesos",
        "C7": "CONDICIÓN",
        "F7": inspeccion.get_condicion_general_display().upper(),
        "C9": "DATOS DEL EQUIPO",
        "C10": "PLANTA",
        "F10": inspeccion.planta,
        "I10": "PROCESO",
        "L10": inspeccion.proceso,
        "C12": "EQUIPO",
        "F12": inspeccion.faja.nombre,
        "I12": "TAG-NUMBER",
        "L12": inspeccion.faja.tag,
        "C14": "ETAPA",
        "F14": inspeccion.etapa,
        "I14": "CONDICIÓN DEL EQUIPO",
        "L14": inspeccion.condicion_equipo,
        "C16": "FECHA DE INSPECCIÓN",
        "F16": inspeccion.fecha_inspeccion,
        "I16": "FECHA DE REPORTE",
        "L16": inspeccion.fecha_reporte,
        "C18": "INSPECTOR DE CAMPO",
        "F18": inspeccion.inspector_campo_nombre,
        "I18": "SUPERVISOR DE CAMPO",
        "L18": inspeccion.supervisor_campo_nombre,
        "C20": "ANALISTA QUE ELABORA",
        "F20": inspeccion.analista_elabora_nombre,
        "I20": "ANALISTA QUE VALIDA",
        "L20": inspeccion.analista_valida_nombre,
    }
    for coordinate, value in values.items():
        worksheet[coordinate] = value if value not in (None, "") else "—"
    worksheet["F16"].number_format = "dd-mm-yyyy"
    worksheet["L16"].number_format = "dd-mm-yyyy"
    for row in worksheet["F7:N8"]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="00B050")

    static_dir = Path(settings.BASE_DIR) / "static" / "inspecciones" / "faja"
    _add_image(
        worksheet,
        static_dir / "cvb0001" / "logo_overall.jpeg",
        "B2",
        210,
        70,
    )
    _add_image(
        worksheet,
        static_dir / "cvb0003" / "logo_mmg_las_bambas.jpeg",
        "S2",
        190,
        70,
    )


def _write_section_title(source, worksheet, row, title):
    cell = _merge_and_write(
        worksheet,
        f"C{row}:U{row}",
        title,
        source["C114"],
    )
    cell.alignment = copy(source["C114"].alignment)
    worksheet.row_dimensions[row].height = 21
    return row + 2


def _write_empalme(source, worksheet, row, title, measurements, summary):
    row = _write_section_title(source, worksheet, row, title)
    header_top = row
    header_bottom = row + 1
    fixed_headers = [
        ("D", "Empalme"),
        ("E", "Bastidor"),
        ("F", "Lado"),
        ("G", "Posición"),
        ("H", "Espesor nominal (mm)"),
    ]
    for column, label in fixed_headers:
        _merge_and_write(
            worksheet,
            f"{column}{header_top}:{column}{header_bottom}",
            label,
            source["D116"],
        )
    _merge_and_write(
        worksheet,
        f"I{header_top}:O{header_top}",
        "Espesor residual (mm)",
        source["I116"],
    )
    _merge_and_write(
        worksheet,
        f"P{header_top}:Q{header_top}",
        "Espesor (mm)",
        source["P116"],
    )
    _merge_and_write(
        worksheet,
        f"R{header_top}:T{header_top}",
        "Desgaste (mínimo)",
        source["R116"],
    )
    subheaders = list("ABCDEFG") + [
        "Mínimo",
        "Promedio",
        "Desgaste (mm)",
        "% desgaste",
        "% residual",
    ]
    for index, label in enumerate(subheaders, start=9):
        cell = worksheet.cell(header_bottom, index, label)
        _copy_style(source["I117"], cell)

    data_start = header_bottom + 1
    if measurements:
        data_end = data_start + len(measurements) - 1
        for column in "DEF":
            worksheet.merge_cells(
                start_row=data_start,
                start_column=ord(column) - 64,
                end_row=data_end,
                end_column=ord(column) - 64,
            )
        first = measurements[0]
        worksheet[f"D{data_start}"] = first.seccion.replace("EMPALME ", "")
        worksheet[f"E{data_start}"] = first.bastidor
        worksheet[f"F{data_start}"] = first.lado
        for coordinate in (f"D{data_start}", f"E{data_start}", f"F{data_start}"):
            _copy_style(source["D119"], worksheet[coordinate])
        for offset, measurement in enumerate(measurements):
            current = data_start + offset
            values = [
                measurement.posicion,
                _number(measurement.espesor_nominal),
                *[_number(getattr(measurement, letter)) for letter in "abcdefg"],
                _number(measurement.minimo),
                _number(measurement.promedio),
                _number(measurement.desgaste),
                _percentage(measurement.porcentaje_desgaste),
                _percentage(measurement.porcentaje_residual),
            ]
            for column, value in enumerate(values, start=7):
                cell = worksheet.cell(current, column, value)
                donor = source["P119"] if column >= 16 else source["I119"]
                _copy_style(donor, cell)
                cell.number_format = "0.00%" if column >= 19 else "0.00"
            worksheet.row_dimensions[current].height = 22
    else:
        data_end = data_start
        _merge_and_write(
            worksheet,
            f"D{data_start}:T{data_start}",
            "No existen mediciones registradas para este empalme.",
            source["D121"],
        )

    summary_start = data_end + 2
    summary_end = summary_start + 2
    cell = _merge_and_write(
        worksheet,
        f"D{summary_start}:T{summary_end}",
        summary["texto"],
        source["D121"],
    )
    cell.alignment = copy(source["D121"].alignment)
    cell.alignment = copy(cell.alignment)
    cell.alignment = cell.alignment.copy(wrapText=True, vertical="center")
    for item in range(summary_start, summary_end + 1):
        worksheet.row_dimensions[item].height = 20
    return summary_end + 2


def _write_photos(source, worksheet, row, title, photos):
    row = _write_section_title(source, worksheet, row, title)
    valid_photos = [(photo, _safe_image_path(photo)) for photo in photos]
    valid_photos = [(photo, path) for photo, path in valid_photos if path]
    if not valid_photos:
        _merge_and_write(
            worksheet,
            f"C{row}:U{row}",
            "No se registraron fotografías en esta sección.",
            source["C150"],
        )
        worksheet.row_dimensions[row].height = 24
        return row + 2

    for index in range(0, len(valid_photos), 2):
        pair = valid_photos[index:index + 2]
        image_row = row
        text_row = row + 13
        for reserved_row in range(image_row, text_row):
            worksheet.row_dimensions[reserved_row].height = 16
        for position, (photo, path) in enumerate(pair):
            start_col = "D" if position == 0 else "M"
            text_range = (
                f"D{text_row}:K{text_row + 2}"
                if position == 0
                else f"M{text_row}:T{text_row + 2}"
            )
            _add_image(worksheet, path, f"{start_col}{image_row}", 390, 245)
            details = [value.strip() for value in (photo.titulo, photo.descripcion) if value and value.strip()]
            if details:
                cell = _merge_and_write(
                    worksheet,
                    text_range,
                    "\n".join(details),
                    source["C150"],
                )
                cell.alignment = cell.alignment.copy(wrapText=True, vertical="center")
        row = text_row + 4
    return row


def _write_tramos(source, worksheet, row, measurements, summary):
    row = _write_section_title(
        source,
        worksheet,
        row,
        "DATOS DE MEDICIÓN DE ESPESORES POR UT EN PUNTOS DE MEDICIÓN",
    )
    headers = [
        "Tramo",
        "Medición",
        "Bastidor",
        "Nominal",
        "A",
        "B",
        "C",
        "D",
        "E",
        "F",
        "G",
        "Mínimo",
        "Promedio",
        "Desgaste",
        "% desgaste",
        "% residual",
    ]
    for index, label in enumerate(headers, start=4):
        cell = worksheet.cell(row, index, label)
        _copy_style(source["D155"], cell)
    data_start = row + 1
    for offset, measurement in enumerate(measurements):
        current = data_start + offset
        values = [
            measurement.seccion,
            measurement.punto,
            measurement.bastidor,
            _number(measurement.espesor_nominal),
            *[_number(getattr(measurement, letter)) for letter in "abcdefg"],
            _number(measurement.minimo),
            _number(measurement.promedio),
            _number(measurement.desgaste),
            _percentage(measurement.porcentaje_desgaste),
            _percentage(measurement.porcentaje_residual),
        ]
        for column, value in enumerate(values, start=4):
            cell = worksheet.cell(current, column, value)
            donor = source["O158"] if column >= 15 else source["H158"]
            _copy_style(donor, cell)
            if column <= 6:
                cell.number_format = "@"
            elif column >= 18:
                cell.number_format = "0.00%"
            else:
                cell.number_format = "0.00"
        worksheet.row_dimensions[current].height = 20

    summary_row = data_start + len(measurements)
    for label, key, current in (
        ("ESPESOR MÍNIMO (mm)", "minimos", summary_row),
        ("ESPESOR PROMEDIO (mm)", "promedios", summary_row + 1),
    ):
        _merge_and_write(
            worksheet,
            f"D{current}:G{current}",
            label,
            source["D173"],
        )
        for offset, letter in enumerate("abcdefg", start=8):
            cell = worksheet.cell(current, offset, summary[key].get(letter))
            _copy_style(source["H173"], cell)
            cell.number_format = "0.00"
        for column in range(15, 20):
            _copy_style(source["O173"], worksheet.cell(current, column))
    return summary_row + 3


def _write_notes(source, worksheet, row, inspeccion):
    for title, value in (
        ("OBSERVACIONES", inspeccion.observaciones),
        ("CONCLUSIONES Y RECOMENDACIONES", inspeccion.recomendaciones),
    ):
        row = _write_section_title(source, worksheet, row, title)
        cell = _merge_and_write(
            worksheet,
            f"C{row}:U{row + 4}",
            value or "—",
            source["F40"],
        )
        cell.alignment = cell.alignment.copy(wrapText=True, vertical="top", horizontal="left")
        for current in range(row, row + 5):
            worksheet.row_dimensions[current].height = 20
        row += 6
    return row


def generar_reporte_faja_cvb0001_excel(
    inspeccion,
    empalme_e01,
    empalme_e02,
    tramos,
    resumen_e01,
    resumen_e02,
    resumen_tramos,
    fotos_e01,
    fotos_e02,
    fotos_puntos,
):
    workbook = load_workbook(TEMPLATE_PATH)
    source = workbook[SOURCE_SHEET]
    worksheet = workbook.create_sheet(OUTPUT_SHEET, 0)

    for column in range(1, 24):
        letter = get_column_letter(column)
        worksheet.column_dimensions[letter].width = source.column_dimensions[letter].width
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_view.zoomScale = 70

    _write_header(source, worksheet, inspeccion)
    row = 24
    row = _write_empalme(
        source,
        worksheet,
        row,
        "DATOS DE MEDICIÓN – EMPALME E-01 TOP COVER",
        empalme_e01,
        resumen_e01,
    )
    if any(_safe_image_path(photo) for photo in fotos_e01):
        worksheet.row_breaks.append(Break(id=row - 1))
    row = _write_photos(
        source,
        worksheet,
        row,
        "INSPECCIÓN VISUAL – EMPALME E-01 TOP COVER",
        fotos_e01,
    )
    worksheet.row_breaks.append(Break(id=row - 1))
    row = _write_empalme(
        source,
        worksheet,
        row,
        "DATOS DE MEDICIÓN – EMPALME E-02 BOTTOM COVER",
        empalme_e02,
        resumen_e02,
    )
    row = _write_photos(
        source,
        worksheet,
        row,
        "INSPECCIÓN VISUAL – EMPALME E-02 BOTTOM COVER",
        fotos_e02,
    )
    worksheet.row_breaks.append(Break(id=row - 1))
    row = _write_tramos(source, worksheet, row, tramos, resumen_tramos)
    row = _write_photos(
        source,
        worksheet,
        row,
        "INSPECCIÓN VISUAL – PUNTOS DE MEDICIÓN",
        fotos_puntos,
    )
    row = _write_notes(source, worksheet, row, inspeccion)

    for old_sheet in list(workbook.worksheets):
        if old_sheet is not worksheet:
            workbook.remove(old_sheet)
    workbook.active = 0
    worksheet.freeze_panes = "D24"
    worksheet.print_area = f"A1:W{row}"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.print_options.horizontalCentered = True
    worksheet.page_margins.left = 0.25
    worksheet.page_margins.right = 0.25
    worksheet.page_margins.top = 0.3
    worksheet.page_margins.bottom = 0.3
    worksheet.page_margins.header = 0.15
    worksheet.page_margins.footer = 0.15

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
