from datetime import date
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from inspecciones.forms import (
    CalibracionUTFajaCVB0003FormSet,
    LifeShaftInspeccionForm,
)
from inspecciones.models import (
    CalibracionUTFajaCVB0003,
    Faja,
    Inspeccion,
    LifeShaftInspeccion,
)
from inspecciones.reportes.cvb0003.exporters.faja import (
    generar_excel_faja_cvb0003_master,
)
from inspecciones.reportes.cvb0003.exporters.life_shaft import (
    generar_excel_life_shaft_cvb0003_master,
)


class CalibracionUTCVB0003Tests(TestCase):
    def setUp(self):
        self.usuario = get_user_model().objects.create_superuser(
            "ut-cvb0003",
            "ut@example.com",
            "test",
        )
        self.faja = Faja.objects.create(nombre="CVB003", tag="CVB0003")
        self.client.force_login(self.usuario)

    def inspeccion(self, tipo):
        return Inspeccion.objects.create(
            faja=self.faja,
            tipo=tipo,
            codigo_reporte=f"TEMP-{tipo}",
            fecha_programada=date(2026, 8, 12),
            fecha_inspeccion=date(2026, 8, 12),
            inspector=self.usuario,
            supervisor=self.usuario,
            analista=self.usuario,
            creado_por=self.usuario,
        )

    def test_faja_crea_ocho_conjuntos_guarda_y_exporta(self):
        inspeccion = self.inspeccion(Inspeccion.Tipo.FAJA)
        CalibracionUTFajaCVB0003.crear_estructura(inspeccion)
        calibraciones = list(inspeccion.calibraciones_ut_faja_cvb0003.all())
        self.assertEqual(len(calibraciones), 8)

        data = {
            "calibraciones-ut-TOTAL_FORMS": "8",
            "calibraciones-ut-INITIAL_FORMS": "8",
            "calibraciones-ut-MIN_NUM_FORMS": "0",
            "calibraciones-ut-MAX_NUM_FORMS": "1000",
        }
        for indice, calibracion in enumerate(calibraciones):
            prefijo = f"calibraciones-ut-{indice}"
            data.update(
                {
                    f"{prefijo}-id": str(calibracion.pk),
                    f"{prefijo}-marca_equipo": "SIUI" if indice == 0 else "Olympus",
                    f"{prefijo}-modelo_equipo": "Smartor" if indice == 0 else "Epoch 6Lt",
                    f"{prefijo}-frecuencia_mhz": "2" if indice == 0 else "1",
                    f"{prefijo}-rango_mm": "30-90",
                    f"{prefijo}-metodo_empleado": "Pulso - eco",
                    f"{prefijo}-acoplante": "Echo gel",
                    f"{prefijo}-rectificacion": "Full",
                    f"{prefijo}-velocidad_ms": "5900" if indice == 0 else calibracion.velocidad_ms,
                    f"{prefijo}-retardo_us": "0.500" if indice == 0 else calibracion.retardo_us,
                    f"{prefijo}-tipo_scan": "A Scan",
                }
            )
        formset = CalibracionUTFajaCVB0003FormSet(
            data,
            instance=inspeccion,
            prefix="calibraciones-ut",
        )
        self.assertTrue(formset.is_valid(), formset.errors)
        formset.save()

        primera = inspeccion.calibraciones_ut_faja_cvb0003.get(numero=1)
        self.assertEqual(primera.marca_equipo, "SIUI")
        self.assertEqual(primera.modelo_equipo, "Smartor")

        formulario = self.client.get(reverse("formulario_faja", args=[inspeccion.pk]))
        self.assertEqual(formulario.status_code, 200)
        self.assertEqual(formulario.content.count(b'class="ut-accordion"'), 8)
        self.assertContains(formulario, "SIUI")

        reporte = self.client.get(reverse("reporte_faja", args=[inspeccion.pk]))
        self.assertContains(reporte, "SIUI")
        self.assertContains(reporte, "Smartor")

        output = generar_excel_faja_cvb0003_master(inspeccion)
        workbook = load_workbook(BytesIO(output.getvalue()), data_only=False)
        worksheet = workbook["REPORTE DE INSPECCION CV0003"]
        self.assertEqual(worksheet["I141"].value, "SIUI")
        self.assertEqual(worksheet["I142"].value, "Smartor")
        self.assertEqual(worksheet["I143"].value, "2")
        self.assertEqual(worksheet["P143"].value, "5900")
        self.assertEqual(worksheet["P144"].value, "0.500")
        workbook.close()

    def test_life_shaft_guarda_siete_campos_y_los_exporta(self):
        inspeccion = self.inspeccion(Inspeccion.Tipo.LIFE_SHAFT)
        shaft = LifeShaftInspeccion.objects.create(
            inspeccion=inspeccion,
            numero=1,
            nombre="LIFE SHAFT #01",
        )
        form = LifeShaftInspeccionForm(
            {
                "nombre": shaft.nombre,
                "tag": "LS-01",
                "ubicacion": "Zona motriz",
                "condicion": Inspeccion.Condicion.NORMAL,
                "tipo_medicion": "NORMAL",
                "observacion_visual": "",
                "observacion_medicion": "",
                "recomendaciones": "",
                "marca_equipo": "SIUI",
                "tipo_haz": "COMPLETA",
                "frecuencia_mhz": "2 MHz",
                "ancho_banda": "0.5-12 MHz",
                "amortiguamiento": "60",
                "velocidad_ms": "5900",
                "retardo_us": "0.500",
            },
            instance=shaft,
        )
        self.assertTrue(form.is_valid(), form.errors)
        form.save()
        shaft.refresh_from_db()
        self.assertEqual(shaft.marca_equipo, "SIUI")
        self.assertEqual(shaft.retardo_us, "0.500")

        reporte = self.client.get(reverse("reporte_life_shaft", args=[inspeccion.pk]))
        self.assertContains(reporte, "SIUI")
        self.assertContains(reporte, "0.500")

        output = generar_excel_life_shaft_cvb0003_master(inspeccion)
        workbook = load_workbook(BytesIO(output.getvalue()), data_only=False)
        valores = {
            str(cell.value)
            for worksheet in workbook.worksheets
            for row in worksheet.iter_rows()
            for cell in row
            if cell.value is not None
        }
        workbook.close()
        self.assertIn("SIUI", valores)
        self.assertIn("0.500", valores)
