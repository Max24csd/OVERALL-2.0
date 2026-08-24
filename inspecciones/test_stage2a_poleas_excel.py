from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from shutil import rmtree
from tempfile import mkdtemp
from zipfile import ZipFile

from PIL import Image as PILImage
from django.contrib.auth import get_user_model
from django.core.files.base import ContentFile
from django.test import TestCase, override_settings
from openpyxl import load_workbook

from inspecciones.models import (
    Faja,
    FaseCampana,
    FotoPolea,
    Inspeccion,
    MedicionPolea,
    MedicionPoleaCampana,
    PoleaInspeccion,
    TipoMedicionComponente,
)
from inspecciones.reportes.campaign_utils import agregar_mediciones_campana_bloque
from inspecciones.reportes.cvb0003.exporters.poleas import (
    _sha256,
    _structure_signature,
    generar_excel_poleas_cvb0003_master,
)
from inspecciones.reportes.cvb0003.mappings.poleas import (
    MASTER_PATH,
    MASTER_SHA256,
    POLEA_BLOCKS,
)


APPROVED_SOURCE = Path.home() / "Downloads" / "20260807-VTUT-CVB0003-POLEAS (1).xlsx"
APPROVED_SOURCE_SHA256 = "19221FFDA1152EE22F5B62AB6655213FFF80B98C332AF92B972BEBD71F02069A"


class PoleasExcelStage2ATests(TestCase):
    def setUp(self):
        self.media_root = mkdtemp(prefix="cvb0003-poleas-excel-")
        self.settings_override = override_settings(MEDIA_ROOT=self.media_root)
        self.settings_override.enable()
        self.user = get_user_model().objects.create_user("excel-auditor")
        self.faja = Faja.objects.create(nombre="CVB003", tag="CVB0003")
        self.inspection = Inspeccion.objects.create(
            faja=self.faja,
            tipo=Inspeccion.Tipo.POLEAS,
            codigo_reporte="TEMP-CVB0003-POLEAS",
            fecha_programada=date(2026, 8, 10),
            fecha_inspeccion=date(2026, 8, 10),
            fecha_reporte=date(2026, 8, 11),
            inspector=self.user,
            supervisor=self.user,
            analista=self.user,
            creado_por=self.user,
            observaciones="Observación vigente de prueba.",
            recomendaciones="Mantener el plan de inspecciones.",
        )
        self.poleas = {}
        for number in range(1, 10):
            polea = PoleaInspeccion.objects.create(
                inspeccion=self.inspection,
                numero=number,
                nombre=f"POLEA #{number:02d}",
                orden=number,
                tipo_medicion=TipoMedicionComponente.NORMAL,
                condicion=Inspeccion.Condicion.NORMAL,
            )
            self.poleas[number] = polea
            for point in range(1, 6):
                MedicionPolea.objects.create(
                    polea=polea,
                    punto=point,
                    orden=point,
                    posicion="PUNTOS SENTIDO RADIAL",
                    a=Decimal("30.00") - Decimal(point),
                    b=Decimal("29.50") - Decimal(point),
                    c=Decimal("29.00") - Decimal(point),
                    d=Decimal("28.50") - Decimal(point),
                    e=Decimal("28.00") - Decimal(point),
                    f=Decimal("27.50") - Decimal(point),
                    g=Decimal("27.00") - Decimal(point),
                )

    def tearDown(self):
        self.settings_override.disable()
        rmtree(self.media_root, ignore_errors=True)

    def blocks(self):
        return [
            agregar_mediciones_campana_bloque(
                {
                    "polea": polea,
                    "mediciones": list(polea.mediciones.order_by("orden", "punto")),
                    "fotografias": list(polea.fotografias.order_by("creada_en", "id")),
                },
                polea,
            )
            for polea in PoleaInspeccion.objects.filter(
                inspeccion=self.inspection
            ).order_by("orden", "numero")
        ]

    def export(self):
        output = generar_excel_poleas_cvb0003_master(self.inspection, self.blocks())
        self.assertIsNone(ZipFile(BytesIO(output.getvalue())).testzip())
        return load_workbook(BytesIO(output.getvalue()), data_only=False)

    def campaign(self, number, include_end=True):
        polea = self.poleas[number]
        polea.tipo_medicion = TipoMedicionComponente.CAMPANA
        polea.save(update_fields=["tipo_medicion"])
        for point in range(1, 6):
            MedicionPoleaCampana.objects.create(
                polea=polea,
                fase=FaseCampana.INICIO,
                punto=point,
                orden=point,
                a=Decimal("40.00") - Decimal(point),
                b=Decimal("39.00") - Decimal(point),
            )
            if include_end:
                MedicionPoleaCampana.objects.create(
                    polea=polea,
                    fase=FaseCampana.FIN,
                    punto=point,
                    orden=point,
                    a=Decimal("35.00") - Decimal(point),
                    b=Decimal("34.00") - Decimal(point),
                )
        return polea

    def test_todas_normales_ocultan_solo_slot_b_y_conservan_estructura(self):
        master = load_workbook(MASTER_PATH)
        exported = self.export()
        self.assertEqual(_structure_signature(master), _structure_signature(exported))
        worksheet = exported["Hoja1"]
        for mapping in POLEA_BLOCKS.values():
            self.assertFalse(worksheet.row_dimensions[mapping["slot_a"][0]].hidden)
            self.assertTrue(worksheet.row_dimensions[mapping["slot_b"][0]].hidden)
        mapping = POLEA_BLOCKS[1]
        data_row = mapping["slot_a"][0] + mapping["data_offset"]
        self.assertEqual(worksheet[f"AC{data_row}"].value, 29.0)
        self.assertIn("NORMAL", worksheet[f"D{mapping['slot_a'][0]}"].value)

    def test_polea_03_inicio_historico_muestra_fin_vacio_sin_usar_normal(self):
        self.campaign(3, include_end=False)
        workbook = self.export()
        worksheet = workbook["Hoja1"]
        mapping = POLEA_BLOCKS[3]
        self.assertFalse(worksheet.row_dimensions[mapping["slot_a"][0]].hidden)
        self.assertFalse(worksheet.row_dimensions[mapping["slot_b"][0]].hidden)
        start_row = mapping["slot_a"][0] + mapping["data_offset"]
        end_row = mapping["slot_b"][0] + mapping["data_offset"]
        self.assertEqual(worksheet[f"AC{start_row}"].value, 39.0)
        self.assertIsNone(worksheet[f"AC{end_row}"].value)
        self.assertIn("INICIO DE CAMPAÑA", worksheet[f"D{mapping['slot_a'][0]}"].value)
        self.assertIn("FIN DE CAMPAÑA", worksheet[f"D{mapping['slot_b'][0]}"].value)

    def test_reporte_mixto_inicio_fin_y_fotografia_en_slot_fijo(self):
        polea = self.campaign(2, include_end=True)
        image_buffer = BytesIO()
        PILImage.new("RGB", (160, 320), color=(25, 90, 160)).save(
            image_buffer, format="PNG"
        )
        FotoPolea.objects.create(
            polea=polea,
            imagen=ContentFile(image_buffer.getvalue(), name="vertical.png"),
            descripcion="Fotografía vertical de prueba",
            subida_por=self.user,
        )
        workbook = self.export()
        worksheet = workbook["Hoja1"]
        mapping = POLEA_BLOCKS[2]
        start_row = mapping["slot_a"][0] + mapping["data_offset"]
        end_row = mapping["slot_b"][0] + mapping["data_offset"]
        self.assertEqual(worksheet[f"AC{start_row}"].value, 39.0)
        self.assertEqual(worksheet[f"AC{end_row}"].value, 34.0)
        self.assertFalse(worksheet.row_dimensions[mapping["slot_b"][0]].hidden)
        self.assertTrue(worksheet.row_dimensions[POLEA_BLOCKS[1]["slot_b"][0]].hidden)
        static_images = [
            image for image in worksheet._images
            if image.anchor._from.row + 1 <= 101
        ]
        dynamic_images = [
            image for image in worksheet._images
            if image.anchor._from.row + 1 > 101
        ]
        self.assertEqual(len(static_images), 4)
        self.assertEqual(len(dynamic_images), 1)
        image = dynamic_images[0]
        self.assertLessEqual(image.anchor._from.row + 1, mapping["photo_rows"][1])
        self.assertGreaterEqual(image.anchor._from.row + 1, mapping["photo_rows"][0])
        self.assertGreater(image.width, 0)
        self.assertGreater(image.height, 0)

    def test_master_y_original_no_cambian_y_archivo_abre_sin_reparacion(self):
        master_before = _sha256(MASTER_PATH)
        source_before = _sha256(APPROVED_SOURCE)
        workbook = self.export()
        self.assertIn("Hoja1", workbook.sheetnames)
        self.assertEqual(_sha256(MASTER_PATH), master_before)
        self.assertEqual(master_before, MASTER_SHA256)
        self.assertEqual(source_before, APPROVED_SOURCE_SHA256)
        self.assertEqual(_sha256(APPROVED_SOURCE), source_before)
