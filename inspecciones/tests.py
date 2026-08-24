import base64
from io import BytesIO
import tempfile
from datetime import date
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.test import override_settings
from django.urls import reverse
from openpyxl import load_workbook

from .forms import InspeccionForm
from .models import Faja, FotoInspeccion, Inspeccion, Medicion
from .views import analizar_mediciones_empalme, crear_estructura_mediciones_faja


class FormularioFajaCVB0001Tests(TestCase):
    def setUp(self):
        self.usuario = get_user_model().objects.create_superuser(
            username="admin_cvb0001",
            email="admin@example.com",
            password="prueba-segura",
        )
        self.faja = Faja.objects.create(
            nombre="Faja Overland 01",
            tag="CVB0001",
        )
        self.inspeccion = Inspeccion.objects.create(
            faja=self.faja,
            tipo=Inspeccion.Tipo.FAJA,
            codigo_reporte="REPORTE-CVB0001-TEST",
            fecha_programada=date(2026, 8, 6),
            inspector=self.usuario,
            supervisor=self.usuario,
            analista=self.usuario,
            creado_por=self.usuario,
        )
        self.client.force_login(self.usuario)

    def test_estructura_e02_es_idempotente_y_conserva_mediciones(self):
        crear_estructura_mediciones_faja(self.inspeccion)
        fila_existente = self.inspeccion.mediciones.get(
            seccion="EMPALME E-02",
            posicion="-1 m",
        )
        fila_existente.a = Decimal("12.34")
        fila_existente.save()
        self.inspeccion.mediciones.filter(
            seccion="EMPALME E-02",
            posicion="+1 m",
        ).delete()

        crear_estructura_mediciones_faja(self.inspeccion)
        crear_estructura_mediciones_faja(self.inspeccion)

        filas_e02 = self.inspeccion.mediciones.filter(seccion="EMPALME E-02")
        self.assertEqual(filas_e02.count(), 2)
        self.assertEqual(
            filas_e02.get(posicion="-1 m").a,
            Decimal("12.34"),
        )

    def test_campos_de_cabecera_se_guardan_y_se_muestran_en_reporte(self):
        formulario = InspeccionForm(
            data={
                "fecha_inspeccion": "2026-08-01",
                "fecha_reporte": "2026-08-02",
                "inspector_campo_nombre": "Inspectora Uno",
                "supervisor_campo_nombre": "Supervisor Dos",
                "analista_elabora_nombre": "Analista Tres",
                "analista_valida_nombre": "Analista Cuatro",
                "condicion_general": Inspeccion.Condicion.NORMAL,
                "condicion_equipo": "En uso",
                "circunstancias": "",
                "antecedentes": "",
                "observaciones": "",
                "recomendaciones": "",
            },
            instance=self.inspeccion,
        )
        self.assertTrue(formulario.is_valid(), formulario.errors)
        formulario.save()
        self.inspeccion.refresh_from_db()
        self.assertEqual(self.inspeccion.fecha_inspeccion, date(2026, 8, 1))
        self.assertEqual(self.inspeccion.analista_valida_nombre, "Analista Cuatro")

        respuesta = self.client.get(
            reverse("reporte_faja", args=[self.inspeccion.id])
        )
        self.assertEqual(respuesta.status_code, 200)
        self.assertContains(respuesta, "Inspectora Uno")
        self.assertContains(respuesta, "02/08/2026")

    def test_resumen_encuentra_letra_posicion_y_porcentaje(self):
        crear_estructura_mediciones_faja(self.inspeccion)
        fila = self.inspeccion.mediciones.get(
            seccion="EMPALME E-01",
            posicion="-1 m",
        )
        fila.espesor_nominal = Decimal("20.00")
        fila.a = Decimal("18.00")
        fila.d = Decimal("16.52")
        fila.save()

        resumen = analizar_mediciones_empalme(
            self.inspeccion.mediciones.filter(seccion="EMPALME E-01"),
            "E-01",
        )
        self.assertEqual(resumen["minimo"], Decimal("16.52"))
        self.assertEqual(resumen["letra"], "D")
        self.assertEqual(resumen["posicion"], "-1 m")
        self.assertEqual(resumen["porcentaje_residual"], Decimal("82.600"))
        self.assertIn("a un metro antes del empalme", resumen["texto"])

    def test_fotografias_se_clasifican_sin_mezclarse(self):
        for seccion in FotoInspeccion.Seccion.values:
            FotoInspeccion.objects.create(
                inspeccion=self.inspeccion,
                seccion=seccion,
                titulo=seccion,
                imagen=f"pruebas/{seccion}.jpg",
                subida_por=self.usuario,
            )
        self.assertEqual(
            self.inspeccion.fotografias.filter(
                seccion=FotoInspeccion.Seccion.EMPALME_E01
            ).count(),
            1,
        )
        self.assertEqual(
            self.inspeccion.fotografias.filter(
                seccion=FotoInspeccion.Seccion.EMPALME_E02
            ).count(),
            1,
        )
        self.assertEqual(
            self.inspeccion.fotografias.filter(
                seccion=FotoInspeccion.Seccion.PUNTOS_MEDICION
            ).count(),
            1,
        )

    def test_formulario_abre_sin_excepciones_y_separa_empalmes(self):
        respuesta = self.client.get(
            reverse("formulario_faja", args=[self.inspeccion.id])
        )
        self.assertEqual(respuesta.status_code, 200)
        self.assertTemplateUsed(
            respuesta,
            "inspecciones/formulario_faja_cvb0001.html",
        )
        self.assertEqual(len(respuesta.context["empalme_e01"]), 2)
        self.assertEqual(len(respuesta.context["empalme_e02"]), 2)
        self.assertContains(respuesta, "INSPECCIÓN VISUAL – EMPALME E-02")

    def test_descarga_excel_es_valida_dinamica_y_con_fotografia(self):
        crear_estructura_mediciones_faja(self.inspeccion)
        self.inspeccion.inspector_campo_nombre = "Nombre dinámico Excel"
        self.inspeccion.save(update_fields=["inspector_campo_nombre"])
        imagen_png = base64.b64decode(
            "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII="
        )
        with tempfile.TemporaryDirectory() as media_root:
            with override_settings(MEDIA_ROOT=media_root):
                FotoInspeccion.objects.create(
                    inspeccion=self.inspeccion,
                    seccion=FotoInspeccion.Seccion.EMPALME_E01,
                    titulo="Foto dinámica",
                    descripcion="Descripción dinámica",
                    imagen=SimpleUploadedFile(
                        "foto_prueba.png",
                        imagen_png,
                        content_type="image/png",
                    ),
                    subida_por=self.usuario,
                )
                respuesta = self.client.get(
                    reverse(
                        "exportar_excel_faja_cvb0001",
                        args=[self.inspeccion.id],
                    )
                )

        self.assertEqual(respuesta.status_code, 200)
        self.assertEqual(
            respuesta["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("REPORTE_INSPECCION_", respuesta["Content-Disposition"])
        workbook = load_workbook(BytesIO(respuesta.content), data_only=True)
        self.assertEqual(workbook.sheetnames, ["REPORTE DE INSPECCION"])
        worksheet = workbook["REPORTE DE INSPECCION"]
        self.assertEqual(worksheet.page_setup.orientation, "landscape")
        self.assertEqual(worksheet.page_setup.fitToWidth, 1)
        self.assertIn(self.inspeccion.codigo_reporte, worksheet["F2"].value)
        valores = [
            str(cell.value)
            for row in worksheet.iter_rows()
            for cell in row
            if cell.value is not None
        ]
        self.assertIn("Nombre dinámico Excel", valores)
        self.assertTrue(any("EMPALME E-02" in value for value in valores))
        self.assertFalse(any("Juler Sanchez" in value for value in valores))
        self.assertGreaterEqual(len(worksheet._images), 1)
