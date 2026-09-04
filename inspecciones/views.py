from decimal import Decimal
from io import BytesIO
from pathlib import Path
import re

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.conf import settings
from django.db import connection, models as django_models, transaction
from django.db.models import Count, Q
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    Image as PdfImage,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from .forms import (
    CalibracionUTFajaCVB0003FormSet,
    FotoFajaCVB0003FormSet,
    FotoInspeccionFormSet,
    FotoLifeShaftForm,
    FotoPoleaFormSet,
    InspeccionForm,
    LifeShaftInspeccionForm,
    MedicionEmpalmeCVB0003FormSet,
    MedicionFormSet,
    MedicionLifeShaftFormSet,
    MedicionPoleaFormSet,
    MedicionTramoCVB0003FormSet,
    PoleaInspeccionForm,
)
from .nueva_parada_forms import (
    CrearIntermitenteParadaForm,
    NuevaParadaForm,
)
from .models import (
    CalibracionUTFajaCVB0003,
    FotoFajaCVB0003,
    FotoInspeccion,
    FotoLifeShaft,
    Inspeccion,
    LifeShaftInspeccion,
    Medicion,
    MedicionEmpalmeCVB0003,
    MedicionLifeShaft,
    MedicionPolea,
    MedicionTramoCVB0003,
    PoleaInspeccion,
    Faja,
    Parada,
)
from accounts.models import (
    AccesoParada,
    HistorialAsignacionParada,
    PerfilUsuario,
)
from .parada_forms import GestionAsignacionParadaForm
from .presentation_scope import (
    PARADA_ACTUAL_ID,
    INSPECCIONES_HISTORICAS_OFICIALES_IDS,
)
from .services.excel_export import generar_reporte_faja_cvb0001_excel
from .reportes.campaign_utils import (
    agregar_mediciones_campana_bloque,
    es_campana,
    formsets_life_shaft_campana,
    formsets_polea_campana,
    guardar_formsets_campana,
)
from .reportes.cvb0003.code_utils import sincronizar_codigo_cvb0003
from .reportes.cvb0003.permissions import (
    es_inspeccion_cvb0003,
    puede_acceder_inspeccion_cvb0003,
)
from .reportes.cvb0003.workflow import cambiar_estado_con_historial
from .reportes.cvb0003.history import (
    clave_pk_form,
    historial_componente_visible,
    historial_faja,
    historial_mediciones_genericas,
    modo_campana_seleccionado,
    preparar_formset_historico,
    validar_formset_historico,
)


def obtener_rol(usuario):
    if usuario.is_superuser:
        return "Administrador"

    grupo = usuario.groups.first()
    return grupo.name if grupo else "Sin rol"

def usuario_vigente_en_sistema(usuario):
    """
    Verifica la vigencia general del usuario.

    - Sin perfil: se mantiene compatible con usuarios antiguos.
    - Permanente: siempre vigente.
    - Contrato: respeta fecha inicio/fin.
    - Intermitente: la vigencia real se valida por parada.
    """
    if usuario.is_superuser:
        return True

    try:
        perfil = usuario.perfil_sistema
    except PerfilUsuario.DoesNotExist:
        return True

    if perfil.tipo_vinculo == PerfilUsuario.TipoVinculo.PERMANENTE:
        return True

    if perfil.tipo_vinculo == PerfilUsuario.TipoVinculo.CONTRATO:
        return perfil.contrato_vigente()

    if perfil.tipo_vinculo == PerfilUsuario.TipoVinculo.INTERMITENTE:
        return True

    return False


def usuario_intermitente_puede_acceder(usuario, inspeccion, rol):
    """
    Un usuario intermitente solamente puede trabajar dentro
    de una parada con un acceso temporal vigente.
    """
    try:
        perfil = usuario.perfil_sistema
    except PerfilUsuario.DoesNotExist:
        return True

    if perfil.tipo_vinculo != PerfilUsuario.TipoVinculo.INTERMITENTE:
        return True

    if not inspeccion.parada_id:
        return False

    acceso = (
        AccesoParada.objects
        .filter(
            parada_id=inspeccion.parada_id,
            usuario_id=usuario.id,
            rol=rol,
            activo=True,
        )
        .order_by("-id")
        .first()
    )

    if not acceso:
        return False

    return acceso.esta_vigente()


def usuario_asignado_a_inspeccion(usuario, inspeccion, rol):
    """
    Verifica pertenencia al equipo autorizado de la parada.

    - En paradas nuevas, AccesoParada es la fuente de verdad.
    - En reportes históricos/paradas antiguas sin accesos, conserva
      compatibilidad con inspector/supervisor/analista/cliente principal.
    """
    principales = {
        "Inspector": inspeccion.inspector_id,
        "Supervisor": inspeccion.supervisor_id,
        "Analista": inspeccion.analista_id,
        "Cliente": inspeccion.cliente_id,
    }

    if inspeccion.parada_id:
        accesos_rol = AccesoParada.objects.filter(
            parada_id=inspeccion.parada_id,
            rol=rol,
        )

        if accesos_rol.exists():
            return accesos_rol.filter(
                usuario_id=usuario.id,
                activo=True,
            ).exists()

    return principales.get(rol) == usuario.id


def usuario_puede_abrir_inspeccion(usuario, inspeccion):
    rol = obtener_rol(usuario)

    # Administrador conserva acceso total.
    if usuario.is_superuser or rol == "Administrador":
        return True

    if rol not in [
        "Inspector",
        "Supervisor",
        "Analista",
        "Cliente",
    ]:
        return False

    # El usuario debe estar vigente en el sistema.
    if not usuario_vigente_en_sistema(usuario):
        return False

    # Cliente: excepción estricta de SOLO LECTURA para los históricos
    # oficiales que pertenecen al mismo usuario cliente.
    if (
        rol == "Cliente"
        and inspeccion.id in INSPECCIONES_HISTORICAS_OFICIALES_IDS
        and inspeccion.cliente_id == usuario.id
    ):
        return True

    # Debe ser el responsable asignado a esta inspección.
    if not usuario_asignado_a_inspeccion(
        usuario,
        inspeccion,
        rol,
    ):
        return False

    # Si es intermitente, además necesita acceso temporal
    # vigente específicamente para la parada.
    if not usuario_intermitente_puede_acceder(
        usuario,
        inspeccion,
        rol,
    ):
        return False

    # CVB003 conserva sus restricciones especiales.
    if es_inspeccion_cvb0003(inspeccion):
        return puede_acceder_inspeccion_cvb0003(
            usuario,
            inspeccion,
            "ver",
        )

    # El cliente solamente puede abrir reportes publicados.
    if rol == "Cliente":
        return (
            inspeccion.estado
            == Inspeccion.Estado.PUBLICADO
        )

    return True


def puede_editar_inspeccion(usuario, inspeccion):
    rol = obtener_rol(usuario)

    if usuario.is_superuser or rol == "Administrador":
        return True

    # Primero debe poder abrir realmente la inspección.
    if not usuario_puede_abrir_inspeccion(
        usuario,
        inspeccion,
    ):
        return False

    # CVB003 conserva además sus permisos específicos.
    if es_inspeccion_cvb0003(inspeccion):
        return puede_acceder_inspeccion_cvb0003(
            usuario,
            inspeccion,
            "editar",
        )

    if rol == "Inspector":
        return inspeccion.estado in [
            Inspeccion.Estado.BORRADOR,
            Inspeccion.Estado.DEVUELTO,
        ]

    if rol == "Supervisor":
        return (
            inspeccion.estado
            == Inspeccion.Estado.EN_REVISION
        )

    if rol == "Analista":
        return inspeccion.estado in [
            Inspeccion.Estado.REVISADO,
            Inspeccion.Estado.APROBADO,
        ]

    return False


def obtener_permisos_flujo(usuario, inspeccion):
    """Variables que usan las plantillas para mostrar los botones."""
    rol = obtener_rol(usuario)
    return {
        "rol": rol,
        "puede_editar": puede_editar_inspeccion(usuario, inspeccion),
        "puede_enviar_revision": (
            rol in ["Inspector", "Administrador"]
            and inspeccion.estado in [
                Inspeccion.Estado.BORRADOR,
                Inspeccion.Estado.DEVUELTO,
            ]
        ),
        "puede_revisar_supervisor": (
            rol in ["Supervisor", "Administrador"]
            and inspeccion.estado == Inspeccion.Estado.EN_REVISION
        ),
        "puede_revisar_analista": (
            rol in ["Analista", "Administrador"]
            and inspeccion.estado == Inspeccion.Estado.REVISADO
        ),
        "puede_publicar": (
            rol in ["Analista", "Administrador"]
            and inspeccion.estado == Inspeccion.Estado.APROBADO
        ),
    }


def _contexto_workflow_ui_cvb0003(inspeccion):
    """Contexto visual CVB003; no altera permisos ni transiciones."""
    return {
        "ultima_devolucion": inspeccion.historial.filter(
            accion__in=[
                "DEVOLVER_SUPERVISOR",
                "DEVOLVER_ANALISTA",
            ],
        ).first(),
    }


def _aplicar_accion_flujo(request, inspeccion, accion, comentario=""):
    """Cambia el estado de una inspección y registra fechas/comentarios."""
    rol = obtener_rol(request.user)
    comentario = (comentario or "").strip()

    if es_inspeccion_cvb0003(inspeccion):
        acciones_permiso = {
            "enviar_supervisor": "enviar_revision",
            "aprobar_supervisor": "aprobar_supervisor",
            "devolver_supervisor": "devolver_supervisor",
            "aprobar_analista": "aprobar_analista",
            "devolver_analista": "devolver_analista",
            "publicar": "publicar",
        }
        accion_permiso = acciones_permiso.get(accion)
        if not accion_permiso or not puede_acceder_inspeccion_cvb0003(
            request.user,
            inspeccion,
            accion_permiso,
        ):
            return False, "No tienes permiso para ejecutar esta acción."

        transiciones = {
            "enviar_supervisor": (
                Inspeccion.Estado.EN_REVISION,
                "ENVIAR_A_REVISION",
                "Informe enviado al supervisor.",
            ),
            "aprobar_supervisor": (
                Inspeccion.Estado.REVISADO,
                "APROBAR_SUPERVISOR",
                "Informe aprobado por el supervisor y enviado al analista.",
            ),
            "devolver_supervisor": (
                Inspeccion.Estado.DEVUELTO,
                "DEVOLVER_SUPERVISOR",
                "Informe devuelto al inspector.",
            ),
            "aprobar_analista": (
                Inspeccion.Estado.APROBADO,
                "APROBAR_ANALISTA",
                "Informe aprobado por el analista.",
            ),
            "devolver_analista": (
                Inspeccion.Estado.DEVUELTO,
                "DEVOLVER_ANALISTA",
                "Informe devuelto al inspector.",
            ),
            "publicar": (
                Inspeccion.Estado.PUBLICADO,
                "PUBLICAR",
                "Reporte final publicado para el cliente.",
            ),
        }
        nuevo_estado, accion_historial, mensaje = transiciones[accion]
        try:
            cambiar_estado_con_historial(
                inspeccion,
                request.user,
                nuevo_estado,
                accion_historial,
                comentario,
            )
        except ValueError as error:
            return False, str(error)
        return True, mensaje

    if accion == "enviar_supervisor":
        if rol not in ["Inspector", "Administrador"]:
            return False, "Solo el inspector o administrador puede enviar a revisión."

        if inspeccion.estado not in [
            Inspeccion.Estado.BORRADOR,
            Inspeccion.Estado.DEVUELTO,
        ]:
            return False, "La inspección no está disponible para envío."

        inspeccion.estado = Inspeccion.Estado.EN_REVISION
        inspeccion.fecha_envio_revision = timezone.now()
        inspeccion.motivo_devolucion = ""
        inspeccion.comentarios_revision = ""
        mensaje = "Informe enviado al supervisor."

    elif accion == "aprobar_supervisor":
        if rol not in ["Supervisor", "Administrador"]:
            return False, "Solo el supervisor o administrador puede aprobar esta etapa."

        if inspeccion.estado != Inspeccion.Estado.EN_REVISION:
            return False, "El informe no está pendiente del supervisor."

        inspeccion.estado = Inspeccion.Estado.REVISADO
        inspeccion.fecha_revision_supervisor = timezone.now()
        inspeccion.comentarios_revision = comentario
        inspeccion.motivo_devolucion = ""
        mensaje = "Informe aprobado por el supervisor y enviado al analista."

    elif accion == "devolver_supervisor":
        if rol not in ["Supervisor", "Administrador"]:
            return False, "Solo el supervisor o administrador puede devolver el informe."

        if inspeccion.estado != Inspeccion.Estado.EN_REVISION:
            return False, "El informe no está pendiente del supervisor."

        if not comentario:
            return False, "Escribe el motivo de devolución."

        inspeccion.estado = Inspeccion.Estado.DEVUELTO
        inspeccion.motivo_devolucion = comentario
        inspeccion.comentarios_revision = comentario
        mensaje = "Informe devuelto al inspector."

    elif accion == "aprobar_analista":
        if rol not in ["Analista", "Administrador"]:
            return False, "Solo el analista o administrador puede aprobar el informe."

        if inspeccion.estado != Inspeccion.Estado.REVISADO:
            return False, "El informe todavía no está listo para aprobación del analista."

        inspeccion.estado = Inspeccion.Estado.APROBADO
        inspeccion.fecha_aprobacion_analista = timezone.now()
        inspeccion.comentarios_revision = comentario
        inspeccion.motivo_devolucion = ""
        mensaje = "Informe aprobado por el analista."

    elif accion == "devolver_analista":
        if rol not in ["Analista", "Administrador"]:
            return False, "Solo el analista o administrador puede devolver el informe."

        if inspeccion.estado != Inspeccion.Estado.REVISADO:
            return False, "El informe no está pendiente del analista."

        if not comentario:
            return False, "Escribe el motivo de devolución."

        inspeccion.estado = Inspeccion.Estado.DEVUELTO
        inspeccion.motivo_devolucion = comentario
        inspeccion.comentarios_revision = comentario
        mensaje = "Informe devuelto al inspector."

    elif accion == "publicar":
        if rol not in ["Analista", "Administrador"]:
            return False, "Solo el analista o administrador puede publicar el informe."

        if inspeccion.estado != Inspeccion.Estado.APROBADO:
            return False, "El informe debe estar aprobado antes de publicarse."

        inspeccion.estado = Inspeccion.Estado.PUBLICADO
        inspeccion.fecha_publicacion = timezone.now()
        mensaje = "Reporte final publicado para el cliente."

    else:
        return False, "La acción solicitada no existe."

    inspeccion.save()
    return True, mensaje

def _procesar_accion_flujo(request, inspeccion):
    accion = request.POST.get("workflow_action", "guardar")
    comentario = request.POST.get(
        "comentario_revision",
        "",
    ).strip()

    if accion == "guardar":
        return True, "Los cambios se guardaron correctamente."

    return _aplicar_accion_flujo(
        request,
        inspeccion,
        accion,
        comentario,
    )


def _nombre_ruta_formulario(inspeccion):
    if inspeccion.tipo == Inspeccion.Tipo.POLEAS:
        return "formulario_poleas"

    if inspeccion.tipo == Inspeccion.Tipo.LIFE_SHAFT:
        return "formulario_life_shaft"

    return "formulario_faja"


def _nombre_ruta_reporte(inspeccion):
    if inspeccion.tipo == Inspeccion.Tipo.POLEAS:
        return "reporte_poleas"

    if inspeccion.tipo == Inspeccion.Tipo.LIFE_SHAFT:
        return "reporte_life_shaft"

    return "reporte_faja"


def _agregar_contadores_dashboard(inspeccion):
    """Calcula contadores sin cruzar todas las relaciones en una sola consulta."""
    if inspeccion.tipo == Inspeccion.Tipo.POLEAS:
        conteos = inspeccion.poleas_inspeccionadas.aggregate(
            normales=Count(
                "mediciones",
                filter=Q(tipo_medicion="NORMAL"),
                distinct=True,
            ),
            campana=Count(
                "mediciones_campana",
                filter=Q(tipo_medicion="CAMPANA"),
                distinct=True,
            ),
            fotos=Count("fotografias", distinct=True),
        )
        inspeccion.total_mediciones = conteos["normales"] + conteos["campana"]
        inspeccion.total_fotos = conteos["fotos"]
        return

    if inspeccion.tipo == Inspeccion.Tipo.LIFE_SHAFT:
        conteos = inspeccion.life_shafts.aggregate(
            normales=Count(
                "mediciones",
                filter=Q(tipo_medicion="NORMAL"),
                distinct=True,
            ),
            campana=Count(
                "mediciones_campana",
                filter=Q(tipo_medicion="CAMPANA"),
                distinct=True,
            ),
            fotos=Count("fotografias", distinct=True),
        )
        inspeccion.total_mediciones = conteos["normales"] + conteos["campana"]
        inspeccion.total_fotos = conteos["fotos"]
        return

    inspeccion.total_mediciones = (
        inspeccion.mediciones.count()
        + inspeccion.empalmes_cvb0003.count()
        + inspeccion.tramos_cvb0003.count()
    )
    inspeccion.total_fotos = (
        inspeccion.fotografias.count()
        + inspeccion.fotografias_cvb0003.count()
    )


def _es_inspeccion_cvb0003(inspeccion):
    tag = (getattr(inspeccion.faja, "tag", "") or "").upper()
    tag_normalizado = re.sub(r"[^A-Z0-9]", "", tag)
    return "CVB0003" in tag_normalizado or "CVB003" in tag_normalizado


def _codigo_equipo_dashboard(inspeccion):
    """Normaliza el equipo real para agrupar la parada de Chancado."""
    tag = (getattr(inspeccion.faja, "tag", "") or "").upper()
    tag_normalizado = re.sub(r"[^A-Z0-9]", "", tag)
    equivalencias = (
        (("CVB0010", "CVB010"), "CVB010"),
        (("CVB0001", "CVB001"), "CVB001"),
        (("CVB0003", "CVB003"), "CVB003"),
        (("CVB0004", "CVB004"), "CVB004"),
        (("CVB0006", "CVB006"), "CVB006"),
        (("CVB0007", "CVB007"), "CVB007"),
        (("CVB0011", "CVB011"), "CVB011"),
        (("CVB0015", "CVB015"), "CVB015"),
        (("CVB0017", "CVB017"), "CVB017"),
        (("CVB0018", "CVB018"), "CVB018"),
    )
    for variantes, codigo in equivalencias:
        if any(variante in tag_normalizado for variante in variantes):
            return codigo
    return inspeccion.faja.tag or inspeccion.faja.nombre


def _ordenar_inspecciones_dashboard(inspecciones):
    orden_tipo = {
        Inspeccion.Tipo.FAJA: 0,
        Inspeccion.Tipo.POLEAS: 1,
        Inspeccion.Tipo.LIFE_SHAFT: 2,
    }
    return sorted(
        inspecciones,
        key=lambda inspeccion: (
            orden_tipo.get(inspeccion.tipo, 99),
            inspeccion.id,
        ),
    )


def _ultima_inspeccion_por_tipo(inspecciones):
    seleccionadas = {}
    for inspeccion in inspecciones:
        fecha_tecnica = inspeccion.fecha_inspeccion or inspeccion.fecha_programada
        clave = (fecha_tecnica, inspeccion.id)
        actual = seleccionadas.get(inspeccion.tipo)
        if actual is None or clave > actual[0]:
            seleccionadas[inspeccion.tipo] = (clave, inspeccion)
    return _ordenar_inspecciones_dashboard(
        [seleccion[1] for seleccion in seleccionadas.values()]
    )


def _agrupar_dashboard_por_parada(inspecciones):
    """Agrupa exclusivamente reportes reales por fecha y equipo."""
    grupos = {}
    for inspeccion in inspecciones:
        fecha = inspeccion.fecha_inspeccion or inspeccion.fecha_programada
        grupos.setdefault(fecha, []).append(inspeccion)
    resultado = []
    for fecha, elementos in sorted(
            grupos.items(),
            key=lambda item: item[0] or timezone.datetime.min.date(),
            reverse=True,
        ):
        por_equipo = {}
        for inspeccion in elementos:
            codigo = _codigo_equipo_dashboard(inspeccion)
            por_equipo.setdefault(codigo, []).append(inspeccion)
        resultado.append({
            "fecha": fecha,
            "equipos": [
                {
                    "codigo": codigo,
                    "inspecciones": _ordenar_inspecciones_dashboard(items),
                }
                for codigo, items in sorted(por_equipo.items())
            ],
            "inspecciones": _ordenar_inspecciones_dashboard(elementos),
        })
    return resultado


def _contadores_dashboard_por_rol(rol, resumen):
    mapas = {
        "Administrador": (
            ("Total inspecciones", "total"),
            ("Borradores", "borradores"),
            ("En revisión", "en_revision"),
            ("Devueltos", "devueltas"),
            ("Revisados", "revisadas"),
            ("Aprobados", "aprobadas"),
            ("Publicados", "publicadas"),
        ),
        "Inspector": (
            ("Pendientes", "pendientes"),
            ("Devueltos", "devueltas"),
            ("En revisión", "en_revision"),
            ("Finalizados", "publicadas"),
        ),
        "Supervisor": (
            ("Por revisar", "en_revision"),
            ("Devueltos", "devueltas"),
            ("Revisados", "revisadas"),
        ),
        "Analista": (
            ("Por analizar", "revisadas"),
            ("Aprobados", "aprobadas"),
            ("Por publicar", "aprobadas"),
            ("Publicados", "publicadas"),
        ),
        "Cliente": (
            ("Reportes disponibles", "publicadas"),
        ),
    }
    return [
        {"etiqueta": etiqueta, "valor": resumen.get(clave, 0)}
        for etiqueta, clave in mapas.get(rol, ())
    ]

@login_required
def nueva_parada(request):
    rol = obtener_rol(request.user)

    if rol != "Administrador":
        return HttpResponseForbidden(
            "Solo el Administrador puede crear una nueva parada."
        )

    form = NuevaParadaForm()
    intermitente_form = CrearIntermitenteParadaForm()

    if request.method == "POST":
        accion = request.POST.get(
            "accion",
            "crear_parada",
        )

        # ======================================================
        # CREAR USUARIO INTERMITENTE
        # ======================================================
        if accion == "crear_intermitente":
            intermitente_form = CrearIntermitenteParadaForm(
                request.POST
            )

            if intermitente_form.is_valid():
                try:
                    with transaction.atomic():
                        usuario_nuevo = (
                            intermitente_form.crear_usuario()
                        )

                    rol_nuevo = (
                        intermitente_form.cleaned_data[
                            "rol_intermitente"
                        ]
                    )

                    campo_por_rol = {
                        "Inspector": "inspectores",
                        "Supervisor": "supervisores",
                        "Analista": "analistas",
                        "Cliente": "cliente",
                    }

                    campo_inicial = campo_por_rol[
                        rol_nuevo
                    ]

                    if rol_nuevo == "Cliente":
                        initial = {
                            campo_inicial: usuario_nuevo.id,
                        }
                    else:
                        initial = {
                            campo_inicial: [usuario_nuevo.id],
                        }

                    form = NuevaParadaForm(
                        initial=initial
                    )

                    messages.success(
                        request,
                        (
                            f"Usuario intermitente "
                            f"{usuario_nuevo.username} creado "
                            f"correctamente como {rol_nuevo}. "
                            "Ya aparece disponible en la selección."
                        ),
                    )

                    intermitente_form = (
                        CrearIntermitenteParadaForm()
                    )

                except Exception as error:
                    messages.error(
                        request,
                        (
                            "No se pudo crear el usuario "
                            f"intermitente: {error}"
                        ),
                    )
            else:
                messages.error(
                    request,
                    (
                        "Revisa los datos del nuevo "
                        "usuario intermitente."
                    ),
                )

        # ======================================================
        # CREAR PARADA
        # ======================================================
        elif accion == "crear_parada":
            form = NuevaParadaForm(
                request.POST
            )

            if form.is_valid():
                fecha_inicio = form.cleaned_data[
                    "fecha_inicio"
                ]

                fecha_fin = (
                    form.cleaned_data.get("fecha_fin")
                    or fecha_inicio
                )

                inspectores = list(
                    form.cleaned_data["inspectores"]
                )
                supervisores = list(
                    form.cleaned_data["supervisores"]
                )
                analistas = list(
                    form.cleaned_data["analistas"]
                )
                cliente = form.cleaned_data["cliente"]

                # Los FK históricos/técnicos se conservan usando
                # un responsable principal por rol.
                inspector_principal = inspectores[0]
                supervisor_principal = supervisores[0]
                analista_principal = analistas[0]

                acceso_inicio = form.cleaned_data.get(
                    "acceso_inicio"
                )
                acceso_fin = form.cleaned_data.get(
                    "acceso_fin"
                )

                tags_validos = {
                    "CVB001": [
                        "CVB001",
                        "CVB0001",
                        "0220-CVB-0001",
                        "0220-CVB0001",
                    ],
                    "CVB003": [
                        "CVB003",
                        "CVB0003",
                        "0220-CVB-0003",
                        "0220-CVB0003",
                    ],
                    "CVB004": [
                        "CVB004",
                        "CVB0004",
                        "0220-CVB-0004",
                        "0220-CVB0004",
                    ],
                }

                equipos = {}

                for codigo, tags in tags_validos.items():
                    faja = Faja.objects.filter(
                        tag__in=tags,
                        estado=Faja.Estado.ACTIVA,
                    ).first()

                    if not faja:
                        form.add_error(
                            None,
                            (
                                "No se encontró el equipo "
                                f"activo {codigo}."
                            ),
                        )

                        return render(
                            request,
                            "inspecciones/nueva_parada.html",
                            {
                                "form": form,
                                "intermitente_form":
                                    intermitente_form,
                            },
                        )

                    equipos[codigo] = faja

                for codigo, faja in equipos.items():
                    for tipo in [
                        Inspeccion.Tipo.FAJA,
                        Inspeccion.Tipo.POLEAS,
                        Inspeccion.Tipo.LIFE_SHAFT,
                    ]:
                        existe = Inspeccion.objects.filter(
                            faja=faja,
                            tipo=tipo,
                            fecha_programada=fecha_inicio,
                        ).exists()

                        if existe:
                            form.add_error(
                                None,
                                (
                                    f"Ya existe una inspección "
                                    f"{tipo} para {codigo} en "
                                    f"la fecha "
                                    f"{fecha_inicio:%d/%m/%Y}."
                                ),
                            )

                if form.errors:
                    return render(
                        request,
                        "inspecciones/nueva_parada.html",
                        {
                            "form": form,
                            "intermitente_form":
                                intermitente_form,
                        },
                    )

                with transaction.atomic():
                    parada = form.save(
                        commit=False
                    )

                    parada.planta = "Chancado"
                    parada.estado = (
                        Parada.Estado.PROGRAMADA
                    )
                    parada.creado_por = request.user
                    parada.save()

                    # Rango administrativo para registrar
                    # pertenencia a la parada.
                    inicio_parada = timezone.make_aware(
                        timezone.datetime.combine(
                            fecha_inicio,
                            timezone.datetime.min.time(),
                        ),
                        timezone.get_current_timezone(),
                    )

                    fin_parada = timezone.make_aware(
                        timezone.datetime.combine(
                            fecha_fin,
                            timezone.datetime.max.time(),
                        ),
                        timezone.get_current_timezone(),
                    )

                    equipos_humanos = {
                        "Inspector": inspectores,
                        "Supervisor": supervisores,
                        "Analista": analistas,
                        "Cliente": [cliente],
                    }

                    for rol_acceso, usuarios_rol in (
                        equipos_humanos.items()
                    ):
                        for usuario_acceso in usuarios_rol:
                            try:
                                perfil = (
                                    usuario_acceso
                                    .perfil_sistema
                                )
                            except PerfilUsuario.DoesNotExist:
                                perfil = None

                            es_intermitente = (
                                perfil
                                and perfil.tipo_vinculo
                                == PerfilUsuario
                                .TipoVinculo
                                .INTERMITENTE
                            )

                            inicio_usuario = (
                                acceso_inicio
                                if es_intermitente
                                else inicio_parada
                            )

                            fin_usuario = (
                                acceso_fin
                                if es_intermitente
                                else fin_parada
                            )

                            AccesoParada.objects.create(
                                parada=parada,
                                usuario=usuario_acceso,
                                rol=rol_acceso,
                                fecha_inicio=inicio_usuario,
                                fecha_fin=fin_usuario,
                                activo=True,
                                creado_por=request.user,
                            )

                    # Generar los 9 reportes técnicos.
                    for codigo, faja in equipos.items():
                        for tipo in [
                            Inspeccion.Tipo.FAJA,
                            Inspeccion.Tipo.POLEAS,
                            Inspeccion.Tipo.LIFE_SHAFT,
                        ]:
                            tipo_codigo = {
                                Inspeccion.Tipo.FAJA:
                                    "FAJA",
                                Inspeccion.Tipo.POLEAS:
                                    "POLEAS",
                                Inspeccion.Tipo.LIFE_SHAFT:
                                    "LIFE-SHAFT",
                            }[tipo]

                            codigo_temporal = (
                                f"PARADA-{parada.id}-"
                                f"{codigo}-"
                                f"{tipo_codigo}"
                            )

                            Inspeccion.objects.create(
                                parada=parada,
                                faja=faja,
                                tipo=tipo,
                                codigo_reporte=(
                                    codigo_temporal
                                ),
                                fecha_programada=(
                                    fecha_inicio
                                ),
                                fecha_inspeccion=(
                                    fecha_inicio
                                ),
                                inspector=(
                                    inspector_principal
                                ),
                                supervisor=(
                                    supervisor_principal
                                ),
                                analista=(
                                    analista_principal
                                ),
                                cliente=cliente,
                                estado=(
                                    Inspeccion
                                    .Estado
                                    .BORRADOR
                                ),
                                planta="Chancado",
                                creado_por=request.user,
                            )

                messages.success(
                    request,
                    (
                        "Parada creada correctamente. "
                        "Se generaron los 9 reportes de Chancado "
                        "y se registró todo el equipo de trabajo."
                    ),
                )

                return redirect(
                    "dashboard"
                )

            messages.error(
                request,
                (
                    "Revisa los datos de la "
                    "nueva parada."
                ),
            )

    return render(
        request,
        "inspecciones/nueva_parada.html",
        {
            "form": form,
            "intermitente_form": intermitente_form,
        },
    )


def _crear_accesos_parada(
    *,
    parada,
    request,
    fecha_inicio,
    fecha_fin,
    inspectores,
    supervisores,
    analistas,
    cliente,
    acceso_inicio,
    acceso_fin,
):
    inicio_parada = timezone.make_aware(
        timezone.datetime.combine(
            fecha_inicio,
            timezone.datetime.min.time(),
        ),
        timezone.get_current_timezone(),
    )

    fin_parada = timezone.make_aware(
        timezone.datetime.combine(
            fecha_fin,
            timezone.datetime.max.time(),
        ),
        timezone.get_current_timezone(),
    )

    equipos_humanos = {
        "Inspector": inspectores,
        "Supervisor": supervisores,
        "Analista": analistas,
        "Cliente": [cliente],
    }

    for rol_acceso, usuarios_rol in equipos_humanos.items():
        for usuario_acceso in usuarios_rol:
            try:
                perfil = usuario_acceso.perfil_sistema
            except PerfilUsuario.DoesNotExist:
                perfil = None

            es_intermitente = (
                perfil
                and perfil.tipo_vinculo
                == PerfilUsuario.TipoVinculo.INTERMITENTE
            )

            AccesoParada.objects.create(
                parada=parada,
                usuario=usuario_acceso,
                rol=rol_acceso,
                fecha_inicio=(
                    acceso_inicio
                    if es_intermitente
                    else inicio_parada
                ),
                fecha_fin=(
                    acceso_fin
                    if es_intermitente
                    else fin_parada
                ),
                activo=True,
                creado_por=request.user,
            )


@login_required
def nueva_parada_molienda(request):
    rol = obtener_rol(request.user)

    if rol != "Administrador":
        return HttpResponseForbidden(
            "Solo el Administrador puede crear una nueva parada."
        )

    form = NuevaParadaForm()
    intermitente_form = CrearIntermitenteParadaForm()
    contexto_base = {
        "proceso_label": "Molienda",
        "reportes_label": (
            "Top Cover/Faja y Poleas para Faja 6, "
            "Faja 7, Faja 10, Faja 11, Faja 15, Faja 17 y Faja 18."
        ),
        "volver_url": "dashboard",
    }

    if request.method == "POST":
        accion = request.POST.get("accion", "crear_parada")

        if accion == "crear_intermitente":
            intermitente_form = CrearIntermitenteParadaForm(request.POST)

            if intermitente_form.is_valid():
                try:
                    with transaction.atomic():
                        usuario_nuevo = intermitente_form.crear_usuario()

                    rol_nuevo = intermitente_form.cleaned_data[
                        "rol_intermitente"
                    ]

                    campo_por_rol = {
                        "Inspector": "inspectores",
                        "Supervisor": "supervisores",
                        "Analista": "analistas",
                        "Cliente": "cliente",
                    }
                    campo_inicial = campo_por_rol[rol_nuevo]

                    form = NuevaParadaForm(
                        initial={
                            campo_inicial: (
                                usuario_nuevo.id
                                if rol_nuevo == "Cliente"
                                else [usuario_nuevo.id]
                            ),
                        }
                    )

                    messages.success(
                        request,
                        (
                            f"Usuario intermitente "
                            f"{usuario_nuevo.username} creado "
                            f"correctamente como {rol_nuevo}. "
                            "Ya aparece disponible en la seleccion."
                        ),
                    )
                    intermitente_form = CrearIntermitenteParadaForm()
                except Exception as error:
                    messages.error(
                        request,
                        (
                            "No se pudo crear el usuario "
                            f"intermitente: {error}"
                        ),
                    )
            else:
                messages.error(
                    request,
                    "Revisa los datos del nuevo usuario intermitente.",
                )

        elif accion == "crear_parada":
            form = NuevaParadaForm(request.POST)

            if form.is_valid():
                fecha_inicio = form.cleaned_data["fecha_inicio"]
                fecha_fin = form.cleaned_data.get("fecha_fin") or fecha_inicio
                inspectores = list(form.cleaned_data["inspectores"])
                supervisores = list(form.cleaned_data["supervisores"])
                analistas = list(form.cleaned_data["analistas"])
                cliente = form.cleaned_data["cliente"]
                inspector_principal = inspectores[0]
                supervisor_principal = supervisores[0]
                analista_principal = analistas[0]
                acceso_inicio = form.cleaned_data.get("acceso_inicio")
                acceso_fin = form.cleaned_data.get("acceso_fin")

                configs_molienda = [
                    ("CVB0006", ["CVB006", "CVB0006", "0240-CVB-006", "0240-CVB0006", "0310CVB0006"], [("FAJA", "FAJA"), ("POLEAS", "POLEAS")]),
                    ("CVB0007", ["CVB007", "CVB0007", "0240-CVB-007", "0240-CVB0007", "0310CVB0007"], [("FAJA", "FAJA"), ("POLEAS", "POLEAS")]),
                    ("CVB0010-ENTRANTE", ["CVB0010-ENTRANTE"], [("FAJA", "FAJA-ENTRANTE"), ("POLEAS", "POLEAS-ENTRANTE")]),
                    ("CVB0010-SALIENTE", ["CVB0010-SALIENTE"], [("FAJA", "FAJA-SALIENTE"), ("POLEAS", "POLEAS-SALIENTE")]),
                    ("CVB0011", ["CVB0011"], [("FAJA", "FAJA"), ("POLEAS", "POLEAS")]),
                    ("CVB0015", ["CVB0015"], [("FAJA", "FAJA"), ("POLEAS", "POLEAS")]),
                    ("CVB0017", ["CVB0017"], [("FAJA", "FAJA"), ("POLEAS", "POLEAS")]),
                    ("CVB0018", ["CVB0018"], [("FAJA", "FAJA"), ("POLEAS", "POLEAS")]),
                ]
                equipos_molienda = []

                for codigo_equipo, tags, reportes in configs_molienda:
                    faja = Faja.objects.filter(
                        tag__in=tags,
                        estado=Faja.Estado.ACTIVA,
                    ).first()

                    if not faja:
                        form.add_error(
                            None,
                            f"No se encontro el equipo activo {codigo_equipo}.",
                        )
                        continue

                    equipos_molienda.append((codigo_equipo, faja, reportes))

                if form.errors:
                    return render(
                        request,
                        "inspecciones/nueva_parada.html",
                        {
                            "form": form,
                            "intermitente_form": intermitente_form,
                            **contexto_base,
                        },
                    )

                with transaction.atomic():
                    parada = form.save(commit=False)
                    parada.planta = "Molienda"
                    parada.estado = Parada.Estado.PROGRAMADA
                    parada.creado_por = request.user
                    parada.save()

                    _crear_accesos_parada(
                        parada=parada,
                        request=request,
                        fecha_inicio=fecha_inicio,
                        fecha_fin=fecha_fin,
                        inspectores=inspectores,
                        supervisores=supervisores,
                        analistas=analistas,
                        cliente=cliente,
                        acceso_inicio=acceso_inicio,
                        acceso_fin=acceso_fin,
                    )

                    for codigo_equipo, faja, reportes in equipos_molienda:
                        for tipo_codigo, sufijo in reportes:
                            tipo = getattr(Inspeccion.Tipo, tipo_codigo)
                            Inspeccion.objects.create(
                                parada=parada,
                                faja=faja,
                                tipo=tipo,
                                codigo_reporte=(
                                    f"PARADA-{parada.id}-"
                                    f"{codigo_equipo}-{sufijo}"
                                ),
                                fecha_programada=fecha_inicio,
                                fecha_inspeccion=fecha_inicio,
                                inspector=inspector_principal,
                                supervisor=supervisor_principal,
                                analista=analista_principal,
                                cliente=cliente,
                                estado=Inspeccion.Estado.BORRADOR,
                                planta="Molienda",
                                proceso="Transporte de concentrado",
                                etapa="Operaciones",
                                condicion_equipo="En uso",
                                creado_por=request.user,
                            )

                messages.success(
                    request,
                    (
                        "Parada de Molienda creada correctamente. "
                        "Se generaron los reportes de Molienda: "
                        "Faja/Top Cover y Poleas."
                    ),
                )
                return redirect("dashboard")

            messages.error(request, "Revisa los datos de la nueva parada.")

    return render(
        request,
        "inspecciones/nueva_parada.html",
        {
            "form": form,
            "intermitente_form": intermitente_form,
            **contexto_base,
        },
    )


@login_required
def gestionar_asignaciones_parada(request, parada_id):
    if obtener_rol(request.user) != "Administrador":
        return HttpResponseForbidden(
            "Solo el Administrador puede modificar la parada."
        )

    parada = get_object_or_404(
        Parada,
        id=parada_id,
    )

    inspecciones_parada = list(
        parada.inspecciones.select_related(
            "faja",
            "inspector",
            "supervisor",
            "analista",
            "cliente",
        ).order_by(
            "faja__tag",
            "tipo",
            "id",
        )
    )

    if not inspecciones_parada:
        messages.error(
            request,
            "La parada no contiene inspecciones.",
        )
        return redirect("dashboard")

    referencia = next(
        (
            inspeccion
            for inspeccion in inspecciones_parada
            if inspeccion.estado
            != Inspeccion.Estado.PUBLICADO
        ),
        inspecciones_parada[0],
    )

    # ==========================================================
    # COMPATIBILIDAD: PARADAS CREADAS ANTES DEL MULTIUSUARIO
    # ==========================================================
    inicio_parada = timezone.make_aware(
        timezone.datetime.combine(
            parada.fecha_inicio,
            timezone.datetime.min.time(),
        ),
        timezone.get_current_timezone(),
    )

    fecha_fin_parada = (
        parada.fecha_fin
        or parada.fecha_inicio
    )

    fin_parada = timezone.make_aware(
        timezone.datetime.combine(
            fecha_fin_parada,
            timezone.datetime.max.time(),
        ),
        timezone.get_current_timezone(),
    )

    principales = {
        "Inspector": referencia.inspector,
        "Supervisor": referencia.supervisor,
        "Analista": referencia.analista,
        "Cliente": referencia.cliente,
    }

    # Si una parada antigua aún no tiene accesos para un rol,
    # registrar su responsable principal sin modificar reportes.
    for rol_base, usuario_base in principales.items():
        if not usuario_base:
            continue

        if not AccesoParada.objects.filter(
            parada=parada,
            rol=rol_base,
        ).exists():
            AccesoParada.objects.create(
                parada=parada,
                usuario=usuario_base,
                rol=rol_base,
                fecha_inicio=inicio_parada,
                fecha_fin=fin_parada,
                activo=True,
                creado_por=request.user,
            )

    if request.method == "POST":
        form = GestionAsignacionParadaForm(
            request.POST
        )

        if form.is_valid():
            accion = form.cleaned_data["accion"]
            rol = form.cleaned_data["rol"]
            usuario = form.cleaned_data["usuario"]
            motivo = form.cleaned_data["motivo"]
            fecha_inicio = form.cleaned_data[
                "fecha_inicio"
            ]
            fecha_fin = form.cleaned_data[
                "fecha_fin"
            ]

            campo_por_rol = {
                "Inspector": "inspector",
                "Supervisor": "supervisor",
                "Analista": "analista",
                "Cliente": "cliente",
            }

            campo = campo_por_rol[rol]

            try:
                perfil = usuario.perfil_sistema
            except PerfilUsuario.DoesNotExist:
                perfil = None

            es_intermitente = (
                perfil
                and perfil.tipo_vinculo
                == PerfilUsuario.TipoVinculo.INTERMITENTE
            )

            inicio_usuario = (
                fecha_inicio
                if es_intermitente
                else inicio_parada
            )
            fin_usuario = (
                fecha_fin
                if es_intermitente
                else fin_parada
            )

            with transaction.atomic():
                if accion == "AGREGAR":
                    if AccesoParada.objects.filter(
                        parada=parada,
                        usuario=usuario,
                        rol=rol,
                        activo=True,
                    ).exists():
                        form.add_error(
                            "usuario",
                            (
                                "El usuario ya pertenece "
                                "a este equipo."
                            ),
                        )
                    else:
                        acceso, creado = (
                            AccesoParada.objects.get_or_create(
                                parada=parada,
                                usuario=usuario,
                                rol=rol,
                                defaults={
                                    "fecha_inicio":
                                        inicio_usuario,
                                    "fecha_fin":
                                        fin_usuario,
                                    "activo": True,
                                    "creado_por":
                                        request.user,
                                },
                            )
                        )

                        if not creado:
                            acceso.fecha_inicio = (
                                inicio_usuario
                            )
                            acceso.fecha_fin = (
                                fin_usuario
                            )
                            acceso.activo = True
                            acceso.creado_por = request.user
                            acceso.save(
                                update_fields=[
                                    "fecha_inicio",
                                    "fecha_fin",
                                    "activo",
                                    "creado_por",
                                    "actualizado_en",
                                ]
                            )

                        HistorialAsignacionParada.objects.create(
                            parada=parada,
                            rol=rol,
                            usuario_anterior=None,
                            usuario_nuevo=usuario,
                            motivo=motivo,
                            cambiado_por=request.user,
                        )

                        messages.success(
                            request,
                            (
                                f"{usuario.get_full_name() or usuario.username} "
                                f"fue agregado como {rol}."
                            ),
                        )

                        return redirect(
                            "gestionar_asignaciones_parada",
                            parada_id=parada.id,
                        )

                elif accion == "QUITAR":
                    acceso = (
                        AccesoParada.objects
                        .filter(
                            parada=parada,
                            usuario=usuario,
                            rol=rol,
                            activo=True,
                        )
                        .first()
                    )

                    if not acceso:
                        form.add_error(
                            "usuario",
                            (
                                "El usuario no pertenece "
                                "actualmente a ese equipo."
                            ),
                        )
                    else:
                        otros = list(
                            AccesoParada.objects
                            .filter(
                                parada=parada,
                                rol=rol,
                                activo=True,
                            )
                            .exclude(
                                usuario=usuario
                            )
                            .select_related("usuario")
                        )

                        if not otros:
                            form.add_error(
                                "usuario",
                                (
                                    "No puedes quitar al último "
                                    f"{rol} de la parada."
                                ),
                            )
                        else:
                            acceso.activo = False
                            acceso.save(
                                update_fields=[
                                    "activo",
                                    "actualizado_en",
                                ]
                            )

                            # Si era principal, trasladar el FK
                            # solamente en reportes no publicados.
                            principal_actual = getattr(
                                referencia,
                                campo,
                            )

                            if (
                                principal_actual
                                and principal_actual.id
                                == usuario.id
                            ):
                                nuevo_principal = (
                                    otros[0].usuario
                                )

                                for inspeccion in inspecciones_parada:
                                    if (
                                        inspeccion.estado
                                        == Inspeccion.Estado.PUBLICADO
                                    ):
                                        continue

                                    setattr(
                                        inspeccion,
                                        campo,
                                        nuevo_principal,
                                    )
                                    inspeccion.save(
                                        update_fields=[campo]
                                    )

                            HistorialAsignacionParada.objects.create(
                                parada=parada,
                                rol=rol,
                                usuario_anterior=usuario,
                                usuario_nuevo=None,
                                motivo=motivo,
                                cambiado_por=request.user,
                            )

                            messages.success(
                                request,
                                (
                                    f"{usuario.get_full_name() or usuario.username} "
                                    f"fue retirado del equipo {rol}."
                                ),
                            )

                            return redirect(
                                "gestionar_asignaciones_parada",
                                parada_id=parada.id,
                            )

                elif accion == "CAMBIAR_CLIENTE":
                    cliente_anterior = (
                        AccesoParada.objects
                        .filter(
                            parada=parada,
                            rol="Cliente",
                            activo=True,
                        )
                        .select_related("usuario")
                        .first()
                    )

                    if (
                        cliente_anterior
                        and cliente_anterior.usuario_id
                        == usuario.id
                    ):
                        form.add_error(
                            "usuario",
                            (
                                "El usuario seleccionado ya "
                                "es el cliente actual."
                            ),
                        )
                    else:
                        AccesoParada.objects.filter(
                            parada=parada,
                            rol="Cliente",
                            activo=True,
                        ).update(
                            activo=False
                        )

                        acceso, creado = (
                            AccesoParada.objects.get_or_create(
                                parada=parada,
                                usuario=usuario,
                                rol="Cliente",
                                defaults={
                                    "fecha_inicio":
                                        inicio_usuario,
                                    "fecha_fin":
                                        fin_usuario,
                                    "activo": True,
                                    "creado_por":
                                        request.user,
                                },
                            )
                        )

                        if not creado:
                            acceso.fecha_inicio = (
                                inicio_usuario
                            )
                            acceso.fecha_fin = (
                                fin_usuario
                            )
                            acceso.activo = True
                            acceso.creado_por = request.user
                            acceso.save(
                                update_fields=[
                                    "fecha_inicio",
                                    "fecha_fin",
                                    "activo",
                                    "creado_por",
                                    "actualizado_en",
                                ]
                            )

                        for inspeccion in inspecciones_parada:
                            if (
                                inspeccion.estado
                                == Inspeccion.Estado.PUBLICADO
                            ):
                                continue

                            inspeccion.cliente = usuario
                            inspeccion.save(
                                update_fields=["cliente"]
                            )

                        HistorialAsignacionParada.objects.create(
                            parada=parada,
                            rol="Cliente",
                            usuario_anterior=(
                                cliente_anterior.usuario
                                if cliente_anterior
                                else None
                            ),
                            usuario_nuevo=usuario,
                            motivo=motivo,
                            cambiado_por=request.user,
                        )

                        messages.success(
                            request,
                            "Cliente actualizado correctamente.",
                        )

                        return redirect(
                            "gestionar_asignaciones_parada",
                            parada_id=parada.id,
                        )

    else:
        form = GestionAsignacionParadaForm()

    accesos_activos = (
        AccesoParada.objects
        .filter(
            parada=parada,
            activo=True,
        )
        .select_related(
            "usuario",
            "usuario__perfil_sistema",
        )
        .order_by(
            "rol",
            "usuario__first_name",
            "usuario__last_name",
            "usuario__username",
        )
    )

    equipos = {
        "Inspector": [],
        "Supervisor": [],
        "Analista": [],
        "Cliente": [],
    }

    for acceso in accesos_activos:
        equipos.setdefault(
            acceso.rol,
            [],
        ).append(acceso)

    historial = (
        parada.historial_asignaciones
        .select_related(
            "usuario_anterior",
            "usuario_nuevo",
            "cambiado_por",
        )
        .all()
    )

    return render(
        request,
        "inspecciones/gestionar_asignaciones_parada.html",
        {
            "parada": parada,
            "form": form,
            "equipos": equipos,
            "historial": historial,
            "inspecciones_parada": inspecciones_parada,
        },
    )


@login_required
def dashboard(request):
    rol = obtener_rol(request.user)

    inspecciones_qs = (
        Inspeccion.objects.select_related(
            "parada",
            "faja",
            "inspector",
            "supervisor",
            "analista",
            "cliente",
        )
        .prefetch_related("historial")
    )

    # ==========================================================
    # FILTRO BASE POR ROL / ASIGNACIÓN
    # ==========================================================
    if rol == "Inspector":
        inspecciones_qs = inspecciones_qs.filter(
            Q(inspector_id=request.user.id)
            | Q(
                parada__accesos_temporales__usuario_id=request.user.id,
                parada__accesos_temporales__rol="Inspector",
                parada__accesos_temporales__activo=True,
            )
        ).distinct()

    elif rol == "Supervisor":
        inspecciones_qs = inspecciones_qs.filter(
            Q(supervisor_id=request.user.id)
            | Q(
                parada__accesos_temporales__usuario_id=request.user.id,
                parada__accesos_temporales__rol="Supervisor",
                parada__accesos_temporales__activo=True,
            )
        ).distinct()

    elif rol == "Analista":
        inspecciones_qs = inspecciones_qs.filter(
            Q(analista_id=request.user.id)
            | Q(
                parada__accesos_temporales__usuario_id=request.user.id,
                parada__accesos_temporales__rol="Analista",
                parada__accesos_temporales__activo=True,
            )
        ).distinct()

    elif rol == "Cliente":
        inspecciones_qs = inspecciones_qs.filter(
            Q(cliente_id=request.user.id)
            | Q(
                parada__accesos_temporales__usuario_id=request.user.id,
                parada__accesos_temporales__rol="Cliente",
                parada__accesos_temporales__activo=True,
            ),
            estado=Inspeccion.Estado.PUBLICADO,
        ).distinct()

    elif rol != "Administrador":
        inspecciones_qs = inspecciones_qs.none()

    # ==========================================================
    # CONVERTIR A LISTA
    # ==========================================================
    inspecciones = list(
        inspecciones_qs.order_by(
            "faja__tag",
            "tipo",
            "id",
        )
    )

    # Seguridad adicional:
    # contrato vigente + asignación + acceso intermitente.
    if rol != "Administrador":
        inspecciones = [
            inspeccion
            for inspeccion in inspecciones
            if usuario_puede_abrir_inspeccion(
                request.user,
                inspeccion,
            )
        ]

    # ==========================================================
    # ALCANCE VISUAL DE LA VERSIÓN APROBADA
    # ==========================================================
    # Solo controla lo que aparece en Dashboard.
    # NO elimina registros y NO interviene en history.py,
    # restricciones A-G, formularios, Excel, fotos o workflow.
    inspecciones = [
        inspeccion
        for inspeccion in inspecciones
        if inspeccion.parada_id == PARADA_ACTUAL_ID
    ]

    # ==========================================================
    # CONTADORES GENERALES SOBRE LO QUE REALMENTE PUEDE VER
    # ==========================================================
    estados = Inspeccion.Estado

    resumen = {
        "total": len(inspecciones),
        "borradores": sum(
            1 for i in inspecciones
            if i.estado == estados.BORRADOR
        ),
        "en_revision": sum(
            1 for i in inspecciones
            if i.estado == estados.EN_REVISION
        ),
        "devueltas": sum(
            1 for i in inspecciones
            if i.estado == estados.DEVUELTO
        ),
        "revisadas": sum(
            1 for i in inspecciones
            if i.estado == estados.REVISADO
        ),
        "aprobadas": sum(
            1 for i in inspecciones
            if i.estado == estados.APROBADO
        ),
        "publicadas": sum(
            1 for i in inspecciones
            if i.estado == estados.PUBLICADO
        ),
        "pendientes": sum(
            1 for i in inspecciones
            if i.estado in {
                estados.BORRADOR,
                estados.DEVUELTO,
            }
        ),
    }

    # ==========================================================
    # CONTADORES POR INSPECCIÓN
    # ==========================================================
    for inspeccion in inspecciones:
        _agregar_contadores_dashboard(inspeccion)

        inspeccion.ultima_devolucion_dashboard = next(
            (
                evento
                for evento in inspeccion.historial.all()
                if evento.accion in {
                    "DEVOLVER_SUPERVISOR",
                    "DEVOLVER_ANALISTA",
                }
            ),
            None,
        )

    inspecciones_cvb0003 = [
        inspeccion
        for inspeccion in inspecciones
        if _es_inspeccion_cvb0003(inspeccion)
    ]

    otras_inspecciones = [
        inspeccion
        for inspeccion in inspecciones
        if not _es_inspeccion_cvb0003(inspeccion)
    ]

    inspecciones_chancado = inspecciones

    # ==========================================================
    # PARADA ACTUAL
    # ==========================================================
    paradas_disponibles = {
        inspeccion.parada_id: inspeccion.parada
        for inspeccion in inspecciones_chancado
        if inspeccion.parada_id is not None
    }

    if paradas_disponibles:
        parada_actual = max(
            paradas_disponibles.values(),
            key=lambda parada: (
                parada.fecha_inicio,
                parada.id,
            ),
        )

        inspecciones_actuales = [
            inspeccion
            for inspeccion in inspecciones_chancado
            if inspeccion.parada_id == parada_actual.id
        ]

        inspecciones_actuales = _ordenar_inspecciones_dashboard(
            inspecciones_actuales
        )

    else:
        parada_actual = None

        # Compatibilidad con inspecciones anteriores
        # a la implementación del modelo Parada.
        inspecciones_actuales = _ultima_inspeccion_por_tipo(
            inspecciones_cvb0003
        )

    ids_actuales = {
        inspeccion.id
        for inspeccion in inspecciones_actuales
    }

    inspecciones_existentes = [
        inspeccion
        for inspeccion in inspecciones_chancado
        if inspeccion.id not in ids_actuales
    ]

    fecha_captura_actual = (
        parada_actual.fecha_inicio
        if parada_actual
        else max(
            (
                inspeccion.fecha_inspeccion
                or inspeccion.fecha_programada
                for inspeccion in inspecciones_actuales
                if (
                    inspeccion.fecha_inspeccion
                    or inspeccion.fecha_programada
                )
            ),
            default=None,
        )
    )

    # ==========================================================
    # REPORTES EXISTENTES AGRUPADOS POR EQUIPO
    # ==========================================================
    reportes_existentes_por_equipo = []

    for codigo_equipo in [
        "CVB001",
        "CVB003",
        "CVB004",
    ]:
        reportes_equipo = [
            inspeccion
            for inspeccion in inspecciones_existentes
            if _codigo_equipo_dashboard(inspeccion)
            == codigo_equipo
        ]

        if reportes_equipo:
            reportes_existentes_por_equipo.append(
                {
                    "codigo": codigo_equipo,
                    "inspecciones":
                        _ordenar_inspecciones_dashboard(
                            reportes_equipo
                        ),
                }
            )

    # ==========================================================
    # AGRUPACIONES POR ROL / ESTADO
    # ==========================================================
    grupos_dashboard = {
        "actuales": _agrupar_dashboard_por_parada(
            inspecciones_actuales
        ),

        "historicas": _agrupar_dashboard_por_parada(
            inspecciones_existentes
        ),

        "pendientes_inspector": _agrupar_dashboard_por_parada(
            [
                i for i in inspecciones_chancado
                if i.estado in {
                    estados.BORRADOR,
                    estados.DEVUELTO,
                }
            ]
        ),

        "enviadas_inspector": _agrupar_dashboard_por_parada(
            [
                i for i in inspecciones_chancado
                if i.estado in {
                    estados.EN_REVISION,
                    estados.REVISADO,
                    estados.APROBADO,
                }
            ]
        ),

        "finalizadas": _agrupar_dashboard_por_parada(
            [
                i for i in inspecciones_chancado
                if i.estado == estados.PUBLICADO
            ]
        ),

        "pendientes_supervisor": _agrupar_dashboard_por_parada(
            [
                i for i in inspecciones_chancado
                if i.estado == estados.EN_REVISION
            ]
        ),

        "devueltas": _agrupar_dashboard_por_parada(
            [
                i for i in inspecciones_chancado
                if i.estado == estados.DEVUELTO
            ]
        ),

        "revisadas_supervisor": _agrupar_dashboard_por_parada(
            [
                i for i in inspecciones_chancado
                if i.estado in {
                    estados.REVISADO,
                    estados.APROBADO,
                    estados.PUBLICADO,
                }
            ]
        ),

        "pendientes_analista": _agrupar_dashboard_por_parada(
            [
                i for i in inspecciones_chancado
                if i.estado == estados.REVISADO
            ]
        ),

        "aprobadas_analista": _agrupar_dashboard_por_parada(
            [
                i for i in inspecciones_chancado
                if i.estado == estados.APROBADO
            ]
        ),

        "publicadas": _agrupar_dashboard_por_parada(
            [
                i for i in inspecciones_chancado
                if i.estado == estados.PUBLICADO
            ]
        ),
    }
    # ==========================================================
    # FILTROS - PARADA DISPONIBLE PARA EL DASHBOARD
    # SOLO LECTURA DEL ESTADO; NO MODIFICA CHANCADO
    # ==========================================================
    parada_filtros_actual = None

    if rol == "Administrador":
        parada_filtros_actual = (
            Parada.objects
            .filter(planta__iexact="Filtros")
            .order_by("-fecha_inicio", "-id")
            .first()
        )
    else:
        accesos_filtros = (
            AccesoParada.objects
            .select_related("parada")
            .filter(
                usuario_id=request.user.id,
                rol=rol,
                activo=True,
                parada__planta__iexact="Filtros",
            )
            .order_by("-parada__fecha_inicio", "-parada_id")
        )

        for acceso in accesos_filtros:
            if acceso.esta_vigente():
                parada_filtros_actual = acceso.parada
                break

    reportes_filtros_total = 0

    if parada_filtros_actual:
        reportes_filtros_total = (
            parada_filtros_actual.reportes_filtros.count()
        )

    parada_molienda_actual = None

    if rol == "Administrador":
        parada_molienda_actual = (
            Parada.objects
            .filter(planta__iexact="Molienda")
            .order_by("-fecha_inicio", "-id")
            .first()
        )
    else:
        accesos_molienda = (
            AccesoParada.objects
            .select_related("parada")
            .filter(
                usuario_id=request.user.id,
                rol=rol,
                activo=True,
                parada__planta__iexact="Molienda",
            )
            .order_by("-parada__fecha_inicio", "-parada_id")
        )

        for acceso in accesos_molienda:
            if acceso.esta_vigente():
                parada_molienda_actual = acceso.parada
                break

    inspecciones_molienda_total = 0

    if parada_molienda_actual:
        inspecciones_molienda_total = (
            parada_molienda_actual.inspecciones.count()
        )
    # ==========================================================
    # CLIENTE: ÚLTIMA PARADA HISTÓRICA OFICIAL (SOLO LECTURA)
    # ==========================================================
    historial_cliente_por_equipo = []

    if rol == "Cliente":
        historicos_cliente = list(
            Inspeccion.objects.select_related(
                "faja",
                "inspector",
                "supervisor",
                "analista",
                "cliente",
            )
            .filter(
                id__in=INSPECCIONES_HISTORICAS_OFICIALES_IDS,
                cliente_id=request.user.id,
            )
            .order_by(
                "faja__tag",
                "tipo",
                "id",
            )
        )

        historicos_cliente = [
            inspeccion
            for inspeccion in historicos_cliente
            if usuario_puede_abrir_inspeccion(
                request.user,
                inspeccion,
            )
        ]

        for inspeccion in historicos_cliente:
            _agregar_contadores_dashboard(inspeccion)

        for codigo_equipo in [
            "CVB001",
            "CVB003",
            "CVB004",
        ]:
            reportes_equipo = [
                inspeccion
                for inspeccion in historicos_cliente
                if _codigo_equipo_dashboard(inspeccion)
                == codigo_equipo
            ]

            if reportes_equipo:
                historial_cliente_por_equipo.append(
                    {
                        "codigo": codigo_equipo,
                        "inspecciones":
                            _ordenar_inspecciones_dashboard(
                                reportes_equipo
                            ),
                    }
                )

    return render(
        request,
        "dashboard.html",
        {
            "rol": rol,
            "inspecciones": inspecciones,
            "inspecciones_nuevas_cvb0003": (
                inspecciones_actuales
            ),
            "historial_completo_cvb0003": (
                inspecciones_existentes
            ),
            "reportes_existentes_por_equipo": (
                reportes_existentes_por_equipo
            ),
            "grupos_dashboard": grupos_dashboard,
            "otras_inspecciones": otras_inspecciones,
            "fecha_captura_actual": fecha_captura_actual,
            "parada_actual": parada_actual,
            "parada_filtros_actual": parada_filtros_actual,
            "reportes_filtros_total": reportes_filtros_total,
            "parada_molienda_actual": parada_molienda_actual,
            "inspecciones_molienda_total": inspecciones_molienda_total,
            "historial_cliente_por_equipo": (
                historial_cliente_por_equipo
            ),
            "fecha_historial_cliente": timezone.datetime(
                2026, 8, 3
            ).date(),
            "resumen": resumen,
            "contadores_dashboard": (
                _contadores_dashboard_por_rol(
                    rol,
                    resumen,
                )
            ),
        },
    )


@login_required
def historial_inspecciones(request):
    """
    Historial visible de la versión aprobada.

    Muestra únicamente los 9 reportes históricos oficiales definidos en
    presentation_scope.py. Esta vista NO modifica la base de datos y NO
    participa en el cálculo técnico de restricciones.
    """
    rol = obtener_rol(request.user)

    inspecciones_qs = (
        Inspeccion.objects.select_related(
            "parada",
            "faja",
            "inspector",
            "supervisor",
            "analista",
            "cliente",
        )
        .prefetch_related("historial")
        .filter(
            id__in=INSPECCIONES_HISTORICAS_OFICIALES_IDS
        )
    )

    # Mantener la misma seguridad por rol de la aplicación.
    if rol == "Inspector":
        inspecciones_qs = inspecciones_qs.filter(
            inspector_id=request.user.id
        )

    elif rol == "Supervisor":
        inspecciones_qs = inspecciones_qs.filter(
            supervisor_id=request.user.id
        )

    elif rol == "Analista":
        inspecciones_qs = inspecciones_qs.filter(
            analista_id=request.user.id
        )

    elif rol == "Cliente":
        # Históricos oficiales del mismo cliente: consulta de solo lectura.
        inspecciones_qs = inspecciones_qs.filter(
            cliente_id=request.user.id,
        )

    elif rol != "Administrador":
        inspecciones_qs = inspecciones_qs.none()

    inspecciones = list(
        inspecciones_qs.order_by(
            "faja__tag",
            "tipo",
            "id",
        )
    )

    if rol != "Administrador":
        inspecciones = [
            inspeccion
            for inspeccion in inspecciones
            if usuario_puede_abrir_inspeccion(
                request.user,
                inspeccion,
            )
        ]

    for inspeccion in inspecciones:
        _agregar_contadores_dashboard(inspeccion)

    historial_por_equipo = []

    for codigo_equipo in [
        "CVB001",
        "CVB003",
        "CVB004",
    ]:
        reportes_equipo = [
            inspeccion
            for inspeccion in inspecciones
            if _codigo_equipo_dashboard(inspeccion)
            == codigo_equipo
        ]

        if reportes_equipo:
            historial_por_equipo.append(
                {
                    "codigo": codigo_equipo,
                    "inspecciones":
                        _ordenar_inspecciones_dashboard(
                            reportes_equipo
                        ),
                }
            )

    return render(
        request,
        "inspecciones/historial_inspecciones.html",
        {
            "rol": rol,
            "historial_por_equipo": historial_por_equipo,
            "inspecciones": inspecciones,
            "fecha_historial_referencia": timezone.datetime(
                2026, 8, 3
            ).date(),
        },
    )

def crear_estructura_mediciones_faja(inspeccion):
    if inspeccion.tipo != Inspeccion.Tipo.FAJA:
        return

    filas = [
        {
            "seccion": "EMPALME E-01",
            "punto": "-1 m",
            "bastidor": "174",
            "lado": "TOP COVER",
            "posicion": "-1 m",
            "orden": 1,
        },
        {
            "seccion": "EMPALME E-01",
            "punto": "+1 m",
            "bastidor": "174",
            "lado": "TOP COVER",
            "posicion": "+1 m",
            "orden": 2,
        },
        {
            "seccion": "EMPALME E-02",
            "punto": "-1 m",
            "bastidor": "86-87",
            "lado": "BOTTOM COVER",
            "posicion": "-1 m",
            "orden": 3,
        },
        {
            "seccion": "EMPALME E-02",
            "punto": "+1 m",
            "bastidor": "86-87",
            "lado": "BOTTOM COVER",
            "posicion": "+1 m",
            "orden": 4,
        },
    ]

    posiciones = [
        200,
        195,
        190,
        185,
        180,
        170,
        160,
        150,
        140,
        130,
        120,
        110,
        100,
        90,
        80,
    ]

    for indice, posicion in enumerate(posiciones, start=1):
        filas.append(
            {
                "seccion": "TRAMOS 200-80",
                "punto": str(indice),
                "bastidor": str(posicion),
                "lado": "TOP COVER",
                "posicion": str(posicion),
                "orden": 100 + indice,
            }
        )

    for fila in filas:
        clave = {
            "inspeccion": inspeccion,
            "seccion": fila["seccion"],
            "posicion": fila["posicion"],
            "bastidor": fila["bastidor"],
            "lado": fila["lado"],
        }
        defaults = {
            "punto": fila["punto"],
            "orden": fila["orden"],
            "espesor_nominal": (
                Decimal("19.00")
                if "EMPALME" in fila["seccion"]
                else Decimal("20.00")
            ),
            "condicion": Inspeccion.Condicion.NORMAL,
        }
        Medicion.objects.get_or_create(
            **clave,
            defaults=defaults,
        )


def crear_estructura_poleas(inspeccion):
    tag_faja = (
        inspeccion.faja.tag
        or ""
    ).upper().strip()

    if tag_faja in [
        "CVB0003",
        "0220-CVB-0003",
        "0220-CVB0003",
    ]:
        cantidad_poleas = 9
        prefijo_tag = "CVB0003-P"
        condicion_inicial = Inspeccion.Condicion.NORMAL
    elif tag_faja in [
        "CVB0006",
        "CVB006",
        "0240-CVB-006",
        "0240-CVB0006",
        "0310CVB0006",
        "CVB0007",
        "CVB007",
        "0240-CVB-007",
        "0240-CVB0007",
        "0310CVB0007",
        "CVB0010",
        "CVB010",
        "CVB0010-ENTRANTE",
        "CVB0010-SALIENTE",
        "0320-CVB-0010",
        "0320CVB0010",
    ]:
        cantidad_poleas = 8
        prefijo_tag = f"{tag_faja.replace('-', '')}-P"
        condicion_inicial = "No medido"
    elif tag_faja == "CVB0011":
        cantidad_poleas = 7
        prefijo_tag = "CVB0011-P"
        condicion_inicial = "No medido"
    elif tag_faja == "CVB0015":
        cantidad_poleas = 5
        prefijo_tag = "CVB0015-P"
        condicion_inicial = "No medido"
    elif tag_faja == "CVB0017":
        cantidad_poleas = 2
        prefijo_tag = "CVB0017-P"
        condicion_inicial = "No medido"
    elif tag_faja == "CVB0018":
        cantidad_poleas = 5
        prefijo_tag = "CVB0018-P"
        condicion_inicial = "No medido"
    else:
        cantidad_poleas = 5
        prefijo_tag = "CVB0001-P"
        condicion_inicial = "No medido"

    for numero in range(1, cantidad_poleas + 1):
        polea, creada = (
            PoleaInspeccion.objects.get_or_create(
                inspeccion=inspeccion,
                numero=numero,
                defaults={
                    "orden": numero,
                    "nombre": f"Polea #{numero:02d}",
                    "tag": f"{prefijo_tag}{numero:02d}",
                    "condicion": condicion_inicial,
                },
            )
        )

        cambios = []

        nombre_correcto = (
            f"Polea #{numero:02d}"
        )

        tag_correcto = (
            f"{prefijo_tag}{numero:02d}"
        )

        if polea.nombre != nombre_correcto:
            polea.nombre = nombre_correcto
            cambios.append("nombre")

        if polea.tag != tag_correcto:
            polea.tag = tag_correcto
            cambios.append("tag")

        if polea.orden != numero:
            polea.orden = numero
            cambios.append("orden")

        if cambios:
            polea.save(
                update_fields=cambios
            )

        for punto in range(1, 6):
            MedicionPolea.objects.get_or_create(
                polea=polea,
                punto=punto,
                defaults={
                    "orden": punto,
                },
            )

def crear_estructura_life_shaft(inspeccion):
    tag_faja = (
        inspeccion.faja.tag
        or ""
    ).upper().strip()

    if tag_faja in [
        "CVB0003",
        "0220-CVB-0003",
        "0220-CVB0003",
    ]:
        cantidad_shafts = 5
        prefijo_tag = "CVB0003-LS"
        condicion_inicial = Inspeccion.Condicion.NORMAL
    else:
        cantidad_shafts = 2
        prefijo_tag = "CVB0001-LS"
        condicion_inicial = "No medido"

    for numero in range(1, cantidad_shafts + 1):
        shaft, creado = (
            LifeShaftInspeccion.objects.get_or_create(
                inspeccion=inspeccion,
                numero=numero,
                defaults={
                    "orden": numero,
                    "nombre": f"Life Shaft #{numero:02d}",
                    "tag": f"{prefijo_tag}{numero:02d}",
                    "condicion": condicion_inicial,
                },
            )
        )

        campos_actualizados = []

        nombre_correcto = (
            f"Life Shaft #{numero:02d}"
        )

        tag_correcto = (
            f"{prefijo_tag}{numero:02d}"
        )

        if shaft.nombre != nombre_correcto:
            shaft.nombre = nombre_correcto
            campos_actualizados.append("nombre")

        if shaft.tag != tag_correcto:
            shaft.tag = tag_correcto
            campos_actualizados.append("tag")

        if shaft.orden != numero:
            shaft.orden = numero
            campos_actualizados.append("orden")

        if campos_actualizados:
            shaft.save(
                update_fields=campos_actualizados
            )

        # El Excel de CVB0003 muestra cuatro puntos.
        cantidad_puntos = 4

        for punto in range(
            1,
            cantidad_puntos + 1,
        ):
            MedicionLifeShaft.objects.get_or_create(
                life_shaft=shaft,
                punto=punto,
                defaults={
                    "orden": punto,
                },
            )

def _separar_formularios_medicion(formset):
    empalme_e01 = []
    empalme_e02 = []
    tramos = []

    for form in formset.forms:
        seccion = (form.instance.seccion or "").upper()
        if "E-01" in seccion:
            empalme_e01.append(form)
        elif "E-02" in seccion:
            empalme_e02.append(form)
        else:
            tramos.append(form)

    return empalme_e01, empalme_e02, tramos


def analizar_mediciones_empalme(mediciones, nombre_empalme):
    """Obtiene el punto mínimo y redacta el resumen técnico del empalme."""
    candidatos = []
    for medicion in mediciones:
        for letra in "abcdefg":
            valor = getattr(medicion, letra)
            if valor is not None:
                candidatos.append((valor, letra.upper(), medicion))

    if not candidatos:
        return {
            "disponible": False,
            "texto": (
                f"El empalme {nombre_empalme} aún no cuenta con "
                "mediciones suficientes para generar el resumen automático."
            ),
        }

    minimo, letra, medicion = min(candidatos, key=lambda item: item[0])
    posicion_texto = {
        "-1 m": "a un metro antes del empalme",
        "+1 m": "a un metro después del empalme",
    }.get(medicion.posicion, f"en la posición {medicion.posicion}")
    zona = "zona de carga" if nombre_empalme == "E-01" else ""
    ubicacion = f"bastidor {medicion.bastidor}"
    if zona:
        ubicacion += f", {zona}"

    desgaste = None
    residual = None
    if medicion.espesor_nominal and medicion.espesor_nominal > 0:
        desgaste = medicion.espesor_nominal - minimo
        residual = minimo / medicion.espesor_nominal * Decimal("100")

    texto = (
        f"El empalme {nombre_empalme} se encontró en el {ubicacion}. "
        f"El espesor mínimo en {(medicion.lado or 'lado no indicado').lower()} "
        f"es de {minimo:.2f} mm en el punto {letra}, {posicion_texto}."
    )
    if desgaste is not None and residual is not None:
        texto += (
            f" El desgaste calculado es de {desgaste:.2f} mm y el "
            f"porcentaje residual es {residual:.2f}%."
        )

    return {
        "disponible": True,
        "minimo": minimo,
        "letra": letra,
        "posicion": medicion.posicion,
        "bastidor": medicion.bastidor,
        "lado": medicion.lado,
        "desgaste": desgaste,
        "porcentaje_residual": residual,
        "texto": texto,
    }


def guardar_fotos_clasificadas(formset, seccion, usuario):
    fotografias = formset.save(commit=False)
    for fotografia in fotografias:
        if fotografia.pk is None:
            fotografia.subida_por = usuario
        fotografia.save()
        # La BD histórica de CVB001 conserva esta columna aunque el modelo
        # compartido no la declara; se mantiene la clasificación sin migrarla.
        with connection.cursor() as cursor:
            cursor.execute(
                "UPDATE inspecciones_fotoinspeccion SET seccion = %s WHERE id = %s",
                [seccion, fotografia.pk],
            )
    for fotografia in formset.deleted_objects:
        fotografia.delete()
    formset.save_m2m()

def obtener_tramo_carga_cvb0003(numero):
    if numero <= 16:
        return "TR E-09 @ E-10"

    if numero <= 32:
        return "TR E-10 @ E-11"

    if numero <= 48:
        return "TR E-11 @ E-12"

    if numero <= 64:
        return "TR E-12 @ E-13"

    if numero <= 80:
        return "TR E-13 @ E-14"

    if numero <= 96:
        return "TR E-14 @ E-15"

    return "TR E-15 @ E-16"


def obtener_tramo_retorno_cvb0003(numero):
    if numero <= 16:
        return "TR E-08 @ E-07"

    if numero <= 28:
        return "TR E-07 @ E-06"

    if numero <= 40:
        return "TR E-06 @ E-05"

    if numero <= 52:
        return "TR E-05 @ E-04"

    if numero <= 70:
        return "TR E-04 @ E-03"

    if numero <= 96:
        return "TR E-03 @ E-01"

    return "TR E-01 @ E-16"

def crear_estructura_faja_cvb0003(inspeccion):
    if inspeccion.tipo != Inspeccion.Tipo.FAJA:
        return

    tag = (inspeccion.faja.tag or "").upper().strip()

    if tag not in [
        "CVB0003",
        "0220-CVB-0003",
    ]:
        return

    # =====================================================
    # EMPALMES CVB0003
    # =====================================================

    if not inspeccion.empalmes_cvb0003.exists():
        empalmes = [
            ("ZONA DE CARGA", "E-15", "CARGA BC-244"),
            ("ZONA DE CARGA", "E-14", "CARGA BC-471"),
            ("ZONA DE CARGA", "E-13", "CARGA BC-702"),
            ("ZONA DE CARGA", "E-12", "CARGA BC-932"),
            ("ZONA DE CARGA", "E-11", "CARGA BC-1160"),
            ("ZONA DE CARGA", "E-10", "CARGA BC-1392"),
            ("ZONA DE CARGA", "E-09", "CARGA BC-1622"),

            ("ZONA MOTRIZ", "E-08", "CARGA BC-1622"),

            ("SIN ACCESO", "E-07", "RETORNO BC-1469"),

            ("ZONA RETORNO", "E-06", "RETORNO BC-1240"),
            ("ZONA RETORNO", "E-05", "RETORNO BC-1004"),
            ("ZONA RETORNO", "E-04", "RETORNO BC-777"),
            ("ZONA RETORNO", "E-03", "RETORNO BC-549"),
            ("ZONA RETORNO", "E-01", "RETORNO BC-350"),
            ("ZONA RETORNO", "E-16", "RETORNO BC-175"),
        ]

        posiciones = [
            "P-1",
            "P-2",
            "P-3",
            "P-4",
            "P-5",
            "B.COVER",
        ]

        filas_empalmes = []
        orden = 1

        for zona, empalme, bastidor_lado in empalmes:
            for posicion in posiciones:
                if posicion == "B.COVER":
                    espesor_nominal = Decimal("8.00")
                else:
                    espesor_nominal = Decimal("19.00")

                filas_empalmes.append(
                    MedicionEmpalmeCVB0003(
                        inspeccion=inspeccion,
                        zona=zona,
                        empalme=empalme,
                        bastidor_lado=bastidor_lado,
                        posicion=posicion,
                        espesor_nominal=espesor_nominal,
                        orden=orden,
                    )
                )

                orden += 1

        MedicionEmpalmeCVB0003.objects.bulk_create(
            filas_empalmes
        )

    # =====================================================
    # TRAMOS DE CARGA
    # =====================================================

    existen_carga = inspeccion.tramos_cvb0003.filter(
        tipo=MedicionTramoCVB0003.Tipo.CARGA,
    ).exists()

    if not existen_carga:
        filas_carga = []

        # El Excel muestra crecimiento aproximado de 14.
        bastidor_inicial = 45

        for numero in range(1, 121):
            bastidor = (
                bastidor_inicial
                + ((numero - 1) * 14)
            )

            filas_carga.append(
                MedicionTramoCVB0003(
                    inspeccion=inspeccion,
                    tipo=MedicionTramoCVB0003.Tipo.CARGA,
                    tramo=obtener_tramo_carga_cvb0003(
                        numero
                    ),
                    medicion=numero,
                    bastidor=str(bastidor),
                    espesor_nominal=Decimal("19.00"),
                    orden=numero,
                )
            )

        MedicionTramoCVB0003.objects.bulk_create(
            filas_carga
        )

    # =====================================================
    # TRAMOS DE RETORNO
    # =====================================================

    existen_retorno = inspeccion.tramos_cvb0003.filter(
        tipo=MedicionTramoCVB0003.Tipo.RETORNO,
    ).exists()

    if not existen_retorno:
        filas_retorno = []

        bastidor_inicial = 1600

        for numero in range(1, 103):
            bastidor = (
                bastidor_inicial
                - ((numero - 1) * 14)
            )

            filas_retorno.append(
                MedicionTramoCVB0003(
                    inspeccion=inspeccion,
                    tipo=MedicionTramoCVB0003.Tipo.RETORNO,
                    tramo=obtener_tramo_retorno_cvb0003(
                        numero
                    ),
                    medicion=numero,
                    bastidor=str(bastidor),
                    espesor_nominal=Decimal("19.00"),
                    orden=numero,
                )
            )

        MedicionTramoCVB0003.objects.bulk_create(
            filas_retorno
        )

def crear_estructura_faja_cvb0004(inspeccion):
    """
    Crea automáticamente las mediciones necesarias
    para el formulario de Faja / Top Cover CVB0004.

    Se reutilizan los modelos de CVB0003 para avanzar
    rápido, pero los registros pertenecen únicamente
    a esta inspección.
    """

    # =====================================================
    # 1. EMPALMES DE CVB0004
    # =====================================================

    empalmes = [
        {
            "empalme": "E-01",
            "zona": "CARGA",
            "bastidor": "CARGA",
        },
        {
            "empalme": "E-1",
            "zona": "RETORNO",
            "bastidor": "RETORNO",
        },
        {
            "empalme": "E-2",
            "zona": "RETORNO",
            "bastidor": "RETORNO",
        },
        {
            "empalme": "E-3",
            "zona": "RETORNO",
            "bastidor": "RETORNO",
        },
        {
            "empalme": "E-5",
            "zona": "CARGA",
            "bastidor": "CARGA",
        },
        {
            "empalme": "E-6",
            "zona": "CARGA",
            "bastidor": "CARGA",
        },
        {
            "empalme": "E-7",
            "zona": "CARGA",
            "bastidor": "CARGA",
        },
        {
            "empalme": "E-8",
            "zona": "CARGA",
            "bastidor": "CARGA",
        },
        {
            "empalme": "E-9",
            "zona": "CARGA",
            "bastidor": "CARGA",
        },
        {
            "empalme": "E-10",
            "zona": "CARGA",
            "bastidor": "CARGA",
        },
        {
            "empalme": "E-11",
            "zona": "CARGA",
            "bastidor": "CARGA",
        },
        {
            "empalme": "E-12",
            "zona": "CARGA",
            "bastidor": "CARGA",
        },
        {
            "empalme": "E-14",
            "zona": "RETORNO",
            "bastidor": "RETORNO",
        },
        {
            "empalme": "E-15",
            "zona": "RETORNO",
            "bastidor": "RETORNO",
        },
        {
            "empalme": "E-16",
            "zona": "RETORNO",
            "bastidor": "RETORNO",
        },
    ]

    posiciones = [
        ("P-1", Decimal("19.00")),
        ("P-2", Decimal("19.00")),
        ("P-3", Decimal("19.00")),
        ("P-4", Decimal("19.00")),
        ("P-5", Decimal("19.00")),
        ("B.COVER", Decimal("8.00")),
    ]

    orden = 1

    for datos_empalme in empalmes:
        for posicion, espesor_nominal in posiciones:
            MedicionEmpalmeCVB0003.objects.get_or_create(
                inspeccion=inspeccion,
                empalme=datos_empalme["empalme"],
                posicion=posicion,
                defaults={
                    "orden": orden,
                    "zona": datos_empalme["zona"],
                    "bastidor_lado": datos_empalme["bastidor"],
                    "espesor_nominal": espesor_nominal,
                },
            )

            orden += 1

    # =====================================================
    # 2. TRAMOS ENTRE EMPALMES
    # =====================================================

    grupos_tramos = [
        {
            "tramo": "Tramos entre E-6 al E-5",
            "bastidores": [
                1755,
                1726,
                1698,
                1668,
                1639,
                1610,
            ],
        },
        {
            "tramo": "Tramos entre E-7 al E-6",
            "bastidores": [
                1552,
                1494,
                1465,
                1436,
                1407,
                1378,
            ],
        },
        {
            "tramo": "Tramos entre E-8 al E-7",
            "bastidores": [
                1320,
                1291,
                1262,
                1233,
                1204,
                1175,
                1146,
            ],
        },
        {
            "tramo": "Tramos entre E-9 al E-8",
            "bastidores": [
                1088,
                1059,
                1030,
                1001,
                972,
                943,
                914,
            ],
        },
        {
            "tramo": "Tramos entre E-9 al E-10",
            "bastidores": [
                885,
                860,
                835,
                810,
                785,
                760,
                735,
                710,
                685,
            ],
        },
        {
            "tramo": "Tramos entre E-10 al E-11",
            "bastidores": [
                660,
                635,
                610,
                585,
                560,
                535,
                510,
                485,
            ],
        },
        {
            "tramo": "Tramos entre E-11 al E-12",
            "bastidores": [
                460,
                435,
                410,
                385,
                360,
                335,
                310,
                285,
                260,
                235,
            ],
        },
        {
            "tramo": "Tramos entre E-12 al E-13",
            "bastidores": [
                185,
                160,
                135,
                110,
                85,
                60,
            ],
        },
    ]

    orden_tramo = 1

    for grupo in grupos_tramos:
        for numero_medicion, bastidor in enumerate(
            grupo["bastidores"],
            start=1,
        ):
            MedicionTramoCVB0003.objects.get_or_create(
                inspeccion=inspeccion,
                tipo=MedicionTramoCVB0003.Tipo.CARGA,
                tramo=grupo["tramo"],
                medicion=numero_medicion,
                defaults={
                    "orden": orden_tramo,
                    "bastidor": bastidor,
                    "espesor_nominal": Decimal("19.00"),
                },
            )

            orden_tramo += 1        
def formulario_faja_cvb0003(request, inspeccion):
    # =====================================================
    # CREAR LAS FILAS AUTOMÁTICAS DE CVB0003
    # =====================================================

    crear_estructura_faja_cvb0003(inspeccion)
    CalibracionUTFajaCVB0003.crear_estructura(inspeccion)

    # =====================================================
    # PERMISOS DEL USUARIO
    # =====================================================

    permisos = obtener_permisos_flujo(
        request.user,
        inspeccion,
    )

    # =====================================================
    # FORMULARIO GENERAL DE LA INSPECCIÓN
    # =====================================================

    formulario = InspeccionForm(
        request.POST or None,
        instance=inspeccion,
    )

    calibraciones_ut_formset = CalibracionUTFajaCVB0003FormSet(
        request.POST or None,
        instance=inspeccion,
        prefix="calibraciones-ut",
        queryset=inspeccion.calibraciones_ut_faja_cvb0003.order_by("numero"),
    )
    if not permisos["puede_editar"]:
        for formulario_ut in calibraciones_ut_formset.forms:
            for campo in formulario_ut.fields.values():
                campo.disabled = True

    # =====================================================
    # FORMSET DE EMPALMES
    # =====================================================

    empalmes_formset = MedicionEmpalmeCVB0003FormSet(
        request.POST or None,
        instance=inspeccion,
        prefix="empalmes",
        queryset=(
            inspeccion.empalmes_cvb0003
            .order_by(
                "orden",
                "id",
            )
        ),
    )

    # =====================================================
    # QUERYSET DE TRAMOS DE CARGA
    # =====================================================

    carga_queryset = (
        inspeccion.tramos_cvb0003
        .filter(
            tipo=MedicionTramoCVB0003.Tipo.CARGA,
        )
        .order_by(
            "orden",
            "id",
        )
    )

    # =====================================================
    # QUERYSET DE TRAMOS DE RETORNO
    # =====================================================

    retorno_queryset = (
        inspeccion.tramos_cvb0003
        .filter(
            tipo=MedicionTramoCVB0003.Tipo.RETORNO,
        )
        .order_by(
            "orden",
            "id",
        )
    )

    # =====================================================
    # FORMSET DE TRAMOS DE CARGA
    # =====================================================

    carga_formset = MedicionTramoCVB0003FormSet(
        request.POST or None,
        instance=inspeccion,
        prefix="carga",
        queryset=carga_queryset,
    )

    # =====================================================
    # FORMSET DE TRAMOS DE RETORNO
    # =====================================================

    retorno_formset = MedicionTramoCVB0003FormSet(
        request.POST or None,
        instance=inspeccion,
        prefix="retorno",
        queryset=retorno_queryset,
    )

    # =====================================================
    # FORMSET DE FOTOGRAFÍAS DE EMPALMES
    # =====================================================

    fotos_empalmes_formset = FotoFajaCVB0003FormSet(
        request.POST or None,
        request.FILES or None,
        instance=inspeccion,
        prefix="fotos-empalmes",
        queryset=(
            inspeccion.fotografias_cvb0003
            .filter(
                seccion=FotoFajaCVB0003.Seccion.EMPALMES,
            )
            .order_by(
                "creada_en",
                "id",
            )
        ),
    )

    # =====================================================
    # FORMSET DE FOTOGRAFÍAS DE CARGA
    # =====================================================

    fotos_carga_formset = FotoFajaCVB0003FormSet(
        request.POST or None,
        request.FILES or None,
        instance=inspeccion,
        prefix="fotos-carga",
        queryset=(
            inspeccion.fotografias_cvb0003
            .filter(
                seccion=FotoFajaCVB0003.Seccion.CARGA,
            )
            .order_by(
                "creada_en",
                "id",
            )
        ),
    )

    # =====================================================
    # FORMSET DE FOTOGRAFÍAS DE RETORNO
    # =====================================================

    fotos_retorno_formset = FotoFajaCVB0003FormSet(
        request.POST or None,
        request.FILES or None,
        instance=inspeccion,
        prefix="fotos-retorno",
        queryset=(
            inspeccion.fotografias_cvb0003
            .filter(
                seccion=FotoFajaCVB0003.Seccion.RETORNO,
            )
            .order_by(
                "creada_en",
                "id",
            )
        ),
    )

    formulario_valido = formulario.is_valid() if request.method == "POST" else False
    fecha_tecnica = (
        formulario.cleaned_data.get("fecha_inspeccion")
        if formulario_valido
        else inspeccion.fecha_inspeccion
    )
    historial_empalmes = historial_faja(
        inspeccion, fecha_tecnica, list(empalmes_formset.queryset), "empalme"
    )
    historial_carga = historial_faja(
        inspeccion, fecha_tecnica, list(carga_formset.queryset), "tramo"
    )
    historial_retorno = historial_faja(
        inspeccion, fecha_tecnica, list(retorno_formset.queryset), "tramo"
    )

    for formset, historial in (
        (empalmes_formset, historial_empalmes),
        (carga_formset, historial_carga),
        (retorno_formset, historial_retorno),
    ):
        preparar_formset_historico(
            formset,
            {pk: dato["valores"] for pk, dato in historial.items()},
            clave_pk_form,
        )

    # =====================================================
    # PROCESAR GUARDADO
    # =====================================================

    if request.method == "POST":
        if not permisos["puede_editar"]:
            return HttpResponseForbidden(
                "La inspección no está habilitada para edición."
            )

        empalmes_validos = validar_formset_historico(
            empalmes_formset,
            {pk: dato["valores"] for pk, dato in historial_empalmes.items()},
            clave_pk_form,
        )

        carga_valida = validar_formset_historico(
            carga_formset,
            {pk: dato["valores"] for pk, dato in historial_carga.items()},
            clave_pk_form,
        )

        retorno_valido = validar_formset_historico(
            retorno_formset,
            {pk: dato["valores"] for pk, dato in historial_retorno.items()},
            clave_pk_form,
        )

        fotos_empalmes_validas = (
            fotos_empalmes_formset.is_valid()
        )

        fotos_carga_validas = (
            fotos_carga_formset.is_valid()
        )

        fotos_retorno_validas = (
            fotos_retorno_formset.is_valid()
        )

        calibraciones_ut_validas = calibraciones_ut_formset.is_valid()

        todo_valido = (
            formulario_valido
            and empalmes_validos
            and carga_valida
            and retorno_valido
            and fotos_empalmes_validas
            and fotos_carga_validas
            and fotos_retorno_validas
            and calibraciones_ut_validas
        )

        if todo_valido:
            # ---------------------------------------------
            # GUARDAR DATOS GENERALES
            # ---------------------------------------------

            inspeccion_guardada = formulario.save()

            # ---------------------------------------------
            # GUARDAR TABLAS
            # ---------------------------------------------

            empalmes_formset.save()
            carga_formset.save()
            retorno_formset.save()
            calibraciones_ut_formset.save()

            # ---------------------------------------------
            # GUARDAR FOTOGRAFÍAS POR SECCIÓN
            # ---------------------------------------------

            formsets_fotograficos = [
                (
                    fotos_empalmes_formset,
                    FotoFajaCVB0003.Seccion.EMPALMES,
                ),
                (
                    fotos_carga_formset,
                    FotoFajaCVB0003.Seccion.CARGA,
                ),
                (
                    fotos_retorno_formset,
                    FotoFajaCVB0003.Seccion.RETORNO,
                ),
            ]

            for fotos_formset, seccion in formsets_fotograficos:
                fotografias = fotos_formset.save(
                    commit=False
                )

                for fotografia in fotografias:
                    fotografia.inspeccion = (
                        inspeccion_guardada
                    )

                    fotografia.seccion = seccion

                    if not fotografia.subida_por_id:
                        fotografia.subida_por = (
                            request.user
                        )

                    fotografia.save()

                # Eliminar las fotografías marcadas
                for fotografia_eliminada in (
                    fotos_formset.deleted_objects
                ):
                    fotografia_eliminada.delete()

            # ---------------------------------------------
            # PROCESAR ACCIÓN DEL FLUJO
            # ---------------------------------------------

            correcto, mensaje = _procesar_accion_flujo(
                request,
                inspeccion_guardada,
            )

            if correcto:
                messages.success(
                    request,
                    mensaje,
                )
            else:
                messages.error(
                    request,
                    mensaje,
                )

            return redirect(
                "formulario_faja",
                inspeccion_id=inspeccion_guardada.id,
            )

        # =================================================
        # MOSTRAR ERRORES EN LA TERMINAL
        # =================================================

        print(
            "=========================================="
        )
        print(
            "ERRORES DEL FORMULARIO CVB0003"
        )
        print(
            "=========================================="
        )

        print(
            "FORMULARIO GENERAL:",
            formulario.errors,
        )

        print(
            "EMPALMES:",
            empalmes_formset.errors,
        )

        print(
            "EMPALMES GENERALES:",
            empalmes_formset.non_form_errors(),
        )

        print(
            "TRAMOS DE CARGA:",
            carga_formset.errors,
        )

        print(
            "CARGA GENERALES:",
            carga_formset.non_form_errors(),
        )

        print(
            "TRAMOS DE RETORNO:",
            retorno_formset.errors,
        )

        print(
            "RETORNO GENERALES:",
            retorno_formset.non_form_errors(),
        )

        print(
            "FOTOS DE EMPALMES:",
            fotos_empalmes_formset.errors,
        )

        print(
            "FOTOS DE CARGA:",
            fotos_carga_formset.errors,
        )

        print(
            "FOTOS DE RETORNO:",
            fotos_retorno_formset.errors,
        )

        print(
            "CALIBRACIONES UT:",
            calibraciones_ut_formset.errors,
        )

        messages.error(
            request,
            (
                "No se pudo guardar CVB0003. "
                "Revisa los campos marcados y la terminal."
            ),
        )

    # =====================================================
    # MOSTRAR EL FORMULARIO
    # =====================================================

    return render(
        request,
        "inspecciones/formulario_faja_cvb0003.html",
        {
            "inspeccion": inspeccion,
            "formulario": formulario,

            # Tablas
            "empalmes_formset": empalmes_formset,
            "carga_formset": carga_formset,
            "retorno_formset": retorno_formset,
            "calibraciones_ut_formset": calibraciones_ut_formset,

            "historial_empalmes": list(historial_empalmes.values()),
            "historial_carga": list(historial_carga.values()),
            "historial_retorno": list(historial_retorno.values()),

            # Fotografías
            "fotos_empalmes_formset": (
                fotos_empalmes_formset
            ),
            "fotos_carga_formset": (
                fotos_carga_formset
            ),
            "fotos_retorno_formset": (
                fotos_retorno_formset
            ),

            # Permisos y botones
            **permisos,
            **_contexto_workflow_ui_cvb0003(inspeccion),
        },
    )
def formulario_faja_cvb0004(request, inspeccion):
    """
    Formulario especial de Faja / Top Cover CVB0004.

    Reutiliza:
    - MedicionEmpalmeCVB0003
    - MedicionTramoCVB0003
    - FotoFajaCVB0003

    CVB0004 utiliza:
    1. Tabla de empalmes.
    2. Tabla de tramos de carga.
    """

    # =====================================================
    # CREAR ESTRUCTURA AUTOMÁTICA
    # =====================================================

    crear_estructura_faja_cvb0004(inspeccion)

    # =====================================================
    # PERMISOS Y FLUJO
    # =====================================================

    permisos = obtener_permisos_flujo(
        request.user,
        inspeccion,
    )

    puede_editar = permisos["puede_editar"]

    # =====================================================
    # FORMULARIO GENERAL
    # =====================================================

    formulario = InspeccionForm(
        request.POST or None,
        instance=inspeccion,
    )

    # =====================================================
    # TABLA 1: EMPALMES
    # =====================================================

    empalmes_queryset = (
        inspeccion.empalmes_cvb0003
        .order_by(
            "orden",
            "id",
        )
    )

    empalmes_formset = (
        MedicionEmpalmeCVB0003FormSet(
            request.POST or None,
            instance=inspeccion,
            prefix="empalmes",
            queryset=empalmes_queryset,
        )
    )

    # =====================================================
    # TABLA 2: TRAMOS DE CARGA
    # =====================================================

    carga_queryset = (
        inspeccion.tramos_cvb0003
        .filter(
            tipo=MedicionTramoCVB0003.Tipo.CARGA,
        )
        .order_by(
            "orden",
            "id",
        )
    )

    carga_formset = (
        MedicionTramoCVB0003FormSet(
            request.POST or None,
            instance=inspeccion,
            prefix="carga",
            queryset=carga_queryset,
        )
    )

    # =====================================================
    # FOTOGRAFÍAS DE EMPALMES
    # =====================================================

    fotos_empalmes_formset = (
        FotoFajaCVB0003FormSet(
            request.POST or None,
            request.FILES or None,
            instance=inspeccion,
            prefix="fotos-empalmes",
            queryset=(
                inspeccion.fotografias_cvb0003
                .filter(
                    seccion=(
                        FotoFajaCVB0003
                        .Seccion
                        .EMPALMES
                    ),
                )
                .order_by(
                    "creada_en",
                    "id",
                )
            ),
        )
    )

    # =====================================================
    # FOTOGRAFÍAS DE TRAMOS DE CARGA
    # =====================================================

    fotos_carga_formset = (
        FotoFajaCVB0003FormSet(
            request.POST or None,
            request.FILES or None,
            instance=inspeccion,
            prefix="fotos-carga",
            queryset=(
                inspeccion.fotografias_cvb0003
                .filter(
                    seccion=(
                        FotoFajaCVB0003
                        .Seccion
                        .CARGA
                    ),
                )
                .order_by(
                    "creada_en",
                    "id",
                )
            ),
        )
    )

    historial_empalmes = historial_faja(
        inspeccion, inspeccion.fecha_inspeccion,
        list(empalmes_formset.queryset), "empalme"
    )
    historial_carga = historial_faja(
        inspeccion, inspeccion.fecha_inspeccion,
        list(carga_formset.queryset), "tramo"
    )
    preparar_formset_historico(
        empalmes_formset,
        {pk: dato["valores"] for pk, dato in historial_empalmes.items()},
        clave_pk_form,
    )
    preparar_formset_historico(
        carga_formset,
        {pk: dato["valores"] for pk, dato in historial_carga.items()},
        clave_pk_form,
    )

    # =====================================================
    # PROCESAR GUARDADO
    # =====================================================

    if request.method == "POST":
        if not puede_editar:
            return HttpResponseForbidden(
                "La inspección no está habilitada para edición."
            )

        formulario_valido = formulario.is_valid()

        empalmes_validos = validar_formset_historico(
            empalmes_formset,
            {pk: dato["valores"] for pk, dato in historial_empalmes.items()},
            clave_pk_form,
        )

        carga_valida = validar_formset_historico(
            carga_formset,
            {pk: dato["valores"] for pk, dato in historial_carga.items()},
            clave_pk_form,
        )

        fotos_empalmes_validas = (
            fotos_empalmes_formset.is_valid()
        )

        fotos_carga_validas = (
            fotos_carga_formset.is_valid()
        )

        todo_valido = (
            formulario_valido
            and empalmes_validos
            and carga_valida
            and fotos_empalmes_validas
            and fotos_carga_validas
        )

        if todo_valido:
            # ---------------------------------------------
            # GUARDAR DATOS GENERALES
            # ---------------------------------------------

            inspeccion_guardada = (
                formulario.save()
            )

            # ---------------------------------------------
            # GUARDAR LAS DOS TABLAS
            # ---------------------------------------------

            empalmes_formset.save()
            carga_formset.save()

            # ---------------------------------------------
            # GUARDAR FOTOS DE EMPALMES
            # ---------------------------------------------

            fotografias_empalmes = (
                fotos_empalmes_formset.save(
                    commit=False
                )
            )

            for fotografia in fotografias_empalmes:
                fotografia.inspeccion = (
                    inspeccion_guardada
                )

                fotografia.seccion = (
                    FotoFajaCVB0003
                    .Seccion
                    .EMPALMES
                )

                if not fotografia.subida_por_id:
                    fotografia.subida_por = (
                        request.user
                    )

                fotografia.save()

            for fotografia_eliminada in (
                fotos_empalmes_formset
                .deleted_objects
            ):
                fotografia_eliminada.delete()

            # ---------------------------------------------
            # GUARDAR FOTOS DE CARGA
            # ---------------------------------------------

            fotografias_carga = (
                fotos_carga_formset.save(
                    commit=False
                )
            )

            for fotografia in fotografias_carga:
                fotografia.inspeccion = (
                    inspeccion_guardada
                )

                fotografia.seccion = (
                    FotoFajaCVB0003
                    .Seccion
                    .CARGA
                )

                if not fotografia.subida_por_id:
                    fotografia.subida_por = (
                        request.user
                    )

                fotografia.save()

            for fotografia_eliminada in (
                fotos_carga_formset
                .deleted_objects
            ):
                fotografia_eliminada.delete()

            # ---------------------------------------------
            # GUARDAR / ENVIAR / APROBAR / PUBLICAR
            # ---------------------------------------------

            correcto, mensaje = (
                _procesar_accion_flujo(
                    request,
                    inspeccion_guardada,
                )
            )

            if correcto:
                messages.success(
                    request,
                    mensaje,
                )
            else:
                messages.error(
                    request,
                    mensaje,
                )

            return redirect(
                "formulario_faja",
                inspeccion_id=(
                    inspeccion_guardada.id
                ),
            )

        # =================================================
        # ERRORES PARA REVISAR EN LA TERMINAL
        # =================================================

        print(
            "======================================"
        )
        print(
            "ERRORES DEL FORMULARIO CVB0004"
        )
        print(
            "======================================"
        )

        print(
            "FORMULARIO GENERAL:",
            formulario.errors,
        )

        print(
            "EMPALMES:",
            empalmes_formset.errors,
        )

        print(
            "EMPALMES GENERALES:",
            empalmes_formset.non_form_errors(),
        )

        print(
            "TRAMOS DE CARGA:",
            carga_formset.errors,
        )

        print(
            "CARGA GENERALES:",
            carga_formset.non_form_errors(),
        )

        print(
            "FOTOS DE EMPALMES:",
            fotos_empalmes_formset.errors,
        )

        print(
            "FOTOS DE CARGA:",
            fotos_carga_formset.errors,
        )

        messages.error(
            request,
            (
                "No se pudo guardar CVB0004. "
                "Revisa los campos marcados "
                "y los errores de la terminal."
            ),
        )

    # =====================================================
    # RENDERIZAR FORMULARIO CVB0004
    # =====================================================

    return render(
        request,
        "inspecciones/formulario_faja_cvb0004.html",
        {
            "inspeccion": inspeccion,
            "formulario": formulario,

            # Tablas
            "empalmes_formset": (
                empalmes_formset
            ),
            "carga_formset": (
                carga_formset
            ),
            "historial_empalmes": list(historial_empalmes.values()),
            "historial_carga": list(historial_carga.values()),

            # Fotografías
            "fotos_empalmes_formset": (
                fotos_empalmes_formset
            ),
            "fotos_carga_formset": (
                fotos_carga_formset
            ),

            # Permisos y botones del flujo
            **permisos,
        },
    )
def crear_estructura_faja_cvb0006_molienda(inspeccion):
    if inspeccion.tipo != Inspeccion.Tipo.FAJA:
        return

    tag = (inspeccion.faja.tag or "").upper().strip()
    if tag not in [
        "CVB0006",
        "CVB006",
        "0240-CVB-006",
        "0240-CVB0006",
        "0310CVB0006",
        "CVB0007",
        "CVB007",
        "0240-CVB-007",
        "0240-CVB0007",
        "0310CVB0007",
        "CVB0010",
        "CVB010",
        "CVB0010-ENTRANTE",
        "CVB0010-SALIENTE",
        "0320-CVB-0010",
        "0320CVB0010",
        "CVB0011",
        "CVB0015",
        "CVB0017",
        "CVB0018",
    ]:
        return

    CalibracionUTFajaCVB0003.objects.get_or_create(
        inspeccion=inspeccion,
        numero=1,
        defaults={
            "marca_equipo": "OLYMPUS",
            "modelo_equipo": "6LT PLUS",
            "frecuencia_mhz": "1",
            "rango_mm": "0.2 MHZ - 1.2 MHZ",
            "metodo_empleado": "OS - UT - 0014",
            "acoplante": "Echo gel",
            "rectificacion": "Full",
            "velocidad_ms": "19.37",
            "retardo_us": "2.87",
            "tipo_scan": "A Scan",
        },
    )

    empalmes_config = [
        ("EMPALME E-01", "DESPUES", 1),
        ("EMPALME E-01", "EMPALME", 2),
        ("EMPALME E-01", "ANTES", 3),
        ("EMPALME E-02", "DESPUES", 101),
        ("EMPALME E-02", "EMPALME", 102),
        ("EMPALME E-02", "ANTES", 103),
    ]

    for zona, posicion, orden in empalmes_config:
        MedicionEmpalmeCVB0003.objects.get_or_create(
            inspeccion=inspeccion,
            zona=zona,
            posicion=posicion,
            defaults={
                "empalme": zona.replace("EMPALME ", ""),
                "bastidor_lado": "0240-CVB-0006",
                "espesor_nominal": Decimal("10.00"),
                "orden": orden,
            },
        )

    for numero in range(1, 8):
        MedicionTramoCVB0003.objects.get_or_create(
            inspeccion=inspeccion,
            tipo=MedicionTramoCVB0003.Tipo.CARGA,
            medicion=numero,
            defaults={
                "tramo": "TOP COVER",
                "bastidor": str(numero),
                "espesor_nominal": Decimal("10.00"),
                "orden": numero,
            },
        )


@transaction.atomic
def formulario_faja_cvb0006_molienda(request, inspeccion):
    crear_estructura_faja_cvb0006_molienda(inspeccion)

    permisos = obtener_permisos_flujo(request.user, inspeccion)
    puede_editar = permisos["puede_editar"]

    empalmes_qs = inspeccion.empalmes_cvb0003.order_by("orden", "id")
    top_cover_qs = inspeccion.tramos_cvb0003.filter(
        tipo=MedicionTramoCVB0003.Tipo.CARGA,
    ).order_by("orden", "id")

    fotos_config = {
        "fotos_e01_formset": ("EMPALMES", "fotos-e01"),
        "fotos_top_cover_formset": ("CARGA", "fotos-top-cover"),
        "fotos_e02_formset": ("RETORNO", "fotos-e02"),
    }

    formsets_fotos = {}
    for nombre, (seccion, prefijo) in fotos_config.items():
        kwargs = {
            "instance": inspeccion,
            "prefix": prefijo,
            "queryset": inspeccion.fotografias_cvb0003.filter(
                seccion=seccion
            ),
        }
        if request.method == "POST":
            kwargs.update({"data": request.POST, "files": request.FILES})
        formsets_fotos[nombre] = FotoFajaCVB0003FormSet(**kwargs)

    calibracion_formset = CalibracionUTFajaCVB0003FormSet(
        request.POST or None,
        instance=inspeccion,
        prefix="calibracion",
    )
    empalmes_formset = MedicionEmpalmeCVB0003FormSet(
        request.POST or None,
        instance=inspeccion,
        prefix="empalmes",
        queryset=empalmes_qs,
    )
    top_cover_formset = MedicionTramoCVB0003FormSet(
        request.POST or None,
        instance=inspeccion,
        prefix="top-cover",
        queryset=top_cover_qs,
    )
    fecha_tecnica = inspeccion.fecha_inspeccion
    historial_empalmes = historial_faja(
        inspeccion, fecha_tecnica, list(empalmes_formset.queryset), "empalme"
    )
    historial_top_cover = historial_faja(
        inspeccion, fecha_tecnica, list(top_cover_formset.queryset), "tramo"
    )
    preparar_formset_historico(
        empalmes_formset,
        {pk: dato["valores"] for pk, dato in historial_empalmes.items()},
        clave_pk_form,
    )
    preparar_formset_historico(
        top_cover_formset,
        {pk: dato["valores"] for pk, dato in historial_top_cover.items()},
        clave_pk_form,
    )

    if request.method == "POST":
        if not puede_editar:
            return HttpResponseForbidden(
                "La inspeccion no esta habilitada para edicion."
            )

        formulario = InspeccionForm(request.POST, instance=inspeccion)
        formularios_validos = [
            formulario.is_valid(),
            calibracion_formset.is_valid(),
            validar_formset_historico(
                empalmes_formset,
                {pk: dato["valores"] for pk, dato in historial_empalmes.items()},
                clave_pk_form,
            ),
            validar_formset_historico(
                top_cover_formset,
                {pk: dato["valores"] for pk, dato in historial_top_cover.items()},
                clave_pk_form,
            ),
            *[formset.is_valid() for formset in formsets_fotos.values()],
        ]

        if all(formularios_validos):
            inspeccion_guardada = formulario.save()
            calibracion_formset.save()
            empalmes_formset.save()
            top_cover_formset.save()

            for nombre, (seccion, _prefijo) in fotos_config.items():
                fotografias = formsets_fotos[nombre].save(commit=False)
                for fotografia in fotografias:
                    fotografia.inspeccion = inspeccion_guardada
                    fotografia.seccion = seccion
                    if not fotografia.subida_por_id:
                        fotografia.subida_por = request.user
                    fotografia.save()

                for fotografia_eliminada in (
                    formsets_fotos[nombre].deleted_objects
                ):
                    fotografia_eliminada.delete()

            accion = request.POST.get("workflow_action", "guardar")
            comentario_revision = request.POST.get(
                "comentario_revision",
                "",
            ).strip()

            if accion == "guardar":
                messages.success(
                    request,
                    "Los cambios se guardaron correctamente.",
                )
            else:
                correcto, mensaje = _aplicar_accion_flujo(
                    request,
                    inspeccion_guardada,
                    accion,
                    comentario_revision,
                )
                if correcto:
                    messages.success(request, mensaje)
                else:
                    messages.error(request, mensaje)

            return redirect("formulario_faja", inspeccion_id=inspeccion.id)

        messages.error(
            request,
            "No se pudo guardar. Revisa los campos marcados.",
        )
    else:
        formulario = InspeccionForm(instance=inspeccion)
    return render(
        request,
        "inspecciones/formulario_faja_cvb0006_molienda.html",
        {
            "inspeccion": inspeccion,
            "formulario": formulario,
            "calibracion_formset": calibracion_formset,
            "empalmes_formset": empalmes_formset,
            "empalme_e01_forms": [
                form for form in empalmes_formset
                if form.instance.zona == "EMPALME E-01"
            ],
            "empalme_e02_forms": [
                form for form in empalmes_formset
                if form.instance.zona == "EMPALME E-02"
            ],
            "top_cover_formset": top_cover_formset,
            "molienda_config": _molienda_config(inspeccion),
            **formsets_fotos,
            **permisos,
        },
    )


@login_required
@transaction.atomic
def formulario_faja(request, inspeccion_id):
    inspeccion = get_object_or_404(
        Inspeccion.objects.select_related(
            "faja",
            "inspector",
            "supervisor",
            "analista",
            "cliente",
        ),
        id=inspeccion_id,
    )

    # =====================================================
    # VALIDACIÓN DE ACCESO
    # =====================================================

    if not usuario_puede_abrir_inspeccion(
        request.user,
        inspeccion,
    ):
        return HttpResponseForbidden(
            "No tienes permiso para abrir esta inspección."
        )

    # =====================================================
    # VALIDACIÓN DEL TIPO DE REPORTE
    # =====================================================

    if inspeccion.tipo != Inspeccion.Tipo.FAJA:
        messages.error(
            request,
            "Esta inspección no corresponde al formulario de Faja.",
        )
        return redirect("dashboard")

    # =====================================================
    # CLIENTE: SOLO REPORTE PUBLICADO
    # =====================================================

    rol = obtener_rol(request.user)

    if rol == "Cliente":
        if inspeccion.estado == Inspeccion.Estado.PUBLICADO:
            return redirect(
                "reporte_faja",
                inspeccion_id=inspeccion.id,
            )

        return HttpResponseForbidden(
            "Este reporte todavía no ha sido publicado."
        )

    # =====================================================
    # DETECTAR TAG
    # =====================================================

    tag_faja = (
        inspeccion.faja.tag
        or ""
    ).upper().strip()

    # =====================================================
    # FORMULARIO ESPECIAL CVB0003
    # =====================================================

    if tag_faja in [
        "CVB0003",
        "0220-CVB-0003",
        "0220-CVB0003",
    ]:
        return formulario_faja_cvb0003(
            request,
            inspeccion,
        )

    # =====================================================
    # FORMULARIO ESPECIAL CVB0004
    # =====================================================

    if tag_faja in [
        "CVB0004",
        "0220-CVB-0004",
        "0220-CVB0004",
    ]:
        return formulario_faja_cvb0004(
            request,
            inspeccion,
        )

    # =====================================================
    # FORMULARIO ESPECIAL MOLIENDA CVB0006
    # =====================================================

    if tag_faja in [
        "CVB0006",
        "CVB006",
        "0240-CVB-006",
        "0240-CVB0006",
        "0310CVB0006",
        "CVB0007",
        "CVB007",
        "0240-CVB-007",
        "0240-CVB0007",
        "0310CVB0007",
        "CVB0010",
        "CVB010",
        "CVB0010-ENTRANTE",
        "CVB0010-SALIENTE",
        "0320-CVB-0010",
        "0320CVB0010",
        "CVB0011",
        "CVB0015",
        "CVB0017",
        "CVB0018",
    ]:
        return formulario_faja_cvb0006_molienda(
            request,
            inspeccion,
        )

    # =====================================================
    # FORMULARIO ACTUAL CVB0001 Y OTRAS FAJAS
    # =====================================================

    crear_estructura_mediciones_faja(
        inspeccion
    )

    permisos = obtener_permisos_flujo(
        request.user,
        inspeccion,
    )

    puede_editar = permisos[
        "puede_editar"
    ]

    configuracion_fotos = {
        "fotos_e01_formset": (
            "EMPALME_E01",
            "fotos-e01",
        ),
        "fotos_e02_formset": (
            "EMPALME_E02",
            "fotos-e02",
        ),
        "fotos_puntos_formset": (
            "PUNTOS_MEDICION",
            "fotos-puntos",
        ),
    }
    formsets_fotos = {}
    for nombre, (seccion, prefijo) in configuracion_fotos.items():
        argumentos = {
            "instance": inspeccion,
            "prefix": prefijo,
            "queryset": inspeccion.fotografias.extra(
                where=["seccion = %s"], params=[seccion]
            ),
        }
        if request.method == "POST":
            argumentos.update({"data": request.POST, "files": request.FILES})
        formsets_fotos[nombre] = FotoInspeccionFormSet(**argumentos)

    mediciones_formset = MedicionFormSet(
        request.POST or None,
        instance=inspeccion,
        prefix="mediciones",
    )
    historial_mediciones = historial_mediciones_genericas(
        inspeccion,
        list(mediciones_formset.queryset),
        inspeccion.fecha_inspeccion,
    )
    preparar_formset_historico(
        mediciones_formset,
        {
            pk: dato["valores"]
            for pk, dato in historial_mediciones.items()
        },
    )

    # =====================================================
    # PETICIÓN POST
    # =====================================================

    if request.method == "POST":
        if not puede_editar:
            return HttpResponseForbidden(
                "La inspección no está habilitada para edición."
            )

        formulario = InspeccionForm(
            request.POST,
            instance=inspeccion,
        )

        formulario_valido = (
            formulario.is_valid()
        )

        mediciones_validas = validar_formset_historico(
            mediciones_formset,
            {
                pk: dato["valores"]
                for pk, dato in historial_mediciones.items()
            },
        )

        fotos_validas = all(
            formset.is_valid() for formset in formsets_fotos.values()
        )

        if (
            formulario_valido
            and mediciones_validas
            and fotos_validas
        ):
            inspeccion_guardada = (
                formulario.save()
            )

            mediciones_formset.save()

            for nombre, (seccion, _prefijo) in configuracion_fotos.items():
                guardar_fotos_clasificadas(
                    formsets_fotos[nombre],
                    seccion,
                    request.user,
                )

            accion = request.POST.get(
                "workflow_action",
                "guardar",
            )

            comentario_revision = (
                request.POST.get(
                    "comentario_revision",
                    "",
                ).strip()
            )

            if accion == "guardar":
                messages.success(
                    request,
                    "Los cambios se guardaron correctamente.",
                )
            else:
                correcto, mensaje = (
                    _aplicar_accion_flujo(
                        request,
                        inspeccion_guardada,
                        accion,
                        comentario_revision,
                    )
                )

                if correcto:
                    messages.success(
                        request,
                        mensaje,
                    )
                else:
                    messages.error(
                        request,
                        mensaje,
                    )

            return redirect(
                "formulario_faja",
                inspeccion_id=(
                    inspeccion_guardada.id
                ),
            )

        print(
            "ERRORES DEL FORMULARIO DE FAJA:",
            formulario.errors,
        )

        print(
            "ERRORES DE MEDICIONES:",
            mediciones_formset.errors,
        )

        print(
            "ERRORES GENERALES DEL FORMSET:",
            mediciones_formset.non_form_errors(),
        )

        messages.error(
            request,
            (
                "No se pudo guardar. "
                "Revisa los campos marcados "
                "y la terminal."
            ),
        )

    # =====================================================
    # PETICIÓN GET
    # =====================================================

    else:
        formulario = InspeccionForm(
            instance=inspeccion,
        )

    # =====================================================
    # SEPARAR EMPALMES Y TRAMOS
    # =====================================================

    empalme_e01, empalme_e02, tramos = (
        _separar_formularios_medicion(
            mediciones_formset
        )
    )

    resumen_e01 = analizar_mediciones_empalme(
        inspeccion.mediciones.filter(seccion__iexact="EMPALME E-01"),
        "E-01",
    )
    resumen_e02 = analizar_mediciones_empalme(
        inspeccion.mediciones.filter(seccion__iexact="EMPALME E-02"),
        "E-02",
    )

    # =====================================================
    # SELECCIONAR TEMPLATE
    # =====================================================

    if tag_faja in [
        "CVB0001",
        "0220-CVB-0001",
        "0220-CVB0001",
    ]:
        template_name = (
            "inspecciones/"
            "formulario_faja_cvb0001.html"
        )
    else:
        template_name = (
            "inspecciones/"
            "formulario_faja.html"
        )

    # =====================================================
    # RENDERIZAR
    # =====================================================

    return render(
    request,
    template_name,
    {
        "inspeccion": inspeccion,
        "formulario": formulario,
        "mediciones_formset": mediciones_formset,
        "empalme_e01": empalme_e01,
        "empalme_e02": empalme_e02,
        "tramos": tramos,
        "historial_mediciones": list(historial_mediciones.values()),
        "resumen_e01": resumen_e01,
        "resumen_e02": resumen_e02,
        **formsets_fotos,
        "rangos_fotos": range(1, 6),
        **permisos,
    },
)
def calcular_resumen_tramos(tramos):
    columnas = [
        "a",
        "b",
        "c",
        "d",
        "e",
        "f",
        "g",
    ]

    resumen = {
        "minimos": {},
        "promedios": {},
        "nominal_minimo": None,
        "nominal_promedio": None,
    }

    nominales = [
        float(medicion.espesor_nominal)
        for medicion in tramos
        if medicion.espesor_nominal is not None
    ]

    if nominales:
        resumen["nominal_minimo"] = min(nominales)
        resumen["nominal_promedio"] = (
            sum(nominales) / len(nominales)
        )

    for columna in columnas:
        valores = []

        for medicion in tramos:
            valor = getattr(medicion, columna, None)

            if valor is not None:
                valores.append(float(valor))

        if valores:
            resumen["minimos"][columna] = min(valores)
            resumen["promedios"][columna] = (
                sum(valores) / len(valores)
            )
        else:
            resumen["minimos"][columna] = None
            resumen["promedios"][columna] = None

    return resumen


@login_required
def reporte_faja(request, inspeccion_id):
    inspeccion = get_object_or_404(
        Inspeccion.objects.select_related(
            "faja",
            "inspector",
            "supervisor",
            "analista",
            "cliente",
        ),
        id=inspeccion_id,
        tipo=Inspeccion.Tipo.FAJA,
    )

    if not usuario_puede_abrir_inspeccion(
        request.user,
        inspeccion,
    ):
        return HttpResponseForbidden(
            "No tienes permiso para ver este reporte."
        )

    tag_faja = (
        inspeccion.faja.tag
        or ""
    ).upper().strip()

    if tag_faja in [
        "CVB0006",
        "CVB006",
        "0240-CVB-006",
        "0240-CVB0006",
        "0310CVB0006",
        "CVB0007",
        "CVB007",
        "0240-CVB-007",
        "0240-CVB0007",
        "0310CVB0007",
        "CVB0010",
        "CVB010",
        "CVB0010-ENTRANTE",
        "CVB0010-SALIENTE",
        "0320-CVB-0010",
        "0320CVB0010",
        "CVB0011",
        "CVB0015",
        "CVB0017",
        "CVB0018",
    ]:
        return redirect(
            "formulario_faja",
            inspeccion_id=inspeccion.id,
        )

    if tag_faja in [
        "CVB0001",
        "0220-CVB-0001",
        "0220-CVB0001",
    ]:
        from .reportes.cvb0001.views import reporte_faja_cvb0001

        return reporte_faja_cvb0001(request, inspeccion)

    # =====================================================
    # REPORTE ESPECIAL CVB0003
    # =====================================================

    if tag_faja in [
        "CVB0003",
        "0220-CVB-0003",
        "0220-CVB0003",
    ]:
        crear_estructura_faja_cvb0003(
            inspeccion
        )
        CalibracionUTFajaCVB0003.crear_estructura(inspeccion)

        empalmes = list(
            inspeccion.empalmes_cvb0003
            .order_by(
                "orden",
                "id",
            )
        )

        tramos_carga = list(
            inspeccion.tramos_cvb0003
            .filter(
                tipo=MedicionTramoCVB0003.Tipo.CARGA,
            )
            .order_by(
                "orden",
                "id",
            )
        )

        tramos_retorno = list(
            inspeccion.tramos_cvb0003
            .filter(
                tipo=MedicionTramoCVB0003.Tipo.RETORNO,
            )
            .order_by(
                "orden",
                "id",
            )
        )

        fotos_empalmes = list(
            inspeccion.fotografias_cvb0003
            .filter(
                seccion=FotoFajaCVB0003.Seccion.EMPALMES,
            )
            .order_by(
                "creada_en",
                "id",
            )
        )

        fotos_carga = list(
            inspeccion.fotografias_cvb0003
            .filter(
                seccion=FotoFajaCVB0003.Seccion.CARGA,
            )
            .order_by(
                "creada_en",
                "id",
            )
        )

        fotos_retorno = list(
            inspeccion.fotografias_cvb0003
            .filter(
                seccion=FotoFajaCVB0003.Seccion.RETORNO,
            )
            .order_by(
                "creada_en",
                "id",
            )
        )

        from .reportes.cvb0003.photo_summary import (
            paginas_fotograficas_faja,
            resumen_medicion_faja,
        )

        observacion_medicion_faja = resumen_medicion_faja(
            empalmes,
            tramos_carga,
            tramos_retorno,
        )

        return render(
            request,
            "inspecciones/reporte_faja_cvb0003.html",
            {
                "inspeccion": inspeccion,
                "empalmes": empalmes,
                "tramos_carga": tramos_carga,
                "tramos_retorno": tramos_retorno,
                "fotos_empalmes": fotos_empalmes,
                "fotos_carga": fotos_carga,
                "fotos_retorno": fotos_retorno,
                "paginas_fotos_empalmes": paginas_fotograficas_faja(
                    fotos_empalmes,
                    "EMPALMES",
                ),
                "paginas_fotos_carga": paginas_fotograficas_faja(
                    fotos_carga,
                    "CARGA",
                ),
                "paginas_fotos_retorno": paginas_fotograficas_faja(
                    fotos_retorno,
                    "RETORNO",
                ),
                "observacion_medicion_faja": observacion_medicion_faja,
                "calibraciones_ut": list(
                    inspeccion.calibraciones_ut_faja_cvb0003.order_by("numero")
                ),
                "rol": obtener_rol(request.user),
            },
        )

    # =====================================================
    # REPORTE ESPECIAL CVB0004
    # =====================================================

    if tag_faja in [
        "CVB0004",
        "0220-CVB-0004",
        "0220-CVB0004",
    ]:
        crear_estructura_faja_cvb0004(
            inspeccion
        )

        empalmes = list(
            inspeccion.empalmes_cvb0003
            .order_by(
                "orden",
                "id",
            )
        )

        tramos_carga = list(
            inspeccion.tramos_cvb0003
            .filter(
                tipo=MedicionTramoCVB0003.Tipo.CARGA,
            )
            .order_by(
                "orden",
                "id",
            )
        )

        fotos_empalmes = list(
            inspeccion.fotografias_cvb0003
            .filter(
                seccion=FotoFajaCVB0003.Seccion.EMPALMES,
            )
            .order_by(
                "creada_en",
                "id",
            )
        )

        fotos_carga = list(
            inspeccion.fotografias_cvb0003
            .filter(
                seccion=FotoFajaCVB0003.Seccion.CARGA,
            )
            .order_by(
                "creada_en",
                "id",
            )
        )

        print(
            "CVB0004 REPORTE:",
            f"EMPALMES={len(empalmes)}",
            f"TRAMOS={len(tramos_carga)}",
            f"FOTOS_EMP={len(fotos_empalmes)}",
            f"FOTOS_CARGA={len(fotos_carga)}",
        )

        return render(
            request,
            "inspecciones/reporte_faja_cvb0004.html",
            {
                "inspeccion": inspeccion,

                # Las dos tablas de CVB0004
                "empalmes": empalmes,
                "tramos_carga": tramos_carga,

                # Fotografías
                "fotos_empalmes": fotos_empalmes,
                "fotos_carga": fotos_carga,

                "rol": obtener_rol(request.user),
            },
        )

    # =====================================================
    # REPORTE CVB0001 Y OTRAS FAJAS
    # =====================================================

    mediciones = inspeccion.mediciones.order_by(
        "orden",
        "id",
    )

    empalme_e01 = list(
        mediciones.filter(seccion__iexact="EMPALME E-01")
    )

    empalme_e02 = list(
        mediciones.filter(seccion__iexact="EMPALME E-02")
    )

    tramos = list(
        mediciones.exclude(seccion__icontains="EMPALME")
    )

    resumen_e01 = analizar_mediciones_empalme(empalme_e01, "E-01")
    resumen_e02 = analizar_mediciones_empalme(empalme_e02, "E-02")

    resumen_tramos = calcular_resumen_tramos(
        tramos
    )

    if tag_faja in [
        "CVB0001",
        "0220-CVB-0001",
        "0220-CVB0001",
    ]:
        template_name = (
            "inspecciones/"
            "reporte_faja_cvb0001.html"
        )
    else:
        template_name = (
            "inspecciones/"
            "reporte_faja.html"
        )

    return render(
        request,
        template_name,
        {
            "inspeccion": inspeccion,
            "empalme_e01": empalme_e01,
            "empalme_e02": empalme_e02,
            "tramos": tramos,
            "resumen_tramos": resumen_tramos,
            "resumen_e01": resumen_e01,
            "resumen_e02": resumen_e02,
            "fotos_e01": list(
                inspeccion.fotografias.extra(
                    where=["seccion = %s"], params=["EMPALME_E01"]
                ).order_by("creada_en", "id")
            ),
            "fotos_e02": list(
                inspeccion.fotografias.extra(
                    where=["seccion = %s"], params=["EMPALME_E02"]
                ).order_by("creada_en", "id")
            ),
            "fotos_puntos": list(
                inspeccion.fotografias.extra(
                    where=["seccion = %s"], params=["PUNTOS_MEDICION"]
                ).order_by("creada_en", "id")
            ),
            "rol": obtener_rol(request.user),
        },
    )


@login_required
def exportar_excel_faja_cvb0001(request, inspeccion_id):
    inspeccion = get_object_or_404(
        Inspeccion.objects.select_related(
            "faja",
            "inspector",
            "supervisor",
            "analista",
            "cliente",
        ),
        id=inspeccion_id,
        tipo=Inspeccion.Tipo.FAJA,
    )
    if not usuario_puede_abrir_inspeccion(request.user, inspeccion):
        return HttpResponseForbidden(
            "No tienes permiso para descargar este reporte."
        )

    tag_faja = (inspeccion.faja.tag or "").upper().strip()
    if tag_faja not in {"CVB0001", "0220-CVB-0001", "0220-CVB0001"}:
        return HttpResponseForbidden(
            "La exportación solicitada sólo corresponde a la faja CVB0001."
        )

    crear_estructura_mediciones_faja(inspeccion)
    mediciones = inspeccion.mediciones.order_by("orden", "id")
    empalme_e01 = list(
        mediciones.filter(seccion__iexact="EMPALME E-01")
    )
    empalme_e02 = list(
        mediciones.filter(seccion__iexact="EMPALME E-02")
    )
    tramos = list(
        mediciones.exclude(seccion__icontains="EMPALME")
    )
    resumen_e01 = analizar_mediciones_empalme(empalme_e01, "E-01")
    resumen_e02 = analizar_mediciones_empalme(empalme_e02, "E-02")
    resumen_tramos = calcular_resumen_tramos(tramos)
    fotos_e01 = list(
        inspeccion.fotografias.extra(
            where=["seccion = %s"], params=["EMPALME_E01"]
        ).order_by("creada_en", "id")
    )
    fotos_e02 = list(
        inspeccion.fotografias.extra(
            where=["seccion = %s"], params=["EMPALME_E02"]
        ).order_by("creada_en", "id")
    )
    fotos_puntos = list(
        inspeccion.fotografias.extra(
            where=["seccion = %s"], params=["PUNTOS_MEDICION"]
        ).order_by("creada_en", "id")
    )

    output = generar_reporte_faja_cvb0001_excel(
        inspeccion=inspeccion,
        empalme_e01=empalme_e01,
        empalme_e02=empalme_e02,
        tramos=tramos,
        resumen_e01=resumen_e01,
        resumen_e02=resumen_e02,
        resumen_tramos=resumen_tramos,
        fotos_e01=fotos_e01,
        fotos_e02=fotos_e02,
        fotos_puntos=fotos_puntos,
    )
    codigo_seguro = re.sub(
        r"[^A-Za-z0-9._-]+",
        "_",
        inspeccion.codigo_reporte,
    ).strip("._") or str(inspeccion.id)
    response = HttpResponse(
        output.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = (
        f'attachment; filename="REPORTE_INSPECCION_{codigo_seguro}.xlsx"'
    )
    return response


def _nombre_usuario(usuario):
    if not usuario:
        return ""
    return usuario.get_full_name() or usuario.username


def _molienda_config(inspeccion):
    tag = re.sub(r"[^A-Z0-9]", "", (inspeccion.faja.tag or "").upper())
    if "CVB0010" in tag or "CVB010" in tag:
        variante = "Saliente" if "SALIENTE" in tag else "Entrante"
        return {
            "codigo": "CVB0010",
            "tag_largo": "0320-CVB-0010",
            "faja": f"Faja 10 {variante}",
            "static_key": "cvb0010",
            "diagrama_faja": "inspecciones/faja/cvb0010/diagrama_ut_cvb0010.png",
            "esquema_poleas": "inspecciones/faja/cvb0010/diagrama_ut_cvb0010.png",
            "poleas_count": 8,
            "poleas_ceramicas": {4},
        }
    if "CVB0007" in tag or "CVB007" in tag:
        return {
            "codigo": "CVB0007",
            "tag_largo": "0240-CVB-0007",
            "faja": "Faja 07",
            "static_key": "cvb0007",
            "diagrama_faja": "inspecciones/faja/cvb0007/diagrama_ut_cvb0007.png",
            "esquema_poleas": "inspecciones/poleas/cvb0007/esquema_poleas_cvb0007.png",
            "poleas_count": 8,
            "poleas_ceramicas": {4},
        }
    if "CVB0018" in tag:
        return {
            "codigo": "CVB0018",
            "tag_largo": "0320-CVB-0018",
            "faja": "Faja 18",
            "static_key": "cvb0018",
            "diagrama_faja": "inspecciones/faja/cvb0018/diagrama_ut_cvb0018.png",
            "esquema_poleas": "inspecciones/poleas/cvb0018/esquema_poleas_cvb0018.png",
            "poleas_count": 5,
            "poleas_ceramicas": {3, 4},
        }
    if "CVB0017" in tag:
        return {
            "codigo": "CVB0017",
            "tag_largo": "0320-CVB-0017",
            "faja": "Faja 17",
            "static_key": "cvb0017",
            "diagrama_faja": "inspecciones/faja/cvb0017/diagrama_ut_cvb0017.png",
            "esquema_poleas": "inspecciones/poleas/cvb0017/esquema_poleas_cvb0017.png",
            "poleas_count": 2,
            "poleas_ceramicas": set(),
        }
    if "CVB0015" in tag:
        return {
            "codigo": "CVB0015",
            "tag_largo": "0320-CVB-0015",
            "faja": "Faja 15",
            "static_key": "cvb0015",
            "diagrama_faja": "inspecciones/faja/cvb0015/diagrama_ut_cvb0015.png",
            "esquema_poleas": "inspecciones/poleas/cvb0015/esquema_poleas_cvb0015.png",
            "poleas_count": 5,
            "poleas_ceramicas": {1, 3, 5},
        }
    if "CVB0011" in tag:
        return {
            "codigo": "CVB0011",
            "tag_largo": "0320-CVB-0011",
            "faja": "Faja 11",
            "static_key": "cvb0011",
            "diagrama_faja": "inspecciones/faja/cvb0011/diagrama_ut_cvb0011.png",
            "esquema_poleas": "inspecciones/poleas/cvb0011/esquema_poleas_cvb0011.png",
            "poleas_count": 7,
            "poleas_ceramicas": set(),
        }
    return {
        "codigo": "CVB0006",
        "tag_largo": "0240-CVB-0006",
        "faja": "Faja 06",
        "static_key": "cvb0006",
        "diagrama_faja": "inspecciones/faja/cvb0006/diagrama_ut_cvb0006.png",
        "esquema_poleas": "inspecciones/poleas/cvb0006/esquema_poleas_cvb0006.png",
        "poleas_count": 8,
        "poleas_ceramicas": {4},
    }


def _molienda_master_path(inspeccion):
    masters = (
        settings.BASE_DIR
        / "inspecciones"
        / "reportes"
        / "molienda"
    )
    config = _molienda_config(inspeccion)
    codigo = config["codigo"].lower()
    masters = masters / codigo / "masters"
    if codigo == "cvb0007":
        if inspeccion.tipo == Inspeccion.Tipo.POLEAS:
            return masters / "20260705-VTUT-0310CVB0007-REPORTE POLEAS FAJA 7.xlsx"
        return masters / "20260630-VTUT-0310CVB0007-REPORTE DE TOP COVER FAJA 7.xlsx"
    if codigo == "cvb0010":
        if "SALIENTE" in (inspeccion.codigo_reporte or "").upper():
            return masters / "20260602-VTUT-0320CVB0010-REPORTE DE TOP COVER FAJA 10 SALIENTE.xlsx"
        return masters / "20260602-VTUT-0320CVB0010-REPORTE DE TOP COVER FAJA 10 ENTRANTE.xlsx"
    if codigo == "cvb0011":
        if inspeccion.tipo == Inspeccion.Tipo.POLEAS:
            return masters / "202603630-VTUT-0320CVB0011-REPORTE POLEAS FAJA 11.xlsx"
        return masters / "20260630-VTUT-0320CVB0011-REPORTE DE TOP COVER FAJA 11.xlsx"
    if codigo == "cvb0015":
        if inspeccion.tipo == Inspeccion.Tipo.POLEAS:
            return masters / "20260630-VTUT-0320CVB0015-REPORTE POLEAS FAJA 15.xlsx"
        return masters / "20260630-VTUT0310CVB0015-REPORTE TOP COVER FAJA 15.xlsx"
    if codigo == "cvb0017":
        if inspeccion.tipo == Inspeccion.Tipo.POLEAS:
            return masters / "20260630-VTUT-0320CVB0017-REPORTE POLEAS FAJA 17.xlsx"
        return masters / "20260630VTUT0310CVB0017-REPORTE DE TOP COVER FAJA 17.xlsx"
    if codigo == "cvb0018":
        if inspeccion.tipo == Inspeccion.Tipo.POLEAS:
            return masters / "20260630-VTUT-0320CVB00018-REPORTE POLEAS FAJA 18.xlsx"
        return masters / "20260630-VTUT-0320CVB0018-REPORTE DE TOP COVER FAJA 18.xlsx"
    if inspeccion.tipo == Inspeccion.Tipo.FAJA:
        return masters / "20260630-VTUT-0310CVB0006-REPORTE DE TOP COVER FAJA 6.xlsx"
    return masters / "20260605-VTUT-0310CVB0006-REPORTE POLEAS FAJA 6.xlsx"


def _response_xlsx(workbook, filename):
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _llenar_cabecera_molienda(ws, inspeccion):
    fecha_inspeccion = inspeccion.fecha_inspeccion or inspeccion.fecha_programada
    fecha_reporte = inspeccion.fecha_reporte or fecha_inspeccion
    config = _molienda_config(inspeccion)
    valores = {
        "J6": inspeccion.condicion_general or "",
        "J9": "MOLIENDA",
        "W9": "TRANSPORTE DE CONCENTRADO",
        "J11": config["faja"].upper(),
        "W11": config["tag_largo"],
        "J13": inspeccion.etapa or "Operaciones",
        "W13": inspeccion.condicion_equipo or "EN USO",
        "J15": fecha_inspeccion,
        "W15": fecha_reporte,
        "J17": _nombre_usuario(inspeccion.inspector),
        "W17": _nombre_usuario(inspeccion.supervisor),
        "J19": _nombre_usuario(inspeccion.analista),
        "W19": _nombre_usuario(inspeccion.analista),
        "J22": inspeccion.circunstancias or "",
        "J23": inspeccion.observaciones or "",
        "J26": inspeccion.recomendaciones or "",
    }
    for celda, valor in valores.items():
        _set_excel_value(ws, celda, valor)


def _set_excel_value(ws, cell_ref, value):
    target = ws[cell_ref]
    for merged_range in ws.merged_cells.ranges:
        if cell_ref in merged_range:
            target = ws.cell(
                row=merged_range.min_row,
                column=merged_range.min_col,
            )
            break
    target.value = value


def _valores_ag(fila):
    return [fila.a, fila.b, fila.c, fila.d, fila.e, fila.f, fila.g]


def _escribir_ag(ws, row, start_col, fila):
    for offset, valor in enumerate(_valores_ag(fila)):
        ws.cell(row=row, column=start_col + offset, value=valor)


def _filas_inicio_poleas(ws, poleas):
    """Locate the first data row of each polea table in the master."""
    numeros = {polea.numero for polea in poleas}
    filas = {}
    for row in ws.iter_rows():
        for cell in row:
            valor = str(cell.value or "").upper()
            match = re.search(r"LAGGING DE LA POLEA\s*([0-9]+)", valor)
            if match and int(match.group(1)) in numeros:
                filas.setdefault(int(match.group(1)), cell.row + 4)
    return filas


def _ruta_foto_modelo(foto):
    imagen = getattr(foto, "imagen", None)
    if not imagen:
        return None
    try:
        path = Path(imagen.path)
    except (NotImplementedError, ValueError):
        return None
    return path if path.exists() else None


def _limpiar_imagenes_desde_fila(ws, fila_minima):
    ws._images = [
        image for image in ws._images
        if image.anchor._from.row + 1 < fila_minima
    ]


def _insertar_foto_excel(ws, foto, celda, ancho=300, alto=220):
    ruta = _ruta_foto_modelo(foto)
    if not ruta:
        return
    imagen = ExcelImage(str(ruta))
    imagen.width = ancho
    imagen.height = alto
    ws.add_image(imagen, celda)


@login_required
def exportar_excel_molienda_cvb0006(request, inspeccion_id):
    inspeccion = get_object_or_404(
        Inspeccion.objects.select_related(
            "faja",
            "inspector",
            "supervisor",
            "analista",
            "cliente",
        ).prefetch_related(
            "empalmes_cvb0003",
            "tramos_cvb0003",
            "poleas_inspeccionadas__mediciones",
        ),
        id=inspeccion_id,
    )
    if not usuario_puede_abrir_inspeccion(request.user, inspeccion):
        return HttpResponseForbidden("No tienes permiso para descargar este reporte.")

    tag = (inspeccion.faja.tag or "").upper().strip()
    tags_molienda = {
        "CVB0006",
        "CVB006",
        "0240-CVB-006",
        "0240-CVB0006",
        "0310CVB0006",
        "CVB0007",
        "CVB007",
        "0240-CVB-007",
        "0240-CVB0007",
        "0310CVB0007",
        "CVB0010",
        "CVB010",
        "CVB0010-ENTRANTE",
        "CVB0010-SALIENTE",
        "0320-CVB-0010",
        "0320CVB0010",
        "CVB0011",
        "CVB0015",
        "CVB0017",
        "CVB0018",
    }
    if tag not in tags_molienda:
        return HttpResponseForbidden("Esta descarga solo corresponde a Molienda.")

    if inspeccion.tipo == Inspeccion.Tipo.FAJA:
        crear_estructura_faja_cvb0006_molienda(inspeccion)
        wb = load_workbook(_molienda_master_path(inspeccion))
        ws = wb.active
        _llenar_cabecera_molienda(ws, inspeccion)
        _limpiar_imagenes_desde_fila(ws, 74)

        e01 = list(inspeccion.empalmes_cvb0003.filter(zona="EMPALME E-01").order_by("orden"))
        e02 = list(inspeccion.empalmes_cvb0003.filter(zona="EMPALME E-02").order_by("orden"))
        top = list(inspeccion.tramos_cvb0003.filter(tipo=MedicionTramoCVB0003.Tipo.CARGA).order_by("orden"))

        for row, fila in zip((56, 57, 58), e01):
            _escribir_ag(ws, row, 7, fila)
        for row, fila in zip((63, 64, 65), e02):
            _escribir_ag(ws, row, 7, fila)
        for row, fila in zip(range(60, 67), top):
            _escribir_ag(ws, row, 31, fila)

        fotos_e01 = list(inspeccion.fotografias_cvb0003.filter(seccion="EMPALMES").order_by("creada_en", "id")[:5])
        fotos_top = list(inspeccion.fotografias_cvb0003.filter(seccion="CARGA").order_by("creada_en", "id")[:5])
        fotos_e02 = list(inspeccion.fotografias_cvb0003.filter(seccion="RETORNO").order_by("creada_en", "id")[:5])
        for foto, celda in zip(fotos_top, ["C74", "R74", "AG74", "C84", "R84"]):
            _insertar_foto_excel(ws, foto, celda)
        for foto, celda in zip(fotos_e01, ["F98", "R98", "AK98", "F108", "R108"]):
            _insertar_foto_excel(ws, foto, celda)
        for foto, celda in zip(fotos_e02, ["C124", "R124", "AG124", "C136", "R136"]):
            _insertar_foto_excel(ws, foto, celda)

        return _response_xlsx(wb, f"{inspeccion.codigo_reporte}_TOP_COVER.xlsx")

    crear_estructura_poleas(inspeccion)
    wb = load_workbook(_molienda_master_path(inspeccion))
    ws = wb.active
    _llenar_cabecera_molienda(ws, inspeccion)
    _limpiar_imagenes_desde_fila(ws, 96)
    poleas = list(
        inspeccion.poleas_inspeccionadas.prefetch_related("mediciones").order_by("numero")
    )
    start_rows = _filas_inicio_poleas(ws, poleas)
    for polea in poleas:
        row0 = start_rows.get(polea.numero)
        if not row0:
            continue
        for idx, medicion in enumerate(polea.mediciones.order_by("orden", "punto")[:3]):
            _escribir_ag(ws, row0 + idx, 9, medicion)
        foto = polea.fotografias.order_by("creada_en", "id").first()
        if foto:
            _insertar_foto_excel(ws, foto, f"AE{row0 + 9}", ancho=330, alto=220)

    return _response_xlsx(wb, f"{inspeccion.codigo_reporte}_POLEAS.xlsx")


@login_required
def exportar_pdf_molienda_cvb0006(request, inspeccion_id):
    inspeccion = get_object_or_404(
        Inspeccion.objects.select_related(
            "faja",
            "inspector",
            "supervisor",
            "analista",
            "cliente",
        ).prefetch_related(
            "empalmes_cvb0003",
            "tramos_cvb0003",
            "poleas_inspeccionadas__mediciones",
        ),
        id=inspeccion_id,
    )
    if not usuario_puede_abrir_inspeccion(request.user, inspeccion):
        return HttpResponseForbidden("No tienes permiso para descargar este reporte.")

    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=24,
        leftMargin=24,
        topMargin=24,
        bottomMargin=24,
    )
    styles = getSampleStyleSheet()
    config = _molienda_config(inspeccion)
    story = [
        Paragraph(f"REPORTE {inspeccion.codigo_reporte}", styles["Title"]),
        Paragraph(
            f"MOLIENDA - {config['tag_largo']} - {config['faja']}",
            styles["Heading2"],
        ),
        Paragraph(
            f"Inspector: {_nombre_usuario(inspeccion.inspector)} | "
            f"Supervisor: {_nombre_usuario(inspeccion.supervisor)} | "
            f"Analista: {_nombre_usuario(inspeccion.analista)}",
            styles["Normal"],
        ),
        Spacer(1, 10),
    ]
    if inspeccion.tipo == Inspeccion.Tipo.POLEAS:
        img_path = settings.BASE_DIR / "static" / config["esquema_poleas"]
        story.append(PdfImage(str(img_path), width=760, height=245))
        story.append(Spacer(1, 10))
        for polea in inspeccion.poleas_inspeccionadas.order_by("numero"):
            data = [["Punto", "A", "B", "C", "D", "E", "F", "G", "Prom", "Min"]]
            for m in polea.mediciones.order_by("orden", "punto")[:3]:
                data.append([m.punto, m.a, m.b, m.c, m.d, m.e, m.f, m.g, m.promedio, m.minimo])
            story.append(Paragraph(f"Polea {polea.numero}", styles["Heading3"]))
            story.append(Table(data, repeatRows=1))
            story.append(Spacer(1, 8))
    else:
        img_path = settings.BASE_DIR / "static" / config["diagrama_faja"]
        story.append(PdfImage(str(img_path), width=760, height=177))
        story.append(Spacer(1, 10))
        data = [["Seccion", "Punto", "A", "B", "C", "D", "E", "F", "G", "Prom", "Min"]]
        filas = list(inspeccion.empalmes_cvb0003.order_by("orden")) + list(
            inspeccion.tramos_cvb0003.filter(tipo=MedicionTramoCVB0003.Tipo.CARGA).order_by("orden")
        )
        for fila in filas:
            punto = getattr(fila, "posicion", None) or getattr(fila, "medicion", "")
            data.append([getattr(fila, "zona", "TOP COVER"), punto, fila.a, fila.b, fila.c, fila.d, fila.e, fila.f, fila.g, fila.promedio, fila.minimo])
        story.append(Table(data, repeatRows=1))

    for element in story:
        if isinstance(element, Table):
            element.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#087bbb")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.black),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
            ]))
    doc.build(story)
    output.seek(0)
    response = HttpResponse(output.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{inspeccion.codigo_reporte}.pdf"'
    return response


@login_required
def formulario_poleas(request, inspeccion_id):
    inspeccion = get_object_or_404(
        Inspeccion.objects.select_related(
            "faja",
            "inspector",
            "supervisor",
            "analista",
            "cliente",
        ),
        id=inspeccion_id,
    )

    if not usuario_puede_abrir_inspeccion(
        request.user,
        inspeccion,
    ):
        return HttpResponseForbidden(
            "No tienes permiso para abrir esta inspección."
        )

    if inspeccion.tipo != Inspeccion.Tipo.POLEAS:
        messages.error(
            request,
            "Esta inspección no corresponde al formulario de Poleas.",
        )
        return redirect("dashboard")

    es_cvb0003_cliente = (inspeccion.faja.tag or "").upper().strip() in {
        "CVB0003", "CVB003", "0220-CVB-0003", "0220-CVB0003", "0220-CVB-003",
    }
    if es_cvb0003_cliente and obtener_rol(request.user) == "Cliente":
        if inspeccion.estado == Inspeccion.Estado.PUBLICADO:
            return redirect("reporte_poleas", inspeccion_id=inspeccion.id)
        return HttpResponseForbidden("Este reporte todavÃ­a no ha sido publicado.")

    if (inspeccion.faja.tag or "").upper().strip() in {
        "CVB0001", "CVB001", "0220-CVB-0001", "0220-CVB0001", "0220-CVB-001",
    }:
        from .reportes.cvb0001.poleas_views import formulario_poleas_cvb0001

        return formulario_poleas_cvb0001(request, inspeccion)

    if (inspeccion.faja.tag or "").upper().strip() in {
        "CVB0004", "CVB004", "0220-CVB-0004", "0220-CVB0004", "0220-CVB-004",
    }:
        from .reportes.cvb0004.poleas_views import formulario_poleas_cvb0004

        return formulario_poleas_cvb0004(request, inspeccion)

    crear_estructura_poleas(inspeccion)

    poleas = list(
        inspeccion.poleas_inspeccionadas
        .prefetch_related(
            "mediciones",
            "mediciones_campana",
            "fotografias",
        )
        .order_by(
            "orden",
            "numero",
        )
    )

    tag_faja = (
        inspeccion.faja.tag
        or ""
    ).upper().strip()
    es_cvb0003 = tag_faja in {
        "CVB0003", "CVB003", "0220-CVB-0003", "0220-CVB0003", "0220-CVB-003",
    }

    # CVB0001: Polea 1 cerámica.
    # CVB0003: Poleas 4 y 5 cerámicas.
    if tag_faja in [
        "CVB0001",
        "0220-CVB-0001",
        "0220-CVB0001",
    ]:
        numeros_poleas_ceramicas = {1}
    elif tag_faja in [
        "CVB0003",
        "0220-CVB-0003",
        "0220-CVB0003",
    ]:
        numeros_poleas_ceramicas = {4, 5}
    elif tag_faja in [
        "CVB0006",
        "CVB006",
        "0240-CVB-006",
        "0240-CVB0006",
        "0310CVB0006",
        "CVB0007",
        "CVB007",
        "0240-CVB-007",
        "0240-CVB0007",
        "0310CVB0007",
        "CVB0010",
        "CVB010",
        "CVB0010-ENTRANTE",
        "CVB0010-SALIENTE",
        "0320-CVB-0010",
        "0320CVB0010",
    ]:
        numeros_poleas_ceramicas = _molienda_config(inspeccion)[
            "poleas_ceramicas"
        ]
    elif tag_faja in {"CVB0011", "CVB0015", "CVB0017", "CVB0018"}:
        numeros_poleas_ceramicas = _molienda_config(inspeccion)[
            "poleas_ceramicas"
        ]
    else:
        numeros_poleas_ceramicas = set()

    permisos = obtener_permisos_flujo(
        request.user,
        inspeccion,
    )
    puede_editar = permisos["puede_editar"]

    datos_inspeccion = None

    if request.method == "POST":
        datos_inspeccion = request.POST.copy()
        campo_condicion = "inspeccion-condicion_equipo"

        if not datos_inspeccion.get(campo_condicion):
            datos_inspeccion[campo_condicion] = (
                inspeccion.condicion_equipo
                or "En operación"
            )

    formulario = InspeccionForm(
        datos_inspeccion,
        instance=inspeccion,
        prefix="inspeccion",
    )

    bloques_poleas = []
    todo_valido = False

    if request.method == "POST":
        todo_valido = formulario.is_valid()

        print("=" * 75)
        print("VALIDACIÓN FORMULARIO GENERAL DE POLEAS")
        print("FORMULARIO GENERAL:", todo_valido)
        print("ERRORES GENERALES:", formulario.errors)
        print("=" * 75)

    fecha_tecnica = (
        formulario.cleaned_data.get("fecha_inspeccion")
        if request.method == "POST" and formulario.is_valid()
        else inspeccion.fecha_inspeccion
    )

    for polea in poleas:
        prefijo_polea = f"polea-{polea.id}"
        prefijo_mediciones = f"mediciones-{polea.id}"
        prefijo_fotografias = f"fotografias-{polea.id}"

        formulario_polea = PoleaInspeccionForm(
            request.POST or None,
            instance=polea,
            prefix=prefijo_polea,
        )
        if es_cvb0003:
            formulario_polea.fields["tipo_medicion"].choices = (
                ("NORMAL", "NORMAL"),
                ("CAMPANA", "INICIO DE CAMPAÑA / FIN DE CAMPAÑA"),
            )

        mediciones_formset = MedicionPoleaFormSet(
            request.POST or None,
            instance=polea,
            prefix=prefijo_mediciones,
        )
        campana_formsets = formsets_polea_campana(request, polea)

        historial = historial_componente_visible(
            inspeccion, polea, fecha_tecnica, "poleas_inspeccionadas"
        )
        valores_historicos = historial["valores"] if historial else {}
        preparar_formset_historico(mediciones_formset, valores_historicos)
        preparar_formset_historico(
            campana_formsets["FIN"], valores_historicos
        )

        fotos_formset = FotoPoleaFormSet(
            request.POST or None,
            request.FILES or None,
            instance=polea,
            prefix=prefijo_fotografias,
        )

        if request.method == "POST":
            formulario_polea_valido = formulario_polea.is_valid()
            modo_campana_actual = modo_campana_seleccionado(
                formulario_polea, polea
            )

            if modo_campana_actual:
                inicio_valido = campana_formsets["INICIO"].is_valid()
                fin_valido = validar_formset_historico(
                    campana_formsets["FIN"], valores_historicos
                )
                mediciones_validas = inicio_valido and fin_valido
            elif polea.numero in numeros_poleas_ceramicas:
                mediciones_validas = True
            else:
                mediciones_validas = validar_formset_historico(
                    mediciones_formset, valores_historicos
                )

            fotografias_validas = fotos_formset.is_valid()

            print("=" * 75)
            print(f"VALIDACIÓN POLEA {polea.numero}")
            print("FORMULARIO:", formulario_polea_valido)
            print("ERRORES FORMULARIO:", formulario_polea.errors)
            print("MEDICIONES:", mediciones_validas)

            if polea.numero in numeros_poleas_ceramicas:
                print("MEDICIONES OMITIDAS: POLEA CERÁMICA")
            else:
                print("ERRORES MEDICIONES:", mediciones_formset.errors)
                print(
                    "ERRORES GENERALES MEDICIONES:",
                    mediciones_formset.non_form_errors(),
                )

            print("FOTOGRAFÍAS:", fotografias_validas)
            print("ERRORES FOTOGRAFÍAS:", fotos_formset.errors)
            print(
                "ERRORES GENERALES FOTOGRAFÍAS:",
                fotos_formset.non_form_errors(),
            )

            todo_valido = (
                todo_valido
                and formulario_polea_valido
                and mediciones_validas
                and fotografias_validas
            )

        bloques_poleas.append(
            {
                "polea": polea,
                "formulario_polea": formulario_polea,
                "mediciones_formset": mediciones_formset,
                "campana_inicio_formset": campana_formsets["INICIO"],
                "campana_fin_formset": campana_formsets["FIN"],
                "campana_formsets": (("INICIO", campana_formsets["INICIO"]), ("FIN", campana_formsets["FIN"])),
                "es_campana": modo_campana_seleccionado(
                    formulario_polea, polea
                ),
                "historial": historial,
                "fotos_formset": fotos_formset,
                "fotografias": list(
                    polea.fotografias.order_by(
                        "creada_en",
                        "id",
                    )
                ),
            }
        )

    if request.method == "POST":
        if not puede_editar:
            return HttpResponseForbidden(
                "La inspección no está habilitada para edición."
            )

        if todo_valido:
            # El historial y todos los formsets ya fueron consultados y
            # validados. SQLite queda bloqueado únicamente durante las
            # escrituras que deben confirmarse como una unidad.
            with transaction.atomic():
                inspeccion_guardada = formulario.save(commit=False)
                if es_cvb0003:
                    sincronizar_codigo_cvb0003(inspeccion_guardada)
                    django_models.Model.save(inspeccion_guardada)
                    formulario.save_m2m()
                else:
                    inspeccion_guardada.save()

                for bloque in bloques_poleas:
                    formulario_polea = bloque["formulario_polea"]
                    polea_guardada = formulario_polea.save(commit=False)
                    polea_guardada.inspeccion = inspeccion_guardada
                    polea_guardada.save()

                    if es_campana(polea_guardada):
                        guardar_formsets_campana(
                            polea_guardada,
                            {
                                "INICIO": bloque["campana_inicio_formset"],
                                "FIN": bloque["campana_fin_formset"],
                            },
                            "polea",
                        )
                    elif (
                        polea_guardada.numero
                        not in numeros_poleas_ceramicas
                    ):
                        mediciones_formset = bloque[
                            "mediciones_formset"
                        ]

                        mediciones = mediciones_formset.save(
                            commit=False
                        )

                        for medicion in mediciones:
                            medicion.polea = polea_guardada
                            medicion.save()

                        for medicion_eliminada in (
                            mediciones_formset.deleted_objects
                        ):
                            medicion_eliminada.delete()

                        mediciones_formset.save_m2m()

                    fotos_formset = bloque["fotos_formset"]
                    fotografias = fotos_formset.save(commit=False)

                    for fotografia in fotografias:
                        fotografia.polea = polea_guardada
                        fotografia.codigo_dano = ""

                        if not fotografia.subida_por_id:
                            fotografia.subida_por = request.user

                        fotografia.save()

                    for fotografia_eliminada in (
                        fotos_formset.deleted_objects
                    ):
                        fotografia_eliminada.delete()

                    fotos_formset.save_m2m()

                correcto, mensaje = _procesar_accion_flujo(
                    request,
                    inspeccion_guardada,
                )

            if correcto:
                messages.success(request, mensaje)
            else:
                messages.error(request, mensaje)

            return redirect(
                "formulario_poleas",
                inspeccion_id=inspeccion_guardada.id,
            )

        print("=" * 75)
        print("NO SE GUARDÓ EL FORMULARIO DE POLEAS")
        print("TODO VÁLIDO:", todo_valido)
        print(
            "ERRORES DEL FORMULARIO GENERAL:",
            formulario.errors,
        )

        for bloque in bloques_poleas:
            numero_polea = bloque["polea"].numero

            print("-" * 75)
            print(f"ERRORES POLEA {numero_polea}")
            print(
                "FORMULARIO:",
                bloque["formulario_polea"].errors,
            )

            if numero_polea in numeros_poleas_ceramicas:
                print("MEDICIONES OMITIDAS: POLEA CERÁMICA")
            else:
                print(
                    "MEDICIONES:",
                    bloque["mediciones_formset"].errors,
                )
                print(
                    "MEDICIONES GENERALES:",
                    bloque[
                        "mediciones_formset"
                    ].non_form_errors(),
                )

            print(
                "FOTOGRAFÍAS:",
                bloque["fotos_formset"].errors,
            )
            print(
                "FOTOGRAFÍAS GENERALES:",
                bloque["fotos_formset"].non_form_errors(),
            )

        messages.error(
            request,
            (
                "No se pudo guardar Poleas. "
                "Revisa los campos marcados y "
                "los errores mostrados en la terminal."
            ),
        )

    if tag_faja in [
        "CVB0003",
        "0220-CVB-0003",
        "0220-CVB0003",
    ]:
        template_name = (
            "inspecciones/formulario_poleas_cvb0003.html"
        )
    elif tag_faja in [
        "CVB0006",
        "CVB006",
        "0240-CVB-006",
        "0240-CVB0006",
        "0310CVB0006",
        "CVB0007",
        "CVB007",
        "0240-CVB-007",
        "0240-CVB0007",
        "0310CVB0007",
        "CVB0010",
        "CVB010",
        "CVB0010-ENTRANTE",
        "CVB0010-SALIENTE",
        "0320-CVB-0010",
        "0320CVB0010",
        "CVB0011",
        "CVB0015",
        "CVB0017",
        "CVB0018",
    ]:
        template_name = (
            "inspecciones/formulario_poleas_cvb0006_molienda.html"
        )
    else:
        template_name = (
            "inspecciones/formulario_poleas.html"
        )

    return render(
        request,
        template_name,
        {
            "inspeccion": inspeccion,
            "formulario": formulario,
            "bloques_poleas": bloques_poleas,
            "historiales_cvb0003": [
                bloque["historial"]
                for bloque in bloques_poleas
                if bloque["historial"]
            ],
            "numeros_poleas_ceramicas": (
                numeros_poleas_ceramicas
            ),
            "molienda_config": _molienda_config(inspeccion),
            **permisos,
            **_contexto_workflow_ui_cvb0003(inspeccion),
        },
    )

@login_required
def reporte_poleas(request, inspeccion_id):
    inspeccion = get_object_or_404(
        Inspeccion.objects.select_related(
            "faja",
            "inspector",
            "supervisor",
            "analista",
            "cliente",
        ).prefetch_related(
            "poleas_inspeccionadas__mediciones",
            "poleas_inspeccionadas__mediciones_campana",
            "poleas_inspeccionadas__fotografias",
        ),
        id=inspeccion_id,
        tipo=Inspeccion.Tipo.POLEAS,
    )

    if not usuario_puede_abrir_inspeccion(
        request.user,
        inspeccion,
    ):
        return HttpResponseForbidden(
            "No tienes permiso para ver este reporte."
        )

    if (inspeccion.faja.tag or "").upper().strip() in {
        "CVB0006",
        "CVB006",
        "0240-CVB-006",
        "0240-CVB0006",
        "0310CVB0006",
        "CVB0007",
        "CVB007",
        "0240-CVB-007",
        "0240-CVB0007",
        "0310CVB0007",
        "CVB0010",
        "CVB010",
        "CVB0010-ENTRANTE",
        "CVB0010-SALIENTE",
        "0320-CVB-0010",
        "0320CVB0010",
        "CVB0011",
        "CVB0015",
        "CVB0017",
        "CVB0018",
    }:
        return redirect(
            "formulario_poleas",
            inspeccion_id=inspeccion.id,
        )

    if (inspeccion.faja.tag or "").upper().strip() in {
        "CVB0001", "CVB001", "0220-CVB-0001", "0220-CVB0001", "0220-CVB-001",
    }:
        from .reportes.cvb0001.poleas_views import reporte_poleas_cvb0001

        return reporte_poleas_cvb0001(request, inspeccion)

    if (inspeccion.faja.tag or "").upper().strip() in {
        "CVB0004", "CVB004", "0220-CVB-0004", "0220-CVB0004", "0220-CVB-004",
    }:
        from .reportes.cvb0004.poleas_views import reporte_poleas_cvb0004

        return reporte_poleas_cvb0004(request, inspeccion)

    if (inspeccion.faja.tag or "").upper().strip() in {
        "CVB0003",
        "CVB003",
        "0220-CVB-0003",
        "0220-CVB0003",
        "0220-CVB-003",
    }:
        from .reportes.cvb0003.poleas_views import reporte_poleas_cvb0003

        return reporte_poleas_cvb0003(request, inspeccion)

    bloques = [
        {
            "polea": polea,
            "mediciones": list(
                polea.mediciones.order_by(
                    "orden",
                    "punto",
                )
            ),
            "fotografias": list(
                polea.fotografias.all()
            ),
        }
        for polea in inspeccion.poleas_inspeccionadas.order_by(
            "orden",
            "numero",
        )
    ]

    return render(
        request,
        "inspecciones/reporte_poleas.html",
        {
            "inspeccion": inspeccion,
            "bloques_poleas": bloques,
            "rol": obtener_rol(request.user),
        },
    )


@login_required
@transaction.atomic
def formulario_life_shaft(request, inspeccion_id):
    inspeccion = get_object_or_404(
        Inspeccion.objects.select_related(
            "faja",
            "inspector",
            "supervisor",
            "analista",
            "cliente",
        ),
        id=inspeccion_id,
    )

    if not usuario_puede_abrir_inspeccion(
        request.user,
        inspeccion,
    ):
        return HttpResponseForbidden(
            "No tienes permiso para abrir esta inspección."
        )

    if inspeccion.tipo != Inspeccion.Tipo.LIFE_SHAFT:
        messages.error(
            request,
            "Esta inspección no corresponde al formulario de Life Shaft.",
        )
        return redirect("dashboard")

    es_cvb0003_cliente = (inspeccion.faja.tag or "").upper().strip() in {
        "CVB0003", "CVB003", "0220-CVB-0003", "0220-CVB0003", "0220-CVB-003",
    }
    if es_cvb0003_cliente and obtener_rol(request.user) == "Cliente":
        if inspeccion.estado == Inspeccion.Estado.PUBLICADO:
            return redirect("reporte_life_shaft", inspeccion_id=inspeccion.id)
        return HttpResponseForbidden("Este reporte todavÃ­a no ha sido publicado.")

    if (inspeccion.faja.tag or "").upper().strip() in {
        "CVB0001", "CVB001", "0220-CVB-0001", "0220-CVB0001", "0220-CVB-001",
    }:
        from .reportes.cvb0001.life_shaft_views import formulario_life_shaft_cvb0001

        return formulario_life_shaft_cvb0001(request, inspeccion)

    if (inspeccion.faja.tag or "").upper().strip() in {
        "CVB0004", "CVB004", "0220-CVB-0004", "0220-CVB0004", "0220-CVB-004",
    }:
        from .reportes.cvb0004.life_shaft_views import formulario_life_shaft_cvb0004

        return formulario_life_shaft_cvb0004(request, inspeccion)

    crear_estructura_life_shaft(inspeccion)

    shafts = list(
        inspeccion.life_shafts
        .prefetch_related(
            "mediciones",
            "mediciones_campana",
            "fotografias",
        )
        .order_by(
            "orden",
            "numero",
        )
    )

    tag_faja = (inspeccion.faja.tag or "").upper().strip()
    es_cvb0003 = tag_faja in {
        "CVB0003", "CVB003", "0220-CVB-0003", "0220-CVB0003", "0220-CVB-003",
    }

    permisos = obtener_permisos_flujo(
        request.user,
        inspeccion,
    )
    puede_editar = permisos["puede_editar"]

    datos_inspeccion = None

    if request.method == "POST":
        datos_inspeccion = request.POST.copy()
        campo_condicion = "inspeccion-condicion_equipo"

        if not datos_inspeccion.get(campo_condicion):
            datos_inspeccion[campo_condicion] = (
                inspeccion.condicion_equipo
                or "En operación"
            )

    formulario = InspeccionForm(
        datos_inspeccion,
        instance=inspeccion,
        prefix="inspeccion",
    )

    bloques = []
    todo_valido = (
        formulario.is_valid()
        if request.method == "POST"
        else False
    )
    fecha_tecnica = (
        formulario.cleaned_data.get("fecha_inspeccion")
        if request.method == "POST" and formulario.is_valid()
        else inspeccion.fecha_inspeccion
    )

    for shaft in shafts:
        prefijo_shaft = f"shaft-{shaft.id}"
        prefijo_mediciones = f"mediciones-{shaft.id}"

        formulario_shaft = LifeShaftInspeccionForm(
            request.POST or None,
            instance=shaft,
            prefix=prefijo_shaft,
        )
        if es_cvb0003:
            formulario_shaft.fields["tipo_medicion"].choices = (
                ("NORMAL", "NORMAL"),
                ("CAMPANA", "INICIO DE CAMPAÑA / FIN DE CAMPAÑA"),
            )

        mediciones_formset = MedicionLifeShaftFormSet(
            request.POST or None,
            instance=shaft,
            prefix=prefijo_mediciones,
        )
        campana_formsets = formsets_life_shaft_campana(request, shaft)

        historial = historial_componente_visible(
            inspeccion, shaft, fecha_tecnica, "life_shafts"
        )
        valores_historicos = historial["valores"] if historial else {}
        preparar_formset_historico(mediciones_formset, valores_historicos)
        preparar_formset_historico(
            campana_formsets["FIN"], valores_historicos
        )

        if request.method == "POST":
            shaft_valido = formulario_shaft.is_valid()
            modo_campana_actual = modo_campana_seleccionado(
                formulario_shaft, shaft
            )
            if modo_campana_actual:
                inicio_valido = campana_formsets["INICIO"].is_valid()
                fin_valido = validar_formset_historico(
                    campana_formsets["FIN"], valores_historicos
                )
                mediciones_validas = inicio_valido and fin_valido
            elif shaft_valido:
                mediciones_validas = validar_formset_historico(
                    mediciones_formset, valores_historicos
                )
            else:
                mediciones_validas = False
            todo_valido = todo_valido and shaft_valido and mediciones_validas

        bloques.append(
            {
                "life_shaft": shaft,
                "formulario_shaft": formulario_shaft,
                "mediciones_formset": mediciones_formset,
                "campana_inicio_formset": campana_formsets["INICIO"],
                "campana_fin_formset": campana_formsets["FIN"],
                "campana_formsets": (("INICIO", campana_formsets["INICIO"]), ("FIN", campana_formsets["FIN"])),
                "es_campana": modo_campana_seleccionado(
                    formulario_shaft, shaft
                ),
                "historial": historial,
                "fotografias": list(
                    shaft.fotografias.order_by("id")
                ),
            }
        )

    if request.method == "POST":
        if not puede_editar:
            return HttpResponseForbidden(
                "La inspección no está habilitada para edición."
            )

        if todo_valido:
            inspeccion_guardada = formulario.save()

            for bloque in bloques:
                shaft_guardado = bloque[
                    "formulario_shaft"
                ].save(commit=False)

                shaft_guardado.inspeccion = inspeccion_guardada
                shaft_guardado.save()

                if es_campana(shaft_guardado):
                    guardar_formsets_campana(
                        shaft_guardado,
                        {
                            "INICIO": bloque["campana_inicio_formset"],
                            "FIN": bloque["campana_fin_formset"],
                        },
                        "life_shaft",
                    )
                else:
                    mediciones_formset = bloque["mediciones_formset"]
                    mediciones = mediciones_formset.save(commit=False)
                    for medicion in mediciones:
                        medicion.life_shaft = shaft_guardado
                        medicion.save()

                # Se permiten hasta diez fotografías nuevas
                # por cada Life Shaft en un solo guardado.
                for indice in range(1, 11):
                    archivo = request.FILES.get(
                        f"foto-{shaft_guardado.id}-{indice}"
                    )

                    if not archivo:
                        continue

                    descripcion = request.POST.get(
                        f"descripcion-foto-{shaft_guardado.id}-{indice}",
                        "",
                    ).strip()

                    FotoLifeShaft.objects.create(
                        life_shaft=shaft_guardado,
                        imagen=archivo,
                        codigo_dano="",
                        descripcion=descripcion,
                        subida_por=request.user,
                    )

            fotografias_a_eliminar = request.POST.getlist(
                "eliminar_fotografias"
            )

            if fotografias_a_eliminar:
                FotoLifeShaft.objects.filter(
                    id__in=fotografias_a_eliminar,
                    life_shaft__inspeccion=inspeccion_guardada,
                ).delete()

            correcto, mensaje = _procesar_accion_flujo(
                request,
                inspeccion_guardada,
            )

            if correcto:
                messages.success(request, mensaje)
            else:
                messages.error(request, mensaje)

            return redirect(
                "formulario_life_shaft",
                inspeccion_id=inspeccion_guardada.id,
            )

        print(
            "ERRORES DEL FORMULARIO GENERAL:",
            formulario.errors,
        )

        for bloque in bloques:
            print(
                "LIFE SHAFT:",
                bloque["life_shaft"].numero,
            )
            print(
                "FORMULARIO LIFE SHAFT:",
                bloque["formulario_shaft"].errors,
            )
            print(
                "MEDICIONES:",
                bloque["mediciones_formset"].errors,
            )
            print(
                "ERRORES GENERALES DE MEDICIONES:",
                bloque[
                    "mediciones_formset"
                ].non_form_errors(),
            )

        messages.error(
            request,
            (
                "No se pudo guardar Life Shaft. "
                "Revisa los campos marcados y la terminal."
            ),
        )

    tag_faja = (
        inspeccion.faja.tag
        or ""
    ).upper().strip()

    if tag_faja in [
        "CVB0003",
        "0220-CVB-0003",
        "0220-CVB0003",
    ]:
        template_name = (
            "inspecciones/"
            "formulario_life_shaft_cvb0003.html"
        )
    else:
        template_name = (
            "inspecciones/"
            "formulario_life_shaft.html"
        )

    return render(
        request,
        template_name,
        {
            "inspeccion": inspeccion,
            "formulario": formulario,
            "bloques_shafts": bloques,
            "historiales_cvb0003": [
                bloque["historial"]
                for bloque in bloques
                if bloque["historial"]
            ],
            "rangos_fotos": range(1, 6),
            **permisos,
            **_contexto_workflow_ui_cvb0003(inspeccion),
        },
    )


@login_required
def reporte_life_shaft(request, inspeccion_id):
    inspeccion = get_object_or_404(
        Inspeccion.objects.select_related(
            "faja",
            "inspector",
            "supervisor",
            "analista",
            "cliente",
        ).prefetch_related(
            "life_shafts__mediciones",
            "life_shafts__mediciones_campana",
            "life_shafts__fotografias",
        ),
        id=inspeccion_id,
        tipo=Inspeccion.Tipo.LIFE_SHAFT,
    )

    if not usuario_puede_abrir_inspeccion(
        request.user,
        inspeccion,
    ):
        return HttpResponseForbidden(
            "No tienes permiso para ver este reporte."
        )

    if (inspeccion.faja.tag or "").upper().strip() in {
        "CVB0006",
        "CVB006",
        "0240-CVB-006",
        "0240-CVB0006",
        "0310CVB0006",
        "CVB0007",
        "CVB007",
        "0240-CVB-007",
        "0240-CVB0007",
        "0310CVB0007",
    }:
        return redirect(
            "formulario_poleas",
            inspeccion_id=inspeccion.id,
        )

    if (inspeccion.faja.tag or "").upper().strip() in {
        "CVB0001", "CVB001", "0220-CVB-0001", "0220-CVB0001", "0220-CVB-001",
    }:
        from .reportes.cvb0001.life_shaft_views import reporte_life_shaft_cvb0001

        return reporte_life_shaft_cvb0001(request, inspeccion)

    if (inspeccion.faja.tag or "").upper().strip() in {
        "CVB0004", "CVB004", "0220-CVB-0004", "0220-CVB0004", "0220-CVB-004",
    }:
        from .reportes.cvb0004.life_shaft_views import reporte_life_shaft_cvb0004

        return reporte_life_shaft_cvb0004(request, inspeccion)

    bloques = [
        agregar_mediciones_campana_bloque({
            "life_shaft": shaft,
            "mediciones": list(
                shaft.mediciones.order_by(
                    "orden",
                    "punto",
                )
            ),
            "fotografias": list(
                shaft.fotografias.order_by("id")
            ),
        }, shaft)
        for shaft in inspeccion.life_shafts.order_by(
            "orden",
            "numero",
        )
    ]

    tag_faja = (
        inspeccion.faja.tag
        or ""
    ).upper().strip()

    if tag_faja in [
        "CVB0003",
        "0220-CVB-0003",
        "0220-CVB0003",
    ]:
        template_name = (
            "inspecciones/"
            "reporte_life_shaft_cvb0003.html"
        )
    else:
        template_name = (
            "inspecciones/"
            "reporte_life_shaft.html"
        )

    return render(
        request,
        template_name,
        {
            "inspeccion": inspeccion,
            "bloques_shafts": bloques,
            "rol": obtener_rol(request.user),
        },
    )


@login_required
@require_POST
@transaction.atomic
def cambiar_estado_inspeccion(
    request,
    inspeccion_id,
    accion,
):
    inspeccion = get_object_or_404(
        Inspeccion.objects.select_related(
            "faja",
            "inspector",
            "supervisor",
            "analista",
            "cliente",
        ),
        id=inspeccion_id,
    )

    if not usuario_puede_abrir_inspeccion(request.user, inspeccion):
        return HttpResponseForbidden(
            "No tienes permiso para cambiar esta inspección."
        )

    comentario = request.POST.get(
        "comentario",
        "",
    ).strip()

    mapa_acciones = {
        "enviar_revision": "enviar_supervisor",
        "enviar_supervisor": "enviar_supervisor",
        "aprobar_supervisor": "aprobar_supervisor",
        "devolver_supervisor": "devolver_supervisor",
        "aprobar_analista": "aprobar_analista",
        "devolver_analista": "devolver_analista",
        "publicar": "publicar",
    }

    accion_interna = mapa_acciones.get(accion)

    if not accion_interna:
        messages.error(
            request,
            "La acción solicitada no existe.",
        )

        return redirect(
            _nombre_ruta_reporte(inspeccion),
            inspeccion_id=inspeccion.id,
        )

    if es_inspeccion_cvb0003(inspeccion):
        acciones_permiso = {
            "enviar_supervisor": "enviar_revision",
            "aprobar_supervisor": "aprobar_supervisor",
            "devolver_supervisor": "devolver_supervisor",
            "aprobar_analista": "aprobar_analista",
            "devolver_analista": "devolver_analista",
            "publicar": "publicar",
        }
        if not puede_acceder_inspeccion_cvb0003(
            request.user,
            inspeccion,
            acciones_permiso[accion_interna],
        ):
            return HttpResponseForbidden(
                "No tienes permiso para ejecutar esta acción."
            )

    correcto, mensaje = _aplicar_accion_flujo(
        request,
        inspeccion,
        accion_interna,
        comentario,
    )

    if correcto:
        messages.success(request, mensaje)
    else:
        messages.error(request, mensaje)

    return redirect(
        _nombre_ruta_reporte(inspeccion),
        inspeccion_id=inspeccion.id,
    )
